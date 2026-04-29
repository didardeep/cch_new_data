"""
Telecom Customer Complaint Handling System - Backend
=====================================================
Full backend with auth, chat, tickets, and the original AI chatbot integrated.
"""

import os
import re
import json
import time
import math
import random
import string
import hashlib
from datetime import date, datetime, timezone, timedelta

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from flask_jwt_extended import (
    JWTManager, create_access_token, jwt_required, get_jwt_identity
)
from flask_mail import Mail, Message
from flask_socketio import SocketIO, join_room, emit
from openai import AzureOpenAI, OpenAI
from dotenv import load_dotenv
from types import SimpleNamespace
import urllib.request
import urllib.parse
import urllib.error

import threading
from sqlalchemy import case as sql_case, text
from sqlalchemy.orm import joinedload
from models import db, bcrypt, User, ChatSession, ChatMessage, Ticket, Feedback, SystemSetting, SlaAlert, TelecomSite, KpiData, ParameterChange, ChangeRequest, FlexibleKpiUpload, CoreComponentKpi
from auth_utils import token_required, role_required
# Add this import after other imports
from whatsapp_integration import send_whatsapp_message, format_chat_summary_for_whatsapp, format_ticket_alert_for_whatsapp
import network_prompts
import broadband_prompts
import network_diagnosis
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

# ─── App Setup ────────────────────────────────────────────────────────────────
def _get_jwt_secret():
    """Return a deterministic 32+ byte JWT signing key.

    Order of precedence:
      1. JWT_SECRET env var, used directly when >= 32 chars.
      2. JWT_SECRET env var, upgraded via SHA-256 to a 64-char hex key when
         shorter (avoids PyJWT's InsecureKeyLengthWarning on Python 3.14
         and the 422 token-rejection that follows).
      3. Deterministic fallback derived from the legacy default string so
         tokens stay valid across restarts when no JWT_SECRET is configured
         (a random os.urandom would invalidate every token on each restart).
    """
    raw = os.environ.get("JWT_SECRET")
    if raw:
        if len(raw) >= 32:
            return raw
        print("⚠️ JWT_SECRET is shorter than 32 bytes; deriving a stronger key via SHA-256.")
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()
    # No JWT_SECRET set — derive a deterministic 64-char key so tokens
    # survive server restarts. Operators should still set JWT_SECRET in .env.
    print("⚠️ JWT_SECRET not set in environment; using a deterministic dev fallback. Set JWT_SECRET in .env for production.")
    return hashlib.sha256(b"telecom-cch-default-jwt-secret-change-me").hexdigest()

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", os.urandom(24).hex())
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:ROOT@localhost:5432/telecom_cch"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
# Raise DB connection pool limits — default SQLAlchemy 5+10 is far too small
# once the background schedulers, daily jobs, and dashboard reads all run
# concurrently. Also recycle stale connections and pre-ping so the pool
# never hands out a dead socket.
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_size": 30,
    "max_overflow": 60,
    "pool_timeout": 30,
    "pool_recycle": 1800,
    "pool_pre_ping": True,
}
# Use _get_jwt_secret() so short JWT_SECRET values from .env are upgraded to a
# SHA-256 derived 64-char key. PyJWT on Python 3.14 emits InsecureKeyLength
# warnings for HMAC keys < 32 bytes, and verification can become flaky; the
# derivation gives a deterministic strong key from whatever the user set.
app.config["JWT_SECRET_KEY"] = _get_jwt_secret()
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = timedelta(hours=24)
# JWT error handler messages are JSON — see JWTManager handlers below.
app.config["JWT_ERROR_MESSAGE_KEY"] = "msg"
app.config["MAX_CONTENT_LENGTH"] = 500 * 1024 * 1024  # 500MB max for large Excel uploads

# Flask-Mail Configuration
app.config["MAIL_SERVER"] = os.environ.get("MAIL_SERVER", "smtp.gmail.com")
app.config["MAIL_PORT"] = int(os.environ.get("MAIL_PORT", 587))
app.config["MAIL_USE_TLS"] = os.environ.get("MAIL_USE_TLS", "True").lower() in ("true", "1", "yes")
app.config["MAIL_USERNAME"] = os.environ.get("MAIL_USERNAME")
app.config["MAIL_PASSWORD"] = os.environ.get("MAIL_PASSWORD")
app.config["MAIL_DEFAULT_SENDER"] = os.environ.get("MAIL_DEFAULT_SENDER", os.environ.get("MAIL_USERNAME"))

db.init_app(app)
bcrypt.init_app(app)
jwt = JWTManager(app)

# ─── JWT error handlers ───────────────────────────────────────────────────────
# By default flask-jwt-extended returns 422 with no useful body when token
# decoding fails. Replace each handler so the frontend gets a clear reason
# (helps the user know whether they need to re-login vs. it's a bug).
@jwt.expired_token_loader
def _jwt_expired(jwt_header, jwt_payload):
    return jsonify({"error": "token_expired", "msg": "Session expired — please log in again."}), 401

@jwt.invalid_token_loader
def _jwt_invalid(reason):
    # Triggered for bad signature, malformed token, etc. Most common after a
    # JWT_SECRET change on the server: every old token becomes invalid and
    # the user must re-login once.
    return jsonify({"error": "token_invalid", "msg": f"Invalid auth token: {reason}. Please log out and log in again."}), 401

@jwt.unauthorized_loader
def _jwt_missing(reason):
    return jsonify({"error": "auth_required", "msg": f"Authentication required: {reason}"}), 401

@jwt.needs_fresh_token_loader
def _jwt_needs_fresh(jwt_header, jwt_payload):
    return jsonify({"error": "fresh_token_required", "msg": "Fresh login required for this action."}), 401

@jwt.revoked_token_loader
def _jwt_revoked(jwt_header, jwt_payload):
    return jsonify({"error": "token_revoked", "msg": "Token has been revoked."}), 401

mail = Mail(app)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="threading")
# CORS_ORIGINS: comma-separated allowed frontend origins.
# Dev default: localhost:3000. Production: set CORS_ORIGINS=https://yourdomain.com
_cors_origins = [
    o.strip()
    for o in os.environ.get("CORS_ORIGINS", "http://localhost:3000,http://localhost:3001,http://127.0.0.1:3000").split(",")
    if o.strip()
]
CORS(
    app,
    resources={r"/api/*": {"origins": _cors_origins}},
    supports_credentials=True,
)

# ─── Global JSON error handlers ──────────────────────────────────────────────
# Flask returns HTML by default for 404/405/500 — override so API callers
# always get a JSON body instead of an HTML error page.
@app.errorhandler(404)
def handle_404(e):
    return jsonify({"error": "Not found", "detail": str(e)}), 404

@app.errorhandler(405)
def handle_405(e):
    return jsonify({"error": "Method not allowed", "detail": str(e)}), 405

@app.errorhandler(500)
def handle_500(e):
    return jsonify({"error": "Internal server error", "detail": str(e)}), 500

# ─── LLM Configuration (Azure OpenAI / OpenAI-compatible) ───────────────────
def _build_llm_client():
    azure_api_key = os.environ.get("AZURE_OPENAI_API_KEY")
    azure_endpoint = os.environ.get("AZURE_OPENAI_ENDPOINT")
    if azure_api_key and azure_endpoint:
        print(">>> Using Azure OpenAI configuration")
        return (
            AzureOpenAI(
                api_key=azure_api_key,
                azure_endpoint=azure_endpoint,
                api_version=os.environ.get("AZURE_OPENAI_API_VERSION", "2023-07-01-preview"),
            ),
            os.environ.get("AZURE_OPENAI_DEPLOYMENT", "gpt-4o-mini"),
        )

    gemini_api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("OPENAI_API_KEY")
    gemini_base_url = os.environ.get("GEMINI_BASE_URL")
    if gemini_api_key and gemini_base_url:
        print(">>> Using Gemini OpenAI-compatible configuration")
        return (
            OpenAI(
                api_key=gemini_api_key,
                base_url=gemini_base_url.rstrip("/") + "/",
            ),
            os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"),
        )

    openai_api_key = os.environ.get("OPENAI_API_KEY")
    if openai_api_key:
        print(">>> Using OpenAI configuration")
        return (
            OpenAI(api_key=openai_api_key),
            os.environ.get("OPENAI_MODEL", "gpt-4o-mini"),
        )

    raise RuntimeError(
        "No LLM credentials found. Configure either AZURE_OPENAI_API_KEY + AZURE_OPENAI_ENDPOINT, "
        "or GEMINI_API_KEY + GEMINI_BASE_URL, or OPENAI_API_KEY."
    )


client, DEPLOYMENT_NAME = _build_llm_client()

# ─── Init AI prompt modules ───────────────────────────────────────────────────
# Initialise both modules with the shared Azure OpenAI client so they can
# make API calls without importing app-level globals themselves.
# TELECOM_MENU is defined further below — modules are re-inited after it.
network_prompts.init(client, DEPLOYMENT_NAME, {})
broadband_prompts.init(client, DEPLOYMENT_NAME, db, User)
network_diagnosis.init(client, DEPLOYMENT_NAME, db, {
    "User": User, "Ticket": Ticket, "ChatSession": ChatSession,
    "KpiData": KpiData, "TelecomSite": TelecomSite,
})

broadband_prompts.register_routes(app)
network_diagnosis.register_routes(app)

def _user_brief(u, off_days=None):
    return {
        "id": u.id,
        "name": u.name,
        "email": u.email,
        "phone": u.phone_number,
        "role": u.role,
        "employee_id": u.employee_id,
        "is_online": bool(u.is_online),
        "off_days": off_days or [],
    }


def _build_duty_roster(target_date):
    managers = User.query.filter_by(role="manager").order_by(User.name.asc(), User.id.asc()).all()
    agents = User.query.filter_by(role="human_agent").order_by(User.name.asc(), User.id.asc()).all()

    if len(managers) < 3:
        return None, "At least 3 managers are required to form 3 teams"

    resources = managers + agents
    total = len(resources)

    day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    weekday_idx = target_date.weekday()  # Mon=0

    # Assign off-days in groups of 3 (extra members get last group's off-days)
    off_days_map = {}
    group_count = max(total // 3, 1)
    for g in range(group_count):
        off1 = (g * 2) % 7
        off2 = (off1 + 3) % 7
        names = [day_names[off1], day_names[off2]]
        for member in resources[g * 3:(g + 1) * 3]:
            off_days_map[member.id] = names
    # Assign any remaining members to the last group's off-days
    for member in resources[group_count * 3:]:
        off_days_map[member.id] = off_days_map.get(resources[group_count * 3 - 1].id, [])

    available = [u for u in resources if day_names[weekday_idx] not in off_days_map.get(u.id, [])]
    managers_available = [u for u in available if u.role == "manager"]
    if len(managers_available) < 3:
        return None, "Not enough managers available today to cover 3 shifts. Adjust admin resources."

    total_available = len(available)
    team_size = total_available // 3
    rotation = target_date.timetuple().tm_yday % len(managers_available)
    rotated = managers_available[rotation:] + managers_available[:rotation]
    lead_managers = rotated[:3]
    lead_ids = {m.id for m in lead_managers}
    pool = [u for u in available if u.id not in lead_ids]

    teams = []
    for m in lead_managers:
        teams.append({
            "id": f"team-{m.id}",
            "name": f"Team {m.name}",
            "manager": _user_brief(m, off_days_map.get(m.id, [])),
            "agents": [],
        })

    for idx, member in enumerate(pool):
        teams[idx % 3]["agents"].append(_user_brief(member, off_days_map.get(member.id, [])))

    off_today = [
        _user_brief(u, off_days_map.get(u.id, []))
        for u in resources
        if day_names[weekday_idx] in off_days_map.get(u.id, [])
    ]

    rotation_mod = target_date.timetuple().tm_yday % 3
    shift_times = [
        {"name": "Shift 1", "time": "00:00-08:00"},
        {"name": "Shift 2", "time": "08:00-16:00"},
        {"name": "Shift 3", "time": "16:00-00:00"},
    ]
    shifts = []
    for i in range(3):
        team = teams[(i + rotation_mod) % 3]
        shifts.append({"shift": shift_times[i], "team": team})

    return {
        "shift_times": shift_times,
        "teams": teams,
        "shifts": shifts,
        "meta": {
            "total_resources": total,
            "team_size": team_size,
            "managers": len(managers),
            "agents": len(agents),
            "rotation_index": rotation_mod,
            "off_today_count": len(off_today),
            "off_today": off_today,
        },
    }, None


@app.route("/api/cto/duty-roster")
@jwt_required()
def cto_duty_roster():
    date_str = request.args.get("date")
    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else datetime.utcnow().date()
    except ValueError:
        return jsonify({"error": "Invalid date format, use YYYY-MM-DD"}), 400
    roster, err = _build_duty_roster(target_date)
    if err:
        return jsonify({"error": err}), 400
    return jsonify(roster)


# ─── Nearest-Tower Lookup (loaded once at startup) ────────────────────────────
_SITE_DATA = []


def _load_site_data():
    """Legacy Excel loader (disabled). Site data is now loaded via admin upload into DB."""
    return


def _haversine(lat1, lon1, lat2, lon2):
    """Calculate distance in km between two lat/lon points."""
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ── Strict KPI mapping: exact KPI names from DB for each problem type ──────
# Only KPIs DIRECTLY related to diagnosing each problem are included.
PROBLEM_KPI_EXACT = {
    "internet_signal": [
        "LTE DL - Usr Ave Throughput",       # Primary: user download speed
        "DL PRB Utilization (1BH)",           # Congestion indicator
        "Average Latency Downlink",           # Latency
        "Ave RRC Connected Ue",               # User load causing congestion
        "Average NI of Carrier-",             # Interference degrading speed
        "Availability",                       # Site up/down
    ],
    "call_drop": [
        "E-RAB Call Drop Rate_1",             # Primary: drop rate
        "LTE Intra-Freq HO Success Rate",     # Handover failures cause drops
        "DL PRB Utilization (1BH)",           # Congestion causes drops
        "Ave RRC Connected Ue",               # Overload causes drops
        "LTE DL - Usr Ave Throughput",        # Degraded service indicator
        "Availability",                       # Site availability
    ],
    "call_failure": [
        "LTE Call Setup Success Rate",        # Primary: CSSR
        "LTE RRC Setup Success Rate",         # RRC failures block calls
        "LTE E-RAB Setup Success Rate",       # E-RAB setup failures
        "DL PRB Utilization (1BH)",           # Congestion blocks calls
        "VoLTE Traffic Erlang",               # Voice traffic load
        "Availability",                       # Site availability
    ],
}

# Legacy keyword-based fallback for matching arbitrary KPI names
NETWORK_PROBLEM_KPI_KEYWORDS = {
    "internet_signal": ["throughput", "latency", "prb", "volume", "interference", "availability", "ue", "user"],
    "call_failure": ["call setup", "rrc setup", "e-rab setup", "csfb", "volte", "availability", "prb"],
    "call_drop": ["drop", "handover", "ho success", "prb", "ue", "throughput", "availability"],
}


def _normalize_problem_text(ticket: Ticket) -> str:
    return " ".join([
        (ticket.category or "").strip().lower(),
        (ticket.subcategory or "").strip().lower(),
        (ticket.description or "").strip().lower(),
    ])


def _detect_network_problem_type(ticket: Ticket) -> str:
    text = _normalize_problem_text(ticket)
    if any(x in text for x in ["call drop", "calls drop", "dropped call", "call disconnect", "drop rate"]):
        return "call_drop"
    if any(x in text for x in ["call failure", "calls fail", "unable to make call", "call not connecting", "call setup"]):
        return "call_failure"
    if "call / sms failures" in text:
        return "call_failure"
    return "internet_signal"


def _problem_type_label(problem_type: str) -> str:
    labels = {
        "internet_signal": "Internet Speed",
        "call_failure": "Call Failure",
        "call_drop": "Call Drop",
    }
    return labels.get(problem_type, "Internet Speed")


def _filter_kpi_names_for_problem(kpi_names, problem_type: str):
    """Select only KPIs strictly related to the problem type.
    Uses exact name matching first, then keyword fallback."""
    exact = PROBLEM_KPI_EXACT.get(problem_type, PROBLEM_KPI_EXACT["internet_signal"])
    # First: exact match against the predefined list
    available = set(kpi_names)
    selected = [k for k in exact if k in available]
    # Fallback: keyword match if exact match found nothing
    if not selected:
        keys = NETWORK_PROBLEM_KPI_KEYWORDS.get(problem_type, NETWORK_PROBLEM_KPI_KEYWORDS["internet_signal"])
        selected = [name for name in sorted(kpi_names) if any(k in (name or "").lower() for k in keys)]
    if not selected:
        selected = sorted(kpi_names)[:8]
    return selected


def _period_key_for_row(row: KpiData, period: str, has_hour_data: bool = False) -> str:
    if period == "month":
        return row.date.strftime("%Y-%m")
    if period == "week":
        iso = row.date.isocalendar()
        return f"{iso[0]}-W{iso[1]:02d}"
    if period == "hour":
        # Only show hourly if actual hour variation exists, otherwise fall back to daily
        if has_hour_data and (row.hour or 0) != 0:
            return f"{row.date.strftime('%m/%d')} {row.hour:02d}h"
        return row.date.strftime("%Y-%m-%d")
    return row.date.strftime("%Y-%m-%d")


def _build_period_stats(rows, period: str):
    has_hour = any((r.hour or 0) != 0 for r in rows) if rows else False
    agg = {}
    for r in rows:
        if r.value is None:
            continue
        key = _period_key_for_row(r, period, has_hour)
        agg.setdefault(key, []).append(float(r.value))
    if not agg:
        return "no data"
    ordered = sorted(agg.items(), key=lambda kv: kv[0])
    latest_key, latest_vals = ordered[-1]
    flat_vals = [v for _, vals in ordered for v in vals]
    avg_all = round(sum(flat_vals) / len(flat_vals), 4)
    return f"{latest_key}: avg={round(sum(latest_vals)/len(latest_vals), 4)}, overall_avg={avg_all}, min={round(min(flat_vals), 4)}, max={round(max(flat_vals), 4)}"


def _build_kpi_summary_text(rows, selected_kpis, data_level_label: str) -> str:
    from collections import defaultdict
    grouped = defaultdict(list)
    for r in rows:
        if r.kpi_name in selected_kpis and r.value is not None:
            grouped[r.kpi_name].append(r)
    if not grouped:
        return f"No {data_level_label} KPI data available."

    parts = []
    for kpi_name in sorted(grouped.keys()):
        kpi_rows = grouped[kpi_name]
        monthly = _build_period_stats(kpi_rows, "month")
        weekly = _build_period_stats(kpi_rows, "week")
        daily = _build_period_stats(kpi_rows, "day")
        hourly = _build_period_stats(kpi_rows, "hour")
        parts.append(
            f"- {kpi_name}\n"
            f"  Monthly: {monthly}\n"
            f"  Weekly: {weekly}\n"
            f"  Daily: {daily}\n"
            f"  Hourly: {hourly}"
        )
    return "\n".join(parts)


# ─── UPDATED: RCA text formatting helpers ────────────────────────────────────

def _normalize_ai_lines(text: str):
    """
    Split raw AI text into clean, complete lines.
    - Strips leading bullet / number markers only
    - Removes stray 'Crux:' prefix labels (all variants)
    - Deduplicates while preserving order
    - Does NOT strip **bold** from mid-line content (titles like **Title**: body are kept)
    - Does NOT truncate content — full sentences are preserved
    """
    lines = []
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line:
            continue
        # Remove leading list markers: "1.", "2)", "-", "*", "•"
        line = re.sub(r"^\s*(?:[-*•]|\d+[.)]\s*)\s*", "", line).strip()
        # Remove "**Crux:**", "Crux:", "**Crux**:", etc. from the front only
        line = re.sub(r"^\*{0,2}[Cc]rux\*{0,2}\s*:?\s*", "", line).strip()
        # Only strip ** if they wrap the ENTIRE line AND there's no colon separator
        # i.e. "**Some Title**" alone → "Some Title"
        # but "**Title**: body text" is kept as-is (bold title with body)
        if re.match(r"^\*\*[^*]+\*\*$", line):
            line = re.sub(r"^\*\*(.*?)\*\*$", r"\1", line).strip()
        if not line:
            continue
        lines.append(line)

    # Deduplicate while preserving order
    unique, seen = [], set()
    for line in lines:
        key = line.lower()
        if key not in seen:
            seen.add(key)
            unique.append(line)

    # Drop the last line if it looks truncated (no closing punctuation).
    # This happens when the AI response is cut by max_tokens mid-sentence.
    if unique and not re.search(r'[.!?)>]$', unique[-1].rstrip()):
        unique = unique[:-1]

    return unique


def _force_numbered_points(raw_text: str, min_points: int, max_points: int, prefix: str = "", fallback_points=None):
    """
    Parse AI output into clean numbered points.
    - Full sentences preserved — no character truncation
    - No 'Crux:' prefix injected (prefix param kept for API compat but defaults to "")
    - Bold markdown (**Title**: body) preserved as-is for frontend rendering
    - Picks most technically relevant lines first
    """
    lines = _normalize_ai_lines(raw_text)

    # Score lines: prefer those with technical/diagnostic keywords and longer content
    ranked = sorted(
        lines,
        key=lambda l: (
            1 if re.search(r"\b(root cause|cause|kpi|trend|alarm|site|cell|impact|action|recommend)\b", l, re.IGNORECASE) else 0,
            len(l),
        ),
        reverse=True,
    )
    picked = []
    for line in ranked:
        if len(line) < 12:
            continue
        picked.append(line)  # No truncation — keep full sentence
        if len(picked) >= max_points:
            break

    for line in (fallback_points or []):
        if len(picked) >= min_points:
            break
        if line and line not in picked:
            picked.append(line)

    picked = picked[:max_points]
    if not picked:
        picked = ["Analysis could not be generated from available data."]

    # prefix param ignored (was used for "**Crux:** " — removed)
    return "\n".join(f"{idx + 1}. {line}" for idx, line in enumerate(picked))


def _strip_markdown_for_pdf(text: str) -> str:
    """Remove **bold** and *italic* markdown markers for plain-text PDF output."""
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)
    text = re.sub(r"\*(.*?)\*", r"\1", text)
    return text


def _format_points_for_pdf(raw_text: str) -> str:
    """
    Convert numbered markdown points to clean plain-text for PDF sections.
    Input:  "1. **Site Offline**: The site is off air due to power failure.\n2. ..."
    Output: "1. Site Offline: The site is off air due to power failure.\n\n2. ..."
    """
    lines = _normalize_ai_lines(raw_text)
    plain_lines = []
    for i, line in enumerate(lines[:5], 1):
        plain = _strip_markdown_for_pdf(line)
        plain_lines.append(f"{i}. {plain}")
    return "\n\n".join(plain_lines)

def _infer_issue_flags(text: str):
    t = (text or "").lower()
    flags = set()
    if any(k in t for k in ["prb", "utilization", "congestion", "overload", "traffic spike"]):
        flags.add("congestion")
    if any(k in t for k in ["sinr", "rsrp", "rsrq", "weak signal", "coverage", "low signal"]):
        flags.add("coverage")
    if any(k in t for k in ["interference", "pilot pollution", "overshoot", "over-shoot"]):
        flags.add("interference")
    if any(k in t for k in ["latency", "packet loss", "delay", "jitter"]):
        flags.add("latency")
    if any(k in t for k in ["handover", "rlf", "drop", "call drop"]):
        flags.add("handover")
    if any(k in t for k in ["call setup", "cssr", "paging", "accessibility", "sdcch", "call failure"]):
        flags.add("access")
    if "off air" in t or "off_air" in t:
        flags.add("off_air")
    return flags


def _clean_ai_response(text):
    """Strip all markdown formatting from AI response: **, ##, ###, ---, •, ━, ─"""
    import re
    if not text:
        return text
    t = text
    t = t.replace('**', '')
    t = t.replace('***', '')
    t = re.sub(r'^#{1,4}\s*', '', t, flags=re.MULTILINE)
    t = re.sub(r'^[\-─━═]{3,}$', '', t, flags=re.MULTILINE)
    t = re.sub(r'^[•●◦▪]\s*', '', t, flags=re.MULTILINE)
    t = re.sub(r'\n{3,}', '\n\n', t)
    return t.strip()


def _build_parameter_recommendations(problem_type, root_cause, trend_summary, nearest):
    """Build concrete RF parameter change recommendations using actual site values."""
    bw = nearest.get("bandwidth_mhz") if nearest else None
    gain = nearest.get("antenna_gain_dbi") if nearest else None
    eirp = nearest.get("rf_power_eirp_dbm") if nearest else None
    height = nearest.get("antenna_height_agl_m") if nearest else None
    tilt = nearest.get("e_tilt_degree") if nearest else None
    crs = nearest.get("crs_gain") if nearest else None

    def _v(val, delta):
        """Safely adjust a numeric value."""
        try: return round(float(val) + delta, 1)
        except: return "N/A"

    recs = []

    # Generate exactly 4 unique parameter recommendations — one per parameter
    # Each addresses a different aspect of the root cause
    params = [
        ("Bandwidth", bw, "MHz",
         {"Internet Speed": (5 if bw and float(bw)<=10 else 10, "increase PRB capacity, reducing congestion and improving DL throughput by 15-25%"),
          "Call Drop": (5, "free resources for active bearers, reducing E-RAB drops under load by ~10%"),
          "Call Failure": (5, "free PRB resources for call setup signaling, improving CSSR by ~5%")}),
        ("E-tilt", tilt, "°",
         {"Internet Speed": (-1, "extend coverage footprint, improving RSRP for cell-edge users and increasing throughput by ~10%"),
          "Call Drop": (1, "reduce overshoot into neighboring cells, minimizing inter-cell interference and reducing drop rate by ~20%"),
          "Call Failure": (-0.5, "optimize coverage-to-interference ratio, improving RRC Setup Success Rate by ~5%")}),
        ("EIRP", eirp, "dBm",
         {"Internet Speed": (2, "boost signal strength at cell edge, reducing latency and improving user throughput by ~12%"),
          "Call Drop": (-1, "reduce pilot pollution in overlapping zones, improving handover success rate by ~15%"),
          "Call Failure": (2, "improve signal quality for call setup, increasing E-RAB Setup Success Rate by ~8%")}),
        ("CRS Gain", crs, "",
         {"Internet Speed": (3, "improve reference signal quality and channel estimation, boosting throughput by ~8%"),
          "Call Drop": (3, "strengthen reference signals during mobility, reducing Radio Link Failures by ~12%"),
          "Call Failure": (3, "improve cell detection and measurement accuracy, enhancing RRC success rate by ~6%")}),
    ]

    pt = problem_type if problem_type in ("Internet Speed","Call Drop","Call Failure") else "Internet Speed"
    for name, val, unit, actions in params:
        if val is not None:
            delta, effect = actions.get(pt, actions["Internet Speed"])
            new_val = _v(val, delta)
            direction = "Increase" if delta > 0 else "Decrease"
            recs.append(f"**{name} {direction}**: Current {name} is {val}{unit}. {direction} to {new_val}{unit} to {effect}.")

    # Add Antenna Height as 5th option if space and value available
    if len(recs) < 4 and height is not None:
        recs.append(f"**Antenna Height Adjustment**: Current height is {height} m. Adjust to {_v(height, -2 if pt=='Call Drop' else 2)} m to {'reduce overshooting and improve HO success' if pt=='Call Drop' else 'improve line-of-sight coverage and signal quality'}.")

    # Add Antenna Gain if still space
    if len(recs) < 4 and gain is not None:
        recs.append(f"**Antenna Gain Upgrade**: Current gain is {gain} dBi. Increase to {_v(gain, 1)} dBi to improve signal levels for edge users.")

    if not recs:
        recs = [
            "**Bandwidth**: Increase carrier bandwidth by 5-10 MHz to reduce congestion.",
            "**E-tilt**: Optimize electrical tilt to balance coverage vs interference.",
            "**EIRP**: Adjust transmit power to improve cell-edge signal quality.",
        ]

    return recs[:4]


def _kpi_degradation_points(rows, selected_kpis, max_points=3):
    """Build fallback RCA points from KPI trend data by detecting significant shifts."""
    from collections import defaultdict
    grouped = defaultdict(list)
    for r in rows:
        if r.kpi_name in selected_kpis and r.value is not None:
            grouped[r.kpi_name].append(r)
    points = []
    for kpi_name, kpi_rows in grouped.items():
        kpi_rows.sort(key=lambda r: (r.date, r.hour))
        vals = [float(r.value) for r in kpi_rows if r.value is not None]
        if len(vals) < 2:
            continue
        overall = sum(vals) / len(vals)
        latest_vals = [float(r.value) for r in kpi_rows[-min(8, len(kpi_rows)):] if r.value is not None]
        latest = sum(latest_vals) / max(len(latest_vals), 1)
        if overall == 0:
            continue
        delta_pct = (latest - overall) / overall * 100
        if abs(delta_pct) < 20:
            continue
        direction = "drop" if delta_pct < 0 else "increase"
        points.append((
            abs(delta_pct),
            f"**KPI Shift**: {kpi_name} shows a {direction} to {round(latest, 3)} vs baseline {round(overall, 3)} ({round(delta_pct, 1)}%)."
        ))
    points.sort(key=lambda x: -x[0])
    return [p for _, p in points[:max_points]]


def _recommendation_has_params(text: str):
    t = (text or "").lower()
    param_hits = 0
    for p in ["bandwidth", "antenna gain", "eirp", "rf power", "antenna height", "e-tilt", "tilt", "crs gain"]:
        if p in t:
            param_hits += 1
    has_numbers = bool(re.search(r"\d+(\.\d+)?", t))
    return param_hits >= 2 and has_numbers


def _filter_rca_lines(lines):
    generic_patterns = [
        r"\bsite status\b",
        r"\bproblem classification\b",
        r"\bproblem type\b",
        r"\bprimary impact domain\b",
        r"\bonly related kpi\b",
        r"\btrend evidence\b",
        r"\baction required\b",
        r"\bfocus area\b",
    ]
    out = []
    for line in lines:
        l = line.lower()
        if any(re.search(p, l) for p in generic_patterns):
            continue
        # Keep lines that mention KPI evidence or alarms/solutions
        if re.search(r"\b(kpi|throughput|latency|prb|sinr|rsrp|rsrq|cqi|drop|handover|cssr|paging|rlf|alarm|off air|on air)\b", l):
            out.append(line)
            continue
        # Keep lines with concrete numbers/percentages
        if re.search(r"\d+(\.\d+)?", l):
            out.append(line)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# TICKET ROOT CAUSE ANALYSIS & FINAL RECOMMENDATIONS
# ─────────────────────────────────────────────────────────────────────────────
# Full workflow (called from agent ticket detail page):
#   1. Agent clicks "Root Cause Analysis" → POST /api/agent/tickets/:id/root-cause
#   2. Agent clicks "Recommendation"      → POST /api/agent/tickets/:id/recommendation
# Both routes are registered via network_diagnosis.register_routes(app).
# The functions below provide the core logic used by those routes.
# ─────────────────────────────────────────────────────────────────────────────


def generate_root_cause_analysis(ticket, session, client, deployment):
    """
    Generate AI-powered root cause analysis for a ticket using Senior Telecom
    RAN Optimization Expert prompt with full KPI trend data and RF parameters.
    """
    if not session or not session.latitude or not session.longitude:
        return {"error": "Customer location not available"}

    nearest_list = find_nearest_sites(session.latitude, session.longitude, n=1)
    if not nearest_list:
        return {"error": "No site data available for nearest-site lookup."}

    nearest = nearest_list[0]
    nearest_site_id = nearest["site_id"]
    nearest_zone = nearest.get("zone") or nearest.get("province", "")
    dist_km = nearest["distance_km"]
    site_status = (nearest.get("site_status") or "on_air").lower()
    site_name = nearest.get("site_name") or nearest_site_id
    cell_id = nearest.get("cell_id") or nearest.get("cell_name") or "N/A"

    problem_type = _detect_network_problem_type(ticket)
    problem_type_label = _problem_type_label(problem_type)

    # Fetch KPI data for nearest site
    site_rows = KpiData.query.filter_by(site_id=nearest_site_id, data_level="site").all()
    cell_rows = KpiData.query.filter_by(site_id=nearest_site_id, data_level="cell").all()
    all_kpis = {r.kpi_name for r in site_rows + cell_rows}
    selected_kpis = _filter_kpi_names_for_problem(all_kpis, problem_type)
    site_kpi_text = _build_kpi_summary_text(site_rows, selected_kpis, "site-level")
    cell_kpi_text = _build_kpi_summary_text(cell_rows, selected_kpis, "cell-level")

    # Build RF parameters block from all site params including extra_params
    rf_params_lines = [
        f"Bandwidth (MHz): {nearest.get('bandwidth_mhz')}",
        f"Antenna Gain (dBi): {nearest.get('antenna_gain_dbi')}",
        f"RF Power (EIRP) [dBm]: {nearest.get('rf_power_eirp_dbm')}",
        f"Antenna Height (AGL) (M): {nearest.get('antenna_height_agl_m')}",
        f"E-tilt (Degree): {nearest.get('e_tilt_degree')}",
        f"CRS Gain: {nearest.get('crs_gain')}",
    ]
    extra = nearest.get("extra_params") or {}
    for k, v in extra.items():
        rf_params_lines.append(f"{k}: {v}")
    rf_parameters_block = "\n".join(rf_params_lines)

    # Extract signal diagnosis data (RSRP, SINR, RSRQ) from chat messages if available
    signal_diagnosis_block = ""
    if session:
        try:
            signal_msgs = ChatMessage.query.filter_by(session_id=session.id, sender='bot').order_by(ChatMessage.created_at.desc()).limit(20).all()
            for msg in signal_msgs:
                content = msg.content or ""
                if any(kw in content.lower() for kw in ['rsrp', 'sinr', 'rsrq', 'signal diagnosis', 'signal:']):
                    signal_diagnosis_block = f"\nCUSTOMER SIGNAL READINGS (from device screenshot):\n{content}\n"
                    break
        except Exception:
            pass

    # Build KPI trend block
    kpi_trend_block = f"SITE-LEVEL KPI TRENDS:\n{site_kpi_text}\n\nCELL-LEVEL KPI TRENDS:\n{cell_kpi_text}"
    if signal_diagnosis_block:
        kpi_trend_block += signal_diagnosis_block

    # Determine observation window
    dates = sorted({r.date for r in site_rows + cell_rows if r.date})
    if dates:
        obs_window = f"{dates[0]} to {dates[-1]} ({(dates[-1]-dates[0]).days} days)"
    else:
        obs_window = "Last 7 Days"

    site_status_display = "ON AIR" if site_status == "on_air" else "OFF AIR"

    prompt = f"""You are a Senior RAN Optimization Expert with 20+ years of hands-on field experience investigating network performance issues across Ericsson, Huawei, and Nokia LTE networks. You think like a real RF engineer — you look at KPI data the way a doctor reads lab results: each number tells a story, and the combination of multiple KPI movements reveals the underlying disease.

A customer has escalated a complaint that the AI chatbot could not resolve. You must analyze the KPI trends from the nearest cell site, correlate them with the site's RF parameters, and determine what is physically happening in the radio environment that is causing the customer's problem.

CUSTOMER COMPLAINT: {ticket.description}
PROBLEM TYPE: {problem_type_label}

NEAREST SITE:
  Site: {site_name} | Cell: {cell_id} | Status: {site_status_display} | Distance: {dist_km} km from customer
  Observation Window: {obs_window}

RF PARAMETERS (live values from database):
{rf_parameters_block}

KPI TREND DATA:
{kpi_trend_block}

FORMATTING RULES:
- Plain text only. No asterisks, no hash symbols, no dashes as separators, no bullet points, no bold markers.
- Write exactly 4 numbered points (1. 2. 3. 4.)
- Each point: 3-4 clear sentences. Start with a short title followed by colon, then explanation.
- Cite actual KPI values and RF parameter values from the data above. No assumptions.

Write these 4 points:

1. Primary KPI Degradation: Identify the main degraded KPI that directly reflects the customer's {problem_type_label} complaint. Cite its exact values (daily avg, min, max vs baseline). Explain what this means in RF terms and how it connects to the customer's experience.

2. Supporting Evidence: Cite a second correlated KPI that confirms the diagnosis. Explain the cause-effect chain between the KPIs (e.g., high PRB util with low users = SINR problem not congestion). {'Use the customer signal readings (RSRP, SINR) from the screenshot as evidence.' if signal_diagnosis_block else ''}

3. RF Parameter Analysis: Explain how the current site configuration — Bandwidth, E-tilt, EIRP, Antenna Height, CRS Gain (use actual values from DB) — contributes to the problem. Connect each relevant parameter to the degraded KPIs.

4. Root Cause Conclusion: State the single physical root cause that explains all degraded KPIs together (e.g., "overshooting due to low E-tilt at high antenna height" or "capacity exhaustion on narrow bandwidth"). State whether the issue is sudden, gradual, or chronic based on KPI trend comparison."""

    try:
        response = client.chat.completions.create(
            model=deployment,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
            max_tokens=4000,
        )
        analysis = _clean_ai_response(response.choices[0].message.content)
        if not analysis or len(analysis) < 50:
            analysis = "AI analysis returned insufficient data. Please retry."
    except Exception as e:
        analysis = f"RCA generation failed: {str(e)}. Please check OpenAI configuration and retry."

    return {
        "analysis": analysis,
        "analysis_pdf": _format_points_for_pdf(analysis),
        "site_id": nearest_site_id,
        "site_zone": nearest_zone,
        "site_status": site_status,
        "distance_km": dist_km,
        "problem_type": problem_type_label,
        "selected_kpis": selected_kpis,
    }


def generate_final_recommendations(ticket, session, root_cause, trend_summary, client, deployment):
    """
    Generate AI-powered final recommendations using Senior Telecom RAN Optimization
    Expert prompt with full RF parameter database and root cause analysis.
    """
    nearest = None
    if session and session.latitude and session.longitude:
        nearest_list = find_nearest_sites(session.latitude, session.longitude, n=1)
        if nearest_list:
            nearest = nearest_list[0]

    problem_type = _problem_type_label(_detect_network_problem_type(ticket))

    # Build RF parameters block
    rf_params_lines = []
    if nearest:
        rf_params_lines = [
            f"Bandwidth (MHz): {nearest.get('bandwidth_mhz')}",
            f"Antenna Gain (dBi): {nearest.get('antenna_gain_dbi')}",
            f"RF Power (EIRP) [dBm]: {nearest.get('rf_power_eirp_dbm')}",
            f"Antenna Height (AGL) (M): {nearest.get('antenna_height_agl_m')}",
            f"E-tilt (Degree): {nearest.get('e_tilt_degree')}",
            f"CRS Gain: {nearest.get('crs_gain')}",
        ]
        extra = nearest.get("extra_params") or {}
        for k, v in extra.items():
            rf_params_lines.append(f"{k}: {v}")
    rf_parameters_block = "\n".join(rf_params_lines) if rf_params_lines else "Not available"

    site_name = nearest.get("site_name", nearest.get("site_id", "Unknown")) if nearest else "Unknown"
    cell_id = nearest.get("cell_id", "N/A") if nearest else "N/A"
    site_status = (nearest.get("site_status") or "on_air").upper().replace("_", " ") if nearest else "ON AIR"

    prompt = f"""You are a Senior RAN Optimization Expert writing field recommendations after completing a root cause analysis. You have the root cause findings and the complete RF parameter database for this site. Your job is to identify ONLY the 2-3 RF parameters whose adjustment will directly resolve the identified root cause, and propose specific changes with engineering justification.

SITE: {site_name} | Cell: {cell_id} | Status: {site_status}
CUSTOMER COMPLAINT: {ticket.description}
PROBLEM TYPE: {problem_type}

ROOT CAUSE ANALYSIS FINDINGS:
{root_cause if root_cause else 'Root cause analysis not available.'}

TREND EVIDENCE:
{trend_summary if trend_summary else 'No trend data available.'}

COMPLETE RF PARAMETER DATABASE (all current live values):
{rf_parameters_block}

FORMATTING RULES:
- Plain text only. No asterisks, no hash symbols, no dashes as separators, no bullet points, no bold markers.
- Write exactly 3 numbered points (1. 2. 3.)
- Each point: 4-5 clear sentences. Start with a title followed by colon.
- Use EXACT current values from the database. Do NOT invent values.

Pick ONLY the 2-3 RF parameters most relevant to the root cause. Do NOT recommend unrelated parameters.

Write these 3 recommendations:

1. [Most critical parameter] Change: State parameter name and current value from DB. Recommend new value within 3GPP standards. Explain why this fixes the root cause — what happens to the radio signal. State which KPIs improve and by how much. Note any risk to adjacent cells.

2. [Second parameter] Change: Same structure — current value, recommended value, RF rationale linking to root cause, expected KPI recovery, rollback plan.

3. [Third parameter or Validation]: Either a third parameter change, or a post-optimization validation step (drive test, 48hr monitoring, neighbor impact check)."""

    try:
        response = client.chat.completions.create(
            model=deployment,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
            max_tokens=4000,
        )
        recommendation = _clean_ai_response(response.choices[0].message.content)
        if not recommendation or len(recommendation) < 50:
            recommendation = "AI recommendations returned insufficient data. Please retry."
    except Exception as e:
        recommendation = f"Recommendation generation failed: {str(e)}. Please check OpenAI configuration and retry."

    return {
        "recommendation": recommendation,
        "recommendation_pdf": _format_points_for_pdf(recommendation),
    }


# ─────────────────────────────────────────────────────────────────────────────


def find_nearest_sites(lat, lon, n=3):
    """Return the n nearest telecom sites with averaged RF parameters across all cells."""
    if lat is None or lon is None:
        return []

    try:
        from sqlalchemy import text as sa_text
        rows = db.session.execute(sa_text("""
            SELECT site_id, MIN(site_name) AS site_name, MIN(cell_id) AS cell_id,
                   MIN(province) AS province, MIN(commune) AS commune,
                   zone, city, state, site_status, alarms, solution, standard_solution_step,
                   AVG(latitude) AS latitude, AVG(longitude) AS longitude,
                   ROUND(AVG(bandwidth_mhz)::numeric, 1) AS bandwidth_mhz,
                   ROUND(AVG(antenna_gain_dbi)::numeric, 1) AS antenna_gain_dbi,
                   ROUND(AVG(rf_power_eirp_dbm)::numeric, 1) AS rf_power_eirp_dbm,
                   ROUND(AVG(antenna_height_agl_m)::numeric, 1) AS antenna_height_agl_m,
                   ROUND(AVG(e_tilt_degree)::numeric, 1) AS e_tilt_degree,
                   ROUND(AVG(crs_gain)::numeric, 1) AS crs_gain
            FROM telecom_sites
            WHERE latitude IS NOT NULL AND longitude IS NOT NULL
            GROUP BY site_id, zone, city, state, site_status, alarms, solution, standard_solution_step
        """)).fetchall()
    except Exception:
        return []

    if not rows:
        return []

    scored = []
    for r in rows:
        dist = _haversine(lat, lon, float(r.latitude or 0), float(r.longitude or 0))
        scored.append((dist, r))

    scored.sort(key=lambda x: x[0])
    results = []
    for dist, r in scored[:n]:
        status = (r.site_status or "on_air").lower()
        # Fetch extra_params for this site
        site_extra = {}
        try:
            site_obj = TelecomSite.query.filter_by(site_id=r.site_id).first()
            if site_obj and site_obj.extra_params:
                site_extra = site_obj.extra_params
        except:
            pass

        results.append({
            "site_id": r.site_id,
            "site_name": getattr(r, 'site_name', None) or r.site_id,
            "cell_id": getattr(r, 'cell_id', None),
            "cell_name": getattr(r, 'cell_id', None),
            "province": getattr(r, 'province', None) or r.zone or "",
            "commune": getattr(r, 'commune', None) or "",
            "zone": r.zone or getattr(r, 'province', None) or "",
            "city": r.city or "",
            "state": r.state or "",
            "latitude": float(r.latitude or 0),
            "longitude": float(r.longitude or 0),
            "site_status": status,
            "status": status,
            "alarms": r.alarms or "None",
            "alarm": r.alarms or "None",
            "solution": r.solution or "No action required",
            "standard_solution_step": r.standard_solution_step or "",
            "distance_km": round(dist, 2),
            # RF Parameters (averaged across all cells of this site)
            "bandwidth_mhz": float(r.bandwidth_mhz) if r.bandwidth_mhz else None,
            "antenna_gain_dbi": float(r.antenna_gain_dbi) if r.antenna_gain_dbi else None,
            "rf_power_eirp_dbm": float(r.rf_power_eirp_dbm) if r.rf_power_eirp_dbm else None,
            "antenna_height_agl_m": float(r.antenna_height_agl_m) if r.antenna_height_agl_m else None,
            "e_tilt_degree": float(r.e_tilt_degree) if r.e_tilt_degree else None,
            "crs_gain": float(r.crs_gain) if r.crs_gain else None,
            "extra_params": site_extra,
        })
    return results


# Excel loading disabled; site data comes from admin upload into DB.

# ═══════════════════════════════════════════════════════════════════════════════
#  CHATBOT CODE 
# ═══════════════════════════════════════════════════════════════════════════════

TELECOM_MENU = {
    "1": {
        "name": "Mobile Services (Prepaid / Postpaid)",
        "icon": "",
        "description": "Covers all issues related to mobile phone services including voice calls, SMS, mobile data, SIM cards, prepaid recharges, postpaid billing, roaming, number portability, and mobile network coverage.",
        "subprocesses": {
            "1": {"name": "Billing & Payment Issues", "semantic_scope": "Unexpected charges, wrong bill amount, double billing, payment failed but money deducted, recharge not credited, balance deducted without usage, auto-renewal charged, EMI issues on phone, refund not received for telecom services, incorrect tax on bill, bill dispute"},
            "2": {"name": "Network / Signal Problems", "semantic_scope": "No signal, weak signal, call drops, poor network coverage, network congestion, unable to make/receive calls, tower issue, dead zone, indoor coverage problem, 4G/5G not available, network outage in area"},
            "3": {"name": "SIM Card & Activation", "semantic_scope": "New SIM not activated, SIM blocked, SIM damaged, SIM swap, eSIM activation, lost SIM replacement, SIM not detected, PUK locked, KYC verification pending, Aadhaar linking with SIM, SIM upgrade to 4G/5G"},
            "4": {"name": "Data Plan & Recharge Issues", "semantic_scope": "Data not working after recharge, wrong plan activated, data exhausted too quickly, unable to recharge, recharge failed but amount debited, validity not extended, data speed throttled, unlimited plan not giving unlimited data, add-on pack issues, coupon/promo code not working"},
            "5": {"name": "International Roaming", "semantic_scope": "Roaming not working abroad, high roaming charges, incoming calls charged during roaming, data roaming activation, roaming pack not applied, unable to call from foreign country, roaming bill shock, ISD/STD calling issues"},
            "6": {"name": "Mobile Number Portability (MNP)", "semantic_scope": "Want to switch operator, MNP request rejected, porting delay, UPC code not received, number lost during porting, services disrupted after porting, porting to another network, port-out issues"},
            "7": {"name": "Call / SMS Failures", "semantic_scope": "Unable to make calls, calls not connecting, one-way audio, SMS not being delivered, SMS not received, OTP not coming, call going to voicemail, DND (Do Not Disturb) issues, spam calls, call forwarding not working, conference call issues"},
            "8": {"name": "Others", "semantic_scope": ""},
        },
    },
    "2": {
        "name": "Broadband / Internet Services",
        "icon": "",
        "description": "Covers all issues related to wired/wireless broadband, fiber internet, DSL connections, WiFi, and home/office internet services.",
        "subprocesses": {
            "1": {"name": "Slow Speed / No Connectivity", "semantic_scope": "Internet too slow, speed not matching plan, buffering while streaming, downloads very slow, no internet connection, WiFi connected but no internet, speed drops at night, latency/ping too high, speed test showing low results, bandwidth issue"},
            "2": {"name": "Frequent Disconnections", "semantic_scope": "Internet keeps disconnecting, connection drops every few minutes, unstable connection, intermittent connectivity, WiFi drops frequently, connection resets, have to restart router repeatedly, disconnects during video calls"},
            "3": {"name": "Billing & Plan Issues", "semantic_scope": "Wrong broadband bill, overcharged, plan upgrade/downgrade issues, FUP limit reached, auto-debit failed, payment not reflected, want to change plan, hidden charges, installation charges disputed, security deposit refund"},
            "4": {"name": "WiFi Signal Issues", "semantic_scope": "Weak WiFi signal, poor coverage, low bars, signal drops far from router, WiFi dead spots, need better in-home coverage, WiFi slow but wired is fast"},
            "5": {"name": "Router / Equipment Problems", "semantic_scope": "Router not working, WiFi router faulty, modem blinking red, ONT device issue, router overheating, need router replacement, firmware update problem, WiFi range too short, LAN port not working, equipment return"},
            "6": {"name": "Others", "semantic_scope": ""},
        },
    },
    "3": {
        "name": "DTH / Cable TV Services",
        "icon": "",
        "description": "Covers all issues related to Direct-To-Home television, cable TV, set-top boxes, and TV channel subscriptions.",
        "subprocesses": {
            "1": {"name": "Channel Not Working / Missing", "semantic_scope": "Channel not showing, channel removed from pack, channel black screen, paid channel not available, regional channel missing, HD channel not working, channel list changed, favorite channel gone"},
            "2": {"name": "Set-Top Box Issues", "semantic_scope": "Set-top box not turning on, remote not working, set-top box hanging/freezing, recording not working, set-top box overheating, display error on box, need set-top box replacement, software update stuck, box showing boot loop"},
            "3": {"name": "Billing & Subscription", "semantic_scope": "Wrong DTH bill, subscription expired, auto-renewal issue, pack change charges, NCF charges too high, channel added without consent, refund not received, wallet recharge failed, monthly charges incorrect"},
            "4": {"name": "Signal / Picture Quality", "semantic_scope": "No signal on TV, picture breaking/pixelating, rain causing signal loss, dish alignment needed, weak signal, audio out of sync, color distortion, signal loss at certain times, frozen picture, horizontal lines on TV"},
            "5": {"name": "Package / Plan Changes", "semantic_scope": "Want to change channel pack, upgrade to HD, add premium channels, downgrade plan, customize channel selection, regional pack addition, sports pack subscription, plan comparison, best value pack"},
            "6": {"name": "Others", "semantic_scope": ""},
        },
    },
    "4": {
        "name": "Landline / Fixed Line Services",
        "icon": "",
        "description": "Covers all issues related to traditional landline phone services, fixed-line connections, and wired telephone services.",
        "subprocesses": {
            "1": {"name": "No Dial Tone / Dead Line", "semantic_scope": "Landline not working, no dial tone, line dead, phone silent, no sound when picking up receiver, line suddenly stopped working, connection cut off, cable damaged"},
            "2": {"name": "Call Quality Issues (Noise / Echo)", "semantic_scope": "Static noise on landline, echo during calls, crackling sound, voice breaking, cross-connection hearing other conversations, humming noise, low volume on calls, distorted audio"},
            "3": {"name": "Billing & Charges", "semantic_scope": "Landline bill too high, calls charged incorrectly, wrong number dialed charges, rental overcharged, payment not updated, metered vs unlimited plan dispute, ISD charges on landline"},
            "4": {"name": "New Connection / Disconnection", "semantic_scope": "Want new landline connection, disconnection request, temporary suspension, connection shifting to new address, reconnection after disconnection, transfer of ownership"},
            "5": {"name": "Fault Repair Request", "semantic_scope": "Cable cut in area, junction box damaged, overhead wire fallen, underground cable fault, technician visit needed, repeated fault in same line, wet cable causing issues, maintenance request"},
            "6": {"name": "Others", "semantic_scope": ""},
        },
    },
    "5": {
        "name": "Enterprise / Business Solutions",
        "icon": "",
        "description": "Covers all issues related to business/corporate telecom solutions including leased lines, SLA-based services, bulk connections, cloud telephony, and managed network services.",
        "subprocesses": {
            "1": {"name": "SLA Breach / Service Downtime", "semantic_scope": "Service level agreement not met, uptime guarantee violated, business internet down, prolonged outage affecting business, compensation for downtime, SLA penalty claim, response time exceeded"},
            "2": {"name": "Leased Line / Dedicated Connection", "semantic_scope": "Leased line down, dedicated bandwidth not delivered, point-to-point link failure, MPLS circuit issue, last mile connectivity problem, fiber cut affecting leased line, jitter/latency on dedicated line"},
            "3": {"name": "Bulk / Corporate Plan Issues", "semantic_scope": "Corporate plan benefits not applied, bulk SIM management, employee connection issues, CUG (Closed User Group) problem, corporate billing discrepancy, group plan changes"},
            "4": {"name": "Cloud / VPN / MPLS Issues", "semantic_scope": "VPN tunnel down, MPLS network unreachable, cloud connectivity slow, SD-WAN issue, site-to-site VPN failure, enterprise cloud access problem, managed WiFi for office not working"},
            "5": {"name": "Technical Support Escalation", "semantic_scope": "Need senior technician, previous complaint not resolved, multiple complaints on same issue, want to escalate to manager, technical team not responding, critical issue needs immediate attention"},
            "6": {"name": "Others", "semantic_scope": ""},
        },
    },
}

# Re-init network_prompts now that TELECOM_MENU is available
network_prompts.init(client, DEPLOYMENT_NAME, TELECOM_MENU)
# Re-init network_diagnosis with full model references after all imports are resolved
network_diagnosis.init(client, DEPLOYMENT_NAME, db, {
    "User": User, "Ticket": Ticket, "ChatSession": ChatSession,
    "KpiData": KpiData, "TelecomSite": TelecomSite,
})


def get_subprocess_details(sector_key: str) -> str:
    sector = TELECOM_MENU[sector_key]
    details = []
    for k, v in sector["subprocesses"].items():
        if isinstance(v, dict) and v["name"] != "Others":
            details.append(f'SUBPROCESS: "{v["name"]}"\n  Typical issues: {v["semantic_scope"]}')
    return "\n\n".join(details)


def get_subprocess_name(sector_key: str, subprocess_key: str) -> str:
    sector = TELECOM_MENU.get(sector_key, {})
    sp = sector.get("subprocesses", {}).get(subprocess_key, {})
    if isinstance(sp, dict):
        return sp.get("name", "Others")
    return sp if isinstance(sp, str) else "Others"


def is_telecom_related(query, sector_name=None, subprocess_name=None):
    return network_prompts.is_telecom_related(query, sector_name, subprocess_name)

def identify_subprocess(query, sector_key):
    return network_prompts.identify_subprocess(query, sector_key)

def detect_greeting(text):
    return network_prompts.detect_greeting(text)

def classify_user_response(text):
    return network_prompts.classify_user_response(text)

def detect_language(text):
    return network_prompts.detect_language(text)

def _friendly_ai_error(err):
    return network_prompts._friendly_ai_error(err)




def generate_resolution(query, sector_name, subprocess_name, language):
    return network_prompts.generate_resolution(query, sector_name, subprocess_name, language)


def generate_single_solution(sector_name, subprocess_name, language, user_query="",
                              previous_solutions=None, attempt=1, original_query="",
                              diagnosis_summary="", sector_key=None,
                              billing_context=None, connection_context=None):
    """Routes to broadband or mobile prompt based on sector_key."""
    prev_block = ""
    if previous_solutions:
        prev_block = (
            "\n\nIMPORTANT: The following solutions have ALREADY been provided and did NOT work. "
            "Do NOT repeat them. Provide a DIFFERENT approach:\n"
            + "\n---\n".join(previous_solutions[-10:])
        )
    query_block = f'\n\nThe user described their specific issue as: "{user_query}"' if user_query else ""
    context_block = ""
    if original_query and original_query != user_query:
        context_block = f"\n\nOriginal issue description: \"{original_query}\"\nThe user's follow-up message is: \"{user_query}\""
    diagnosis_block = ""
    if diagnosis_summary:
        diagnosis_block = (
            f"\n\nSIGNAL DIAGNOSIS RESULTS: {diagnosis_summary}\n"
            "Use this diagnosis data to tailor your solution precisely. "
            "If RSRP < -100 dBm or SINR < 0 dB: cell-edge coverage — suggest band change, VoLTE/VoWiFi, SIM re-provisioning. "
            "If RSRP -100 to -85 dBm: moderate signal — APN reconfiguration, preferred network type, VoLTE toggle. "
            "If RSRP > -85 dBm and SINR > 5 dB: signal adequate — account/provisioning issues."
        )

    # Route to broadband prompt
    if broadband_prompts.is_broadband_sector(sector_key) or (sector_name and "broadband" in sector_name.lower()):
        system_prompt = broadband_prompts.build_broadband_prompt(
            subprocess_name=subprocess_name,
            language=language,
            attempt=attempt,
            billing_context=billing_context,
            connection_context=connection_context,
            query_block=query_block,
            context_block=context_block,
            prev_block=prev_block,
        )
    else:
        # Mobile / generic prompt
        system_prompt = network_prompts.build_mobile_system_prompt(
            sector_name=sector_name,
            subprocess_name=subprocess_name,
            language=language,
            attempt=attempt,
            query_block=query_block,
            context_block=context_block,
            diagnosis_block=diagnosis_block,
            prev_block=prev_block,
        )

    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": (
                    f"You are an expert telecom customer support agent. The user has an issue "
                    f"under the sector: '{sector_name}' and subprocess: '{subprocess_name}'.\n\n"
                    f"This is solution attempt #{attempt}.\n\n"
                    "IMPORTANT: Base this solution on BOTH the selected dropdown context "
                    "(sector/subprocess) and the user's latest query. "
                    "If they conflict, prioritize the latest query while staying within telecom scope.\n\n"
                    "Provide ONE focused, actionable solution at a time with steps that explain how to perform that action. "
                    "Be concise and specific. Do not provide multiple alternative solutions -- just one.\n"
                    "Do NOT include any URLs, links, or website references in your response.\n"
                    "STRICT RULE: Do NOT suggest the user to 'contact customer support', 'call customer care', "
                    "'raise a ticket', 'reach out to support', 'visit a service center', or any form of escalation. "
                    "Only provide self-help troubleshooting steps that the user can do on their own.\n"
                    "Acknowledge the issue briefly and give the steps."
                    + query_block
                    + context_block
                    + diagnosis_block
                    + prev_block +
                    f"\n\nIMPORTANT: Respond entirely in {language}. "
                    "Keep the tone professional, empathetic, and helpful."
                )},
                {"role": "user", "content": user_query if user_query else f"I have an issue with {subprocess_name} in {sector_name}"},
            ],
            temperature=0.5,
            max_tokens=500,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        return _friendly_ai_error(e)


def translate_text(text, target_language):
    return network_prompts.translate_text(text, target_language)


def generate_chat_summary(messages_list, sector_name, subprocess_name):
    return network_prompts.generate_chat_summary(messages_list, sector_name, subprocess_name)


def analyze_signal_screenshot(image_base64):
    return network_prompts.analyze_signal_screenshot(image_base64)



def generate_ref_number():
    ts = hex(int(time.time()))[2:].upper()
    rand = "".join(random.choices(string.ascii_uppercase + string.digits, k=4))
    return f"TC-{ts}-{rand}"


def validate_password(password):
    """Return an error string if the password is invalid, else None."""
    import re
    if len(password) < 7:
        return "Password must be at least 7 characters long"
    if not re.search(r"[A-Z]", password):
        return "Password must contain at least 1 uppercase letter"
    if not re.search(r"[!@#$%^&*()_+\-=\[\]{};':\"\\|,.<>/?`~]", password):
        return "Password must contain at least 1 special character"
    return None


# ─── Priority Ranking System ──────────────────────────────────────────────────

# Numeric rank for each priority level — higher is more urgent.
PRIORITY_RANK = {"critical": 4, "high": 3, "medium": 2, "low": 1}

# Customer tier → priority floor.
# A Platinum customer's ticket is always at least Critical regardless of content.
USER_TYPE_PRIORITY = {
    "platinum": "critical",
    "gold":     "high",
    "silver":   "medium",
    "bronze":   "low",
}

VALID_USER_TYPES = set(USER_TYPE_PRIORITY.keys())


# ─── Subprocess → base severity ──────────────────────────────────────────────
# The subprocess name is selected by the chatbot flow and is the most reliable
# signal for how urgent a ticket is.  Map every known subprocess to a base
# severity; keyword scanning then upgrades (never downgrades) from that base.
#
# MOBILE
#   Critical  – nothing at subprocess level (needs query-text signal)
#   High      – Network/Signal, SIM issues, Call/SMS failures, MNP
#   Medium    – Billing, Data plan, Roaming
#   Low       – Others
#
# BROADBAND
#   Critical  – (needs query-text: "no internet at all")
#   High      – Slow/No Connectivity, Frequent Disconnections, Router/Equipment
#   Medium    – Billing, New Connection, IP/DNS
#   Low       – Others
#
# DTH
#   Critical  – (needs query-text: "stb dead", "dish fallen")
#   High      – Signal/Picture Quality, Set-Top Box Issues
#   Medium    – Channel Missing, Billing, Package Changes
#   Low       – Others
#
# LANDLINE
#   Critical  – No Dial Tone / Dead Line
#   High      – Fault Repair Request
#   Medium    – Call Quality, Billing, New Connection
#   Low       – Others
#
# ENTERPRISE
#   Critical  – SLA Breach/Downtime, Leased Line, Cloud/VPN/MPLS
#   High      – Technical Support Escalation, Bulk/Corporate Issues
#   Medium    – Others within enterprise
#   Low       – (nothing — enterprise is never low by default)

SUBPROCESS_BASE_SEVERITY = {
    # ── Mobile ────────────────────────────────────────────────────────────────
    "Network / Signal Problems":          "high",
    "SIM Card & Activation":              "high",
    "Mobile Number Portability (MNP)":    "high",
    "Call / SMS Failures":                "high",
    "Data Plan & Recharge Issues":        "medium",
    "Billing & Payment Issues":           "medium",
    "International Roaming":              "medium",

    # ── Broadband ─────────────────────────────────────────────────────────────
    "Slow Speed / No Connectivity":       "high",
    "Frequent Disconnections":            "high",
    "Router / Equipment Problems":        "high",
    "IP Address / DNS Issues":            "medium",
    "Billing & Plan Issues":              "medium",
    "New Connection / Installation":      "medium",

    # ── DTH ───────────────────────────────────────────────────────────────────
    "Signal / Picture Quality":           "high",
    "Set-Top Box Issues":                 "high",
    "Channel Not Working / Missing":      "medium",
    "Billing & Subscription":             "medium",
    "Package / Plan Changes":             "low",

    # ── Landline ──────────────────────────────────────────────────────────────
    "No Dial Tone / Dead Line":           "critical",
    "Fault Repair Request":               "high",
    "Call Quality Issues (Noise / Echo)": "medium",
    "Billing & Charges":                  "medium",
    "New Connection / Disconnection":     "medium",

    # ── Enterprise ────────────────────────────────────────────────────────────
    "SLA Breach / Service Downtime":      "critical",
    "Leased Line / Dedicated Connection": "critical",
    "Cloud / VPN / MPLS Issues":          "critical",
    "Technical Support Escalation":       "high",
    "Bulk / Corporate Plan Issues":       "high",
}

# ─── Query-text upgrade keywords ─────────────────────────────────────────────
# These can RAISE severity above the subprocess base but never lower it.
_UPGRADE_TO_CRITICAL = [
    "no network", "complete outage", "area outage", "entire area",
    "no internet at all", "totally down", "stb dead", "box dead",
    "dish fallen", "dish damaged", "emergency", "urgent", "critical",
    "fraud", "unauthorized", "sla breach", "business down", "production down",
    "vpn down", "leased line down", "mpls down",
]
_UPGRADE_TO_HIGH = [
    "not working", "no signal", "dead", "down", "failed", "outage",
    "cannot call", "no internet", "no service", "unable to connect",
    "disconnecting", "router dead", "modem dead",
]
_UPGRADE_TO_MEDIUM = [
    "slow", "intermittent", "billing", "wrong charge", "refund",
    "overcharged", "pixelat", "delay", "quality",
]


def _detect_severity_llm(query_text: str, subprocess_name: str, sector_name: str) -> dict | None:
    """
    Uses Azure OpenAI to semantically classify ticket severity.
    Returns {"severity": "...", "reasoning": "..."} or None on failure.
    """
    if not query_text or not query_text.strip():
        return None

    context_parts = []
    if sector_name:
        context_parts.append(f"Telecom sector: {sector_name}")
    if subprocess_name:
        context_parts.append(f"Issue subcategory: {subprocess_name}")
    context_block = "\n".join(context_parts)

    system_prompt = (
        "You are a severity classifier for a telecom customer complaint ticketing system.\n\n"
        "Given the customer's complaint and its telecom context, determine the urgency/severity level.\n\n"
        "SEVERITY LEVELS (choose exactly one):\n"
        "- critical: Complete service outage, safety/fraud/emergency, SLA breach affecting business, "
        "total loss of connectivity, equipment destroyed/fallen, production/business down\n"
        "- high: Service significantly impaired but not total outage — e.g. frequent drops, "
        "SIM/device not working, call/SMS failures, major equipment fault, porting failures\n"
        "- medium: Degraded service quality — e.g. slow speeds, intermittent issues, billing disputes, "
        "wrong charges, plan/package problems, minor quality issues\n"
        "- low: General inquiries, informational requests, minor cosmetic issues, plan comparisons, "
        "feature questions, non-urgent requests\n\n"
        f"CONTEXT:\n{context_block}\n\n"
        "Analyze the customer's complaint semantically — consider the INTENT and IMPACT, "
        "not just keyword matches.\n\n"
        'Respond with ONLY valid JSON: {"severity": "<level>", "reasoning": "<one sentence>"}'
    )

    try:
        response = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": query_text},
            ],
            temperature=0,
            max_tokens=80,
            timeout=5,
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()
        result = json.loads(raw)
        severity = result.get("severity", "").lower().strip()
        if severity not in PRIORITY_RANK:
            print(f"[Severity-LLM] Invalid severity '{severity}' returned, falling back")
            return None
        return {"severity": severity, "reasoning": result.get("reasoning", "")}
    except Exception as e:
        print(f"[Severity-LLM] Failed: {e}")
        return None


def _detect_severity(query_text: str, subprocess_name: str, sector_name: str = "") -> str:
    """
    Returns one of: critical / high / medium / low.

    Strategy:
    1. Try LLM-based semantic classification first.
    2. If LLM fails, fall back to rule-based logic.
    """
    # ── Attempt 1: LLM-based semantic classification ──
    llm_result = _detect_severity_llm(query_text, subprocess_name, sector_name)
    if llm_result is not None:
        print(f"[Severity] LLM classified as '{llm_result['severity']}' — {llm_result.get('reasoning', '')}")
        return llm_result["severity"]

    # ── Attempt 2: Rule-based fallback ──
    print("[Severity] Using rule-based fallback")
    base = SUBPROCESS_BASE_SEVERITY.get((subprocess_name or "").strip(), "low")
    text = (query_text or "").lower()

    if any(w in text for w in _UPGRADE_TO_CRITICAL):
        text_sev = "critical"
    elif any(w in text for w in _UPGRADE_TO_HIGH):
        text_sev = "high"
    elif any(w in text for w in _UPGRADE_TO_MEDIUM):
        text_sev = "medium"
    else:
        text_sev = "low"

    return base if PRIORITY_RANK.get(base, 1) >= PRIORITY_RANK.get(text_sev, 1) else text_sev


def _compute_final_priority(user_type: str | None, severity: str) -> str:
    """
    Final ticket priority = the higher of (user-type floor, issue severity).

    Examples:
      Platinum user  + medium severity  → critical   (user floor wins)
      Bronze user    + critical severity → critical   (severity wins)
      Gold user      + high severity     → high       (both equal)
      Silver user    + low severity      → medium     (user floor wins)
    """
    user_floor = USER_TYPE_PRIORITY.get((user_type or "bronze").lower(), "low")
    rank_floor = PRIORITY_RANK.get(user_floor, 1)
    rank_sev   = PRIORITY_RANK.get(severity, 1)
    return user_floor if rank_floor >= rank_sev else severity


# Keep backward-compatible alias so any other callers still work.
def auto_assign_priority(query_text, subprocess_name):
    """Simple priority assignment based on keywords."""
    text = (query_text + " " + subprocess_name).lower()
    if any(w in text for w in ["urgent", "critical", "emergency", "business down", "sla breach", "escalat"]):
        return "critical"
    if any(w in text for w in ["not working", "failed", "no signal", "dead", "down", "outage"]):
        return "high"
    if any(w in text for w in ["slow", "intermittent", "billing", "wrong charge", "refund"]):
        return "medium"
    return "low"


# ─── Agent Routing Constants & Helpers ─────────────────────────────────────────

SECTOR_TO_DOMAIN = {
    "Mobile Services (Prepaid / Postpaid)": "mobile",
    "Broadband / Internet Services": "broadband",
    "DTH / Cable TV Services": "dth",
    "Landline / Fixed Line Services": "landline",
    "Enterprise / Business Solutions": "enterprise",
}

VALID_EXPERT_DOMAINS = {"mobile", "broadband", "dth", "landline", "enterprise", "fiber"}

# Maps complaint subcategories → agent expertise for routing
SUBPROCESS_TO_EXPERTISE = {
    # Mobile subcategories
    "Billing & Payment Issues": "GENERAL",
    "Network / Signal Problems": "NETWORK_RF",
    "SIM Card & Activation": "GENERAL",
    "Data Plan & Recharge Issues": "GENERAL",
    "International Roaming": "GENERAL",
    "Mobile Number Portability (MNP)": "GENERAL",
    "Call / SMS Failures": "VoLTE",
    # Broadband subcategories
    "Slow Speed / No Connectivity": "NETWORK_OPTIMIZATION",
    "Frequent Disconnections": "NETWORK_RF",
    "Billing & Plan Issues": "GENERAL",
    "New Connection / Installation": "GENERAL",
    "Router / Equipment Problems": "GENERAL",
    # DTH subcategories
    "Channel Not Working / Missing": "GENERAL",
    "Set-Top Box Issues": "GENERAL",
    "Billing & Subscription": "GENERAL",
    "Signal / Picture Quality": "NETWORK_RF",
    "Package / Plan Changes": "GENERAL",
    # Landline subcategories
    "No Dial Tone / Dead Line": "NETWORK_RF",
    "Call Quality Issues (Noise / Echo)": "VoLTE",
    "Billing & Charges": "GENERAL",
    "New Connection / Disconnection": "GENERAL",
    "Fault Repair Request": "NETWORK_RF",
    # Enterprise subcategories
    "SLA Breach / Service Downtime": "NETWORK_OPTIMIZATION",
    "Leased Line / Dedicated Connection": "TRANSPORT",
    "Bulk / Corporate Plan Issues": "GENERAL",
    "Cloud / VPN / MPLS Issues": "TRANSPORT",
    "Technical Support Escalation": "NETWORK_OPTIMIZATION",
}


def _resolve_expertise(subprocess_name: str) -> str:
    """Map a complaint subcategory to agent expertise."""
    return SUBPROCESS_TO_EXPERTISE.get(subprocess_name or "", "GENERAL")


def _resolve_ticket_domain(sector_name: str) -> str:
    """Return the domain slug for a sector name, defaulting to 'mobile'."""
    return SECTOR_TO_DOMAIN.get(sector_name or "", "mobile")


def _open_ticket_count(agent_id: int) -> int:
    """Count open (pending/in_progress) tickets assigned to an agent."""
    return Ticket.query.filter(
        Ticket.assigned_to == agent_id,
        Ticket.status.in_(["pending", "in_progress"]),
    ).count()


def _extract_city_from_address(address: str | None) -> str:
    """
    Extract a clean city name from a free-text address string.
    Scans all known expert cities and returns the first one found as a
    substring (case-insensitive).
    """
    if not address:
        return ""
    address_lower = address.strip().lower()
    cities = {
        (u.location or "").strip().lower()
        for u in User.query.filter_by(role="human_agent").with_entities(User.location).all()
        if u.location
    }
    for city in sorted(cities, key=len, reverse=True):
        if city and city in address_lower:
            return city
    return address_lower


def _find_best_expert(domain: str, city: str | None, priority: str = "low", expertise: str = None) -> "User | None":
    """
    Agent routing with 4 factors:
    1. Domain = complaint category (mobile/broadband/dth/landline/enterprise)
    2. Expertise = complaint subcategory (NETWORK_RF/VoLTE/TRANSPORT/GENERAL etc.)
    3. Location = nearest to customer (for network issues)
    4. Capacity = agents under bandwidth limit get priority

    Tier order:
      1. Same domain + same expertise + same city + under capacity
      2. Same domain + same expertise + under capacity
      3. Same domain + same city + under capacity
      4. Same domain + under capacity
      5. Same domain (ignore capacity for non-urgent)
      6. Any agent under capacity (global fallback)
      7. Any agent (last resort for non-urgent)
    """
    city_norm = _extract_city_from_address(city)
    expertise_norm = (expertise or "").strip().upper()

    all_agents = User.query.filter_by(role="human_agent").all()
    if not all_agents:
        return None

    domain_agents = [a for a in all_agents if (a.domain or "").lower() == (domain or "").lower()]

    def _under_capacity(agent):
        return _open_ticket_count(agent.id) < (agent.bandwidth_capacity or 10)

    def _same_city(agent):
        return city_norm and (agent.location or "").strip().lower() == city_norm

    def _same_expertise(agent):
        return expertise_norm and (getattr(agent, 'expertise', '') or "").strip().upper() == expertise_norm

    def _load(agent):
        return _open_ticket_count(agent.id)

    # Tier 1 – same domain + same expertise + same city + under capacity
    if expertise_norm and city_norm:
        pool = [a for a in domain_agents if _same_expertise(a) and _same_city(a) and _under_capacity(a)]
        if pool: return min(pool, key=_load)

    # Tier 2 – same domain + same expertise + under capacity
    if expertise_norm:
        pool = [a for a in domain_agents if _same_expertise(a) and _under_capacity(a)]
        if pool: return min(pool, key=_load)

    # Tier 3 – same domain + same city + under capacity
    if city_norm:
        pool = [a for a in domain_agents if _same_city(a) and _under_capacity(a)]
        if pool: return min(pool, key=_load)

    # Tier 4 – same domain + under capacity
    pool = [a for a in domain_agents if _under_capacity(a)]
    if pool: return min(pool, key=_load)

    # Tier 5 – same domain, ignore capacity (for non-urgent only)
    is_urgent = PRIORITY_RANK.get(priority, 1) >= PRIORITY_RANK.get("high", 3)
    if not is_urgent and domain_agents:
        return min(domain_agents, key=_load)

    # Tier 6 – any agent, under capacity
    under = [a for a in all_agents if _under_capacity(a)]
    if under: return min(under, key=_load)

    # Tier 7 – any agent (last resort for non-urgent)
    if not is_urgent:
        return min(all_agents, key=_load)

    return None


def _manager_priority_load(manager_id: int) -> dict:
    """Return a breakdown of open tickets assigned to a manager, keyed by priority."""
    OPEN_STATUSES = ["pending", "in_progress", "manager_escalated"]
    counts = {"critical": 0, "high": 0, "medium": 0, "low": 0, "total": 0}
    tickets = Ticket.query.filter(
        Ticket.assigned_to == manager_id,
        Ticket.status.in_(OPEN_STATUSES),
    ).with_entities(Ticket.priority).all()
    for (p,) in tickets:
        level = (p or "low").lower()
        counts[level] = counts.get(level, 0) + 1
        counts["total"] += 1
    return counts


def _find_best_manager(priority: str) -> "User | None":
    """
    Assign an escalated ticket to the most suitable manager using
    priority-driven load balancing.
    """
    managers = User.query.filter_by(role="manager").all()
    if not managers:
        return None
    if len(managers) == 1:
        return managers[0]

    level = (priority or "low").lower()
    loads = {m.id: _manager_priority_load(m.id) for m in managers}

    return min(managers, key=lambda m: (loads[m.id].get(level, 0), loads[m.id]["total"]))


# Broadband routes (/api/broadband/*) are registered via broadband_prompts.register_routes(app)
# See broadband_prompts.py



# AUTH ROUTES


@app.route("/api/auth/register", methods=["POST"])
def register():
    data = request.json
    name = data.get("name", "").strip()
    email = data.get("email", "").strip().lower()
    phone_number = data.get("phone_number", "").strip()  # ← NEW
    password = data.get("password", "")

    if not name or not email or not password:
        return jsonify({"error": "Name, email, and password are required"}), 400

    pw_err = validate_password(password)
    if pw_err:
        return jsonify({"error": pw_err}), 400

    # ← NEW: Validate phone number
    if not phone_number:
        return jsonify({"error": "Phone number is required"}), 400

    if User.query.filter_by(email=email).first():
        return jsonify({"error": "Email already registered"}), 409

    user_type = (data.get("user_type") or "bronze").strip().lower()
    if user_type not in VALID_USER_TYPES:
        user_type = "bronze"

    user = User(name=name, email=email, phone_number=phone_number, role="customer", user_type=user_type)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()

    token = create_access_token(identity=str(user.id))
    return jsonify({"token": token, "user": user.to_dict()}), 201


@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.json
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")

    user = User.query.filter_by(email=email).first()
    if not user or not user.check_password(password):
        return jsonify({"error": "Invalid email or password"}), 401

    token = create_access_token(identity=str(user.id))
    return jsonify({"token": token, "user": user.to_dict()})


@app.route("/api/auth/me", methods=["GET"])
@jwt_required()
def get_me():
    user = db.session.get(User, int(get_jwt_identity()))
    if not user:
        return jsonify({"error": "User not found"}), 404
    data = request.json or {}
    if 'name' in data and data['name'].strip():
        user.name = data['name'].strip()
    if 'phone_number' in data:
        user.phone_number = (data['phone_number'] or '').strip() or None
    if 'email' in data and data['email'].strip():
        new_email = data['email'].strip().lower()
        if new_email != user.email:
            existing = User.query.filter_by(email=new_email).first()
            if existing:
                return jsonify({"error": "Email already in use"}), 409
            user.email = new_email
    db.session.commit()
    return jsonify({"user": user.to_dict()})


@app.route("/api/user/settings", methods=["PUT"])
@jwt_required()
def update_user_settings():
    """Update user profile (name, email, phone)."""
    user = db.session.get(User, int(get_jwt_identity()))
    if not user:
        return jsonify({"error": "User not found"}), 404
    data = request.json or {}
    if "name" in data and data["name"].strip():
        user.name = data["name"].strip()
    if "phone_number" in data:
        user.phone_number = (data["phone_number"] or "").strip() or None
    if "email" in data and data["email"].strip():
        new_email = data["email"].strip().lower()
        if new_email != user.email:
            existing = User.query.filter_by(email=new_email).first()
            if existing:
                return jsonify({"error": "Email already in use"}), 409
            user.email = new_email
    db.session.commit()
    return jsonify({"user": user.to_dict()})


@app.route("/api/user/password", methods=["PUT"])
@jwt_required()
def update_user_password():
    """Change user password."""
    user = db.session.get(User, int(get_jwt_identity()))
    if not user:
        return jsonify({"error": "User not found"}), 404
    data = request.json or {}
    current_password = data.get("current_password", "")
    new_password = data.get("new_password", "")
    if not current_password or not new_password:
        return jsonify({"error": "Both current and new password are required"}), 400
    if not user.check_password(current_password):
        return jsonify({"error": "Current password is incorrect"}), 401
    if len(new_password) < 6:
        return jsonify({"error": "New password must be at least 6 characters"}), 400
    user.set_password(new_password)
    db.session.commit()
    return jsonify({"message": "Password updated successfully"})


# CHATBOT ROUTES 
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/menu", methods=["GET"])
def get_menu():
    menu = {}
    for key, sector in TELECOM_MENU.items():
        menu[key] = {"name": sector["name"], "icon": sector["icon"]}
    return jsonify({"menu": menu})


@app.route("/api/subprocesses", methods=["POST"])
def get_subprocesses():
    data = request.json
    sector_key = data.get("sector_key")
    language = data.get("language", "English")
    if sector_key not in TELECOM_MENU:
        return jsonify({"error": "Invalid sector"}), 400
    sector = TELECOM_MENU[sector_key]
    subprocesses = {}
    for k, v in sector["subprocesses"].items():
        subprocesses[k] = v["name"] if isinstance(v, dict) else v
    if language.lower() not in ("english", "en"):
        translated = {}
        for k, v in subprocesses.items():
            translated[k] = translate_text(v, language)
        subprocesses = translated
    return jsonify({"sector_name": sector["name"], "subprocesses": subprocesses})


@app.route("/api/resolve", methods=["POST"])
def resolve_complaint():
    data = request.json
    query = data.get("query", "").strip()
    sector_key = data.get("sector_key")
    subprocess_key = data.get("subprocess_key")
    selected_subprocess = data.get("selected_subprocess", "").strip()
    language = data.get("language", "English")
    if not query:
        return jsonify({"error": "Please enter your complaint/query."}), 400
    sector = TELECOM_MENU.get(sector_key, {})
    sector_name = sector.get("name", "Telecom")
    subprocess_name = selected_subprocess or get_subprocess_name(sector_key, subprocess_key)
    if not is_telecom_related(query, sector_name=sector_name, subprocess_name=subprocess_name):
        msg = (
            "I'm sorry, but I can only assist with **telecom-related** complaints. "
            "Your query doesn't appear to be telecom-related. Please try again."
        )
        translated_msg = translate_text(msg, language)
        return jsonify({"resolution": translated_msg, "is_telecom": False})
    if subprocess_name == "Others":
        subprocess_name = identify_subprocess(query, sector_key)
    resolution = generate_resolution(query, sector_name, subprocess_name, language)
    return jsonify({
        "resolution": resolution,
        "is_telecom": True,
        "identified_subprocess": subprocess_name,
    })


@app.route("/api/resolve-step", methods=["POST"])
def resolve_step():
    """Generate a single solution step. Used in the iterative resolution flow."""
    data = request.json
    sector_key = data.get("sector_key")
    subprocess_key = data.get("subprocess_key")
    selected_subprocess = data.get("selected_subprocess", "").strip()
    user_query = data.get("query", "").strip()
    language = data.get("language", "English")
    previous_solutions = data.get("previous_solutions", [])
    attempt = data.get("attempt", 1)
    original_query = data.get("original_query", "")
    diagnosis_summary = data.get("diagnosis_summary", "")

    # Broadband diagnostic context — passed from frontend when billing/connection
    # check results are available (stored in bb_* columns on chat_sessions)
    billing_context = data.get("billing_context", None)
    connection_context = data.get("connection_context", None)

    sector = TELECOM_MENU.get(sector_key, {})
    sector_name = sector.get("name", "Telecom")
    subprocess_name = selected_subprocess or get_subprocess_name(sector_key, subprocess_key)

    # If user provided a query, check if it's telecom-related
    if user_query:
        if not is_telecom_related(user_query, sector_name=sector_name, subprocess_name=subprocess_name):
            msg = (
                "I'm sorry, but I can only assist with **telecom-related** complaints. "
                "Your query doesn't appear to be telecom-related. Please try again."
            )
            translated_msg = translate_text(msg, language)
            return jsonify({"resolution": translated_msg, "is_telecom": False})

    solution = generate_single_solution(
        sector_name, subprocess_name, language,
        user_query=user_query,
        previous_solutions=previous_solutions,
        attempt=attempt,
        original_query=original_query,
        diagnosis_summary=diagnosis_summary,
        sector_key=sector_key,
        billing_context=billing_context,
        connection_context=connection_context,
    )
    return jsonify({
        "resolution": solution,
        "is_telecom": True,
        "attempt": attempt,
    })


@app.route("/api/detect-language", methods=["POST"])
def detect_lang():
    data = request.json
    text = data.get("text", "")
    language = detect_language(text)
    return jsonify({"language": language})


@app.route("/api/detect-greeting", methods=["POST"])
def detect_greeting_route():
    data = request.json
    text = data.get("text", "")
    is_greeting = detect_greeting(text)
    return jsonify({"is_greeting": is_greeting})


@app.route("/api/classify-response", methods=["POST"])
def classify_response_route():
    """Classify user response: satisfied? mentions signal/network?"""
    data = request.json
    text = data.get("text", "")
    result = classify_user_response(text)
    return jsonify(result)


# ═══════════════════════════════════════════════════════════════════════════════
# CHAT SESSION ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/chat/session", methods=["POST"])
@jwt_required()
def create_chat_session():
    user_id = int(get_jwt_identity())

    # ── Clean up empty/abandoned sessions before creating a new one ──────────
    # Delete active sessions that have zero messages and are older than 1 hour
    from datetime import timedelta
    stale_cutoff = datetime.now(timezone.utc) - timedelta(hours=1)
    stale_sessions = ChatSession.query.filter(
        ChatSession.user_id == user_id,
        ChatSession.status == "active",
        ChatSession.last_message_at < stale_cutoff,
    ).all()
    for s in stale_sessions:
        msg_count = ChatMessage.query.filter_by(session_id=s.id).count()
        if msg_count == 0:
            db.session.delete(s)
    db.session.flush()
    # ─────────────────────────────────────────────────────────────────────────

    session = ChatSession(user_id=user_id, status="active")
    db.session.add(session)
    db.session.commit()
    return jsonify({"session": session.to_dict()}), 201


@app.route("/api/admin/cleanup-sessions", methods=["POST"])
@jwt_required()
def cleanup_old_sessions():
    """Admin route: delete resolved sessions older than 30 days."""
    from datetime import timedelta
    cutoff = datetime.now(timezone.utc) - timedelta(days=30)
    old_sessions = ChatSession.query.filter(
        ChatSession.status == "resolved",
        ChatSession.resolved_at < cutoff,
    ).all()
    count = 0
    for s in old_sessions:
        ChatMessage.query.filter_by(session_id=s.id).delete()
        db.session.delete(s)
        count += 1
    db.session.commit()
    return jsonify({"deleted": count, "message": f"Deleted {count} old resolved sessions"})


@app.route("/api/chat/session/<int:session_id>/message", methods=["POST"])
@jwt_required()
def add_chat_message(session_id):
    user_id = int(get_jwt_identity())
    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404

    data = request.json
    msg = ChatMessage(
        session_id=session_id,
        sender=data.get("sender", "user"),
        content=data.get("content", ""),
        content_json=data.get("payload"),   # ← saves full card/UI payload for restoration
    )
    db.session.add(msg)

    # Update session metadata
    if data.get("sector_name"):
        session.sector_name = data["sector_name"]
    if data.get("subprocess_name"):
        session.subprocess_name = data["subprocess_name"]
    if data.get("query_text"):
        session.query_text = data["query_text"]
    if data.get("resolution"):
        session.resolution = data["resolution"]
    if data.get("language"):
        session.language = data["language"]
    if data.get("current_step"):            # ← saves step so resume works correctly
        session.current_step = data["current_step"]
    session.last_message_at = datetime.now(timezone.utc)  # ← keeps session timestamp fresh

    db.session.commit()

    # Push to WebSocket so the agent sees customer messages in real-time
    msg_dict = msg.to_dict()
    msg_dict["session_id"] = session_id
    socketio.emit("new_message", msg_dict, room=f"session_{session_id}")

    return jsonify({"message": msg_dict})

@app.route("/api/chat/session/<int:session_id>/location", methods=["POST"])
@jwt_required()
def save_session_location(session_id):
    """Save customer's GPS location for network signal complaints."""
    user_id = int(get_jwt_identity())
    session = ChatSession.query.get(session_id)

    if not session:
        return jsonify({"error": "Session not found"}), 404
    if session.user_id != user_id:
        return jsonify({"error": "Unauthorized"}), 403

    # Product decision: regardless of what the client sends, the customer's
    # location is always recorded as Phnom Penh — Chakto Mukh, Cambodia.
    # (The 3 nearest sites are still computed by formula from telecom_sites.)
    # Siem Reap — Svay Dankum commune (near Angkor Wat)
    DEFAULT_LATITUDE  = 13.3633
    DEFAULT_LONGITUDE = 103.8564
    session.latitude             = DEFAULT_LATITUDE
    session.longitude            = DEFAULT_LONGITUDE
    session.location_description = "Siem Reap, Svay Dankum"
    session.state_province       = "Siem Reap"
    session.country              = "Cambodia"

    db.session.commit()

    return jsonify({
        "message": "Location saved successfully",
        "latitude":  session.latitude,
        "longitude": session.longitude,
        "location_description": session.location_description,
        "state_province": session.state_province,
        "country": session.country,
    }), 200


@app.route("/api/chat/session/<int:session_id>/analyze-signal", methods=["POST"])
@jwt_required()
def analyze_signal(session_id):
    """Analyze a signal screenshot using Azure OpenAI Vision."""
    user_id = int(get_jwt_identity())
    session = ChatSession.query.get(session_id)

    if not session:
        return jsonify({"error": "Session not found"}), 404
    if session.user_id != user_id:
        return jsonify({"error": "Unauthorized"}), 403

    data = request.json
    image_base64 = data.get("image")

    if not image_base64:
        return jsonify({"error": "No image provided"}), 400

    # Limit ~5MB base64
    if len(image_base64) > 7_000_000:
        return jsonify({"error": "Image too large. Please upload a smaller screenshot."}), 400

    try:
        result = analyze_signal_screenshot(image_base64)

        # If signal is red (Poor), do not include nearest tower sites in user-facing diagnosis

        # Save diagnosis as a bot message for chat history
        diagnosis_text = (
            f"Signal Diagnosis Results: "
            f"RSRP: {result.get('rsrp', 'N/A')} dBm ({result.get('rsrp_label', 'Unknown')}), "
            f"SINR: {result.get('sinr', 'N/A')} dB ({result.get('sinr_label', 'Unknown')}), "
            f"Cell ID: {result.get('cell_id', 'N/A')}"
        )
        if result.get("nearest_sites"):
            diagnosis_text += "\n\nNearest Sites:\n"
            for s in result["nearest_sites"]:
                diagnosis_text += (
                    f"- {s['site_id']} | Status: {s['status']} | "
                    f"Alarm: {s['alarm']} | Distance: {s['distance_km']} km\n"
                )

        msg = ChatMessage(session_id=session_id, sender="bot", content=diagnosis_text)
        db.session.add(msg)
        session.diagnosis_ran = True
        db.session.commit()

        return jsonify({"diagnosis": result}), 200
    except Exception as e:
        return jsonify({"error": f"Failed to analyze screenshot: {str(e)}"}), 500


@app.route("/api/chat/session/<int:session_id>/resolve", methods=["PUT"])
@jwt_required()
def resolve_session(session_id):
    user_id = int(get_jwt_identity())
    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404

    session.status = "resolved"
    session.resolved_at = datetime.now(timezone.utc)
    db.session.commit()

    # Generate summary
    try:
        msgs = [{"sender": m.sender, "content": m.content} for m in session.messages]
        session.summary = generate_chat_summary(msgs, session.sector_name, session.subprocess_name)
        db.session.commit()
    except Exception:
        pass

    # ← NEW: Send WhatsApp message
    try:
        user = User.query.get(user_id)
        if user and user.phone_number:
            whatsapp_msg = format_chat_summary_for_whatsapp(session, user.name)
            result = send_whatsapp_message(user.phone_number, whatsapp_msg)
            if result["success"]:
                print(f"✅ WhatsApp sent to {user.phone_number}: {result['message_sid']}")
            else:
                print(f"⚠️  WhatsApp failed: {result['error']}")
    except Exception as e:
        print(f"⚠️  WhatsApp error: {e}")

    return jsonify({"session": session.to_dict(), "summary": session.summary})


@app.route("/api/chat/session/<int:session_id>", methods=["DELETE"])
@jwt_required()
def delete_chat_session(session_id):
    """Delete a chat session and its messages (customer clearing from dashboard)."""
    user_id = int(get_jwt_identity())
    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404
    if session.user_id != user_id:
        return jsonify({"error": "Unauthorized"}), 403
    # Don't allow deleting sessions that have tickets
    ticket = Ticket.query.filter_by(chat_session_id=session_id).first()
    if ticket:
        return jsonify({"error": "Cannot delete session with an active ticket"}), 409
    # Delete messages first, then session
    ChatMessage.query.filter_by(session_id=session_id).delete()
    Feedback.query.filter_by(chat_session_id=session_id).delete()
    db.session.delete(session)
    db.session.commit()
    return jsonify({"ok": True})


def send_ticket_assignment_email(agent, ticket, session):
    """Send a styled HTML email to the assigned agent with ticket details."""
    if not agent or not agent.email:
        return

    sla_deadline_str = ticket.sla_deadline.strftime('%B %d, %Y at %I:%M %p UTC') if ticket.sla_deadline else 'N/A'
    description_preview = (ticket.description[:300] + '...') if ticket.description and len(ticket.description) > 300 else (ticket.description or 'N/A')

    html_body = f"""
    <div style="font-family:'Segoe UI',Arial,sans-serif;max-width:600px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,0.08);">
        <div style="background:linear-gradient(135deg,#00338d 0%,#004fc4 100%);padding:24px 30px;text-align:center;">
            <h1 style="color:#fff;margin:0;font-size:20px;font-weight:600;">New Ticket Assigned</h1>
            <p style="color:rgba(255,255,255,0.8);margin:4px 0 0;font-size:13px;">A support ticket has been assigned to you</p>
        </div>
        <div style="padding:28px 30px;">
            <p style="margin:0 0 20px;font-size:15px;color:#1e293b;">Hello <strong>{agent.name}</strong>,</p>
            <p style="margin:0 0 20px;font-size:14px;color:#475569;line-height:1.6;">
                A new customer support ticket has been assigned to you. Please review the details below:
            </p>
            <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:20px;margin-bottom:20px;">
                <table style="width:100%;border-collapse:collapse;font-size:14px;">
                    <tr><td style="padding:8px 0;color:#94a3b8;width:140px;">Reference</td><td style="padding:8px 0;color:#1e293b;font-weight:600;">{ticket.reference_number}</td></tr>
                    <tr><td style="padding:8px 0;color:#94a3b8;">Category</td><td style="padding:8px 0;color:#1e293b;">{ticket.category or 'N/A'}</td></tr>
                    <tr><td style="padding:8px 0;color:#94a3b8;">Issue Type</td><td style="padding:8px 0;color:#1e293b;">{ticket.subcategory or 'N/A'}</td></tr>
                    <tr><td style="padding:8px 0;color:#94a3b8;">Priority</td><td style="padding:8px 0;color:#1e293b;font-weight:700;">{ticket.priority.upper() if ticket.priority else 'N/A'}</td></tr>
                    <tr><td style="padding:8px 0;color:#94a3b8;">SLA Hours</td><td style="padding:8px 0;color:#1e293b;">{ticket.sla_hours or 'N/A'} hours</td></tr>
                    <tr><td style="padding:8px 0;color:#94a3b8;">SLA Deadline</td><td style="padding:8px 0;color:#dc2626;font-weight:600;">{sla_deadline_str}</td></tr>
                </table>
            </div>
            <div style="border-left:3px solid #2563eb;background:#eff6ff;border-radius:0 10px 10px 0;padding:16px 20px;margin-bottom:20px;">
                <h3 style="color:#1e40af;font-size:13px;text-transform:uppercase;letter-spacing:0.08em;margin:0 0 10px;">Customer Description</h3>
                <p style="color:#1e293b;font-size:14px;line-height:1.7;margin:0;">{description_preview}</p>
            </div>
            <div style="background:#eff6ff;border:1px solid #93c5fd;border-radius:8px;padding:14px 18px;">
                <p style="margin:0;color:#1e40af;font-size:14px;font-weight:600;">Please review this ticket and begin working on it at your earliest convenience.</p>
            </div>
        </div>
        <div style="background:#f8fafc;border-top:1px solid #e2e8f0;padding:14px 30px;text-align:center;">
            <p style="color:#94a3b8;font-size:12px;margin:0;">Customer Handling &mdash; Automated Ticket Assignment</p>
        </div>
    </div>
    """

    try:
        msg = Message(
            subject=f"New Ticket Assigned - {ticket.reference_number}",
            recipients=[agent.email],
            html=html_body,
        )
        mail.send(msg)
        print(f"Assignment email sent to agent {agent.name} at {agent.email}")
    except Exception as e:
        print(f"Agent assignment email failed: {e}")


@app.route("/api/chat/session/<int:session_id>/escalate", methods=["PUT"])
@jwt_required()
def escalate_session(session_id):
    user_id = int(get_jwt_identity())
    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404

    session.status = "escalated"

    # Generate summary
    msgs = [{"sender": m.sender, "content": m.content} for m in session.messages]
    session.summary = generate_chat_summary(msgs, session.sector_name, session.subprocess_name)

    # Derive ticket domain and customer city
    ticket_domain = _resolve_ticket_domain(session.sector_name)
    ticket_city = session.location_description  # may be None

    # --- Priority calculation ---
    # 1. Issue severity from query keywords
    severity = _detect_severity(session.query_text, session.subprocess_name, session.sector_name)
    # 2. Customer tier floor (look up the session's owner)
    customer = db.session.get(User, user_id)
    customer_type = (customer.user_type or "bronze") if customer else "bronze"
    # 3. Final priority = max(severity, user-type floor)
    priority = _compute_final_priority(customer_type, severity)

    sla_targets = get_sla_targets()
    sla_h = sla_targets.get(priority, 48)
    now_utc = datetime.now(timezone.utc)
    sla_deadline = now_utc + timedelta(hours=sla_h)

    ticket_expertise = _resolve_expertise(session.subprocess_name)
    assigned_agent = _find_best_expert(ticket_domain, ticket_city, priority, expertise=ticket_expertise)

    # Enrich ticket description with signal diagnosis if available
    ticket_description = session.query_text or ""
    if session.diagnosis_ran:
        try:
            sig_msg = ChatMessage.query.filter_by(session_id=session_id, sender='bot').filter(
                ChatMessage.content.ilike('%rsrp%')
            ).order_by(ChatMessage.created_at.desc()).first()
            if sig_msg and sig_msg.content:
                ticket_description += f"\n\n[Signal Diagnosis: {sig_msg.content}]"
        except Exception:
            pass

    # Create ticket
    ref = generate_ref_number()
    ticket = Ticket(
        chat_session_id=session_id,
        user_id=user_id,
        reference_number=ref,
        category=session.sector_name,
        subcategory=session.subprocess_name,
        domain=ticket_domain,
        description=ticket_description,
        status="pending",
        severity=severity,
        priority=priority,
        assigned_to=assigned_agent.id if assigned_agent else None,
        sla_hours=sla_h,
        sla_deadline=sla_deadline,
    )
    db.session.add(ticket)
    db.session.commit()

    # Send WhatsApp message for ticket
    try:
        user = User.query.get(user_id)
        if user and user.phone_number:
            whatsapp_msg = format_ticket_alert_for_whatsapp(ticket, user.name, session)
            result = send_whatsapp_message(user.phone_number, whatsapp_msg)
            if result["success"]:
                print(f"✅ WhatsApp ticket alert sent to {user.phone_number}")
            else:
                print(f"⚠️  WhatsApp failed: {result['error']}")
    except Exception as e:
        print(f"⚠️  WhatsApp error: {e}")

    # Send email notification to assigned agent
    if assigned_agent:
        send_ticket_assignment_email(assigned_agent, ticket, session)

    agent_info = None
    if assigned_agent:
        agent_info = {
            "name": assigned_agent.name,
            "email": assigned_agent.email,
            "phone": assigned_agent.phone_number,
            "employee_id": assigned_agent.employee_id,
        }

    return jsonify({
        "session": session.to_dict(),
        "ticket": ticket.to_dict(),
        "assigned_agent": agent_info,
    })

@app.route("/api/chat/session/<int:session_id>/send-summary-email", methods=["POST"])
@jwt_required()
def send_summary_email(session_id):
    """Send chat summary to the user's email saved in DB."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404

    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404

    if session.user_id != user_id:
        return jsonify({"error": "Unauthorized"}), 403

    if not session.summary:
        return jsonify({"error": "No summary available for this session"}), 400

    # Build email HTML
    html_body = f"""
    <div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.08);">
        <!-- Header -->
        <div style="background: linear-gradient(135deg, #00338d 0%, #004fc4 100%); padding: 24px 30px; text-align: center;">
            <h1 style="color: #ffffff; margin: 0; font-size: 20px; font-weight: 600;">Customer Handling</h1>
            <p style="color: rgba(255,255,255,0.8); margin: 4px 0 0; font-size: 13px;">Chat Summary Report</p>
        </div>

        <!-- Body -->
        <div style="padding: 30px;">
            <p style="color: #1e293b; font-size: 15px; margin: 0 0 20px;">Hello <strong>{user.name}</strong>,</p>
            <p style="color: #64748b; font-size: 14px; line-height: 1.6; margin: 0 0 24px;">
                Here is the summary of your recent support chat session:
            </p>

            <!-- Session Details -->
            <div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 10px; padding: 20px; margin-bottom: 20px;">
                <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                    <tr>
                        <td style="padding: 8px 0; color: #94a3b8; width: 130px;">Session ID</td>
                        <td style="padding: 8px 0; color: #1e293b; font-weight: 500;">#{session.id}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0; color: #94a3b8;">Category</td>
                        <td style="padding: 8px 0; color: #1e293b; font-weight: 500;">{session.sector_name or 'N/A'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0; color: #94a3b8;">Issue Type</td>
                        <td style="padding: 8px 0; color: #1e293b; font-weight: 500;">{session.subprocess_name or 'N/A'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0; color: #94a3b8;">Status</td>
                        <td style="padding: 8px 0;">
                            <span style="background: {'#ecfdf5' if session.status == 'resolved' else '#fef3c7'}; color: {'#047857' if session.status == 'resolved' else '#b45309'}; padding: 3px 10px; border-radius: 12px; font-size: 12px; font-weight: 600;">
                                {session.status.upper()}
                            </span>
                        </td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0; color: #94a3b8;">Language</td>
                        <td style="padding: 8px 0; color: #1e293b; font-weight: 500;">{session.language or 'English'}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0; color: #94a3b8;">Date</td>
                        <td style="padding: 8px 0; color: #1e293b; font-weight: 500;">{session.created_at.strftime('%B %d, %Y at %I:%M %p') if session.created_at else 'N/A'}</td>
                    </tr>
                </table>
            </div>

            <!-- Summary -->
            <div style="border-left: 3px solid #10b981; background: #f0fdf4; border-radius: 0 10px 10px 0; padding: 16px 20px; margin-bottom: 20px;">
                <h3 style="color: #047857; font-size: 13px; text-transform: uppercase; letter-spacing: 0.08em; margin: 0 0 10px;">Chat Summary</h3>
                <p style="color: #1e293b; font-size: 14px; line-height: 1.7; margin: 0;">{session.summary}</p>
            </div>

            <!-- Your Query -->
            {f'''
            <div style="background: #eff6ff; border: 1px solid #bfdbfe; border-radius: 10px; padding: 16px 20px; margin-bottom: 20px;">
                <h3 style="color: #2563eb; font-size: 13px; text-transform: uppercase; letter-spacing: 0.08em; margin: 0 0 8px;">Your Query</h3>
                <p style="color: #1e293b; font-size: 14px; line-height: 1.6; margin: 0;">{session.query_text}</p>
            </div>
            ''' if session.query_text else ''}

            <!-- Ticket Info -->
            {f'''
            <div style="background: #fffbeb; border: 1px solid #fde68a; border-radius: 10px; padding: 16px 20px; margin-bottom: 20px;">
                <h3 style="color: #b45309; font-size: 13px; text-transform: uppercase; letter-spacing: 0.08em; margin: 0 0 8px;">Escalation Ticket</h3>
                <p style="color: #1e293b; font-size: 14px; margin: 0;">Reference: <strong>{session.ticket.reference_number}</strong></p>
            </div>
            ''' if session.ticket else ''}

            <p style="color: #64748b; font-size: 13px; line-height: 1.6; margin: 20px 0 0;">
                If you have further questions, feel free to start a new chat session anytime.
            </p>
        </div>

        <!-- Footer -->
        <div style="background: #f8fafc; border-top: 1px solid #e2e8f0; padding: 16px 30px; text-align: center;">
            <p style="color: #94a3b8; font-size: 12px; margin: 0;">Customer Handling &mdash; AI-Powered Support</p>
        </div>
    </div>
    """

    email_ok = False
    whatsapp_ok = False

    # Send email
    try:
        msg = Message(
            subject=f"Chat Summary - {session.sector_name or 'Telecom Support'} (Session #{session.id})",
            recipients=[user.email],
            html=html_body,
        )
        mail.send(msg)
        email_ok = True
    except Exception as e:
        print(f"⚠️  Email failed: {e}")

    # Send WhatsApp
    if user.phone_number:
        try:
            whatsapp_msg = format_chat_summary_for_whatsapp(session, user.name)
            result = send_whatsapp_message(user.phone_number, whatsapp_msg)
            if result["success"]:
                whatsapp_ok = True
                print(f"✅ WhatsApp summary sent to {user.phone_number}")
            else:
                print(f"⚠️  WhatsApp failed: {result['error']}")
        except Exception as e:
            print(f"⚠️  WhatsApp error: {e}")

    # Build response message
    parts = []
    if email_ok:
        parts.append(f"email ({user.email})")
    if whatsapp_ok:
        parts.append(f"WhatsApp ({user.phone_number})")

    if parts:
        return jsonify({"message": f"Summary sent to {' and '.join(parts)}", "email_sent": email_ok, "whatsapp_sent": whatsapp_ok}), 200
    else:
        return jsonify({"error": "Failed to send summary. Please try again later.", "email_sent": False, "whatsapp_sent": False}), 500


@app.route("/api/chat/session/<int:session_id>", methods=["GET"])
@jwt_required()
def get_chat_session(session_id):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    session = db.session.get(ChatSession, session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404
    # Allow the session owner, managers, CTOs, and admins
    if session.user_id != user_id and (not user or user.role not in ("manager", "cto", "admin", "human_agent")):
        return jsonify({"error": "Unauthorized"}), 403
    return jsonify({
        "session": session.to_dict(),
        "messages": [m.to_dict() for m in session.messages],
    })


@app.route("/api/chat/session/<int:session_id>/status", methods=["GET"])
@jwt_required()
def get_session_status(session_id):
    """Lightweight poll endpoint: return session status + latest bot message."""
    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404
    ticket = Ticket.query.filter_by(chat_session_id=session_id).order_by(Ticket.created_at.desc()).first()
    # Latest bot message (so the chatbot can show it)
    latest_bot_msg = None
    for m in reversed(session.messages):
        if m.sender == "bot":
            latest_bot_msg = m.content
            break
    return jsonify({
        "session_status": session.status,
        "ticket_status": ticket.status if ticket else None,
        "ticket_reference": ticket.reference_number if ticket else None,
        "latest_bot_message": latest_bot_msg,
    })


@app.route("/api/chat/session/<int:session_id>/presence", methods=["POST"])
@jwt_required()
def session_presence(session_id):
    """Update customer presence for a chat session."""
    user_id = int(get_jwt_identity())
    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404
    if session.user_id != user_id:
        return jsonify({"error": "Unauthorized"}), 403
    data = request.json or {}
    session.customer_present = bool(data.get("present"))
    db.session.commit()
    return jsonify({"customer_present": session.customer_present})


# ═══════════════════════════════════════════════════════════════════════════════
# CUSTOMER ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/customer/dashboard", methods=["GET"])
@jwt_required()
def customer_dashboard():
    user_id = int(get_jwt_identity())

    # Single query for all chat stats
    chat_stats = db.session.query(
        ChatSession.status, db.func.count(ChatSession.id)
    ).filter_by(user_id=user_id).group_by(ChatSession.status).all()
    stat_map = dict(chat_stats)
    total = sum(stat_map.values())

    pending_tickets = Ticket.query.filter_by(user_id=user_id).filter(
        Ticket.status.in_(["pending", "in_progress"])
    ).count()

    # Fetch recent sessions with user eagerly loaded + feedback ratings in one join
    recent_sessions = db.session.query(ChatSession, Feedback.rating).outerjoin(
        Feedback, db.and_(
            Feedback.chat_session_id == ChatSession.id,
            Feedback.user_id == user_id,
        )
    ).filter(
        ChatSession.user_id == user_id
    ).options(
        joinedload(ChatSession.user)
    ).order_by(ChatSession.created_at.desc()).all()

    sessions_data = []
    for s, rating in recent_sessions:
        sd = s.to_dict()
        sd["rating"] = rating
        sessions_data.append(sd)

    return jsonify({
        "stats": {
            "total_chats": total,
            "resolved": stat_map.get("resolved", 0),
            "escalated": stat_map.get("escalated", 0),
            "active": stat_map.get("active", 0),
            "pending_tickets": pending_tickets,
        },
        "recent_sessions": sessions_data,
    })


@app.route("/api/customer/active-session", methods=["GET"])
@jwt_required()
def customer_active_session():
    """Return the most recent active chat session for the current user, with messages."""
    user_id = int(get_jwt_identity())
    session = ChatSession.query.filter_by(user_id=user_id, status="active").order_by(
        ChatSession.created_at.desc()
    ).first()
    if not session:
        return jsonify({"session": None, "messages": []})
    return jsonify({
        "session": session.to_dict(),
        "messages": [m.to_dict() for m in session.messages],
    })


@app.route("/api/customer/pending-feedback", methods=["GET"])
@jwt_required()
def customer_pending_feedback():
    """Return ALL resolved sessions (with tickets) that the user hasn't given feedback for."""
    user_id = int(get_jwt_identity())
    # Subquery: session IDs that already have feedback from this user
    feedback_session_ids = db.session.query(Feedback.chat_session_id).filter(
        Feedback.user_id == user_id,
        Feedback.chat_session_id.isnot(None),
    ).subquery()

    # Only sessions that were escalated and resolved (i.e. have a ticket)
    sessions = db.session.query(ChatSession, Ticket).join(
        Ticket, Ticket.chat_session_id == ChatSession.id
    ).filter(
        ChatSession.user_id == user_id,
        ChatSession.status == "resolved",
        ~ChatSession.id.in_(feedback_session_ids),
    ).order_by(ChatSession.resolved_at.desc()).all()

    result = []
    for s, t in sessions:
        sd = s.to_dict()
        sd["ticket_id"] = t.id
        sd["ticket_priority"] = t.priority
        sd["assigned_agent"] = t.assignee.name if t.assignee else ""
        result.append(sd)

    return jsonify({
        "sessions": result,
    })


@app.route("/api/customer/sessions", methods=["GET"])
@jwt_required()
def customer_sessions():
    user_id = int(get_jwt_identity())
    sessions = ChatSession.query.filter_by(user_id=user_id).order_by(
        ChatSession.created_at.desc()
    ).all()
    return jsonify({"sessions": [s.to_dict() for s in sessions]})


@app.route("/api/customer/tickets", methods=["GET"])
@jwt_required()
def customer_tickets():
    user_id = int(get_jwt_identity())
    tickets = Ticket.query.filter_by(user_id=user_id).order_by(
        Ticket.created_at.desc()
    ).all()
    return jsonify({"tickets": [t.to_dict() for t in tickets]})


# ═══════════════════════════════════════════════════════════════════════════════
# FEEDBACK ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/feedback", methods=["POST"])
@jwt_required()
def submit_feedback():
    user_id = int(get_jwt_identity())
    data = request.json
    chat_session_id = data.get("chat_session_id")
    if chat_session_id:
        existing = Feedback.query.filter_by(
            user_id=user_id,
            chat_session_id=chat_session_id,
        ).first()
        if existing:
            return jsonify({"error": "Feedback already submitted"}), 409
    fb = Feedback(
        user_id=user_id,
        chat_session_id=chat_session_id,
        rating=data.get("rating", 0),
        comment=data.get("comment", ""),
    )
    db.session.add(fb)
    db.session.commit()
    return jsonify({"feedback": fb.to_dict()}), 201


@app.route("/api/feedback/list", methods=["GET"])
@jwt_required()
def list_feedback():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role == "customer":
        feedbacks = Feedback.query.filter_by(user_id=user_id).order_by(Feedback.created_at.desc()).all()
    else:
        feedbacks = Feedback.query.order_by(Feedback.created_at.desc()).all()
    return jsonify({"feedbacks": [f.to_dict() for f in feedbacks]})


# ═══════════════════════════════════════════════════════════════════════════════
# MANAGER / CTO ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/manager/dashboard", methods=["GET"])
@jwt_required()
def manager_dashboard():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    # Chat stats — single GROUP BY
    chat_stats = db.session.query(
        ChatSession.status, db.func.count(ChatSession.id)
    ).group_by(ChatSession.status).all()
    chat_map = dict(chat_stats)
    total_chats = sum(chat_map.values())
    resolved_chats = chat_map.get("resolved", 0)
    escalated_chats = chat_map.get("escalated", 0)
    active_chats = chat_map.get("active", 0)

    # Ticket stats — single GROUP BY for status
    ticket_stats = db.session.query(
        Ticket.status, db.func.count(Ticket.id)
    ).group_by(Ticket.status).all()
    ts_map = dict(ticket_stats)
    total_tickets = sum(ts_map.values())
    pending_tickets = ts_map.get("pending", 0)
    in_progress_tickets = ts_map.get("in_progress", 0)
    resolved_tickets = ts_map.get("resolved", 0)
    escalated_tickets = ts_map.get("escalated", 0)
    manager_escalated_tickets = ts_map.get("manager_escalated", 0)

    # Critical/high pending tickets — single query
    urgent_stats = db.session.query(
        Ticket.priority, db.func.count(Ticket.id)
    ).filter_by(status="pending").filter(
        Ticket.priority.in_(["critical", "high"])
    ).group_by(Ticket.priority).all()
    urgent_map = dict(urgent_stats)
    critical_tickets = urgent_map.get("critical", 0)
    high_tickets = urgent_map.get("high", 0)

    # Feedback — single aggregation query
    fb_agg = db.session.query(
        db.func.count(Feedback.id),
        db.func.avg(sql_case((Feedback.rating > 0, Feedback.rating))),
        db.func.sum(sql_case((Feedback.rating >= 4, 1), else_=0)),
    ).first()
    total_feedback = fb_agg[0] or 0
    avg_rating = fb_agg[1] or 0
    satisfied_count = fb_agg[2] or 0
    csat_score = round((satisfied_count / max(total_feedback, 1)) * 100, 1)

    total_users = User.query.filter_by(role="customer").count()

    # Category breakdown
    categories = db.session.query(
        ChatSession.sector_name, db.func.count(ChatSession.id)
    ).group_by(ChatSession.sector_name).all()

    return jsonify({
        "stats": {
            "total_chats": total_chats,
            "resolved_chats": resolved_chats,
            "escalated_chats": escalated_chats,
            "active_chats": active_chats,
            "total_tickets": total_tickets,
            "pending_tickets": pending_tickets,
            "in_progress_tickets": in_progress_tickets,
            "resolved_tickets": resolved_tickets,
            "escalated_tickets": escalated_tickets,
            "manager_escalated_tickets": manager_escalated_tickets,
            "critical_tickets": critical_tickets,
            "high_tickets": high_tickets,
            "total_feedback": total_feedback,
            "avg_rating": round(float(avg_rating), 1),
            "csat_score": csat_score,
            "total_customers": total_users,
        },
        "category_breakdown": [{"name": c[0] or "Unknown", "count": c[1]} for c in categories],
    })


@app.route("/api/manager/tickets", methods=["GET"])
@jwt_required()
def manager_tickets():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    status   = request.args.get("status")
    priority = request.args.get("priority")
    category = request.args.get("category")
    search   = request.args.get("search")

    query = Ticket.query

    # Managers see all open/active tickets (not just ones assigned to them).
    # CTO and Admin retain a full view of all tickets.

    if status:
        query = query.filter_by(status=status)
    if priority:
        query = query.filter_by(priority=priority)
    if category:
        query = query.filter_by(category=category)
    if search:
        query = query.join(User, Ticket.user_id == User.id).filter(
            db.or_(
                User.name.ilike(f"%{search}%"),
                User.email.ilike(f"%{search}%"),
                Ticket.reference_number.ilike(f"%{search}%"),
                Ticket.description.ilike(f"%{search}%"),
            )
        )

    # Sort by priority rank (critical first) then by creation time
    priority_order = sql_case(
        {"critical": 1, "high": 2, "medium": 3, "low": 4},
        value=Ticket.priority,
        else_=5,
    )
    tickets = query.order_by(priority_order, Ticket.created_at.asc()).all()
    return jsonify({"tickets": [t.to_dict() for t in tickets]})


@app.route("/api/manager/parameter-changes", methods=["GET"])
@jwt_required()
def manager_parameter_changes():
    user_id = int(get_jwt_identity())
    user = db.session.get(User, user_id)
    if user.role not in ("manager", "cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    status = request.args.get("status")
    query = ParameterChange.query.options(
        joinedload(ParameterChange.ticket),
        joinedload(ParameterChange.agent),
        joinedload(ParameterChange.reviewer),
    )
    if status:
        query = query.filter_by(status=status)

    changes = query.order_by(ParameterChange.created_at.desc()).all()
    return jsonify({"changes": [c.to_dict() for c in changes]})


@app.route("/api/manager/parameter-changes/<int:change_id>/review", methods=["PUT"])
@jwt_required()
def manager_review_parameter_change(change_id):
    user_id = int(get_jwt_identity())
    user = db.session.get(User, user_id)
    if user.role not in ("manager", "cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    change = db.session.get(ParameterChange, change_id)
    if not change:
        return jsonify({"error": "Change request not found"}), 404

    data = request.json or {}
    decision = data.get("decision")
    if decision not in ("approved", "disapproved"):
        return jsonify({"error": "Invalid decision"}), 400

    change.status = decision
    change.manager_note = (data.get("note") or "").strip()
    change.reviewed_at = datetime.now(timezone.utc)
    change.reviewed_by = user_id

    db.session.commit()
    return jsonify({"change": change.to_dict()})


# ═══════════════════════════════════════════════════════════════════════════════
# CHANGE WORKFLOW (ITIL) — Manager Endpoints
# ═══════════════════════════════════════════════════════════════════════════════

def _cr_auth(user):
    return user and user.role in ("manager", "cto", "admin")


@app.route("/api/manager/change-requests", methods=["GET"])
@jwt_required()
def manager_list_change_requests():
    user_id = int(get_jwt_identity())
    user    = db.session.get(User, user_id)
    if not _cr_auth(user):
        return jsonify({"error": "Unauthorized"}), 403

    status  = request.args.get("status")
    query   = ChangeRequest.query
    if status:
        if status == "needs_action":
            query = query.filter(ChangeRequest.status.in_(["created", "invalid", "validated", "classified", "implemented", "rolled_back"]))
        else:
            query = query.filter_by(status=status)

    crs = query.order_by(ChangeRequest.created_at.desc()).all()
    all_crs = ChangeRequest.query.all()
    stats = {
        "total":        len(all_crs),
        "needs_action": sum(1 for c in all_crs if c.status in ("created","invalid","validated","classified","implemented","rolled_back")),
        "approved":     sum(1 for c in all_crs if c.status in ("approved","implementing")),
        "closed":       sum(1 for c in all_crs if c.status in ("closed","rejected","auto_rejected")),
    }
    return jsonify({"change_requests": [c.to_dict() for c in crs], "stats": stats})


@app.route("/api/manager/change-requests/<int:cr_id>", methods=["GET"])
@jwt_required()
def manager_get_change_request(cr_id):
    user_id = int(get_jwt_identity())
    user    = db.session.get(User, user_id)
    if not _cr_auth(user):
        return jsonify({"error": "Unauthorized"}), 403
    cr = db.session.get(ChangeRequest, cr_id)
    if not cr:
        return jsonify({"error": "Change request not found"}), 404
    return jsonify({"cr": cr.to_dict()})


@app.route("/api/manager/change-requests/<int:cr_id>/validate", methods=["PUT"])
@jwt_required()
def manager_validate_cr(cr_id):
    """Stage 1: Manager validates whether the CR is acceptable."""
    user_id = int(get_jwt_identity())
    user    = db.session.get(User, user_id)
    if not _cr_auth(user):
        return jsonify({"error": "Unauthorized"}), 403

    cr = db.session.get(ChangeRequest, cr_id)
    if not cr:
        return jsonify({"error": "Change request not found"}), 404
    if cr.status not in ("created", "invalid"):
        return jsonify({"error": f"CR cannot be validated in status '{cr.status}'"}), 409

    data     = request.json or {}
    decision = (data.get("decision") or "").strip()
    remark   = (data.get("remark")   or "").strip()
    if decision not in ("valid", "invalid"):
        return jsonify({"error": "decision must be 'valid' or 'invalid'"}), 400

    now = datetime.now(timezone.utc)
    cr.validation_remark = remark
    cr.validated_by      = user_id
    cr.validated_at      = now
    cr.updated_at        = now

    if decision == "valid":
        cr.status = "validated"
    else:
        cr.rejection_count += 1
        if cr.rejection_count >= 2:
            cr.status = "auto_rejected"
        else:
            cr.status = "invalid"

    db.session.commit()
    return jsonify({"cr": cr.to_dict()})


@app.route("/api/manager/change-requests/<int:cr_id>/classify", methods=["PUT"])
@jwt_required()
def manager_classify_cr(cr_id):
    """Stage 2: Manager classifies the change type."""
    user_id = int(get_jwt_identity())
    user    = db.session.get(User, user_id)
    if not _cr_auth(user):
        return jsonify({"error": "Unauthorized"}), 403

    cr = db.session.get(ChangeRequest, cr_id)
    if not cr:
        return jsonify({"error": "Change request not found"}), 404
    if cr.status != "validated":
        return jsonify({"error": f"CR must be validated before classification (current: {cr.status})"}), 409

    data        = request.json or {}
    change_type = (data.get("change_type") or "").strip()
    note        = (data.get("note")        or "").strip()
    if change_type not in ("standard", "normal", "emergency"):
        return jsonify({"error": "change_type must be standard, normal, or emergency"}), 400

    now = datetime.now(timezone.utc)
    cr.change_type         = change_type
    cr.classification_note = note
    cr.classified_by       = user_id
    cr.classified_at       = now
    cr.status              = "classified"
    cr.updated_at          = now
    db.session.commit()
    return jsonify({"cr": cr.to_dict()})


@app.route("/api/manager/change-requests/<int:cr_id>/approve", methods=["PUT"])
@jwt_required()
def manager_approve_cr(cr_id):
    """Stage 3: Manager approves or rejects the classified CR."""
    user_id = int(get_jwt_identity())
    user    = db.session.get(User, user_id)
    if not _cr_auth(user):
        return jsonify({"error": "Unauthorized"}), 403

    cr = db.session.get(ChangeRequest, cr_id)
    if not cr:
        return jsonify({"error": "Change request not found"}), 404
    if cr.status != "classified":
        return jsonify({"error": f"CR must be classified before approval (current: {cr.status})"}), 409

    data     = request.json or {}
    decision = (data.get("decision") or "").strip()
    remark   = (data.get("remark")   or "").strip()
    if decision not in ("approved", "rejected"):
        return jsonify({"error": "decision must be 'approved' or 'rejected'"}), 400

    now = datetime.now(timezone.utc)
    cr.approval_remark = remark
    cr.approved_by     = user_id
    cr.approved_at     = now
    cr.status          = decision
    cr.updated_at      = now

    if cr.parameter_change_id:
        pc = db.session.get(ParameterChange, cr.parameter_change_id)
        if pc:
            pc.status       = "approved" if decision == "approved" else "disapproved"
            pc.reviewed_by  = user_id
            pc.reviewed_at  = now
            pc.manager_note = remark

    db.session.commit()
    return jsonify({"cr": cr.to_dict()})


@app.route("/api/manager/change-requests/<int:cr_id>/close", methods=["PUT"])
@jwt_required()
def manager_close_cr(cr_id):
    """Stage 5: Manager closes the CR after implementation (success or rollback)."""
    user_id = int(get_jwt_identity())
    user    = db.session.get(User, user_id)
    if not _cr_auth(user):
        return jsonify({"error": "Unauthorized"}), 403

    cr = db.session.get(ChangeRequest, cr_id)
    if not cr:
        return jsonify({"error": "Change request not found"}), 404
    if cr.status not in ("implemented", "rolled_back"):
        return jsonify({"error": f"CR can only be closed after implementation (current: {cr.status})"}), 409

    data = request.json or {}
    now  = datetime.now(timezone.utc)
    cr.closure_notes = (data.get("notes") or "").strip()
    cr.closed_at     = now
    cr.status        = "closed"
    cr.updated_at    = now
    db.session.commit()
    return jsonify({"cr": cr.to_dict()})


@app.route("/api/agent/change-requests/ticket/<int:ticket_id>", methods=["GET"])
@jwt_required()
def agent_get_cr_for_ticket(ticket_id):
    user_id = int(get_jwt_identity())
    user    = db.session.get(User, user_id)
    if not user or user.role != "human_agent":
        return jsonify({"error": "Unauthorized"}), 403
    cr = (ChangeRequest.query
          .filter_by(ticket_id=ticket_id, raised_by=user_id)
          .order_by(ChangeRequest.created_at.desc())
          .first())
    return jsonify({"cr": cr.to_dict() if cr else None})


@app.route("/api/manager/tickets/<int:ticket_id>/escalation-review", methods=["PUT"])
@jwt_required()
def manager_escalation_review(ticket_id):
    """Manager approves or rejects an escalated ticket."""
    user_id = int(get_jwt_identity())
    user = db.session.get(User, user_id)
    if not user or user.role != "manager":
        return jsonify({"error": "Unauthorized"}), 403

    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        return jsonify({"error": "Ticket not found"}), 404
    if ticket.assigned_to != user_id:
        return jsonify({"error": "Ticket not assigned to you"}), 403
    if ticket.status != "manager_escalated":
        return jsonify({"error": "Ticket is not in escalated state"}), 409

    data = request.json or {}
    decision = (data.get("decision") or "").strip().lower()
    note = (data.get("note") or "").strip()

    if decision not in ("approved", "rejected"):
        return jsonify({"error": "decision must be 'approved' or 'rejected'"}), 400

    pending_changes = ParameterChange.query.filter_by(ticket_id=ticket_id, status="pending").all()
    for pc in pending_changes:
        pc.status = "approved" if decision == "approved" else "disapproved"
        pc.manager_note = note
        pc.reviewed_at = datetime.now(timezone.utc)
        pc.reviewed_by = user_id

    if decision == "approved":
        ticket.status = "in_progress"
        if note:
            ticket.resolution_notes = note
    else:
        if ticket.escalated_by:
            ticket.assigned_to = ticket.escalated_by
        ticket.status = "in_progress"
        if note:
            ticket.resolution_notes = (
                f"[Manager rejected escalation: {note}]\n" + (ticket.resolution_notes or "")
            )

    db.session.commit()
    return jsonify({"ticket": ticket.to_dict(), "decision": decision})


@app.route("/api/manager/tickets/<int:ticket_id>", methods=["PUT"])
@jwt_required()
def update_ticket(ticket_id):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    ticket = Ticket.query.get(ticket_id)
    if not ticket:
        return jsonify({"error": "Ticket not found"}), 404

    # Managers can only update tickets assigned to them
    if user.role == "manager" and ticket.assigned_to != user_id:
        return jsonify({"error": "Ticket not assigned to you"}), 403

    data = request.json
    if "status" in data:
        old_status = ticket.status
        ticket.status = data["status"]
        if data["status"] == "resolved":
            ticket.resolved_at = datetime.now(timezone.utc)
        # Track reopening: resolved/closed → any open state
        if old_status in ("resolved", "closed") and data["status"] in ("pending", "in_progress"):
            ticket.reopened_count = (ticket.reopened_count or 0) + 1
            ticket.last_reopened_at = datetime.now(timezone.utc)
    if "priority" in data:
        ticket.priority = data["priority"]
    if "assigned_to" in data and user.role in ("cto", "admin"):
        # Only CTO/admin can re-assign; managers cannot move tickets between themselves
        ticket.assigned_to = data["assigned_to"]
    if "resolution_notes" in data:
        ticket.resolution_notes = data["resolution_notes"]

    db.session.commit()
    return jsonify({"ticket": ticket.to_dict()})


@app.route("/api/manager/chats", methods=["GET"])
@jwt_required()
def manager_chats():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    status = request.args.get("status")
    query = ChatSession.query
    if status:
        query = query.filter_by(status=status)

    sessions = query.order_by(ChatSession.created_at.desc()).all()
    return jsonify({"sessions": [s.to_dict() for s in sessions]})


@app.route("/api/manager/users", methods=["GET"])
@jwt_required()
def manager_users():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403
    managers = User.query.filter(User.role.in_(["manager"])).all()
    return jsonify({"managers": [u.to_dict() for u in managers]})


# ═══════════════════════════════════════════════════════════════════════════════
# CTO-SPECIFIC ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/cto/overview", methods=["GET"])
@jwt_required()
def cto_overview():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "cto":
        return jsonify({"error": "Unauthorized"}), 403

    # Resolution rate
    total = ChatSession.query.count() or 1
    resolved = ChatSession.query.filter_by(status="resolved").count()
    resolution_rate = round((resolved / total) * 100, 1)

    # Avg rating
    avg_rating = db.session.query(db.func.avg(Feedback.rating)).filter(Feedback.rating > 0).scalar() or 0

    # Tickets by priority
    priorities = db.session.query(
        Ticket.priority, db.func.count(Ticket.id)
    ).group_by(Ticket.priority).all()

    # Monthly trends (last 6 months)
    six_months_ago = datetime.now(timezone.utc) - timedelta(days=180)
    monthly = db.session.query(
        db.func.date_trunc("month", ChatSession.created_at).label("month"),
        db.func.count(ChatSession.id),
    ).filter(ChatSession.created_at >= six_months_ago).group_by("month").order_by("month").all()

    # Network issue stats
    worst_cell_count = overutilized_count = 0
    try:
        from network_issues import NetworkIssueTicket, OverutilizedTicket
        worst_cell_count = NetworkIssueTicket.query.filter(NetworkIssueTicket.status.in_(["open", "in_progress"])).count()
        overutilized_count = OverutilizedTicket.query.filter(OverutilizedTicket.status.in_(["open", "in_progress"])).count()
    except Exception:
        pass

    return jsonify({
        "resolution_rate": resolution_rate,
        "avg_rating": round(float(avg_rating), 1),
        "total_customers": User.query.filter_by(role="customer").count(),
        "total_sessions": total,
        "priority_breakdown": [{"priority": p[0], "count": p[1]} for p in priorities],
        "monthly_trends": [{"month": m[0].isoformat() if m[0] else "", "count": m[1]} for m in monthly],
        "worst_cell_tickets": worst_cell_count,
        "overutilized_tickets": overutilized_count,
    })


def _require_cto_user():
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role not in ("cto", "admin"):
        return None
    return user


# ── CTO KPI Cache (avoids re-querying 25M rows on every page load) ─────────
import time as _time
_cto_cache = {}   # key → {"data": ..., "ts": epoch}
_CTO_CACHE_TTL = 300  # 5 minutes

def _cache_get(key):
    entry = _cto_cache.get(key)
    if entry and (_time.time() - entry["ts"]) < _CTO_CACHE_TTL:
        return entry["data"]
    return None

def _cache_set(key, data):
    _cto_cache[key] = {"data": data, "ts": _time.time()}


def _cache_clear(*keys):
    """Drop cached entries for specified keys, or clear everything if no keys given."""
    if not keys:
        _cto_cache.clear()
        return
    for k in keys:
        _cto_cache.pop(k, None)


def _latest_site_values_for_kpi(kpi_name):
    """Latest value per site. Uses site-level data; falls back to cell-level (averaged per site)
    for sites that have no site-level data."""
    date_floor = date.today() - timedelta(days=30)

    # Site-level first
    rows = db.session.execute(db.text("""
        SELECT DISTINCT ON (site_id) site_id, date, value
        FROM kpi_data
        WHERE data_level = 'site' AND kpi_name = :kpi AND value IS NOT NULL AND date >= :floor
        ORDER BY site_id, date DESC
    """), {"kpi": kpi_name, "floor": date_floor}).fetchall()

    result = {r.site_id: {"date": r.date, "value": float(r.value)} for r in rows}

    # Cell-level fallback for sites not in result
    cell_rows = db.session.execute(db.text("""
        SELECT site_id, MAX(date) AS date, AVG(value) AS value
        FROM kpi_data
        WHERE data_level = 'cell' AND kpi_name = :kpi AND value IS NOT NULL AND date >= :floor
        GROUP BY site_id
    """), {"kpi": kpi_name, "floor": date_floor}).fetchall()

    for r in cell_rows:
        if r.site_id not in result:
            result[r.site_id] = {"date": r.date, "value": float(r.value)}

    return result


def _site_values_near_date(kpi_name, target_date):
    """Return {site_id: value} for the closest date <= target_date per site.
    Falls back to cell-level (averaged per site) for sites without site-level data."""
    date_floor = target_date - timedelta(days=14)

    # Site-level
    rows = db.session.execute(db.text("""
        SELECT DISTINCT ON (site_id) site_id, value
        FROM kpi_data
        WHERE data_level = 'site' AND kpi_name = :kpi AND value IS NOT NULL
          AND date >= :floor AND date <= :target
        ORDER BY site_id, date DESC
    """), {"kpi": kpi_name, "floor": date_floor, "target": target_date}).fetchall()

    result = {r.site_id: float(r.value) for r in rows}

    # Cell-level fallback
    cell_rows = db.session.execute(db.text("""
        SELECT site_id, AVG(value) AS value
        FROM kpi_data
        WHERE data_level = 'cell' AND kpi_name = :kpi AND value IS NOT NULL
          AND date >= :floor AND date <= :target
        GROUP BY site_id
    """), {"kpi": kpi_name, "floor": date_floor, "target": target_date}).fetchall()

    for r in cell_rows:
        if r.site_id not in result:
            result[r.site_id] = float(r.value)

    return result


def _series_for_kpi_patterns(patterns, days=30):
    """Daily average series. Uses site-level; merges cell-level for sites without site data."""
    from sqlalchemy import func as sa_func
    date_floor = date.today() - timedelta(days=days)

    # Sites with site-level data
    site_level_sites = set(r[0] for r in db.session.query(
        db.distinct(KpiData.site_id)
    ).filter(
        KpiData.data_level == "site",
        KpiData.kpi_name.in_(patterns),
        KpiData.date >= date_floor,
    ).all())

    # Site-level series
    site_rows = db.session.query(
        KpiData.date,
        sa_func.avg(KpiData.value).label("avg_val"),
    ).filter(
        KpiData.data_level == "site",
        KpiData.value.isnot(None),
        KpiData.kpi_name.in_(patterns),
        KpiData.date >= date_floor,
    ).group_by(KpiData.date).order_by(KpiData.date).all()

    # Cell-level fallback for sites without site data
    cell_rows = []
    if site_level_sites:
        cell_rows = db.session.query(
            KpiData.date,
            sa_func.avg(KpiData.value).label("avg_val"),
        ).filter(
            KpiData.data_level == "cell",
            KpiData.value.isnot(None),
            KpiData.kpi_name.in_(patterns),
            KpiData.date >= date_floor,
            ~KpiData.site_id.in_(site_level_sites),
        ).group_by(KpiData.date).order_by(KpiData.date).all()
    else:
        # No site-level data at all — use all cell data
        cell_rows = db.session.query(
            KpiData.date,
            sa_func.avg(KpiData.value).label("avg_val"),
        ).filter(
            KpiData.data_level == "cell",
            KpiData.value.isnot(None),
            KpiData.kpi_name.in_(patterns),
            KpiData.date >= date_floor,
        ).group_by(KpiData.date).order_by(KpiData.date).all()

    # Merge by date
    merged = {}
    for r in site_rows:
        merged.setdefault(r.date, []).append(float(r.avg_val))
    for r in cell_rows:
        merged.setdefault(r.date, []).append(float(r.avg_val))

    return [{"date": d.isoformat(), "value": round(sum(v) / len(v), 2)}
            for d, v in sorted(merged.items())]


def _latest_average_for_patterns(patterns):
    series = _series_for_kpi_patterns(patterns)
    return round(series[-1]["value"], 2) if series else 0


@app.route("/api/cto/map-data", methods=["GET"])
@jwt_required()
def cto_map_data():
    user = _require_cto_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 403

    sites = TelecomSite.query.with_entities(
        TelecomSite.site_id,
        TelecomSite.latitude,
        TelecomSite.longitude,
        TelecomSite.zone,
        TelecomSite.site_status,
        TelecomSite.alarms,
        TelecomSite.solution,
    ).all()

    return jsonify({
        "sites": [
            {
                "site_id": site.site_id,
                "lat": site.latitude,
                "lng": site.longitude,
                "zone": site.zone or "",
                "status": (site.site_status or "active").lower(),
                "alarm": site.alarms or "",
                "solution": site.solution or "",
            }
            for site in sites
            if (
                site.latitude is not None and
                site.longitude is not None and
                float(site.latitude) != 0 and
                float(site.longitude) != 0
            )
        ]
    })


@app.route("/api/cto/ticket-heatmap", methods=["GET"])
@jwt_required()
def cto_ticket_heatmap():
    """Return ticket counts grouped by state/province with centroid lat/lng
    for rendering as Leaflet circle markers (no GeoJSON needed).
    """
    user = _require_cto_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 403

    from sqlalchemy import func as sa_func

    # Query tickets joined with chat sessions to get state/province + avg location
    rows = db.session.query(
        ChatSession.state_province,
        ChatSession.country,
        sa_func.count(Ticket.id).label("total"),
        sa_func.sum(
            db.case((Ticket.status == "resolved", 1), else_=0)
        ).label("resolved"),
        sa_func.avg(ChatSession.latitude).label("avg_lat"),
        sa_func.avg(ChatSession.longitude).label("avg_lng"),
    ).join(
        ChatSession, Ticket.chat_session_id == ChatSession.id
    ).filter(
        ChatSession.state_province.isnot(None),
        ChatSession.state_province != "",
    ).group_by(
        ChatSession.state_province,
        ChatSession.country,
    ).all()

    # Detect dominant country
    country_counts = {}
    for _, country, total, _, _, _ in rows:
        c = (country or "").strip()
        if c:
            country_counts[c] = country_counts.get(c, 0) + (total or 0)
    detected_country = max(country_counts, key=country_counts.get) if country_counts else ""

    state_data = []
    for state, country, total, resolved, avg_lat, avg_lng in rows:
        lat = float(avg_lat) if avg_lat else None
        lng = float(avg_lng) if avg_lng else None
        state_data.append({
            "state": (state or "").strip(),
            "total": total or 0,
            "resolved": resolved or 0,
            "pending": (total or 0) - (resolved or 0),
            "lat": lat,
            "lng": lng,
        })

    return jsonify({
        "state_data": state_data,
        "detected_country": detected_country,
    })


@app.route("/api/cto/technical-kpi", methods=["GET"])
@jwt_required()
def cto_technical_kpi():
    user = _require_cto_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 403

    kpi_defs = [
        {
            "key": "accessibility",
            "label": "Accessibility (CSSR)",
            "names": [
                "LTE Call Setup Success Rate",
            ],
            "unit": "%",
        },
        {
            "key": "retainability",
            "label": "Retainability (E-RAB Drop)",
            "names": [
                "E-RAB Call Drop Rate_1",
            ],
            "unit": "%",
        },
        {
            "key": "downlink_throughput",
            "label": "Downlink Throughput",
            "names": [
                "LTE DL - Cell Ave Throughput",
                "LTE DL - Usr Ave Throughput",
            ],
            "unit": "Mbps",
        },
        {
            "key": "prb_utilization",
            "label": "PRB Utilization",
            "names": [
                "DL PRB Utilization (1BH)",
                "UL PRB Utilization (1BH",
            ],
            "unit": "%",
        },
        {
            "key": "downlink_volume",
            "label": "Downlink Volume",
            "names": [
                "DL Data Total Volume",
            ],
            "unit": "GB",
        },
        {
            "key": "uplink_volume",
            "label": "Uplink Volume",
            "names": [
                "UL Data Total Volume",
            ],
            "unit": "GB",
        },
    ]

    # Check cache first
    cached = _cache_get("technical_kpi")
    if cached:
        return jsonify(cached)

    # Bulk query: all KPI names in one shot, grouped by kpi_name + date
    # Only last 45 days — enough for 30-day trend + buffer
    # Uses site-level data first; for sites missing site-level, falls back to
    # cell-level data averaged per site.
    from sqlalchemy import func as sa_func
    all_names = []
    for item in kpi_defs:
        all_names.extend(item["names"])

    date_floor = date.today() - timedelta(days=45)

    # Step 1: Find sites that have site-level data
    sites_with_site_level = set(r[0] for r in db.session.query(
        db.distinct(KpiData.site_id)
    ).filter(
        KpiData.data_level == "site",
        KpiData.kpi_name.in_(all_names),
        KpiData.date >= date_floor,
    ).all())

    # Step 2: Find sites that only have cell-level data (no site-level)
    sites_cell_only = set(r[0] for r in db.session.query(
        db.distinct(KpiData.site_id)
    ).filter(
        KpiData.data_level == "cell",
        KpiData.kpi_name.in_(all_names),
        KpiData.date >= date_floor,
        ~KpiData.site_id.in_(sites_with_site_level) if sites_with_site_level else db.true(),
    ).all()) if True else set()

    # Step 3: Query site-level data
    bulk_site = db.session.query(
        KpiData.kpi_name,
        KpiData.date,
        sa_func.avg(KpiData.value).label("avg_val"),
    ).filter(
        KpiData.data_level == "site",
        KpiData.value.isnot(None),
        KpiData.kpi_name.in_(all_names),
        KpiData.date >= date_floor,
    ).group_by(KpiData.kpi_name, KpiData.date).order_by(KpiData.date).all()

    # Step 4: Query cell-level data for sites missing site-level, averaged per site then overall
    bulk_cell = []
    if sites_cell_only:
        bulk_cell = db.session.query(
            KpiData.kpi_name,
            KpiData.date,
            sa_func.avg(KpiData.value).label("avg_val"),
        ).filter(
            KpiData.data_level == "cell",
            KpiData.value.isnot(None),
            KpiData.kpi_name.in_(all_names),
            KpiData.date >= date_floor,
            KpiData.site_id.in_(sites_cell_only),
        ).group_by(KpiData.kpi_name, KpiData.date).order_by(KpiData.date).all()

    # Merge: combine site-level and cell-level (fallback) into one result
    # Cell-level values are merged into the same kpi_name+date buckets
    merged = {}
    for kpi_name, kpi_date, avg_val in bulk_site:
        merged.setdefault((kpi_name, kpi_date), []).append(float(avg_val))
    for kpi_name, kpi_date, avg_val in bulk_cell:
        merged.setdefault((kpi_name, kpi_date), []).append(float(avg_val))

    bulk = [(kn, dt, sum(vals) / len(vals)) for (kn, dt), vals in sorted(merged.items(), key=lambda x: x[0][1])]

    # Map kpi_name → key
    name_to_key = {}
    for item in kpi_defs:
        for n in item["names"]:
            name_to_key[n] = item["key"]

    # Aggregate per key per date
    key_date = {}  # key → {date_str: [values]}
    for kpi_name, kpi_date, avg_val in bulk:
        k = name_to_key.get(kpi_name)
        if not k:
            continue
        d_str = kpi_date.isoformat()
        key_date.setdefault(k, {}).setdefault(d_str, []).append(float(avg_val))

    cards = []
    chart_series = {}
    for item in kpi_defs:
        k = item["key"]
        date_vals = key_date.get(k, {})
        series = []
        for day in sorted(date_vals.keys()):
            vals = date_vals[day]
            series.append({"date": day, "value": round(sum(vals) / len(vals), 2)})
        cards.append({
            "key": k,
            "label": item["label"],
            "value": round(series[-1]["value"], 2) if series else 0,
            "unit": item.get("unit", ""),
        })
        chart_series[k] = series[-30:]

    # ── KPI Forecast (simple linear regression) ──────────────────────
    def _linear_forecast(series, forecast_days=7):
        vals = [p["value"] for p in series]
        n = len(vals)
        if n < 2:
            return None
        x_mean = (n - 1) / 2.0
        y_mean = sum(vals) / n
        num = sum((i - x_mean) * (v - y_mean) for i, v in enumerate(vals))
        den = sum((i - x_mean) ** 2 for i in range(n))
        slope = num / den if den else 0
        intercept = y_mean - slope * x_mean
        predictions = [round(max(0, slope * (n + i) + intercept), 2) for i in range(forecast_days)]
        return slope, predictions

    forecast = {}
    forecast_keys = ["accessibility", "retainability", "downlink_throughput", "prb_utilization"]
    for k in forecast_keys:
        s = chart_series.get(k, [])
        if len(s) < 2:
            continue
        result_fc = _linear_forecast(s)
        if not result_fc:
            continue
        slope, preds = result_fc
        current = s[-1]["value"]
        predicted_7d = preds[-1] if preds else current
        # Determine direction: for PRB, up is bad; for others, down is bad
        if k == "prb_utilization":
            direction = "up" if slope > 0.01 else ("down" if slope < -0.01 else "stable")
        else:
            direction = "up" if slope > 0.01 else ("down" if slope < -0.01 else "stable")
        forecast[k] = {
            "current": current,
            "predicted_7d": predicted_7d,
            "direction": direction,
            "daily_change": round(slope, 3),
        }

    # ── Per-site latest values for the site breakdown table ──────────
    # Uses site-level data; for sites with only cell-level data, averages cells per site
    date_floor_site = date.today() - timedelta(days=7)

    # Site-level latest
    site_sub = db.session.query(
        KpiData.site_id,
        KpiData.kpi_name,
        sa_func.max(KpiData.date).label("max_date"),
    ).filter(
        KpiData.data_level == "site",
        KpiData.kpi_name.in_(all_names),
        KpiData.value.isnot(None),
        KpiData.date >= date_floor_site,
    ).group_by(KpiData.site_id, KpiData.kpi_name).subquery()

    site_bulk = db.session.query(
        KpiData.site_id, KpiData.kpi_name, KpiData.value,
    ).join(site_sub, db.and_(
        KpiData.site_id == site_sub.c.site_id,
        KpiData.kpi_name == site_sub.c.kpi_name,
        KpiData.date == site_sub.c.max_date,
    )).filter(
        KpiData.data_level == "site",
        KpiData.kpi_name.in_(all_names),
    ).all()

    site_pivot = {}
    for site_id, kpi_name, value in site_bulk:
        k = name_to_key.get(kpi_name)
        if k:
            site_pivot.setdefault(site_id, {})[k] = round(float(value), 2)

    # Cell-level fallback: for sites missing from site_pivot, average their cells
    if sites_cell_only:
        cell_sub = db.session.query(
            KpiData.site_id,
            KpiData.kpi_name,
            sa_func.max(KpiData.date).label("max_date"),
        ).filter(
            KpiData.data_level == "cell",
            KpiData.kpi_name.in_(all_names),
            KpiData.value.isnot(None),
            KpiData.date >= date_floor_site,
            KpiData.site_id.in_(sites_cell_only),
        ).group_by(KpiData.site_id, KpiData.kpi_name).subquery()

        cell_latest = db.session.query(
            KpiData.site_id,
            KpiData.kpi_name,
            sa_func.avg(KpiData.value).label("avg_val"),
        ).join(cell_sub, db.and_(
            KpiData.site_id == cell_sub.c.site_id,
            KpiData.kpi_name == cell_sub.c.kpi_name,
            KpiData.date == cell_sub.c.max_date,
        )).filter(
            KpiData.data_level == "cell",
            KpiData.kpi_name.in_(all_names),
        ).group_by(KpiData.site_id, KpiData.kpi_name).all()

        for site_id, kpi_name, avg_val in cell_latest:
            k = name_to_key.get(kpi_name)
            if k and site_id not in site_pivot:
                site_pivot.setdefault(site_id, {})[k] = round(float(avg_val), 2)

    site_list = []
    for sid in sorted(site_pivot.keys()):
        row = {"site_id": sid}
        row.update(site_pivot[sid])
        site_list.append(row)

    # Sort by worst accessibility, limit to 50
    site_list.sort(key=lambda r: r.get("accessibility", 100))
    site_list = site_list[:50]

    # ── Packet Loss ──────────────────────────────────────────────────
    # Primary: use network_kpi_timeseries.packet_loss (agent upload)
    # Fallback: derive from E-RAB Call Drop Rate_1 in kpi_data
    pl_floor = date.today() - timedelta(days=14)
    pl_series = []
    pl_worst = []

    from sqlalchemy import inspect as sa_inspect
    has_nkt = sa_inspect(db.engine).has_table("network_kpi_timeseries")
    if has_nkt:
        # Check if packet_loss data exists
        nkt_count = db.session.execute(db.text(
            "SELECT count(*) FROM network_kpi_timeseries WHERE packet_loss IS NOT NULL AND timestamp::date >= :floor"
        ), {"floor": pl_floor}).scalar()
    else:
        nkt_count = 0

    if nkt_count > 0:
        # ── Primary: real packet_loss from network_kpi_timeseries ──
        pl_series_rows = db.session.execute(db.text("""
            SELECT timestamp::date AS d, AVG(packet_loss) AS avg_val
            FROM network_kpi_timeseries
            WHERE packet_loss IS NOT NULL AND timestamp::date >= :floor
            GROUP BY d ORDER BY d
        """), {"floor": pl_floor}).fetchall()
        pl_series = [{"date": r.d.isoformat(), "value": round(float(r.avg_val), 2)} for r in pl_series_rows]

        pl_site_rows = db.session.execute(db.text("""
            SELECT DISTINCT ON (site_id) site_id, packet_loss
            FROM network_kpi_timeseries
            WHERE packet_loss IS NOT NULL AND timestamp::date >= :floor
            ORDER BY site_id, timestamp DESC
        """), {"floor": pl_floor}).fetchall()
        pl_worst = sorted(
            [{"site_id": r.site_id, "value": round(float(r.packet_loss), 2)} for r in pl_site_rows],
            key=lambda r: r["value"], reverse=True
        )
    else:
        # ── Fallback: derive from E-RAB Call Drop Rate_1 ──
        pl_name = "E-RAB Call Drop Rate_1"
        pl_series_rows = db.session.query(
            KpiData.date,
            sa_func.avg(KpiData.value).label("avg_val"),
        ).filter(
            KpiData.data_level == "site",
            KpiData.kpi_name == pl_name,
            KpiData.value.isnot(None),
            KpiData.date >= pl_floor,
        ).group_by(KpiData.date).order_by(KpiData.date).all()
        pl_series = [{"date": r.date.isoformat(), "value": round(float(r.avg_val), 2)} for r in pl_series_rows]

        pl_site_rows = db.session.execute(db.text("""
            SELECT DISTINCT ON (site_id) site_id, value
            FROM kpi_data
            WHERE data_level = 'site' AND kpi_name = :kpi AND value IS NOT NULL AND date >= :floor
            ORDER BY site_id, date DESC
        """), {"kpi": pl_name, "floor": pl_floor}).fetchall()
        pl_worst = sorted(
            [{"site_id": r.site_id, "value": round(float(r.value), 2)} for r in pl_site_rows],
            key=lambda r: r["value"], reverse=True
        )

    pl_avg = round(pl_series[-1]["value"], 2) if pl_series else 0
    packet_loss = {
        "avg": pl_avg,
        "series": pl_series,
        "worst_sites": pl_worst,
        "source": "network_kpi_timeseries" if nkt_count > 0 else "kpi_data (E-RAB Call Drop Rate)",
    }

    # ── Individual KPI series for specific trend charts ────────────────
    # Frontend shows dedicated trend graphs for these specific metrics.
    individual_kpi_names = [
        "LTE Call Setup Success Rate",
        "E-RAB Call Drop Rate_1",
    ]
    individual_series = {}
    for kpi_name, kpi_date, avg_val in bulk:
        if kpi_name in individual_kpi_names:
            key = kpi_name.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "")
            individual_series.setdefault(key, []).append({
                "date": kpi_date.isoformat() if hasattr(kpi_date, 'isoformat') else str(kpi_date),
                "value": round(float(avg_val), 2),
            })
    # Sort each series by date and take last 30
    for k in individual_series:
        individual_series[k] = sorted(individual_series[k], key=lambda x: x["date"])[-30:]

    result = {
        "cards": cards,
        "series": chart_series,
        "individual_series": individual_series,
        "sites": site_list,
        "packet_loss": packet_loss,
        "forecast": forecast,
    }
    _cache_set("technical_kpi", result)
    return jsonify(result)


@app.route("/api/cto/core-kpi", methods=["GET"])
@jwt_required()
def cto_core_kpi():
    """Return Core Network KPI data — component-based from CoreComponentKpi table.
    Falls back to FlexibleKpiUpload(core) if no component data exists.
    Shows data by component_type/component_id instead of site_id.
    """
    import datetime as _dt
    from sqlalchemy import func as sa_func
    from models import CoreComponentKpi

    user = _require_cto_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 403

    def _iso(d):
        return d.isoformat() if d and hasattr(d, 'isoformat') else (str(d) if d else None)

    # ── Check CoreComponentKpi table first (component-based data) ──
    has_core_comp = db.session.query(CoreComponentKpi.id).limit(1).scalar() is not None

    if has_core_comp:
        return _core_kpi_from_components()

    # ── Fallback: FlexibleKpiUpload (core) — site-based ──
    has_flex = db.session.query(FlexibleKpiUpload.id).filter(
        FlexibleKpiUpload.kpi_type == "core",
        FlexibleKpiUpload.kpi_name.isnot(None),
    ).limit(1).scalar() is not None

    if has_flex:
        return _core_kpi_from_flexible()

    return jsonify({"available": False, "message": "No Core KPI data uploaded yet."})


def _core_kpi_from_components():
    """Build rich Core KPI response from CoreComponentKpi — CTO-level analytics.
    Auto-categorises each KPI as success_rate / utilization / latency / capacity.
    Computes health scores, identifies anomalies, builds comparison data.
    """
    from sqlalchemy import func as sa_func

    def _iso(d):
        return d.isoformat() if d and hasattr(d, 'isoformat') else (str(d) if d else None)

    def _categorise(name):
        nl = name.lower()
        if any(k in nl for k in ("success", "availability", "accuracy")):
            return "success_rate"
        if any(k in nl for k in ("cpu", "memory", "utilization", "prb", "tps usage", "queue depth")):
            return "utilization"
        if any(k in nl for k in ("latency", "(ms)", "(s)", "lag", "jitter", "response time")):
            return "latency"
        if any(k in nl for k in ("failure", "drop", "loss", "timeout", "contention", "error")):
            return "failure_rate"
        return "capacity"  # sessions, counts, throughput, etc.

    # ── Per-KPI scale detection: if max value <= 1.5 for a percentage-type
    # KPI, treat stored values as fractions (0-1) and multiply by 100 for display.
    max_rows = db.session.query(
        CoreComponentKpi.kpi_name,
        sa_func.max(CoreComponentKpi.value),
    ).filter(CoreComponentKpi.value.isnot(None)).group_by(CoreComponentKpi.kpi_name).all()
    kpi_scale = {}  # kpi_name → multiplier
    for kn, mx in max_rows:
        if mx is None:
            kpi_scale[kn] = 1.0
            continue
        cat = _categorise(kn)
        if cat in ("success_rate", "failure_rate", "utilization") and float(mx) <= 1.5:
            kpi_scale[kn] = 100.0
        else:
            kpi_scale[kn] = 1.0

    def _scale(kn, v):
        if v is None:
            return None
        return float(v) * kpi_scale.get(kn, 1.0)

    # ── Date range ──
    date_range = db.session.query(
        sa_func.min(CoreComponentKpi.date),
        sa_func.max(CoreComponentKpi.date),
    ).one()
    min_date, latest_date = date_range
    if not min_date:
        return jsonify({"available": False, "message": "No valid dates found."})

    # ── Component types + instance counts ──
    comp_types_raw = db.session.query(
        CoreComponentKpi.component_type,
        sa_func.count(sa_func.distinct(CoreComponentKpi.component_id)),
    ).group_by(CoreComponentKpi.component_type).all()
    comp_types = sorted([r[0] for r in comp_types_raw])
    type_instance_count = {r[0]: r[1] for r in comp_types_raw}
    total_nodes = sum(type_instance_count.values())

    # ── All KPI names and their categories ──
    all_kpis = [r[0] for r in db.session.query(CoreComponentKpi.kpi_name).distinct().all()]
    kpi_category = {k: _categorise(k) for k in all_kpis}

    # ── Latest values: component_type × kpi_name → avg across instances ──
    latest_rows = db.session.query(
        CoreComponentKpi.component_type,
        CoreComponentKpi.kpi_name,
        sa_func.avg(CoreComponentKpi.value).label("avg_val"),
    ).filter(
        CoreComponentKpi.date == latest_date,
        CoreComponentKpi.value.isnot(None),
    ).group_by(
        CoreComponentKpi.component_type,
        CoreComponentKpi.kpi_name,
    ).all()

    # Build per-type KPI dict (apply scale)
    type_kpis = {}  # {MME: {kpi_name: avg_val}}
    for ct, kn, av in latest_rows:
        type_kpis.setdefault(ct, {})[kn] = round(_scale(kn, av), 2)

    # ── Compute health score per component type ──
    # Health = avg of success_rate KPIs (higher=better).
    # If no success rates, fall back to 100 - avg(failure_rates) - avg(utilization > 80 penalty).
    type_health = {}
    for ct in comp_types:
        kpis = type_kpis.get(ct, {})
        success_vals = [v for k, v in kpis.items() if kpi_category.get(k) == "success_rate" and v is not None]
        failure_vals = [v for k, v in kpis.items() if kpi_category.get(k) == "failure_rate" and v is not None]
        util_vals    = [v for k, v in kpis.items() if kpi_category.get(k) == "utilization" and v is not None]

        if success_vals:
            health = round(sum(success_vals) / len(success_vals), 1)
        elif failure_vals:
            health = round(100 - (sum(failure_vals) / len(failure_vals)), 1)
        else:
            health = 100.0

        # Penalise high utilisation (>80% is a concern)
        if util_vals:
            over80 = [v for v in util_vals if v > 80]
            if over80:
                health = max(0, health - len(over80) * 2)

        cpu_val = next((v for k, v in kpis.items() if "cpu" in k.lower()), None)
        mem_val = next((v for k, v in kpis.items() if "memory" in k.lower()), None)

        type_health[ct] = {
            "health": round(health, 1),
            "cpu": cpu_val,
            "memory": mem_val,
            "instances": type_instance_count.get(ct, 0),
            "success_rates": {k: v for k, v in kpis.items() if kpi_category.get(k) == "success_rate"},
            "failure_rates": {k: v for k, v in kpis.items() if kpi_category.get(k) == "failure_rate"},
            "utilization":   {k: v for k, v in kpis.items() if kpi_category.get(k) == "utilization"},
            "latency":       {k: v for k, v in kpis.items() if kpi_category.get(k) == "latency"},
            "capacity":      {k: v for k, v in kpis.items() if kpi_category.get(k) == "capacity"},
        }

    # ── Overall hero metrics ──
    all_health = [v["health"] for v in type_health.values()]
    overall_health = round(sum(all_health) / len(all_health), 1) if all_health else 0

    all_cpu = [v["cpu"] for v in type_health.values() if v["cpu"] is not None]
    avg_cpu = round(sum(all_cpu) / len(all_cpu), 1) if all_cpu else None

    all_mem = [v["memory"] for v in type_health.values() if v["memory"] is not None]
    avg_mem = round(sum(all_mem) / len(all_mem), 1) if all_mem else None

    # Average of all success-rate KPIs across all components
    all_sr = []
    for th in type_health.values():
        all_sr.extend(th["success_rates"].values())
    avg_success = round(sum(all_sr) / len(all_sr), 2) if all_sr else None

    # ── Trends: daily avg for CPU, Memory, Health across all components ──
    trend_rows = db.session.query(
        CoreComponentKpi.kpi_name,
        CoreComponentKpi.date,
        sa_func.avg(CoreComponentKpi.value).label("avg_val"),
    ).filter(
        CoreComponentKpi.value.isnot(None),
    ).group_by(CoreComponentKpi.kpi_name, CoreComponentKpi.date).all()

    # Build daily trend for CPU, Memory, and overall success (apply scale)
    daily = {}  # date → {cpu: [], mem: [], success: []}
    for kn, dt, av in trend_rows:
        d_str = _iso(dt)
        bucket = daily.setdefault(d_str, {"cpu": [], "mem": [], "success": [], "failure": []})
        cat = kpi_category.get(kn, "")
        val = _scale(kn, av)
        if val is None:
            continue
        if "cpu" in kn.lower():
            bucket["cpu"].append(val)
        elif "memory" in kn.lower():
            bucket["mem"].append(val)
        if cat == "success_rate":
            bucket["success"].append(val)
        elif cat == "failure_rate":
            bucket["failure"].append(val)

    trend = []
    for d_str in sorted(daily.keys()):
        b = daily[d_str]
        cpu_avg = round(sum(b["cpu"]) / len(b["cpu"]), 2) if b["cpu"] else None
        mem_avg = round(sum(b["mem"]) / len(b["mem"]), 2) if b["mem"] else None
        sr_avg  = round(sum(b["success"]) / len(b["success"]), 2) if b["success"] else None
        fr_avg  = round(sum(b["failure"]) / len(b["failure"]), 2) if b["failure"] else None
        health_d = sr_avg if sr_avg is not None else (round(100 - fr_avg, 2) if fr_avg is not None else None)
        trend.append({"date": d_str, "cpu": cpu_avg, "memory": mem_avg, "success_rate": sr_avg, "health": health_d})

    # ── Per-component instance latest (for drilldown table) ──
    inst_rows = db.session.query(
        CoreComponentKpi.component_type,
        CoreComponentKpi.component_id,
        CoreComponentKpi.kpi_name,
        sa_func.avg(CoreComponentKpi.value).label("avg_val"),
    ).filter(
        CoreComponentKpi.date == latest_date,
    ).group_by(
        CoreComponentKpi.component_type,
        CoreComponentKpi.component_id,
        CoreComponentKpi.kpi_name,
    ).all()

    # Build per-instance health (apply scale)
    inst_pivot = {}
    for ct, cid, kn, av in inst_rows:
        key = f"{ct}:{cid}"
        inst_pivot.setdefault(key, {"component_type": ct, "component_id": cid, "kpis": {}})
        sv = _scale(kn, av)
        if sv is not None:
            inst_pivot[key]["kpis"][kn] = round(sv, 2)

    component_table = []
    for key in sorted(inst_pivot.keys()):
        inst = inst_pivot[key]
        kpis = inst["kpis"]
        sr = [v for k, v in kpis.items() if kpi_category.get(k) == "success_rate"]
        fr = [v for k, v in kpis.items() if kpi_category.get(k) == "failure_rate"]
        health = round(sum(sr) / len(sr), 1) if sr else (round(100 - sum(fr) / len(fr), 1) if fr else 100)
        cpu = next((v for k, v in kpis.items() if "cpu" in k.lower()), None)
        mem = next((v for k, v in kpis.items() if "memory" in k.lower()), None)
        component_table.append({
            "component_type": inst["component_type"],
            "component_id": inst["component_id"],
            "health": health,
            "cpu": cpu,
            "memory": mem,
            "kpi_count": len(kpis),
        })

    # ── Anomalies: instances with health < 95 or CPU > 80 or Memory > 80 ──
    anomalies = []
    for row in component_table:
        reasons = []
        if row["health"] < 95:
            reasons.append(f"Health {row['health']}%")
        if row["cpu"] is not None and row["cpu"] > 80:
            reasons.append(f"CPU {row['cpu']}%")
        if row["memory"] is not None and row["memory"] > 80:
            reasons.append(f"Memory {row['memory']}%")
        if reasons:
            anomalies.append({**row, "reasons": reasons})
    anomalies.sort(key=lambda r: r["health"])

    # ── Component type comparison (for bar chart) ──
    type_comparison = []
    for ct in comp_types:
        th = type_health[ct]
        type_comparison.append({
            "type": ct,
            "health": th["health"],
            "cpu": th["cpu"],
            "memory": th["memory"],
            "instances": th["instances"],
            "success_count": len(th["success_rates"]),
            "total_kpis": sum(len(v) for v in [th["success_rates"], th["failure_rates"], th["utilization"], th["latency"], th["capacity"]]),
        })

    return jsonify({
        "available": True,
        "data_source": "component",
        "date_range": {"from": _iso(min_date), "to": _iso(latest_date)},
        "hero": {
            "overall_health": overall_health,
            "avg_cpu": avg_cpu,
            "avg_memory": avg_mem,
            "avg_success_rate": avg_success,
            "total_nodes": total_nodes,
            "total_kpis": len(all_kpis),
            "component_types": len(comp_types),
        },
        "type_health": type_health,
        "type_comparison": type_comparison,
        "trend": trend,
        "components": component_table,
        "anomalies": anomalies,
    })


def _core_kpi_from_flexible():
    """Fallback: Build Core KPI from FlexibleKpiUpload(core) — site-based."""
    import datetime as _dt
    from sqlalchemy import func as sa_func

    def _iso(d):
        return d.isoformat() if d and hasattr(d, 'isoformat') else (str(d) if d else None)

    def _parse_col_date(col_name):
        s = col_name.replace("_", "-")[:10]
        try:
            return _dt.datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            return None

    LABEL_MAP = {
        "CPU Utilization": "CPU Usage",
        "Authentication Success Rate": "Auth Success Rate",
        "Attach Success Rate": "4G Attach Success",
        "PDP Bearer Setup Success Rate": "4G Bearer Success",
    }
    kpi_names = [r[0] for r in db.session.query(
        FlexibleKpiUpload.kpi_name
    ).filter(
        FlexibleKpiUpload.kpi_type == "core",
        FlexibleKpiUpload.kpi_name.isnot(None),
    ).distinct().all()]

    db_to_key = {}
    cols_meta = {}
    for kn in kpi_names:
        key = kn.lower().replace(" ", "_")
        db_to_key[kn] = key
        cols_meta[key] = {"db_name": kn, "label": LABEL_MAP.get(kn, kn)}

    has_row_date = db.session.query(FlexibleKpiUpload.id).filter(
        FlexibleKpiUpload.kpi_type == "core",
        FlexibleKpiUpload.row_date.isnot(None),
    ).limit(1).scalar() is not None

    if has_row_date:
        date_range = db.session.query(
            sa_func.min(FlexibleKpiUpload.row_date),
            sa_func.max(FlexibleKpiUpload.row_date),
        ).filter(
            FlexibleKpiUpload.kpi_type == "core",
            FlexibleKpiUpload.row_date.isnot(None),
        ).one()
        min_date, latest_date = date_range
        if not min_date:
            return jsonify({"available": False, "message": "No valid dates found."})

        bulk_rows = db.session.query(
            FlexibleKpiUpload.kpi_name,
            FlexibleKpiUpload.row_date,
            sa_func.avg(FlexibleKpiUpload.num_value).label("avg_val"),
        ).filter(
            FlexibleKpiUpload.kpi_type == "core",
            FlexibleKpiUpload.kpi_name.isnot(None),
            FlexibleKpiUpload.row_date.isnot(None),
        ).group_by(FlexibleKpiUpload.kpi_name, FlexibleKpiUpload.row_date).all()

        all_dates = sorted(set(r[1] for r in bulk_rows if r[1]))
        per_kpi_all = {k: [] for k in cols_meta}
        per_kpi_latest = {k: None for k in cols_meta}
        trend_pivot = {_iso(d): {"date": _iso(d)} for d in all_dates}

        for kpi_db_name, row_date, avg_val in bulk_rows:
            key = db_to_key.get(kpi_db_name)
            if not key or avg_val is None:
                continue
            val = round(float(avg_val), 2)
            per_kpi_all[key].append(val)
            if row_date == latest_date:
                per_kpi_latest[key] = val
            d_str = _iso(row_date)
            if d_str in trend_pivot:
                trend_pivot[d_str][key] = val

        site_rows = db.session.query(
            FlexibleKpiUpload.site_id, FlexibleKpiUpload.kpi_name, FlexibleKpiUpload.num_value,
        ).filter(FlexibleKpiUpload.kpi_type == "core", FlexibleKpiUpload.row_date == latest_date).all()
    else:
        date_cols = [r[0] for r in db.session.query(FlexibleKpiUpload.column_name).filter_by(kpi_type="core").distinct().all()]
        date_cols_parsed = {dc: _parse_col_date(dc) for dc in date_cols}
        date_cols_parsed = {k: v for k, v in date_cols_parsed.items() if v}
        sorted_date_cols = sorted(date_cols_parsed.items(), key=lambda x: x[1])
        if not sorted_date_cols:
            return jsonify({"available": False, "message": "No valid date columns found."})
        min_date = sorted_date_cols[0][1]
        latest_date = sorted_date_cols[-1][1]
        latest_col = sorted_date_cols[-1][0]

        bulk_rows = db.session.query(
            FlexibleKpiUpload.kpi_name, FlexibleKpiUpload.column_name,
            sa_func.avg(FlexibleKpiUpload.num_value).label("avg_val"),
        ).filter(FlexibleKpiUpload.kpi_type == "core", FlexibleKpiUpload.kpi_name.isnot(None),
        ).group_by(FlexibleKpiUpload.kpi_name, FlexibleKpiUpload.column_name).all()

        per_kpi_all = {k: [] for k in cols_meta}
        per_kpi_latest = {k: None for k in cols_meta}
        trend_pivot = {_iso(d): {"date": _iso(d)} for _, d in sorted_date_cols}

        for kpi_db_name, col_name, avg_val in bulk_rows:
            key = db_to_key.get(kpi_db_name)
            if not key or avg_val is None:
                continue
            val = round(float(avg_val), 2)
            per_kpi_all[key].append(val)
            if col_name == latest_col:
                per_kpi_latest[key] = val
            d = date_cols_parsed.get(col_name)
            if d:
                d_str = _iso(d)
                if d_str in trend_pivot:
                    trend_pivot[d_str][key] = val

        site_rows = db.session.query(
            FlexibleKpiUpload.site_id, FlexibleKpiUpload.kpi_name, FlexibleKpiUpload.num_value,
        ).filter(FlexibleKpiUpload.kpi_type == "core", FlexibleKpiUpload.column_name == latest_col).all()

    summary = {}
    for key, meta in cols_meta.items():
        vals = per_kpi_all[key]
        avg_all = round(sum(vals) / len(vals), 2) if vals else None
        summary[key] = {"label": meta["label"], "avg": avg_all, "latest_avg": per_kpi_latest[key]}

    trend = sorted(trend_pivot.values(), key=lambda x: x["date"])

    pivot = {}
    for site_id, kpi_name, num_value in site_rows:
        key = db_to_key.get(kpi_name)
        if key:
            pivot.setdefault(site_id, {})[key] = num_value

    site_table = []
    for site_id in sorted(pivot.keys()):
        row = {"site_id": site_id}
        for key in cols_meta:
            val = pivot[site_id].get(key)
            row[key] = round(float(val), 2) if val is not None else None
        site_table.append(row)

    return jsonify({
        "available": True,
        "data_source": "site",
        "total_sites": len(site_table),
        "date_range": {"from": _iso(min_date), "to": _iso(latest_date)},
        "columns": [{"key": k, "label": m["label"]} for k, m in cols_meta.items()],
        "summary": summary,
        "trend": trend,
        "sites": site_table,
    })


@app.route("/api/cto/business-kpi", methods=["GET"])
@jwt_required()
def cto_business_kpi():
    """Business KPI dashboard.
    Users from FlexibleKpiUpload(kpi_type='business') or KpiData('Site Users').
    Revenue from FlexibleKpiUpload(kpi_type='revenue') or KpiData('Site Revenue').
    Merges both sources by site_id.
    """
    user = _require_cto_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 403

    cached = _cache_get("business_kpi")
    if cached:
        return jsonify(cached)

    return _business_kpi_merged()


def _business_kpi_merged():
    """
    Build business KPI response by merging:
      - revenue data  → flexible_kpi_uploads WHERE kpi_type='revenue'
      - users data    → flexible_kpi_uploads WHERE kpi_type='business'

    Response shape (consumed by BusinessKPI.jsx frontend):
    {
      "revenue_rows": [
        {
          "site_id": str,
          "utilization": float,          # Site Utilization %
          "opex": float,                 # total OPEX for this site
          "monthly_revenue": {           # dict of month_label → SUM(revenue)
            "Feb": 1071.34,
            "Mar": 563.71,
            ...                          # expands automatically as new months arrive
          }
        },
        ...
      ],
      "users_rows": [
        {
          "site_id": str,
          "monthly_total_users": {       # dict of month_label → SUM(total users)
            "Feb": 811,
            "Mar": 831,
            ...
          }
        },
        ...
      ]
    }

    The frontend (BusinessKPI.jsx deriveKpis()) computes ALL KPI card values,
    trend charts, site-health scores, declining/overloaded tables from these
    two arrays — no pre-computed summary is needed from the backend.

    Column detection is fully dynamic:
      Revenue file: "Feb Avg", "Mar Avg", "Feb Total", "Mar Total", "OPEX",
                    "Total Revenue", "Site Utilization", ...
      Users file:   "Feb avg users", "Mar avg users", "Feb Total Users",
                    "March Total Users", ...
    Adding new month columns to either file automatically extends all trends.
    """
    import re as _re

    # ────────────────────────────────────────────────────────────────────────
    # 1.  CONSTANTS & HELPERS
    # ────────────────────────────────────────────────────────────────────────
    MONTH_ORDER = ["jan", "feb", "mar", "apr", "may", "jun",
                   "jul", "aug", "sep", "oct", "nov", "dec"]
    MONTH_LABELS = {
        "jan": "Jan", "feb": "Feb", "mar": "Mar", "apr": "Apr",
        "may": "May", "jun": "Jun", "jul": "Jul", "aug": "Aug",
        "sep": "Sep", "oct": "Oct", "nov": "Nov", "dec": "Dec",
    }
    _MONTH_PAT = {
        "jan": r"jan(?:uary)?",   "feb": r"feb(?:ruary)?",
        "mar": r"mar(?:ch)?",     "apr": r"apr(?:il)?",
        "may": r"may",            "jun": r"jun(?:e)?",
        "jul": r"jul(?:y)?",      "aug": r"aug(?:ust)?",
        "sep": r"sep(?:t(?:ember)?)?", "oct": r"oct(?:ober)?",
        "nov": r"nov(?:ember)?",  "dec": r"dec(?:ember)?",
    }

    def _norm(s):
        return (s or "").lower().strip()

    def _month_token(col):
        cl = _norm(col)
        for m in MONTH_ORDER:
            if _re.search(rf"(^|[^a-z]){_MONTH_PAT[m]}([^a-z]|$)", cl):
                return m
        return None

    def _is_opex(col):
        return "opex" in _norm(col)

    def _is_util(col):
        cl = _norm(col)
        return "util" in cl or "utilization" in cl

    def _is_total_rev_col(col):
        cl = _norm(col)
        return ("total" in cl and "revenue" in cl) or cl in (
            "total_revenue", "total revenue", "totalrevenue"
        )

    def _is_id_like(col):
        cl = _norm(col)
        return any(k in cl for k in (
            "site_id", "site id", "siteid", "abs_id", "abs id", "absid",
            "pcid", "bandwidth", "latitude", "longitude", "lat", "lon", "lng",
            "antenna", "rf_power", "eirp", "tilt", "azimuth", "crs", "gain",
            "subscribers", "subscriber count", "zone", "state", "country",
        )) or cl in ("id",)

    def _is_rev_total_month(col):          # "Feb Total", "Mar Total" (revenue)
        cl = _norm(col)
        return "total" in cl and "opex" not in cl and "user" not in cl and "subs" not in cl

    def _is_rev_avg_month(col):            # "Feb Avg", "Mar Avg"
        cl = _norm(col)
        return "avg" in cl and "opex" not in cl and "user" not in cl and "subs" not in cl

    def _is_total_users(col):              # "Feb Total Users", "March Total Users"
        cl = _norm(col)
        return "total" in cl and ("user" in cl or "subs" in cl)

    def _is_avg_users(col):                # "Feb avg users", "March avg users"
        cl = _norm(col)
        return "avg" in cl and ("user" in cl or "subs" in cl)

    # ────────────────────────────────────────────────────────────────────────
    # 2.  LOAD RAW ROWS FROM DB
    #     One row per (site_id, column_name) in flexible_kpi_uploads.
    #     We need EVERY row (not just distinct site_id) because some sites
    #     have multiple sectors/cells — the frontend sums them.
    # ────────────────────────────────────────────────────────────────────────
    def _load_kpi_rows(kpi_type):
        """Return list of dicts {site_id, column_name, column_type, num_value, str_value}."""
        sql = db.text("""
            SELECT site_id, column_name, column_type, num_value, str_value
            FROM flexible_kpi_uploads
            WHERE kpi_type = :t
            ORDER BY site_id, id
        """)
        rows = db.session.execute(sql, {"t": kpi_type}).fetchall()
        return rows  # list of Row objects

    rev_raw = _load_kpi_rows("revenue")
    biz_raw = _load_kpi_rows("business")

    # ────────────────────────────────────────────────────────────────────────
    # 3.  DISCOVER COLUMN NAMES (from distinct column_name values)
    # ────────────────────────────────────────────────────────────────────────
    rev_numeric_cols = set()
    biz_numeric_cols = set()

    for _, col_name, col_type, num_val, _ in rev_raw:
        if col_type == "numeric" and num_val is not None:
            rev_numeric_cols.add(col_name)
    for _, col_name, col_type, num_val, _ in biz_raw:
        if col_type == "numeric" and num_val is not None:
            biz_numeric_cols.add(col_name)

    # ── Revenue: monthly revenue columns ──
    rev_total_month_map = {}   # month_token → col_name  ("Feb Total", "Mar Total")
    rev_avg_month_map   = {}   # month_token → col_name  ("Feb Avg",   "Mar Avg")
    for c in rev_numeric_cols:
        if _is_opex(c) or _is_util(c) or _is_id_like(c) or _is_total_rev_col(c):
            continue
        m = _month_token(c)
        if not m:
            continue
        if _is_rev_total_month(c) and m not in rev_total_month_map:
            rev_total_month_map[m] = c
        elif _is_rev_avg_month(c) and m not in rev_avg_month_map:
            rev_avg_month_map[m] = c
        else:
            rev_avg_month_map.setdefault(m, c)

    # Prefer "Total" columns (monthly sums match Excel); fall back to "Avg"
    rev_month_map  = rev_total_month_map if rev_total_month_map else rev_avg_month_map
    rev_month_cols = [rev_month_map[m] for m in MONTH_ORDER if m in rev_month_map]

    # ── Revenue: OPEX column(s) ──
    opex_single_col = None
    opex_month_map  = {}
    for c in rev_numeric_cols:
        if not _is_opex(c):
            continue
        m = _month_token(c)
        if m:
            opex_month_map.setdefault(m, c)
        elif opex_single_col is None:
            opex_single_col = c
    opex_month_cols = [opex_month_map[m] for m in MONTH_ORDER if m in opex_month_map]

    # ── Revenue: utilization column ──
    util_col = next((c for c in rev_numeric_cols if _is_util(c)), None)

    # ── Revenue: single "Total Revenue" column (if present) ──
    total_rev_col = next((c for c in rev_numeric_cols if _is_total_rev_col(c)), None)

    # ── Users: monthly user columns ──
    user_total_month_map = {}   # "Feb Total Users", "March Total Users"
    user_avg_month_map   = {}   # "Feb avg users",   "March avg users"
    for c in biz_numeric_cols:
        if _is_util(c) or _is_id_like(c):
            continue
        m = _month_token(c)
        if not m:
            continue
        if _is_total_users(c) and m not in user_total_month_map:
            user_total_month_map[m] = c
        elif _is_avg_users(c) and m not in user_avg_month_map:
            user_avg_month_map[m] = c

    user_month_map  = user_total_month_map if user_total_month_map else user_avg_month_map
    if not user_month_map:
        # Last resort: any month-tagged column in business upload
        for c in biz_numeric_cols:
            if _is_util(c) or _is_id_like(c):
                continue
            m = _month_token(c)
            if m and m not in user_month_map:
                user_month_map[m] = c
    user_month_cols = [user_month_map[m] for m in MONTH_ORDER if m in user_month_map]

    print(f"[BUSINESS KPI] rev_month_map={rev_month_map}, "
          f"user_month_map={user_month_map}, "
          f"opex_single={opex_single_col!r}, opex_months={opex_month_cols}, "
          f"util_col={util_col!r}, total_rev_col={total_rev_col!r}")

    # ────────────────────────────────────────────────────────────────────────
    # 4.  PIVOT RAW ROWS → per-site dicts
    #     We store ALL numeric values keyed by (site_id → {col_name: value}).
    #     Multiple sector rows for the same site_id are SUMMED so the
    #     frontend receives one entry per unique site and can re-sum correctly.
    # ────────────────────────────────────────────────────────────────────────
    # Identify columns that are SITE-LEVEL totals (repeated on every sector
    # row) — these must use MAX (i.e. pick one copy), never SUM.
    _site_level_cols = set()
    for c in rev_numeric_cols | biz_numeric_cols:
        if _is_opex(c) or _is_total_rev_col(c) or _is_util(c):
            _site_level_cols.add(c)
        # Monthly "Total" columns (e.g. "Feb Total", "Mar Total") are site-level
        m = _month_token(c)
        if m and _is_rev_total_month(c):
            _site_level_cols.add(c)
        # Monthly "Total Users" columns are also site-level
        if m and _is_total_users(c):
            _site_level_cols.add(c)

    def _pivot_rows(raw_rows, numeric_cols_set):
        """
        Returns a dict:  site_id → { col_name: float, ... }
        Site-level columns (OPEX, Total Revenue, utilization, monthly Totals)
        use MAX across sectors (pick one copy — they are repeated identically).
        Per-sector columns (e.g. Feb Avg, Mar Avg) are AVERAGED across sectors.
        """
        out = {}
        counts = {}  # site_id → { col_name: count } for averaging
        for site_id, col_name, col_type, num_val, str_val in raw_rows:
            if site_id not in out:
                out[site_id] = {}
                counts[site_id] = {}
            if col_type == "numeric" and num_val is not None:
                v = float(num_val)
                prev = out[site_id].get(col_name, 0.0)
                cnt  = counts[site_id].get(col_name, 0)
                if col_name in _site_level_cols:
                    # Site-level total — take MAX (all rows have the same value)
                    out[site_id][col_name] = max(prev, v)
                else:
                    # Per-sector value — accumulate for averaging
                    out[site_id][col_name] = prev + v
                    counts[site_id][col_name] = cnt + 1
            elif col_type == "text" and str_val is not None:
                out[site_id][col_name] = str_val

        # Average per-sector columns (not site-level)
        for sid in out:
            for col in list(out[sid].keys()):
                if col not in _site_level_cols and col in counts.get(sid, {}):
                    n = counts[sid][col]
                    if n > 1:
                        out[sid][col] = out[sid][col] / n
        return out

    rev_by_site = _pivot_rows(rev_raw, rev_numeric_cols)
    biz_by_site = _pivot_rows(biz_raw, biz_numeric_cols)

    all_site_ids_rev  = set(rev_by_site.keys())
    all_site_ids_biz  = set(biz_by_site.keys())
    all_site_ids      = sorted(all_site_ids_rev | all_site_ids_biz)

    # ── Fallback to legacy KpiData if flexible tables are empty ──
    if not all_site_ids:
        kpi_users = _latest_site_values_for_kpi("Site Users")
        kpi_rev   = _latest_site_values_for_kpi("Site Revenue")
        # Return minimal shape; frontend will show zeros gracefully
        revenue_rows = [
            {
                "site_id": sid,
                "utilization": 0.0,
                "opex": 0.0,
                "monthly_revenue": {"Latest": round(float(info.get("value", 0) or 0), 2)},
            }
            for sid, info in kpi_rev.items()
        ]
        users_rows = [
            {
                "site_id": sid,
                "monthly_total_users": {"Latest": int(round(float(info.get("value", 0) or 0)))},
            }
            for sid, info in kpi_users.items()
        ]
        result = {"revenue_rows": revenue_rows, "users_rows": users_rows}
        _cache_set("business_kpi", result)
        return jsonify(result)

    # ────────────────────────────────────────────────────────────────────────
    # 5.  PULL PRB UTILIZATION FROM KpiData
    #     Prefer this over the "Site Utilization" column in the revenue file
    #     because it comes from live network data.
    # ────────────────────────────────────────────────────────────────────────
    prb_latest = {}
    try:
        from sqlalchemy import text as _t
        r = db.session.execute(_t("""
            SELECT kpi_name FROM kpi_data
            WHERE kpi_name IS NOT NULL AND LOWER(kpi_name) LIKE '%dl%prb%util%'
            GROUP BY kpi_name ORDER BY COUNT(*) DESC LIMIT 1
        """)).fetchone()
        dl_prb_name = r[0] if r else "DL PRB Utilization (1BH)"
    except Exception:
        dl_prb_name = "DL PRB Utilization (1BH)"

    prb_latest = _latest_site_values_for_kpi(dl_prb_name)
    if not prb_latest:
        date_floor = date.today() - timedelta(days=60)
        prb_rows = KpiData.query.filter(
            KpiData.kpi_name.ilike("%prb%util%"),
            KpiData.date >= date_floor,
        ).order_by(KpiData.site_id, KpiData.date.desc()).all()
        for row in prb_rows:
            if row.site_id not in prb_latest and row.value is not None:
                prb_latest[row.site_id] = {"value": float(row.value)}

    def _site_util(sid):
        if isinstance(prb_latest.get(sid), dict):
            v = prb_latest[sid].get("value", 0) or 0
            if v:
                return float(v)
        # Fallback: "Site Utilization" column from revenue upload
        if util_col:
            return float(rev_by_site.get(sid, {}).get(util_col, 0) or 0)
        return 0.0

    def _site_opex(sid):
        d = rev_by_site.get(sid, {})
        if opex_single_col:
            return float(d.get(opex_single_col, 0) or 0)
        return sum(float(d.get(c, 0) or 0) for c in opex_month_cols)

    # ────────────────────────────────────────────────────────────────────────
    # 6.  BUILD revenue_rows — one entry per unique site_id
    #
    #     monthly_revenue dict uses canonical labels ("Feb", "Mar", …)
    #     and contains the SUMMED value across all sectors (already done by
    #     _pivot_rows).
    #
    #     We also record whether this site has users data so the frontend
    #     can enforce the "all three data points present" rule for site
    #     health and KPI-card totals.
    # ────────────────────────────────────────────────────────────────────────
    revenue_rows = []
    for sid in sorted(all_site_ids_rev):
        d = rev_by_site[sid]

        monthly_revenue = {}
        for m in MONTH_ORDER:
            col = rev_month_map.get(m)
            if col and col in d:
                monthly_revenue[MONTH_LABELS[m]] = round(float(d[col] or 0), 4)

        # If no monthly columns but "Total Revenue" column exists, put it under
        # a pseudo-month key so the frontend still has something to render.
        if not monthly_revenue and total_rev_col and total_rev_col in d:
            monthly_revenue["Total"] = round(float(d[total_rev_col] or 0), 4)

        util_val  = round(_site_util(sid), 2)
        opex_val  = round(_site_opex(sid), 4)

        # has_users: True when this site also appears in the business upload.
        # Used by frontend to restrict site-health, low-margin, and overloaded
        # tables to only sites where ALL three data dimensions are present.
        has_users = sid in all_site_ids_biz

        # "Total Revenue" column — the overall total (not monthly sum).
        # Used for ROI = (Total Revenue - OPEX) / OPEX × 100
        site_total_rev = round(float(d.get(total_rev_col, 0) or 0), 4) if total_rev_col else 0.0

        revenue_rows.append({
            "site_id":         sid,
            "utilization":     util_val,
            "opex":            opex_val,
            "monthly_revenue": monthly_revenue,
            "total_revenue":   site_total_rev,
            "has_users":       has_users,
        })

    # ────────────────────────────────────────────────────────────────────────
    # 7.  BUILD users_rows — one entry per unique site_id
    #
    #     monthly_total_users uses the same canonical month labels as
    #     monthly_revenue so the frontend can join them by label.
    #
    #     We also record whether this site has revenue data so the frontend
    #     can enforce completeness checks from the users side too.
    # ────────────────────────────────────────────────────────────────────────
    users_rows = []
    for sid in sorted(all_site_ids_biz):
        d = biz_by_site[sid]

        monthly_total_users = {}
        for m in MONTH_ORDER:
            col = user_month_map.get(m)
            if col and col in d:
                monthly_total_users[MONTH_LABELS[m]] = int(round(float(d[col] or 0)))

        has_revenue = sid in all_site_ids_rev

        users_rows.append({
            "site_id":              sid,
            "monthly_total_users":  monthly_total_users,
            "has_revenue":          has_revenue,
        })

    # ────────────────────────────────────────────────────────────────────────
    # 8.  SERVER-SIDE PRE-COMPUTATION — low_margin_sites & overloaded_sites
    #
    #     These two tables have strict data-presence rules so we compute them
    #     here (backend knows which columns exist) and ship the final sorted
    #     lists to the frontend rather than replicating the column-detection
    #     logic in JS.
    #
    #  LOW MARGIN SITES
    #  ─────────────────
    #  Eligibility:  site must have utilization + users + revenue ALL present.
    #  Definition:   Margin % = (Revenue − OPEX) / Revenue × 100
    #                Sites where Margin % < LOW_MARGIN_THRESHOLD are "low margin".
    #  Sort:         ascending margin % (worst first).
    #  Columns shown: site_id, revenue (current month), opex, margin_pct, users (current month)
    #
    #  OVERLOADED SITES
    #  ─────────────────
    #  Eligibility:  site must have BOTH utilization AND revenue data present.
    #                (users data is NOT required for overloaded sites.)
    #  Definition:   utilization ≥ OVERLOAD_THRESHOLD (80%).
    #  Sort:         utilization DESC, revenue DESC.
    #  Columns shown: site_id, utilization, revenue (current month)
    # ────────────────────────────────────────────────────────────────────────
    LOW_MARGIN_THRESHOLD = 30.0   # margin % below this is "low margin"
    OVERLOAD_THRESHOLD   = 80.0   # utilization % at or above this is "overloaded"

    # Determine latest month label present in revenue data
    _latest_rev_month_label = None
    for m in reversed(MONTH_ORDER):
        if m in rev_month_map:
            _latest_rev_month_label = MONTH_LABELS[m]
            break

    # Helper: get current-month revenue for a site from its revenue_row entry
    # (we re-index revenue_rows by site_id for O(1) lookup)
    rev_row_by_sid = {r["site_id"]: r for r in revenue_rows}
    usr_row_by_sid = {u["site_id"]: u for u in users_rows}

    def _current_rev(sid):
        r = rev_row_by_sid.get(sid, {})
        mr = r.get("monthly_revenue", {})
        if _latest_rev_month_label and _latest_rev_month_label in mr:
            return float(mr[_latest_rev_month_label])
        # Fallback: last value in dict
        vals = [v for v in mr.values() if isinstance(v, (int, float))]
        return float(vals[-1]) if vals else 0.0

    def _current_users(sid):
        u = usr_row_by_sid.get(sid, {})
        mu = u.get("monthly_total_users", {})
        if _latest_rev_month_label and _latest_rev_month_label in mu:
            return int(mu[_latest_rev_month_label])
        vals = [v for v in mu.values() if isinstance(v, (int, float))]
        return int(vals[-1]) if vals else 0

    # ── LOW MARGIN SITES ──────────────────────────────────────────────────
    # Only include sites that have ALL THREE: utilization > 0, users > 0, revenue > 0
    # Uses "Total Revenue" column (overall total) vs OPEX for margin calculation
    low_margin_sites = []
    for r in revenue_rows:
        sid   = r["site_id"]
        util  = r["utilization"]
        opex  = r["opex"]
        # Use "Total Revenue" column for margin comparison with OPEX
        rev   = r.get("total_revenue", 0) or _current_rev(sid)
        users = _current_users(sid) if r["has_users"] else 0

        # STRICT completeness: all three dimensions must be non-zero
        if util <= 0 or users <= 0 or rev <= 0:
            continue

        # Margin % = (Total Revenue − OPEX) / Total Revenue × 100
        margin_pct = round(((rev - opex) / rev) * 100, 2) if rev else 0.0

        if margin_pct < LOW_MARGIN_THRESHOLD:
            low_margin_sites.append({
                "site_id":    sid,
                "revenue":    round(rev, 2),
                "opex":       round(opex, 2),
                "margin_pct": margin_pct,
                "users":      users,
                "utilization": util,
            })

    # Sort: worst margin first
    low_margin_sites.sort(key=lambda x: x["margin_pct"])
    low_margin_sites = low_margin_sites[:20]   # cap at 20 rows

    # ── OVERLOADED SITES ──────────────────────────────────────────────────
    # Only include sites that have BOTH utilization AND revenue present.
    # Users data is NOT required.
    overloaded_sites = []
    for r in revenue_rows:
        sid  = r["site_id"]
        util = r["utilization"]
        rev  = _current_rev(sid)

        # Both utilization and revenue must be present and non-zero
        if util <= 0 or rev <= 0:
            continue

        if util >= OVERLOAD_THRESHOLD:
            overloaded_sites.append({
                "site_id":    sid,
                "utilization": util,
                "revenue":    round(rev, 2),
            })

    # Sort: highest utilization first; break ties by revenue desc
    overloaded_sites.sort(key=lambda x: (-x["utilization"], -x["revenue"]))
    overloaded_sites = overloaded_sites[:20]   # cap at 20 rows

    # ────────────────────────────────────────────────────────────────────────
    # 9.  SERVER-SIDE SUMMARY — current_month, previous_month, ARPU
    #
    #     Pre-compute authoritative month labels and KPI card values so the
    #     frontend card sub-labels are driven by the same column-detection
    #     logic that built revenue_rows / users_rows.
    #
    #     current_month  = latest month label present in revenue data
    #     previous_month = second-to-last month label (or None)
    #     total_revenue  = Sum monthly_revenue[current_month] across all sites
    #     total_users    = Sum monthly_total_users[current_month] across all sites
    #     arpu           = total_revenue / total_users  (0 if no users)
    #
    #     Frontend BusinessKPI.jsx reads summary_kpis to set:
    #       * KPI card values and sub-label month names
    #       * Period label  e.g. "Feb -> Mar"
    #       * ARPU  = "Mar total revenue / Mar total users"
    #     so "Feb Total" column revenue is never confused with the cross-month
    #     "Total Revenue" column.
    # ────────────────────────────────────────────────────────────────────────

    # Discover all canonical month labels present in revenue rows
    _all_rev_labels = set()
    for r in revenue_rows:
        _all_rev_labels.update(r.get("monthly_revenue", {}).keys())
    _LABEL_TO_IDX = {MONTH_LABELS[m]: i for i, m in enumerate(MONTH_ORDER)}
    _sorted_rev_labels = sorted(
        _all_rev_labels,
        key=lambda lb: _LABEL_TO_IDX.get(lb, 99)
    )
    _current_month_label  = _sorted_rev_labels[-1] if _sorted_rev_labels else None
    _previous_month_label = _sorted_rev_labels[-2] if len(_sorted_rev_labels) >= 2 else None

    # Total Revenue — sum across all sites for each of the two months
    _total_rev_current  = 0.0
    _total_rev_previous = 0.0
    for r in revenue_rows:
        mr = r.get("monthly_revenue", {})
        if _current_month_label:
            _total_rev_current  += float(mr.get(_current_month_label,  0) or 0)
        if _previous_month_label:
            _total_rev_previous += float(mr.get(_previous_month_label, 0) or 0)

    # Total Users — sum across all sites, aligned to the same month labels
    _total_usr_current  = 0
    _total_usr_previous = 0
    for u in users_rows:
        mu = u.get("monthly_total_users", {})
        if _current_month_label:
            _total_usr_current  += int(mu.get(_current_month_label,  0) or 0)
        if _previous_month_label:
            _total_usr_previous += int(mu.get(_previous_month_label, 0) or 0)

    # ARPU = Total Revenue (current month) / Total Users (current month)
    _arpu = round(_total_rev_current / _total_usr_current, 6)         if _total_usr_current > 0 else 0.0

    # Revenue Growth % MoM
    _rev_growth = 0.0
    if _previous_month_label and _total_rev_previous > 0:
        _rev_growth = round(
            ((_total_rev_current - _total_rev_previous) / _total_rev_previous) * 100, 2
        )

    _num_sites = len(revenue_rows) or 1

    # Overall "Total Revenue" column sum (the 10-month total, NOT monthly sum)
    _overall_total_rev = sum(r.get("total_revenue", 0) or 0 for r in revenue_rows)
    _overall_total_opex = sum(r.get("opex", 0) or 0 for r in revenue_rows)

    # Revenue at Risk = Σ(prev - current) for sites where current < previous
    _revenue_at_risk = 0.0
    if _previous_month_label and _current_month_label:
        for r in revenue_rows:
            mr = r.get("monthly_revenue", {})
            curr = float(mr.get(_current_month_label, 0) or 0)
            prev = float(mr.get(_previous_month_label, 0) or 0)
            if curr < prev:
                _revenue_at_risk += (prev - curr)

    # Churn Rate = (Users(prev) - Users(current)) / Users(prev) × 100
    _churn_rate = 0.0
    if _previous_month_label and _total_usr_previous > 0:
        _churn_rate = max(0, ((_total_usr_previous - _total_usr_current) / _total_usr_previous) * 100)

    # Network ROI = (Total Revenue - OPEX) / OPEX × 100
    # Uses the "Total Revenue" column value, NOT monthly sums
    _network_roi = round(((_overall_total_rev - _overall_total_opex) / _overall_total_opex) * 100, 2) if _overall_total_opex > 0 else 0.0

    summary_kpis = {
        "current_month":        _current_month_label,
        "previous_month":       _previous_month_label,
        "total_revenue":        round(_total_rev_current,  2),
        "total_revenue_prev":   round(_total_rev_previous, 2),
        "total_users":          _total_usr_current,
        "total_users_prev":     _total_usr_previous,
        "arpu":                 _arpu,
        "revenue_growth":       _rev_growth,
        "num_sites":            _num_sites,
        "avg_users_per_site":   round(_total_usr_current / _num_sites, 2),
        "revenue_at_risk":      round(_revenue_at_risk, 2),
        "churn_rate":           round(_churn_rate, 2),
        "network_roi":          _network_roi,
        "overall_total_revenue": round(_overall_total_rev, 2),
        "overall_total_opex":   round(_overall_total_opex, 2),
        "period_label": (
            f"{_previous_month_label} → {_current_month_label}"
            if _previous_month_label and _current_month_label
            else (_current_month_label or "")
        ),
    }

    # ────────────────────────────────────────────────────────────────────────
    # 10.  DIAGNOSTICS
    # ────────────────────────────────────────────────────────────────────────
    _sample_months_rev = list(revenue_rows[0]["monthly_revenue"].keys()) if revenue_rows else []
    _sample_months_usr = list(users_rows[0]["monthly_total_users"].keys()) if users_rows else []
    _complete_sites = sum(
        1 for r in revenue_rows
        if r["utilization"] > 0 and r["has_users"] and _current_rev(r["site_id"]) > 0
    )
    print(f"[BUSINESS KPI] revenue_rows={len(revenue_rows)}, "
          f"users_rows={len(users_rows)}, "
          f"complete_sites(util+users+rev)={_complete_sites}, "
          f"low_margin={len(low_margin_sites)}, "
          f"overloaded={len(overloaded_sites)}, "
          f"rev_months={_sample_months_rev}, "
          f"user_months={_sample_months_usr}, "
          f"current_month={_current_month_label!r}, "
          f"total_revenue={_total_rev_current:.2f}, "
          f"total_users={_total_usr_current}, "
          f"arpu={_arpu:.6f}")

    # ────────────────────────────────────────────────────────────────────────
    # 11.  ASSEMBLE RESULT & CACHE
    # ────────────────────────────────────────────────────────────────────────
    result = {
        "revenue_rows":      revenue_rows,
        "users_rows":        users_rows,
        "low_margin_sites":  low_margin_sites,
        "overloaded_sites":  overloaded_sites,
        # Authoritative summary consumed directly by BusinessKPI.jsx KPI cards.
        # Ensures "Feb Total" column revenue is not confused with cross-month
        # "Total Revenue" column — ARPU = total_revenue / total_users for the
        # same month, not a mixed-column calculation.
        "summary_kpis":      summary_kpis,
    }
    _cache_set("business_kpi", result)
    return jsonify(result)


@app.route("/api/cto/operational-kpi", methods=["GET"])
@jwt_required()
def cto_operational_kpi():
    user = _require_cto_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 403

    tickets = Ticket.query.all()
    total_tickets = len(tickets)
    resolved_tickets = [t for t in tickets if t.status == "resolved"]
    sla_breaches = len([t for t in tickets if t.sla_breached and t.status != "resolved"])
    sla_compliance = round(((total_tickets - sla_breaches) / total_tickets) * 100, 1) if total_tickets else 0

    resolution_hours = []
    for ticket in resolved_tickets:
        if ticket.created_at and ticket.resolved_at:
            resolution_hours.append((ticket.resolved_at - ticket.created_at).total_seconds() / 3600)
    avg_resolution_time = round(sum(resolution_hours) / len(resolution_hours), 2) if resolution_hours else 0

    csat_raw = db.session.query(db.func.avg(Feedback.rating)).filter(Feedback.rating > 0).scalar() or 0
    csat = round(float(csat_raw), 2)

    status_breakdown = db.session.query(Ticket.status, db.func.count(Ticket.id)).group_by(Ticket.status).all()
    status_data = [{"name": status or "unknown", "value": count} for status, count in status_breakdown]

    agent_workload = db.session.query(
        User.name,
        db.func.count(Ticket.id)
    ).outerjoin(Ticket, Ticket.assigned_to == User.id).filter(User.role == "human_agent").group_by(User.name).all()
    workload_data = [{"agent": name or "Unassigned", "tickets": count} for name, count in agent_workload]

    escalated_count = len([t for t in tickets if t.status in ("escalated", "manager_escalated")])
    escalation_rate = round((escalated_count / total_tickets) * 100, 1) if total_tickets else 0

    breach_alerts = SlaAlert.query.filter_by(recipient_role="cto").count()

    # ── Critical incidents: active tickets sorted by SLA urgency ────────────
    now_utc = datetime.now(timezone.utc)

    def _sla_remaining(t):
        dl = t.sla_deadline
        if dl.tzinfo is None:
            dl = dl.replace(tzinfo=timezone.utc)
        return (dl - now_utc).total_seconds()

    active_with_sla = [t for t in tickets if t.status not in ("resolved",) and t.sla_deadline]
    critical_sorted = sorted(active_with_sla, key=_sla_remaining)[:10]

    critical_incidents = []
    for t in critical_sorted:
        rem = _sla_remaining(t)
        abs_s = abs(rem)
        h = int(abs_s // 3600)
        m = int((abs_s % 3600) // 60)
        s_val = int(abs_s % 60)
        sign = "-" if rem < 0 else ""
        critical_incidents.append({
            "id": t.reference_number,
            "db_id": t.id,
            "service": t.category or "General",
            "subcategory": t.subcategory or "",
            "priority": t.priority or "low",
            "sla_clock": f"{sign}{h:02d}:{m:02d}:{s_val:02d}",
            "sla_remaining": round(rem),
            "status": t.status or "pending",
            "description": (t.description[:200] if t.description else "No description"),
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "sla_hours": t.sla_hours,
            "assigned_to": t.assignee.name if t.assignee else "Unassigned",
            "sla_breached": bool(t.sla_breached),
            "resolution_notes": t.resolution_notes or "",
        })

    # ── Escalation trend: last 7 days daily escalated-ticket counts ─────────
    escalation_trend = []
    for i in range(7):
        day_start = (now_utc - timedelta(days=6 - i)).replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        count = sum(
            1 for t in tickets
            if t.status in ("escalated", "manager_escalated") and t.created_at and
            day_start <= (t.created_at.replace(tzinfo=timezone.utc) if t.created_at.tzinfo is None else t.created_at) < day_end
        )
        escalation_trend.append(count)

    # ── Ticket growth % vs previous period ────────────────────────────────
    period_days = 7
    period_start = now_utc - timedelta(days=period_days)
    prev_period_start = period_start - timedelta(days=period_days)

    def _in_range(t, start, end):
        if not t.created_at:
            return False
        ca = t.created_at.replace(tzinfo=timezone.utc) if t.created_at.tzinfo is None else t.created_at
        return start <= ca < end

    current_period_count = sum(1 for t in tickets if _in_range(t, period_start, now_utc))
    prev_period_count = sum(1 for t in tickets if _in_range(t, prev_period_start, period_start))
    ticket_growth_pct = round(((current_period_count - prev_period_count) / prev_period_count) * 100, 1) if prev_period_count else 0.0

    # ── Resolution time change vs previous period ─────────────────────────
    prev_resolved = [t for t in tickets if t.status == "resolved" and t.created_at and t.resolved_at and _in_range(t, prev_period_start, period_start)]
    prev_res_hours = [(t.resolved_at - t.created_at).total_seconds() / 3600 for t in prev_resolved if t.resolved_at and t.created_at]
    prev_avg_res = round(sum(prev_res_hours) / len(prev_res_hours), 2) if prev_res_hours else 0
    resolution_change = round(avg_resolution_time - prev_avg_res, 2)

    # ── Escalation rate change vs previous period ─────────────────────────
    prev_escalated = sum(1 for t in tickets if t.status in ("escalated", "manager_escalated") and _in_range(t, prev_period_start, period_start))
    prev_esc_rate = round((prev_escalated / prev_period_count) * 100, 1) if prev_period_count else 0.0
    escalation_rate_change = round(escalation_rate - prev_esc_rate, 1)

    # ── Highest breach category ───────────────────────────────────────────
    breach_by_category = {}
    for t in tickets:
        if t.sla_breached:
            cat = t.category or "General"
            breach_by_category[cat] = breach_by_category.get(cat, 0) + 1
    top_breach_category = max(breach_by_category, key=breach_by_category.get) if breach_by_category else ""

    # ── Escalation commentary (dynamic) ───────────────────────────────────
    esc_this_week = sum(escalation_trend[-7:])
    esc_prev_week = sum(escalation_trend[:7]) if len(escalation_trend) >= 14 else 0
    if escalation_rate == 0:
        esc_comment = "No escalations recorded in the current period."
    elif escalation_rate_change < -1:
        esc_comment = f"Escalation rate decreased by {abs(escalation_rate_change)}% compared to last period."
    elif escalation_rate_change > 1:
        esc_comment = f"Escalation rate increased by {escalation_rate_change}% — review agent capacity and routing rules."
    else:
        esc_comment = "Escalation rate is stable compared to last period."

    return jsonify({
        "summary": {
            "total_tickets": total_tickets,
            "sla_compliance": sla_compliance,
            "sla_breaches": sla_breaches,
            "avg_resolution_time": avg_resolution_time,
            "csat": csat,
            "escalation_rate": escalation_rate,
            "breach_alerts": breach_alerts,
            "ticket_growth_pct": ticket_growth_pct,
            "resolution_change": resolution_change,
            "escalation_rate_change": escalation_rate_change,
            "top_breach_category": top_breach_category,
            "escalation_comment": esc_comment,
        },
        "status_breakdown": status_data,
        "agent_workload": workload_data,
        "critical_incidents": critical_incidents,
        "escalation_trend": escalation_trend,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# CDO ENGAGEMENT KPI ENDPOINT
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/cto/cdo-engagement-kpi", methods=["GET"])
@jwt_required()
def cto_cdo_engagement_kpi():
    user = _require_cto_user()
    if not user:
        return jsonify({"error": "Unauthorized"}), 403

    from sqlalchemy import func as sa_func, extract
    now_utc = datetime.now(timezone.utc)

    # ── 1. Resolution Funnel ─────────────────────────────────────────────────
    total_conversations = ChatSession.query.count()
    escalated_session_ids = set(
        r[0] for r in db.session.query(Ticket.chat_session_id).filter(
            Ticket.chat_session_id.isnot(None)
        ).all()
    )
    escalated_count = len(escalated_session_ids)
    ai_resolved = ChatSession.query.filter(
        ChatSession.status == "resolved",
        ~ChatSession.id.in_(escalated_session_ids) if escalated_session_ids else ChatSession.id.isnot(None),
    ).count()
    human_resolved = Ticket.query.filter_by(status="resolved").count()

    ai_resolution_rate = round((ai_resolved / total_conversations) * 100, 1) if total_conversations else 0
    escalation_rate_funnel = round((escalated_count / total_conversations) * 100, 1) if total_conversations else 0

    funnel = {
        "conversations": total_conversations,
        "ai_resolved": ai_resolved,
        "escalated": escalated_count,
        "human_resolved": human_resolved,
        "ai_resolution_rate": ai_resolution_rate,
        "escalation_rate": escalation_rate_funnel,
    }

    # ── 2. Customer Sentiment Analysis ────────────────────────────────────────
    ratings = db.session.query(Feedback.rating, sa_func.count(Feedback.id)).filter(
        Feedback.rating > 0
    ).group_by(Feedback.rating).all()
    rating_map = {int(r): c for r, c in ratings}
    total_feedback = sum(rating_map.values())

    sentiment_labels = {5: "Excellent", 4: "Good", 3: "Neutral", 2: "Poor", 1: "Bad"}
    sentiment = []
    for score in [5, 4, 3, 2, 1]:
        count = rating_map.get(score, 0)
        sentiment.append({
            "label": sentiment_labels[score],
            "count": count,
            "pct": round((count / total_feedback) * 100, 1) if total_feedback else 0,
        })

    positive_score = round(((rating_map.get(5, 0) + rating_map.get(4, 0)) / total_feedback) * 100, 1) if total_feedback else 0
    csat_index = round(sum(r * c for r, c in rating_map.items()) / total_feedback, 2) if total_feedback else 0

    sentiment_data = {
        "total": total_feedback,
        "distribution": sentiment,
        "positive_score": positive_score,
        "csat_index": csat_index,
    }

    # ── 3. Weekly Activity Heatmap ────────────────────────────────────────────
    from models import ChatMessage
    activity_rows = db.session.query(
        extract("dow", ChatMessage.created_at).label("dow"),
        extract("hour", ChatMessage.created_at).label("hr"),
        sa_func.count(ChatMessage.id),
    ).filter(
        ChatMessage.created_at.isnot(None),
    ).group_by("dow", "hr").all()

    # Build heatmap: {day: {hour: count}}
    heatmap = []
    total_activity = sum(r[2] for r in activity_rows)
    max_activity = max((r[2] for r in activity_rows), default=1)
    for dow, hr, cnt in activity_rows:
        heatmap.append({"day": int(dow), "hour": int(hr), "count": cnt})

    # Peak slots and idle calculation
    total_slots = 7 * 24
    active_slots = len(activity_rows)
    idle_pct = round(((total_slots - active_slots) / total_slots) * 100, 1)
    workforce_util = round((active_slots / total_slots) * 100, 1)

    heatmap_data = {
        "cells": heatmap,
        "peak_activity": max_activity,
        "idle_pct": idle_pct,
        "workforce_util": workforce_util,
    }

    # ── 4. Predictive Workload Forecast ───────────────────────────────────────
    # Daily ticket counts for last 14 days + predict next 7
    daily_counts = db.session.query(
        sa_func.date_trunc("day", Ticket.created_at).label("day"),
        sa_func.count(Ticket.id),
    ).filter(
        Ticket.created_at.isnot(None),
        Ticket.created_at >= now_utc - timedelta(days=14),
    ).group_by("day").order_by("day").all()

    # Fill in missing days
    workload_series = []
    for i in range(14):
        d = (now_utc - timedelta(days=13 - i)).date()
        count = next((c for day, c in daily_counts if day.date() == d), 0)
        workload_series.append({"date": d.isoformat(), "count": count})

    # Simple linear forecast for next 7 days
    vals = [p["count"] for p in workload_series]
    n = len(vals)
    if n >= 2:
        x_mean = (n - 1) / 2.0
        y_mean = sum(vals) / n
        num = sum((i - x_mean) * (v - y_mean) for i, v in enumerate(vals))
        den = sum((i - x_mean) ** 2 for i in range(n))
        slope = num / den if den else 0
        intercept = y_mean - slope * x_mean
        forecast_vals = [max(0, round(slope * (n + i) + intercept)) for i in range(7)]
    else:
        forecast_vals = [0] * 7

    forecast_series = []
    for i, v in enumerate(forecast_vals):
        d = (now_utc + timedelta(days=i + 1)).date()
        forecast_series.append({"date": d.isoformat(), "count": v, "type": "forecast"})

    # Combine last 7 actual + 7 forecast
    combined_workload = [
        {**p, "type": "actual"} for p in workload_series[-7:]
    ] + forecast_series

    total_assigned = Ticket.query.count()
    total_completed = Ticket.query.filter_by(status="resolved").count()
    current_rate = round((total_completed / total_assigned) * 100) if total_assigned else 0
    target_rate = 90
    gap_to_target = target_rate - current_rate

    workload_data = {
        "series": combined_workload,
        "current_rate": current_rate,
        "target_rate": target_rate,
        "gap_to_target": gap_to_target,
        "peak_day": max(combined_workload, key=lambda x: x["count"])["date"] if combined_workload else None,
    }

    # ── 5. Alerts ─────────────────────────────────────────────────────────────
    alerts = []
    if escalation_rate_funnel > 30:
        alerts.append({"type": "alert", "message": f"Escalation Rate {escalation_rate_funnel}% exceeds 30% threshold"})
    if ai_resolution_rate < 70:
        alerts.append({"type": "warning", "message": f"AI Resolution Rate {ai_resolution_rate}% below 70% target"})
    if positive_score < 85:
        alerts.append({"type": "critical", "message": f"Sentiment Score {positive_score}% below 85% threshold"})
    if idle_pct > 40:
        alerts.append({"type": "info", "message": f"Idle Time {idle_pct}% — optimization needed"})

    return jsonify({
        "funnel": funnel,
        "sentiment": sentiment_data,
        "heatmap": heatmap_data,
        "workload": workload_data,
        "alerts": alerts,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# SLA ALERT DASHBOARD ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/manager/sla-alerts", methods=["GET"])
@jwt_required()
def manager_sla_alerts():
    user = User.query.get(int(get_jwt_identity()))
    role = (getattr(user, "role", "") or "").strip().lower()
    if role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403
    unread_only = request.args.get("unread_only", "false").lower() == "true"
    q = SlaAlert.query.filter_by(recipient_role="manager")
    if unread_only:
        q = q.filter_by(is_read=False)
    alerts = q.order_by(SlaAlert.created_at.desc()).limit(100).all()
    return jsonify({"alerts": [a.to_dict() for a in alerts]})


@app.route("/api/cto/sla-alerts", methods=["GET"])
@jwt_required()
def cto_sla_alerts():
    user = User.query.get(int(get_jwt_identity()))
    role = (getattr(user, "role", "") or "").strip().lower()
    if role not in ("cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403
    unread_only = request.args.get("unread_only", "false").lower() == "true"
    q = SlaAlert.query.filter_by(recipient_role="cto")
    if unread_only:
        q = q.filter_by(is_read=False)
    alerts = q.order_by(SlaAlert.created_at.desc()).limit(100).all()
    return jsonify({"alerts": [a.to_dict() for a in alerts]})


@app.route("/api/manager/sla-alerts/<int:alert_id>/read", methods=["PUT"])
@jwt_required()
def manager_mark_alert_read(alert_id):
    user = User.query.get(int(get_jwt_identity()))
    role = (getattr(user, "role", "") or "").strip().lower()
    if role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403
    alert = SlaAlert.query.get(alert_id)
    if not alert or (alert.recipient_role or "").strip().lower() != "manager":
        return jsonify({"error": "Alert not found"}), 404
    alert.is_read = True
    db.session.commit()
    return jsonify({"success": True})


@app.route("/api/cto/sla-alerts/<int:alert_id>/read", methods=["PUT"])
@jwt_required()
def cto_mark_alert_read(alert_id):
    user = User.query.get(int(get_jwt_identity()))
    role = (getattr(user, "role", "") or "").strip().lower()
    if role not in ("cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403
    alert = SlaAlert.query.get(alert_id)
    if not alert or (alert.recipient_role or "").strip().lower() != "cto":
        return jsonify({"error": "Alert not found"}), 404
    alert.is_read = True
    db.session.commit()
    return jsonify({"success": True})


@app.route("/api/manager/sla-alerts/read-all", methods=["PUT"])
@jwt_required()
def manager_mark_all_alerts_read():
    user = User.query.get(int(get_jwt_identity()))
    role = (getattr(user, "role", "") or "").strip().lower()
    if role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403
    SlaAlert.query.filter_by(recipient_role="manager", is_read=False).update({"is_read": True})
    db.session.commit()
    return jsonify({"success": True})


@app.route("/api/cto/sla-alerts/read-all", methods=["PUT"])
@jwt_required()
def cto_mark_all_alerts_read():
    user = User.query.get(int(get_jwt_identity()))
    role = (getattr(user, "role", "") or "").strip().lower()
    if role not in ("cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403
    SlaAlert.query.filter_by(recipient_role="cto", is_read=False).update({"is_read": True})
    db.session.commit()
    return jsonify({"success": True})


# ═══════════════════════════════════════════════════════════════════════════════
# CR SLA ALERTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/cto/cr-alerts", methods=["GET"])
@jwt_required()
def cto_cr_alerts():
    """CR SLA alerts for CTO — breach + 90% warnings."""
    from models import CrSlaAlert
    user = User.query.get(int(get_jwt_identity()))
    role = (getattr(user, "role", "") or "").strip().lower()
    if role not in ("cto", "admin"):
        return jsonify({"error": "Unauthorized"}), 403
    unread_only = request.args.get("unread_only", "false").lower() == "true"
    q = CrSlaAlert.query.filter(CrSlaAlert.recipient_role == "cto")
    if unread_only:
        q = q.filter_by(is_read=False)
    alerts = q.order_by(CrSlaAlert.created_at.desc()).limit(100).all()
    return jsonify({"alerts": [a.to_dict() for a in alerts]})


@app.route("/api/manager/cr-alerts", methods=["GET"])
@jwt_required()
def manager_cr_alerts():
    """CR SLA alerts for Manager — 75% warnings + breaches."""
    from models import CrSlaAlert
    user = User.query.get(int(get_jwt_identity()))
    if user.role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403
    unread_only = request.args.get("unread_only", "false").lower() == "true"
    q = CrSlaAlert.query.filter(CrSlaAlert.recipient_role == "manager")
    if unread_only:
        q = q.filter_by(is_read=False)
    alerts = q.order_by(CrSlaAlert.created_at.desc()).limit(100).all()
    return jsonify({"alerts": [a.to_dict() for a in alerts]})


@app.route("/api/cto/cr-alerts/<int:alert_id>/read", methods=["PUT"])
@jwt_required()
def cto_mark_cr_alert_read(alert_id):
    from models import CrSlaAlert
    user = User.query.get(int(get_jwt_identity()))
    if user.role != "cto":
        return jsonify({"error": "Unauthorized"}), 403
    alert = CrSlaAlert.query.get(alert_id)
    if not alert:
        return jsonify({"error": "Alert not found"}), 404
    alert.is_read = True
    db.session.commit()
    return jsonify({"success": True})


@app.route("/api/manager/cr-alerts/<int:alert_id>/read", methods=["PUT"])
@jwt_required()
def manager_mark_cr_alert_read(alert_id):
    from models import CrSlaAlert
    user = User.query.get(int(get_jwt_identity()))
    if user.role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403
    alert = CrSlaAlert.query.get(alert_id)
    if not alert:
        return jsonify({"error": "Alert not found"}), 404
    alert.is_read = True
    db.session.commit()
    return jsonify({"success": True})


@app.route("/api/cto/cr-alerts/read-all", methods=["PUT"])
@jwt_required()
def cto_mark_all_cr_alerts_read():
    from models import CrSlaAlert
    user = User.query.get(int(get_jwt_identity()))
    if user.role != "cto":
        return jsonify({"error": "Unauthorized"}), 403
    CrSlaAlert.query.filter_by(recipient_role="cto", is_read=False).update({"is_read": True})
    db.session.commit()
    return jsonify({"success": True})


@app.route("/api/manager/cr-alerts/read-all", methods=["PUT"])
@jwt_required()
def manager_mark_all_cr_alerts_read():
    from models import CrSlaAlert
    user = User.query.get(int(get_jwt_identity()))
    if user.role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403
    CrSlaAlert.query.filter_by(recipient_role="manager", is_read=False).update({"is_read": True})
    db.session.commit()
    return jsonify({"success": True})


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE ID GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

ROLE_PREFIX = {
    "manager": "MGR",
    "human_agent": "HA",
    "cto": "CTO",
    "admin": "ADM",
}


def generate_employee_id(role):
    prefix = ROLE_PREFIX.get(role)
    if not prefix:
        return None
    existing = User.query.filter(User.employee_id.like(f"{prefix}%")).order_by(User.employee_id.desc()).first()
    if existing and existing.employee_id:
        num = int(existing.employee_id[len(prefix):]) + 1
    else:
        num = 1
    return f"{prefix}{num:05d}"


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/admin/dashboard", methods=["GET"])
@jwt_required()
def admin_dashboard():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    # User counts by role
    user_counts = db.session.query(
        User.role, db.func.count(User.id)
    ).group_by(User.role).all()

    total_users = sum(c[1] for c in user_counts)

    # Chat stats — single GROUP BY query
    chat_stats = db.session.query(
        ChatSession.status, db.func.count(ChatSession.id)
    ).group_by(ChatSession.status).all()
    chat_map = dict(chat_stats)
    total_chats = sum(chat_map.values())
    resolved_chats = chat_map.get("resolved", 0)
    escalated_chats = chat_map.get("escalated", 0)
    active_chats = chat_map.get("active", 0)

    # Ticket status stats — single GROUP BY query
    ticket_status_stats = db.session.query(
        Ticket.status, db.func.count(Ticket.id)
    ).group_by(Ticket.status).all()
    ts_map = dict(ticket_status_stats)
    total_tickets = sum(ts_map.values())
    pending_tickets = ts_map.get("pending", 0)
    in_progress_tickets = ts_map.get("in_progress", 0)
    resolved_tickets = ts_map.get("resolved", 0)

    # Ticket priority stats — single GROUP BY query
    ticket_priority_stats = db.session.query(
        Ticket.priority, db.func.count(Ticket.id)
    ).group_by(Ticket.priority).all()
    tp_map = dict(ticket_priority_stats)
    critical_tickets = tp_map.get("critical", 0)
    high_tickets = tp_map.get("high", 0)

    # Feedback — single query for count, avg, and satisfied
    fb_agg = db.session.query(
        db.func.count(Feedback.id),
        db.func.avg(sql_case((Feedback.rating > 0, Feedback.rating))),
        db.func.sum(sql_case((Feedback.rating >= 4, 1), else_=0)),
    ).first()
    total_feedback = fb_agg[0] or 0
    avg_rating = fb_agg[1] or 0
    satisfied_count = fb_agg[2] or 0
    csat_score = round((satisfied_count / max(total_feedback, 1)) * 100, 1)

    # Resolution rate
    resolution_rate = round((resolved_chats / max(total_chats, 1)) * 100, 1)

    # Category breakdown
    categories = db.session.query(
        ChatSession.sector_name, db.func.count(ChatSession.id)
    ).group_by(ChatSession.sector_name).all()

    return jsonify({
        "stats": {
            "total_users": total_users,
            "total_chats": total_chats,
            "resolved_chats": resolved_chats,
            "escalated_chats": escalated_chats,
            "active_chats": active_chats,
            "total_tickets": total_tickets,
            "pending_tickets": pending_tickets,
            "in_progress_tickets": in_progress_tickets,
            "resolved_tickets": resolved_tickets,
            "critical_tickets": critical_tickets,
            "high_tickets": high_tickets,
            "total_feedback": total_feedback,
            "avg_rating": round(float(avg_rating), 1),
            "csat_score": csat_score,
            "resolution_rate": resolution_rate,
        },
        "user_breakdown": [{"role": r[0], "count": r[1]} for r in user_counts],
        "category_breakdown": [{"name": c[0] or "Unknown", "count": c[1]} for c in categories],
    })


@app.route("/api/admin/users", methods=["GET"])
@jwt_required()
def admin_list_users():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    role_filter = request.args.get("role")
    search = request.args.get("search")

    query = User.query
    if role_filter:
        query = query.filter_by(role=role_filter)
    if search:
        query = query.filter(
            db.or_(
                User.name.ilike(f"%{search}%"),
                User.email.ilike(f"%{search}%"),
            )
        )

    users = query.order_by(User.created_at.desc()).all()
    return jsonify({"users": [u.to_dict() for u in users]})


@app.route("/api/admin/users", methods=["POST"])
@jwt_required()
def admin_create_user():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    data = request.json
    name = data.get("name", "").strip()
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")
    role = data.get("role", "customer").lower()
    phone_number = data.get("phone_number", "").strip()

    if not name or not email or not password:
        return jsonify({"error": "Name, email, and password are required"}), 400
    pw_err = validate_password(password)
    if pw_err:
        return jsonify({"error": pw_err}), 400
    if role not in ("customer", "manager", "human_agent", "cto", "admin"):
        return jsonify({"error": "Invalid role"}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({"error": "Email already registered"}), 409

    emp_id = generate_employee_id(role)
    new_user = User(name=name, email=email, role=role, employee_id=emp_id)
    if phone_number:
        new_user.phone_number = phone_number
    if role == "customer":
        ut = (data.get("user_type") or "bronze").strip().lower()
        new_user.user_type = ut if ut in VALID_USER_TYPES else "bronze"
    new_user.set_password(password)
    db.session.add(new_user)
    db.session.commit()
    return jsonify({"user": new_user.to_dict()}), 201


@app.route("/api/admin/users/upload", methods=["POST"])
@jwt_required()
def admin_upload_users():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Only .xlsx or .xlsm files are supported (.xls is not supported)."}), 400
    ok, err = _validate_ooxml_excel_upload(file)
    if not ok:
        return jsonify({"error": err}), 400

    from openpyxl import load_workbook
    import io

    try:
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb.active
    except Exception as e:
        return jsonify({"error": f"Could not read Excel workbook. Please upload a valid .xlsx/.xlsm file. Details: {str(e)}"}), 400

    # Parse headers from first row
    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
    required = {"name", "email", "role"}
    header_set = set(headers)
    if not required.issubset(header_set):
        missing = required - header_set
        return jsonify({"error": f"Missing required columns: {', '.join(missing)}. Required: Name, Email, Role"}), 400

    col_map = {h: i for i, h in enumerate(headers)}
    valid_roles = {"manager", "human_agent", "cto", "admin"}
    created = 0
    updated = 0
    skipped = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = str(row[col_map["name"]] or "").strip()
        email = str(row[col_map["email"]] or "").strip().lower()
        role = str(row[col_map["role"]] or "").strip().lower().replace(" ", "_")

        if not name or not email:
            skipped.append(f"Row {row_num}: missing name or email")
            continue
        if role not in valid_roles:
            skipped.append(f"Row {row_num}: invalid role '{role}' for {email}")
            continue

        # Check for employee_id column
        emp_id_from_excel = None
        if "employee id" in col_map:
            emp_id_from_excel = str(row[col_map["employee id"]] or "").strip() or None
        elif "employee_id" in col_map:
            emp_id_from_excel = str(row[col_map["employee_id"]] or "").strip() or None

        existing = User.query.filter_by(email=email).first()
        if existing:
            existing.name = name
            existing.role = role
            if emp_id_from_excel:
                existing.employee_id = emp_id_from_excel
            elif not existing.employee_id:
                existing.employee_id = generate_employee_id(role)
            updated += 1
        else:
            emp_id = emp_id_from_excel or generate_employee_id(role)
            new_user = User(name=name, email=email, role=role, employee_id=emp_id)
            new_user.set_password("Welcome@123")
            db.session.add(new_user)
            created += 1

    db.session.commit()
    return jsonify({
        "message": f"Upload complete: {created} created, {updated} updated",
        "created": created,
        "updated": updated,
        "skipped": skipped,
    })


@app.route("/api/admin/users/<int:uid>", methods=["PUT"])
@jwt_required()
def admin_update_user(uid):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    target = User.query.get(uid)
    if not target:
        return jsonify({"error": "User not found"}), 404

    data = request.json
    if "name" in data:
        target.name = data["name"].strip()
    if "email" in data:
        new_email = data["email"].strip().lower()
        existing = User.query.filter_by(email=new_email).first()
        if existing and existing.id != uid:
            return jsonify({"error": "Email already in use"}), 409
        target.email = new_email
    if "role" in data:
        new_role = data["role"].lower()
        if new_role not in ("customer", "manager", "human_agent", "cto", "admin"):
            return jsonify({"error": "Invalid role"}), 400
        if uid == user_id and new_role != "admin":
            return jsonify({"error": "Cannot change your own role"}), 400
        target.role = new_role
        if new_role == "customer":
            target.employee_id = None
        else:
            target.employee_id = generate_employee_id(new_role)
    if "password" in data and data["password"]:
        pw_err = validate_password(data["password"])
        if pw_err:
            return jsonify({"error": pw_err}), 400
        target.set_password(data["password"])
    if "user_type" in data and target.role == "customer":
        ut = (data["user_type"] or "bronze").strip().lower()
        target.user_type = ut if ut in VALID_USER_TYPES else "bronze"
    # Agent-specific fields (expertise, location, domain)
    if "domain" in data and target.role == "human_agent":
        target.domain = (data["domain"] or "").strip()
    if "location" in data and target.role == "human_agent":
        target.location = (data["location"] or "").strip()
    if "expertise" in data and target.role == "human_agent":
        target.expertise = (data["expertise"] or "").strip()
    if "specialization" in data and target.role == "human_agent":
        target.specialization = (data["specialization"] or "").strip()

    db.session.commit()
    return jsonify({"user": target.to_dict()})


@app.route("/api/admin/users/<int:uid>", methods=["DELETE"])
@jwt_required()
def admin_delete_user(uid):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    if uid == user_id:
        return jsonify({"error": "Cannot delete your own account"}), 400

    target = User.query.get(uid)
    if not target:
        return jsonify({"error": "User not found"}), 404

    # Delete associated data
    Feedback.query.filter_by(user_id=uid).delete()
    ChatMessage.query.filter(
        ChatMessage.session_id.in_(
            db.session.query(ChatSession.id).filter_by(user_id=uid)
        )
    ).delete(synchronize_session=False)
    Ticket.query.filter_by(user_id=uid).delete()
    ChatSession.query.filter_by(user_id=uid).delete()
    db.session.delete(target)
    db.session.commit()
    return jsonify({"message": "User deleted"})


@app.route("/api/admin/experts", methods=["GET"])
@jwt_required()
def admin_list_experts():
    """List all domain experts (human_agents) with their domain/location/capacity."""
    user = db.session.get(User, int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    experts = User.query.filter_by(role="human_agent").order_by(User.domain, User.name).all()
    result = []
    for e in experts:
        d = e.to_dict()
        d["open_tickets"] = _open_ticket_count(e.id)
        result.append(d)
    return jsonify({"experts": result})


@app.route("/api/admin/experts/<int:uid>", methods=["PUT"])
@jwt_required()
def admin_update_expert(uid):
    """Update domain, location, or bandwidth_capacity of an expert."""
    user = db.session.get(User, int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    expert = db.session.get(User, uid)
    if not expert or expert.role != "human_agent":
        return jsonify({"error": "Expert not found"}), 404

    data = request.json or {}
    if "domain" in data:
        if data["domain"] not in VALID_EXPERT_DOMAINS:
            return jsonify({"error": f"Invalid domain. Valid: {sorted(VALID_EXPERT_DOMAINS)}"}), 400
        expert.domain = data["domain"]
    if "location" in data:
        expert.location = (data["location"] or "").strip() or None
    if "bandwidth_capacity" in data:
        cap = int(data["bandwidth_capacity"])
        if cap < 1:
            return jsonify({"error": "bandwidth_capacity must be >= 1"}), 400
        expert.bandwidth_capacity = cap
    db.session.commit()
    d = expert.to_dict()
    d["open_tickets"] = _open_ticket_count(expert.id)
    return jsonify({"expert": d})


@app.route("/api/admin/agent-tickets", methods=["GET"])
@jwt_required()
def admin_agent_tickets():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    status = request.args.get("status")
    agent_id = request.args.get("agent_id")
    search = request.args.get("search")

    # Alias User for the assignee join
    AgentUser = db.aliased(User)
    query = (
        Ticket.query
        .join(AgentUser, Ticket.assigned_to == AgentUser.id)
        .filter(AgentUser.role == "human_agent")
    )

    if status:
        query = query.filter(Ticket.status == status)
    if agent_id:
        query = query.filter(Ticket.assigned_to == int(agent_id))
    if search:
        CustomerUser = db.aliased(User)
        query = (
            query
            .join(CustomerUser, Ticket.user_id == CustomerUser.id)
            .filter(db.or_(
                CustomerUser.name.ilike(f"%{search}%"),
                CustomerUser.email.ilike(f"%{search}%"),
                Ticket.reference_number.ilike(f"%{search}%"),
            ))
        )

    tickets = query.order_by(Ticket.created_at.desc()).all()
    agents = User.query.filter_by(role="human_agent").order_by(User.name).all()

    return jsonify({
        "tickets": [t.to_dict() for t in tickets],
        "agents": [{"id": a.id, "name": a.name} for a in agents],
    })


@app.route("/api/admin/agent-alerts", methods=["GET"])
@jwt_required()
def admin_agent_alerts():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    alerts = []

    # 1. New escalations (tickets escalated in last 7 days, assigned to human agents)
    week_ago = datetime.now(timezone.utc) - timedelta(days=7)
    escalated = Ticket.query.join(
        User, Ticket.assigned_to == User.id
    ).filter(
        User.role == "human_agent",
        Ticket.status == "escalated",
        Ticket.created_at >= week_ago,
    ).order_by(Ticket.created_at.desc()).all()
    for t in escalated:
        alerts.append({
            "type": "escalation",
            "severity": "high",
            "title": f"Escalated: {t.reference_number}",
            "message": f"{t.category} - {t.subcategory or 'General'} (Customer: {t.user.name if t.user else 'Unknown'})",
            "time": t.created_at.isoformat() if t.created_at else None,
            "ticket_id": t.id,
        })

    # 2. Critical/high priority pending tickets assigned to agents
    critical = Ticket.query.join(
        User, Ticket.assigned_to == User.id
    ).filter(
        User.role == "human_agent",
        Ticket.priority.in_(["critical", "high"]),
        Ticket.status.in_(["pending", "in_progress"]),
    ).order_by(Ticket.created_at.asc()).all()
    for t in critical:
        alerts.append({
            "type": "critical_ticket",
            "severity": "critical" if t.priority == "critical" else "high",
            "title": f"{t.priority.upper()} priority: {t.reference_number}",
            "message": f"Assigned to {t.assignee.name if t.assignee else 'Unassigned'} - {t.category} ({t.status.replace('_', ' ')})",
            "time": t.created_at.isoformat() if t.created_at else None,
            "ticket_id": t.id,
        })

    # 3. Low ratings (1-2 stars) from last 7 days linked to agent sessions
    low_feedbacks = db.session.query(Feedback, ChatSession, Ticket).join(
        ChatSession, Feedback.chat_session_id == ChatSession.id
    ).join(
        Ticket, Ticket.chat_session_id == ChatSession.id
    ).join(
        User, Ticket.assigned_to == User.id
    ).filter(
        User.role == "human_agent",
        Feedback.rating <= 2,
        Feedback.rating > 0,
        Feedback.created_at >= week_ago,
    ).order_by(Feedback.created_at.desc()).all()
    for fb, session, ticket in low_feedbacks:
        alerts.append({
            "type": "low_rating",
            "severity": "warning",
            "title": f"Low rating ({fb.rating}/5) on {ticket.reference_number}",
            "message": fb.comment or f"Customer rated {fb.rating}/5 for {session.sector_name or 'General'} issue",
            "time": fb.created_at.isoformat() if fb.created_at else None,
            "ticket_id": ticket.id,
        })

    # 4. Overdue tickets (pending for more than 3 days)
    three_days_ago = datetime.now(timezone.utc) - timedelta(days=3)
    overdue = Ticket.query.join(
        User, Ticket.assigned_to == User.id
    ).filter(
        User.role == "human_agent",
        Ticket.status.in_(["pending", "in_progress"]),
        Ticket.created_at <= three_days_ago,
    ).order_by(Ticket.created_at.asc()).all()
    for t in overdue:
        days_old = (datetime.utcnow() - t.created_at).days if t.created_at else 0
        alerts.append({
            "type": "overdue",
            "severity": "warning",
            "title": f"Overdue ({days_old}d): {t.reference_number}",
            "message": f"Assigned to {t.assignee.name if t.assignee else 'Unassigned'} - {t.status.replace('_', ' ')} for {days_old} days",
            "time": t.created_at.isoformat() if t.created_at else None,
            "ticket_id": t.id,
        })

    # Sort all alerts: critical first, then high, then warning, then by time
    severity_order = {"critical": 0, "high": 1, "warning": 2}
    alerts.sort(key=lambda a: (severity_order.get(a["severity"], 3), a["time"] or ""))

    return jsonify({"alerts": alerts, "total": len(alerts)})


@app.route("/api/admin/feedback", methods=["GET"])
@jwt_required()
def admin_feedback():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    feedbacks = Feedback.query.order_by(Feedback.created_at.desc()).all()
    result = []
    for f in feedbacks:
        fd = f.to_dict()
        if f.chat_session:
            fd["session_sector"] = f.chat_session.sector_name
            fd["session_subprocess"] = f.chat_session.subprocess_name
        result.append(fd)
    return jsonify({"feedbacks": result})


# ═══════════════════════════════════════════════════════════════════════════════
# SITE & KPI DATA UPLOAD (Admin)
# ═══════════════════════════════════════════════════════════════════════════════

def haversine(lat1, lon1, lat2, lon2):
    """Calculate distance in km between two lat/lng points."""
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _validate_ooxml_excel_upload(file_storage):
    """Validate that uploaded file is a real OOXML workbook (.xlsx/.xlsm), not renamed/corrupt/temp file."""
    name = (file_storage.filename or "").strip()
    if name.startswith("~$"):
        return False, "You selected an Excel temporary lock file (~$...). Please upload the actual workbook."

    # OOXML files are ZIP containers and start with 'PK'
    try:
        stream = file_storage.stream
        pos = stream.tell()
        magic = stream.read(4)
        stream.seek(pos)
    except Exception:
        magic = b""

    if not magic.startswith(b"PK"):
        return False, (
            "Invalid Excel workbook format. Please upload a real .xlsx/.xlsm file "
            "(not .xls, not CSV renamed to .xlsx, and not a temporary file)."
        )

    return True, None


def _normalize_excel_header(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _find_site_id_column(headers):
    """Find a site identifier column across multiple workbook formats."""
    for i, header in enumerate(headers):
        normalized = _normalize_excel_header(header)
        if normalized in {"siteid", "homesitecode", "cellsiteid"}:
            return i
        if "site" in normalized and "id" in normalized:
            return i
    return None


def _extract_excel_date_columns(headers, skip_indices=None):
    """Return (column_index, date) for headers that can be parsed as dates."""
    skip_indices = set(skip_indices or [])
    date_columns = []
    for col_idx, header in enumerate(headers):
        if col_idx in skip_indices or header is None:
            continue
        try:
            if isinstance(header, datetime):
                date_columns.append((col_idx, header.date()))
            elif isinstance(header, str):
                for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%b-%y", "%d-%b-%Y"):
                    try:
                        date_columns.append((col_idx, datetime.strptime(header.strip(), fmt).date()))
                        break
                    except ValueError:
                        continue
            elif hasattr(header, "date"):
                date_columns.append((col_idx, header.date()))
        except Exception:
            continue
    return date_columns


SHARED_WORKBOOK_KPI_NAMES = ("Site Users", "Site Revenue")


@app.route("/api/admin/upload-sites", methods=["POST"])
@jwt_required()
def admin_upload_sites():
    """Upload site data Excel (site_id, latitude, longitude, zone)."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Only .xlsx or .xlsm files are supported (.xls is not supported)."}), 400
    ok, err = _validate_ooxml_excel_upload(file)
    if not ok:
        return jsonify({"error": err}), 400

    import openpyxl
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active

    # ── Auto-detect columns from first row headers ──────────────────────────
    raw_headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    headers = [h.lower() for h in raw_headers]
    col_map = {}
    extra_col_indices = []  # indices of columns that don't map to known fields

    # Known field matchers: (field_name, matcher_fn)
    _KNOWN_MATCHERS = [
        ("site_id",            lambda h: ("site" in h and "id" in h and "abs" not in h) or h in ("site_id", "siteid")),
        ("site_abs_id",        lambda h: ("site" in h and "abs" in h) or h in ("site_abs_id", "siteabsid", "abs_id", "absid")),
        ("site_name",          lambda h: h in ("site_name", "sitename", "site name") or (h == "site")),
        ("cell_id",            lambda h: h in ("cell_name", "cell name", "cellname", "cell_id", "cell id", "cellid", "cell")),
        ("vendor_name",        lambda h: h in ("vendor", "vendor_name", "vendor name", "vendorname")),
        ("latitude",           lambda h: "lat" in h),
        ("longitude",          lambda h: "lon" in h),
        ("province",           lambda h: h in ("province", "prov", "region") or "province" in h),
        ("commune",            lambda h: h in ("commune", "comm") or "commune" in h),
        ("zone",               lambda h: "zone" in h or "cluster" in h),
        ("city",               lambda h: h in ("city", "ville")),
        ("state",              lambda h: h in ("state",)),
        ("site_status",        lambda h: h == "status" or "site status" in h or "site_status" in h),
        ("alarms",             lambda h: "alarm" in h),
        ("solution",           lambda h: h in ("solution", "standard solution step", "standard solution", "standard_solution_step")),
        ("bandwidth_mhz",      lambda h: "bandwidth" in h or h == "bandwidth_mhz"),
        ("antenna_gain_dbi",   lambda h: ("antenna" in h and "gain" in h) or h == "antenna_gain_dbi"),
        ("rf_power_eirp_dbm",  lambda h: ("rf" in h and "power" in h) or "eirp" in h or h == "rf_power_eirp_dbm"),
        ("antenna_height_agl_m", lambda h: ("antenna" in h and "height" in h) or h == "antenna_height_agl_m"),
        ("e_tilt_degree",      lambda h: "tilt" in h or h == "e_tilt_degree"),
        ("crs_gain",           lambda h: "crs" in h or h == "crs_gain"),
        ("country",            lambda h: h in ("country",)),
        ("technology",         lambda h: h in ("technology", "tech", "tech_type", "technology_type",
                                                "network type", "network_type", "networktype",
                                                "radio", "radio_type", "radio access technology", "rat")),
    ]

    for i, h in enumerate(headers):
        if not h:
            continue
        matched = False
        for field_name, matcher in _KNOWN_MATCHERS:
            if field_name not in col_map and matcher(h):
                col_map[field_name] = i
                matched = True
                break
        if not matched:
            extra_col_indices.append(i)

    required = ["site_id"]
    missing = [k for k in required if k not in col_map]
    if missing:
        return jsonify({"error": f"Missing columns: {', '.join(missing)}. Found headers: {headers}"}), 400

    # Helper to read a string cell value
    def _str_cell(row, key, default=""):
        idx = col_map.get(key)
        if idx is None or idx >= len(row) or row[idx] is None:
            return default
        return str(row[idx]).strip()

    # Helper to read a float cell value
    def _float_cell(row, key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row) or row[idx] is None:
            return None
        try:
            return float(row[idx])
        except (ValueError, TypeError):
            return None

    # Clear existing site data so upload is a full replace
    TelecomSite.query.delete()
    db.session.flush()

    status_map = {
        "active": "on_air", "on_air": "on_air", "on air": "on_air",
        "down": "off_air", "off_air": "off_air", "off air": "off_air",
        "alarm": "off_air",
    }

    created = 0
    skipped = []
    detected_extra_cols = [raw_headers[i] for i in extra_col_indices if i < len(raw_headers) and raw_headers[i]]

    def _clean_key(raw):
        """Normalise a header into a safe JSON key (no newlines, no control chars)."""
        s = str(raw or "").replace("\n", " ").replace("\r", " ").replace("\t", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s or "unknown"

    def _clean_val(v):
        """Coerce a cell value into something JSON-serialisable."""
        if v is None:
            return None
        if isinstance(v, bool):
            return v
        if isinstance(v, (int, float)):
            # Guard against NaN/Inf which are not valid JSON.
            try:
                if math.isnan(v) or math.isinf(v):
                    return None
            except Exception:
                pass
            return v
        # Everything else → trimmed, newline-free string
        s = str(v).replace("\n", " ").replace("\r", " ").strip()
        return s[:500] if len(s) > 500 else s

    # ── Build a list of plain dicts first (no ORM objects) ────────
    # This lets us use bulk_insert_mappings and deduplicate (site_id, cell_id)
    # before ever touching the DB.
    mappings = []
    seen_keys = set()  # (site_id, cell_id_or_None) — enforces unique constraint client-side

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            sid_raw = row[col_map["site_id"]] if col_map.get("site_id") is not None and col_map["site_id"] < len(row) else None
            if sid_raw is None:
                skipped.append(f"Row {row_idx}: empty site_id")
                continue
            sid = str(sid_raw).strip()
            if not sid or sid.lower() == "none":
                skipped.append(f"Row {row_idx}: empty site_id")
                continue
            lat = _float_cell(row, "latitude") or 0.0
            lon = _float_cell(row, "longitude") or 0.0
        except Exception as e:
            skipped.append(f"Row {row_idx}: parse error: {e}")
            continue

        cell_id_val = _str_cell(row, "cell_id") or None
        dedup_key = (sid, cell_id_val)
        if dedup_key in seen_keys:
            skipped.append(f"Row {row_idx}: duplicate site_id={sid} cell_id={cell_id_val}")
            continue
        seen_keys.add(dedup_key)

        raw_status = _str_cell(row, "site_status", "on_air").lower()
        site_status = status_map.get(raw_status, raw_status or "on_air")

        # Build extra_params dict — sanitised keys/values.
        extra_params = {}
        for idx in extra_col_indices:
            if idx < len(row) and row[idx] is not None:
                raw_label = raw_headers[idx] if idx < len(raw_headers) else f"col_{idx}"
                key = _clean_key(raw_label)
                val = _clean_val(row[idx])
                if val is not None and val != "":
                    extra_params[key] = val

        province_val = _str_cell(row, "province")
        commune_val = _str_cell(row, "commune")
        zone_val = _str_cell(row, "zone")
        if not province_val and zone_val:
            province_val = zone_val

        mappings.append({
            "site_id": sid,
            "site_abs_id": _str_cell(row, "site_abs_id") or None,
            "site_name": _str_cell(row, "site_name") or None,
            "cell_id": cell_id_val,
            "vendor_name": _str_cell(row, "vendor_name") or None,
            "latitude": lat,
            "longitude": lon,
            "province": province_val or "",
            "commune": commune_val or "",
            "zone": zone_val or province_val or "",
            "city": _str_cell(row, "city") or None,
            "state": _str_cell(row, "state") or None,
            "country": _str_cell(row, "country") or "",
            "technology": _str_cell(row, "technology") or None,
            "site_status": site_status,
            "alarms": _str_cell(row, "alarms") or "",
            "solution": _str_cell(row, "solution") or "",
            "standard_solution_step": "",
            "bandwidth_mhz": _float_cell(row, "bandwidth_mhz"),
            "antenna_gain_dbi": _float_cell(row, "antenna_gain_dbi"),
            "rf_power_eirp_dbm": _float_cell(row, "rf_power_eirp_dbm"),
            "antenna_height_agl_m": _float_cell(row, "antenna_height_agl_m"),
            "e_tilt_degree": _float_cell(row, "e_tilt_degree"),
            "crs_gain": _float_cell(row, "crs_gain"),
            "extra_params": extra_params if extra_params else None,
        })

    # ── Bulk insert in chunks with per-chunk fallback to row-by-row ──
    CHUNK = 500
    try:
        for i in range(0, len(mappings), CHUNK):
            chunk = mappings[i:i + CHUNK]
            try:
                db.session.bulk_insert_mappings(TelecomSite, chunk)
                db.session.flush()
                created += len(chunk)
            except Exception as chunk_err:
                # Fallback: insert row-by-row so one bad row doesn't kill the chunk
                db.session.rollback()
                app.logger.warning(f"Chunk {i}-{i+len(chunk)} failed ({type(chunk_err).__name__}); retrying row-by-row")
                for r_idx, r in enumerate(chunk):
                    try:
                        db.session.bulk_insert_mappings(TelecomSite, [r])
                        db.session.flush()
                        created += 1
                    except Exception as row_err:
                        db.session.rollback()
                        underlying = getattr(row_err, "orig", None)
                        err_msg = str(underlying) if underlying else str(row_err)
                        skipped.append(f"site_id={r.get('site_id')} cell_id={r.get('cell_id')}: {type(row_err).__name__}: {err_msg[:200]}")
                        # Log the first 5 bad rows in full so we can see the real cause
                        if len([s for s in skipped if 'site_id=' in s]) <= 5:
                            app.logger.error(f"Bad row: {r}")
                            app.logger.error(f"Underlying error: {err_msg}")

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        underlying = getattr(e, "orig", None)
        err_msg = str(underlying) if underlying else str(e)
        return jsonify({
            "error": f"{type(e).__name__}: {err_msg[:500]}",
            "created_before_error": created,
            "skipped_sample": skipped[-10:],
            "total_skipped": len(skipped),
        }), 500

    # Auto-populate city/state/country from lat/lng for sites missing geo data
    try:
        _auto_populate_geo()
    except Exception as e:
        print(f">>> Geo auto-populate failed (non-fatal): {e}")

    clear_analytics_cache()
    return jsonify({
        "created": created, "skipped": skipped, "total": created,
        "detected_columns": list(col_map.keys()),
        "extra_columns": detected_extra_cols,
    })


def _auto_populate_geo():
    """Populate province, commune, city, state, country for telecom_sites.
    Dynamic: uses province/commune from upload if available, infers technology from site_id.
    Copies province→state and commune→city for backward compatibility when those fields are empty.
    No hardcoded country/city mappings — all geo comes from the uploaded data."""
    try:
        sites = TelecomSite.query.filter(
            TelecomSite.latitude.isnot(None),
            TelecomSite.longitude.isnot(None),
        ).all()

        updated = 0
        for s in sites:
            changed = False

            # Infer technology from site_id (e.g. "GUR_LTE_001" → "LTE")
            if not s.technology and s.site_id:
                sid_upper = s.site_id.upper()
                for tech_name in ("5G", "NR", "LTE", "4G", "3G", "UMTS", "2G", "GSM"):
                    if tech_name in sid_upper:
                        s.technology = "5G" if tech_name == "NR" else "4G" if tech_name == "LTE" else tech_name
                        changed = True
                        break

            # Sync province → state and commune → city for backward compat
            if s.province and not s.state:
                s.state = s.province
                changed = True
            if s.commune and not s.city:
                s.city = s.commune
                changed = True
            # Also sync zone from province if zone is empty
            if s.province and not s.zone:
                s.zone = s.province
                changed = True

            if changed:
                updated += 1

        if updated:
            db.session.commit()
            print(f">>> Auto-populated geo data for {updated} sites")
    except Exception as e:
        print(f">>> Geo population error: {e}")
        db.session.rollback()


@app.route("/api/admin/populate-geo", methods=["POST"])
@jwt_required()
def admin_populate_geo():
    """Manually trigger geo-population for all sites missing city/state/country."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    _auto_populate_geo()
    clear_analytics_cache()
    count = TelecomSite.query.filter(TelecomSite.city.isnot(None), TelecomSite.city != '').count()
    total = TelecomSite.query.count()
    return jsonify({"message": f"Geo populated: {count}/{total} sites have city data"})


@app.route("/api/admin/upload-kpi-site-level", methods=["POST"])
@jwt_required()
def admin_upload_kpi_site_level():
    """Upload site-level KPI workbook (27 sheets, sheet name = KPI name).
    Each sheet: Site_ID column, then date columns with values."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Only .xlsx or .xlsm files are supported (.xls is not supported)."}), 400
    ok, err = _validate_ooxml_excel_upload(file)
    if not ok:
        return jsonify({"error": err}), 400

    import io, openpyxl
    from bulk_insert import bulk_insert_from_sheet_site

    # Read file into memory so openpyxl can read it reliably
    raw_bytes = file.read()
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)

    total_inserted = 0
    kpi_summary = []
    errors = []

    try:
        for ws in wb.worksheets:
            kpi_name = ws.title.strip()
            if not kpi_name:
                continue

            rows_iter = ws.iter_rows(values_only=True)
            try:
                headers = next(rows_iter)
            except StopIteration:
                errors.append(f"Sheet '{kpi_name}': empty sheet")
                continue

            if not headers or len(headers) < 2:
                errors.append(f"Sheet '{kpi_name}': insufficient columns")
                continue

            # Auto-detect identifier columns and date columns
            id_col_map = {}
            date_columns = []
            for col_idx, h in enumerate(headers):
                if h is None:
                    continue
                h_str = str(h).strip().lower() if isinstance(h, str) else ""

                # Check if it's an identifier column
                if isinstance(h, str):
                    if ("site" in h_str and "id" in h_str and "abs" not in h_str) or h_str in ("site_id", "siteid"):
                        id_col_map["site_id"] = col_idx
                        continue
                    elif ("site" in h_str and "abs" in h_str) or h_str in ("site_abs_id", "siteabsid", "abs_id"):
                        id_col_map["site_abs_id"] = col_idx
                        continue
                    elif h_str in ("vendor", "vendor_name", "vendor name", "vendorname"):
                        id_col_map["vendor_name"] = col_idx
                        continue
                    elif h_str in ("site_name", "site name", "sitename"):
                        id_col_map["site_name"] = col_idx
                        continue

                # Try to parse as date
                try:
                    if isinstance(h, datetime):
                        date_columns.append((col_idx, h.date()))
                    elif isinstance(h, str):
                        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
                            try:
                                date_columns.append((col_idx, datetime.strptime(h.strip(), fmt).date()))
                                break
                            except ValueError:
                                continue
                    elif hasattr(h, 'date'):
                        date_columns.append((col_idx, h.date()))
                except Exception:
                    continue

            # Default site_id to column 0 if not explicitly found
            if "site_id" not in id_col_map:
                id_col_map["site_id"] = 0

            if not date_columns:
                errors.append(f"Sheet '{kpi_name}': no valid date columns found")
                continue

            sheet_inserted = bulk_insert_from_sheet_site(db, rows_iter, kpi_name, date_columns, id_col_map=id_col_map)
            total_inserted += sheet_inserted
            kpi_summary.append({"name": kpi_name, "rows": sheet_inserted})
            app.logger.info(f"Site-level upload: sheet '{kpi_name}' done — {sheet_inserted} rows")
    except Exception as e:
        app.logger.error(f"Site-level upload error: {e}")
        return jsonify({"error": f"Upload failed: {e}"}), 500
    finally:
        wb.close()

    clear_analytics_cache()
    return jsonify({
        "inserted": total_inserted,
        "kpis_processed": len(kpi_summary),
        "kpi_summary": kpi_summary,
        "errors": errors,
    })


@app.route("/api/admin/upload-shared-site-workbook", methods=["POST"])
@jwt_required()
def admin_upload_shared_site_workbook():
    """Upload the shared telecom_site_dataset workbook format as site-level KPI data."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx or .xls files accepted"}), 400

    import openpyxl
    wb = openpyxl.load_workbook(file, data_only=True)

    total_inserted = 0
    kpi_summary = []
    errors = []

    for ws in wb.worksheets:
        kpi_name = ws.title.strip()
        if not kpi_name:
            continue

        headers = [c.value for c in ws[1]]
        if not headers:
            errors.append(f"Sheet '{kpi_name}': missing header row")
            continue

        site_id_col = _find_site_id_column(headers)
        if site_id_col is None:
            errors.append(f"Sheet '{kpi_name}': no HomeSitecode/site ID column found")
            continue

        date_columns = _extract_excel_date_columns(headers, skip_indices={site_id_col})
        if not date_columns:
            errors.append(f"Sheet '{kpi_name}': no valid date columns found")
            continue

        sheet_inserted = 0
        batch = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            site_id = str(row[site_id_col]).strip() if site_id_col < len(row) and row[site_id_col] else None
            if not site_id or site_id == "None":
                continue

            for col_idx, date_val in date_columns:
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        val = float(row[col_idx])
                    except (ValueError, TypeError):
                        continue
                    batch.append(KpiData(
                        site_id=site_id, kpi_name=kpi_name, date=date_val,
                        hour=0, value=val, data_level="site"
                    ))
                    sheet_inserted += 1

                    if len(batch) >= 2000:
                        db.session.bulk_save_objects(batch)
                        batch = []

        if batch:
            db.session.bulk_save_objects(batch)

        total_inserted += sheet_inserted
        kpi_summary.append({"name": kpi_name, "rows": sheet_inserted})

    db.session.commit()
    return jsonify({
        "message": f"Shared workbook data added to database: {total_inserted} records inserted across {len(kpi_summary)} sheets.",
        "inserted": total_inserted,
        "kpis_processed": len(kpi_summary),
        "kpi_summary": kpi_summary,
        "errors": errors,
    })


@app.route("/api/admin/shared-site-workbook-summary", methods=["GET"])
@jwt_required()
def admin_shared_site_workbook_summary():
    """Return summary for shared workbook KPI data only."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    total_records = db.session.query(db.func.count(KpiData.id)).filter(
        KpiData.data_level == "site",
        KpiData.kpi_name.in_(SHARED_WORKBOOK_KPI_NAMES)
    ).scalar() or 0

    total_sites = db.session.query(db.func.count(db.distinct(KpiData.site_id))).filter(
        KpiData.data_level == "site",
        KpiData.kpi_name.in_(SHARED_WORKBOOK_KPI_NAMES)
    ).scalar() or 0

    return jsonify({
        "total_sites": total_sites,
        "total_records": total_records,
    })


@app.route("/api/admin/delete-shared-site-workbook", methods=["DELETE"])
@jwt_required()
def admin_delete_shared_site_workbook():
    """Delete all shared workbook KPI data only."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    count = KpiData.query.filter(
        KpiData.data_level == "site",
        KpiData.kpi_name.in_(SHARED_WORKBOOK_KPI_NAMES)
    ).count()
    KpiData.query.filter(
        KpiData.data_level == "site",
        KpiData.kpi_name.in_(SHARED_WORKBOOK_KPI_NAMES)
    ).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({"deleted": count})


@app.route("/api/admin/debug-upload", methods=["POST", "OPTIONS"])
def debug_upload():
    """Debug endpoint — no auth, reads Excel structure for troubleshooting."""
    if 'file' not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files['file']
    import io as _io, openpyxl
    raw = f.read()
    wb = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True, read_only=True)
    result = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True, max_row=4))
        result[ws.title] = []
        for ri, row in enumerate(rows):
            result[ws.title].append([f"{type(c).__name__}:{str(c)[:40]}" if c is not None else None for c in row[:10]])
    wb.close()
    return jsonify({"sheets": len(result), "preview": result})


@app.route("/api/admin/upload-kpi-cell-level", methods=["POST"])
@jwt_required()
def admin_upload_kpi_cell_level():
    """Upload cell-level KPI workbook (27 sheets, sheet name = KPI name).
    Each sheet: Site_ID, Cell_ID, Cell_Site_ID columns, then date columns with values."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Only .xlsx or .xlsm files are supported (.xls is not supported)."}), 400
    ok, err = _validate_ooxml_excel_upload(file)
    if not ok:
        return jsonify({"error": err}), 400

    import io, openpyxl
    from bulk_insert import bulk_insert_from_sheet_cell

    raw_bytes = file.read()
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)

    total_inserted = 0
    kpi_summary = []
    errors = []

    try:
        for ws in wb.worksheets:
            kpi_name = ws.title.strip()
            if not kpi_name:
                continue

            rows_iter = ws.iter_rows(values_only=True)
            try:
                headers = next(rows_iter)
            except StopIteration:
                errors.append(f"Sheet '{kpi_name}': empty sheet")
                continue

            if not headers or len(headers) < 3:
                errors.append(f"Sheet '{kpi_name}': insufficient columns")
                continue

            # Auto-detect identifier columns and date columns
            id_col_map = {}
            date_columns = []
            for col_idx, h in enumerate(headers):
                if h is None:
                    continue
                h_str = str(h).strip().lower() if isinstance(h, str) else ""

                # Check if it's an identifier column
                if isinstance(h, str):
                    if ("site" in h_str and "id" in h_str and "abs" not in h_str) or h_str in ("site_id", "siteid"):
                        id_col_map["site_id"] = col_idx
                        continue
                    elif ("site" in h_str and "abs" in h_str) or h_str in ("site_abs_id", "siteabsid", "abs_id"):
                        id_col_map["site_abs_id"] = col_idx
                        continue
                    elif h_str in ("cell_name", "cell name", "cellname", "cell_id", "cell id", "cellid", "cell"):
                        id_col_map["cell_id"] = col_idx
                        continue
                    elif h_str in ("cell_site_id", "cell site id", "cellsiteid"):
                        id_col_map["cell_site_id"] = col_idx
                        continue
                    elif h_str in ("vendor", "vendor_name", "vendor name", "vendorname"):
                        id_col_map["vendor_name"] = col_idx
                        continue
                    elif h_str in ("site_name", "site name", "sitename"):
                        id_col_map["site_name"] = col_idx
                        continue

                # Try to parse as date
                try:
                    if isinstance(h, datetime):
                        date_columns.append((col_idx, h.date()))
                    elif isinstance(h, str):
                        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
                            try:
                                date_columns.append((col_idx, datetime.strptime(h.strip(), fmt).date()))
                                break
                            except ValueError:
                                continue
                    elif hasattr(h, 'date'):
                        date_columns.append((col_idx, h.date()))
                except Exception:
                    continue

            # Defaults if not auto-detected (legacy format: col0=Site_ID, col1=Cell_ID, col2=Cell_Site_ID)
            if "site_id" not in id_col_map:
                id_col_map["site_id"] = 0
            if "cell_id" not in id_col_map:
                id_col_map["cell_id"] = 1
            if "cell_site_id" not in id_col_map and "cell_site_id" not in id_col_map:
                id_col_map["cell_site_id"] = 2

            if not date_columns:
                errors.append(f"Sheet '{kpi_name}': no valid date columns found")
                continue

            sheet_inserted = bulk_insert_from_sheet_cell(db, rows_iter, kpi_name, date_columns, id_col_map=id_col_map)
            total_inserted += sheet_inserted
            kpi_summary.append({"name": kpi_name, "rows": sheet_inserted})
            app.logger.info(f"Cell-level upload: sheet '{kpi_name}' done — {sheet_inserted} rows")
    except Exception as e:
        app.logger.error(f"Cell-level upload error: {e}")
        return jsonify({"error": f"Upload failed: {e}"}), 500
    finally:
        wb.close()

    clear_analytics_cache()
    return jsonify({
        "inserted": total_inserted,
        "kpis_processed": len(kpi_summary),
        "kpi_summary": kpi_summary,
        "errors": errors,
    })


# ── Core Component KPI Upload (MME / SGW / PGW / HSS / PCRF) ─────────────
# Mapping: which KPI belongs to which component type
# All 77 Core KPIs mapped to their component type
CORE_KPI_COMPONENT_MAP = {
    # ── MME (13 KPIs) ──────────────────────────────────────────────────────
    "attach success rate":            "MME",
    "service request success rate":   "MME",
    "service request sr":             "MME",
    "paging success rate":            "MME",
    "csfb success rate":              "MME",
    "srvcc success rate":             "MME",
    "hosr":                           "MME",
    "context setup failure rate":     "MME",
    "tau success rate":               "MME",
    "bearer count":                   "MME",
    "tcp session count":              "MME",
    "sctp association status":        "MME",
    "thread process queue depth":     "MME",
    # ── SGW (12 KPIs) ──────────────────────────────────────────────────────
    "create session success rate":    "SGW",
    "modify bearer success rate":     "SGW",
    "delete session success rate":    "SGW",
    "handover data path success rate":"SGW",
    "sgw relocation success rate":    "SGW",
    "packet loss":                    "SGW",
    "packet loss (ul/dl)":            "SGW",
    "user plane latency":             "SGW",
    "user plane latency (ms)":        "SGW",
    "throughput utilization":         "SGW",
    "bearer count vs capacity":       "SGW",
    "gtp-u tunnel availability":      "SGW",
    # ── PGW (15 KPIs) ──────────────────────────────────────────────────────
    "default bearer setup success rate": "PGW",
    "dedicated bearer success rate":  "PGW",
    "session setup success rate":     "PGW",
    "pdn session setup sr":           "PGW",
    "pcrf pcf interaction success rate": "PGW",
    "packet loss dl ul":              "PGW",
    "throughput per subscriber":      "PGW",
    "throughput per subscriber (%)":  "PGW",
    "jitter":                         "PGW",
    "jitter (ms)":                    "PGW",
    "charging record success rate":   "PGW",
    "online charging latency":        "PGW",
    "online charging latency (ms)":   "PGW",
    "policy enforcement accuracy":    "PGW",
    "policy enforcement accuracy (%)":"PGW",
    "ip allocation success rate":     "PGW",
    "end-to-end core latency":        "PGW",
    "active sessions vs capacity":    "PGW",
    "nat table utilization":          "PGW",
    # ── HSS (19 KPIs) ──────────────────────────────────────────────────────
    "authentication success rate":    "HSS",
    "location update success rate":   "HSS",
    "profile retrieval success rate": "HSS",
    "authentication failure rate":    "HSS",
    "s6a transaction success rate":   "HSS",
    "s6a response latency":           "HSS",
    "s6a response latency (ms)":      "HSS",
    "s6a timeout rate":               "HSS",
    "cx success rate":                "HSS",
    "cx response latency":            "HSS",
    "cx response latency (ms)":       "HSS",
    "db query success rate":          "HSS",
    "db replication lag":             "HSS",
    "db replication lag (ms)":        "HSS",
    "db lock contention rate":        "HSS",
    "db lock contention rate (%)":    "HSS",
    "diameter transaction sr":        "HSS",
    "authentication response time":   "HSS",
    "db utilization":                 "HSS",
    "diameter tps vs capacity":       "HSS",
    "thread worker queue depth":      "HSS",
    "thread worker queue depth (%)":  "HSS",
    "hss availability":               "HSS",
    "geo redundancy sync status":     "HSS",
    "geo redundancy sync status (%)": "HSS",
    "failover time":                  "HSS",
    "failover time (s)":              "HSS",
    # ── PCRF (18 KPIs) ─────────────────────────────────────────────────────
    "policy decision success rate":   "PCRF",
    "session establishment success rate": "PCRF",
    "charging rule install success rate": "PCRF",
    "charging rule update success rate":  "PCRF",
    "gx success rate":                "PCRF",
    "gx response latency":            "PCRF",
    "gx response latency (ms)":       "PCRF",
    "gx timeout rate":                "PCRF",
    "ocs sy gy success rate":         "PCRF",
    "credit control failure rate":    "PCRF",
    "active policy sessions":         "PCRF",
    "policy session retention rate":  "PCRF",
    "dedicated bearer trigger accuracy": "PCRF",
    "diameter tps usage":             "PCRF",
    "diameter message latency":       "PCRF",
    "policy rule provisioning delay": "PCRF",
    "thread queue depth":             "PCRF",
    "thread queue depth (%)":         "PCRF",
    "pcrf availability":              "PCRF",
}

# KPIs shared across multiple components (detected by keyword)
CORE_SHARED_KPIS = {
    "cpu utilization": True,     # All 5 components have this
    "memory utilization": True,  # All 5 components have this
}


def _detect_component_type_from_kpi(kpi_name):
    """Detect component type from KPI name using fuzzy matching.
    Shared KPIs (CPU/Memory Utilization) return None — component type
    must be determined from the component_id column instead."""
    kn = (kpi_name or "").strip().lower()
    # Shared KPIs exist on all components — can't determine type from name alone
    if kn in CORE_SHARED_KPIS:
        return None  # Caller must use component_id to determine type
    if kn in CORE_KPI_COMPONENT_MAP:
        return CORE_KPI_COMPONENT_MAP[kn]
    # Fuzzy fallback
    for key, comp in CORE_KPI_COMPONENT_MAP.items():
        if key in kn or kn in key:
            return comp
    return "UNKNOWN"


@app.route("/api/admin/upload-core-component-kpi", methods=["POST"])
@jwt_required()
def admin_upload_core_component_kpi():
    """Upload Core Component KPI workbook — fully flexible auto-detection.

    Only requirement: sheet name = KPI name.

    Auto-detects the layout. Supports at minimum:
      Layout A (standard):
        Row 0 = title row (skipped if no DATE/TIME header)
        Row 1 = headers: DATE | TIME (HH:MM) | comp1 | comp2 | … | (blank) | SUMMARY …
        Row 2+ = data:   datetime | "HH:MM"  | float | float | …
      Layout B (transposed):
        Col 0 = Component_Type, Col 1 = Component_ID, remaining cols = datetimes
    """
    from bulk_insert import bulk_insert_core_component_rows
    import uuid

    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role not in ("admin", "cto"):
        return jsonify({"error": "Unauthorized"}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Only .xlsx or .xlsm files are supported."}), 400
    ok, err = _validate_ooxml_excel_upload(file)
    if not ok:
        return jsonify({"error": err}), 400

    import io as _io, openpyxl
    import re as _re

    raw_bytes = file.read()
    wb = openpyxl.load_workbook(_io.BytesIO(raw_bytes), data_only=True, read_only=True)

    batch_id = str(uuid.uuid4())[:12]
    total_inserted = 0
    kpi_summary = []
    errors = []

    def _detect_comp_type_from_id_static(comp_id):
        cid = (comp_id or "").strip().upper()
        for prefix in ("MME", "SGW", "PGW", "HSS", "PCRF"):
            if cid.startswith(prefix): return prefix
        return "UNKNOWN"

    # Detect which component type this file is for (from first sheet header col 0)
    _upload_comp_type = None
    try:
        first_ws = wb.worksheets[0]
        first_rows = list(first_ws.iter_rows(values_only=True, max_row=2))
        if first_rows:
            for row in first_rows:
                if row and row[0]:
                    _ct = _detect_comp_type_from_id_static(str(row[0]).strip())
                    if _ct != "UNKNOWN":
                        _upload_comp_type = _ct
                        break
    except Exception:
        pass

    # Delete only this component's data (not all components)
    try:
        with db.engine.connect() as conn:
            if _upload_comp_type:
                conn.execute(text("DELETE FROM core_component_kpi WHERE component_type = :ct"), {"ct": _upload_comp_type})
                app.logger.info(f"Deleted existing {_upload_comp_type} data before upload")
            else:
                conn.execute(text("DELETE FROM core_component_kpi"))
                app.logger.info("Deleted ALL core component data (could not detect specific component type)")
            conn.commit()
    except Exception:
        pass

    def _detect_comp_type_from_id(comp_id):
        """Extract component type from component ID like MME1 → MME, SGW2 → SGW."""
        cid = (comp_id or "").strip().upper()
        for prefix in ("MME", "SGW", "PGW", "HSS", "PCRF"):
            if cid.startswith(prefix):
                return prefix
        return "UNKNOWN"

    def _parse_time_str(ts):
        """Parse time string 'HH:MM' or 'H:MM' → (hour, minute)."""
        m = _re.match(r'^(\d{1,2}):(\d{2})$', (ts or "").strip())
        if m:
            return int(m.group(1)), int(m.group(2))
        return None, None

    def _is_header_row(row):
        """Check if a row looks like a header.
        Returns True if row has DATE/TIME columns OR has component-like ID strings."""
        non_none = [c for c in row[:8] if c is not None]
        if len(non_none) < 2:
            return False
        for c in row[:4]:
            if isinstance(c, str):
                ct = c.strip().upper()
                if len(ct) < 25 and (ct == "DATE" or ct.startswith("TIME") or ct == "TIMESTAMP"):
                    return True
        # Also check if row has component IDs like MME1, SGW2, etc. in later columns
        for c in row[1:8]:
            if isinstance(c, str):
                cs = c.strip().upper()
                for prefix in ("MME", "SGW", "PGW", "HSS", "PCRF"):
                    if cs.startswith(prefix) and len(cs) <= 8:
                        return True
        # Check if row has datetime objects in later columns (transposed format)
        dt_count = sum(1 for c in row[1:8] if isinstance(c, datetime) or (isinstance(c, str) and _re.match(r'\d{4}-\d{2}-\d{2}', (c or '').strip())))
        if dt_count >= 2:
            return True
        return False

    def _find_data_columns(header_row):
        """Find component-ID columns from header row.
        Returns list of (col_idx, component_id) for columns that look like
        component IDs (e.g. MME1, SGW2, HSS1, PCRF1, PGW3).
        Stops at first blank/None column or 'SUMMARY' column."""
        comp_cols = []
        for idx in range(2, len(header_row)):
            h = header_row[idx]
            if h is None:
                break  # blank column = end of data section
            hs = str(h).strip().upper()
            if "SUMMARY" in hs or "STAT" in hs or "AVG" in hs or "MIN" in hs or "MAX" in hs:
                break
            if hs:  # any non-empty string after DATE/TIME is a component ID
                comp_cols.append((idx, str(h).strip()))
        return comp_cols

    try:
        for ws in wb.worksheets:
            kpi_name = ws.title.strip()
            if not kpi_name:
                continue
            # Skip index/info/readme sheets
            if kpi_name.upper() in ("INDEX", "README", "INFO", "DICTIONARY", "METADATA", "SHEET1"):
                continue

            # Read all rows into list for flexibility
            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) < 2:
                errors.append(f"Sheet '{kpi_name}': too few rows")
                continue

            # ── Auto-detect header row ───────────────────────────────────
            # Skip title/info rows until we find the header row with DATE/TIME
            header_idx = None
            for ri, row in enumerate(all_rows):
                if _is_header_row(row):
                    header_idx = ri
                    break

            if header_idx is None:
                # Fallback: treat row 0 or 1 as header
                # Check if row 1 looks like a header (common: row 0 is title)
                if len(all_rows) > 1 and _is_header_row(all_rows[1]):
                    header_idx = 1
                elif _is_header_row(all_rows[0]):
                    header_idx = 0
                else:
                    # Last resort: assume row 0 is header
                    header_idx = 0

            header_row = all_rows[header_idx]
            data_rows = all_rows[header_idx + 1:]
            app.logger.info(f"Sheet '{kpi_name}': header_idx={header_idx}, rows={len(data_rows)}, header[:6]={[str(c)[:30] if c else None for c in header_row[:6]]}")

            if not data_rows:
                errors.append(f"Sheet '{kpi_name}': no data rows after header")
                continue

            # Detect component type from KPI name (fallback for shared KPIs like CPU/Memory)
            auto_comp_type = _detect_component_type_from_kpi(kpi_name) or "UNKNOWN"

            # ── Smart layout detection ───────────────────────────────────
            # Scan header row to classify each column as: identifier, date, component_id, or unknown
            layout = None
            date_col = None
            time_col = None
            comp_cols = []  # [(col_idx, comp_id)]
            date_header_cols = []  # [(col_idx, datetime)]
            comp_id_col = 0  # default: first column

            id_col_indices = []
            for ci, h in enumerate(header_row):
                if h is None: continue
                hs = str(h).strip() if isinstance(h, str) else ""
                hu = hs.upper()

                # Check for DATE/TIME/INTERVAL labels
                if isinstance(h, str) and len(hu) < 25:
                    if hu == "DATE" or hu == "DATES": date_col = ci; continue
                    if hu.startswith("TIME") or hu == "TIMESTAMP": time_col = ci; continue
                    if "INTERVAL" in hu: continue  # skip interval column

                # Check for identifier column names
                if isinstance(h, str) and ci < 6:
                    if any(kw in hu for kw in ["COMPONENT", "NODE", "VENDOR", "SITE", "ID", "NAME", "TYPE"]):
                        id_col_indices.append(ci)
                        continue

                # Check if header is a component ID (MME1, SGW2, PCRF, HSS etc.)
                is_comp = False
                if isinstance(h, str) and len(hs) <= 10:
                    for pfx in ("MME", "SGW", "PGW", "HSS", "PCRF"):
                        if hu.startswith(pfx):
                            comp_cols.append((ci, hs))
                            is_comp = True
                            break

                # Try to parse as date (supports: 10-Apr-26, 2026-04-10, 10/04/2026, etc.)
                if not is_comp and isinstance(h, str):
                    _DATE_FMTS = ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S",
                                  "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%b-%d-%y", "%b %d, %Y")
                    for fmt in _DATE_FMTS:
                        try:
                            dt = datetime.strptime(hs.strip()[:19], fmt)
                            date_header_cols.append((ci, dt))
                            break
                        except (ValueError, IndexError):
                            continue

                # Check for datetime objects
                if isinstance(h, datetime):
                    date_header_cols.append((ci, h))
                elif hasattr(h, 'date') and not isinstance(h, str):
                    date_header_cols.append((ci, datetime.combine(h, datetime.min.time())))

            # Determine layout
            if comp_cols and date_col is not None:
                layout = "A"  # DATE | TIME | MME1 | MME2 ...
            elif comp_cols and date_col is None and len(comp_cols) >= 2:
                layout = "A"  # No explicit DATE col but component IDs found — date might be in col 0
                if date_col is None:
                    # Check if first non-id column has date-like data
                    for row in data_rows[:3]:
                        if row and row[0] is not None:
                            if isinstance(row[0], datetime) or (isinstance(row[0], str) and _re.match(r'\d{4}-\d{2}-\d{2}', str(row[0]))):
                                date_col = 0
                                break
            elif date_header_cols and len(date_header_cols) >= 2:
                layout = "B"  # Transposed: comp_id in rows, dates as column headers
                comp_id_col = id_col_indices[0] if id_col_indices else 0
                # If col 0 header is a component type (e.g., "PCRF", "MME"), use it as forced type
                if comp_cols:
                    header_comp_type = _detect_comp_type_from_id(comp_cols[0][1])
                    if header_comp_type != "UNKNOWN":
                        auto_comp_type = header_comp_type  # Force all rows to this type
            else:
                # Last resort: check if column 0 has dates and remaining have numbers
                has_dates_in_col0 = False
                for row in data_rows[:5]:
                    if row and row[0] is not None:
                        if isinstance(row[0], datetime) or (isinstance(row[0], str) and _re.match(r'\d{4}', str(row[0]).strip())):
                            has_dates_in_col0 = True; break
                if has_dates_in_col0:
                    layout = "A"; date_col = 0
                    # Remaining non-None header columns are component IDs
                    for ci, h in enumerate(header_row):
                        if ci == 0 or h is None: continue
                        hs = str(h).strip()
                        if hs and hs.upper() not in ("SUMMARY", "AVG", "MIN", "MAX", "STAT"):
                            comp_cols.append((ci, hs))

            app.logger.info(f"Sheet '{kpi_name}': layout={layout}, comp_cols={len(comp_cols)}, date_headers={len(date_header_cols)}, date_col={date_col}, time_col={time_col}, id_cols={id_col_indices}, auto_type={auto_comp_type}")
            if layout is None or (layout == "A" and not comp_cols) or (layout == "B" and len(date_header_cols) < 2):
                errors.append(f"Sheet '{kpi_name}': could not detect layout (comp_cols={len(comp_cols)}, date_headers={len(date_header_cols)}, header[:5]={[str(c)[:25] for c in header_row[:5] if c]})")
                continue

            # ── Generate rows ────────────────────────────────────────────
            def _row_gen():
                if layout == "A":
                    # Layout A: DATE | TIME | comp1 | comp2 ...
                    for row in data_rows:
                        if not row or len(row) < 2: continue
                        d_val = None
                        if date_col is not None and date_col < len(row) and row[date_col] is not None:
                            cell = row[date_col]
                            if isinstance(cell, datetime): d_val = cell.date()
                            elif hasattr(cell, 'date'): d_val = cell.date()
                            elif isinstance(cell, str):
                                for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
                                    try: d_val = datetime.strptime(cell.strip()[:10], fmt).date(); break
                                    except (ValueError, IndexError): continue
                        if d_val is None: continue
                        hr, mn = 0, 0
                        if time_col is not None and time_col < len(row) and row[time_col] is not None:
                            tc = row[time_col]
                            if isinstance(tc, str):
                                _h, _m = _parse_time_str(tc)
                                if _h is not None: hr, mn = _h, _m
                            elif isinstance(tc, datetime): hr, mn = tc.hour, tc.minute
                            elif hasattr(tc, 'hour'): hr, mn = tc.hour, tc.minute
                        for col_idx, comp_id in comp_cols:
                            if col_idx >= len(row) or row[col_idx] is None: continue
                            try: val = float(row[col_idx])
                            except (ValueError, TypeError): continue
                            ct = _detect_comp_type_from_id(comp_id)
                            if ct == "UNKNOWN": ct = auto_comp_type if auto_comp_type != "UNKNOWN" else _detect_comp_type_from_id(comp_id)
                            yield (ct, comp_id, kpi_name, d_val, hr, mn, val, batch_id)
                else:
                    # Layout B: row = [comp_id, time, interval?, val_for_date1, val_for_date2, ...]
                    for row in data_rows:
                        if not row or len(row) < 2: continue
                        comp_id = str(row[comp_id_col] if comp_id_col < len(row) else "").strip()
                        if not comp_id: continue
                        # Determine component type: prefer from comp_id (PCRF1→PCRF), then from header/KPI name
                        ct = _detect_comp_type_from_id(comp_id)
                        if ct == "UNKNOWN" and auto_comp_type and auto_comp_type != "UNKNOWN":
                            ct = auto_comp_type
                        # Extract time from time_col if available
                        hr, mn = 0, 0
                        if time_col is not None and time_col < len(row) and row[time_col] is not None:
                            tc = row[time_col]
                            if isinstance(tc, str):
                                _h, _m = _parse_time_str(tc)
                                if _h is not None: hr, mn = _h, _m
                            elif isinstance(tc, datetime): hr, mn = tc.hour, tc.minute
                            elif hasattr(tc, 'hour'): hr, mn = tc.hour, getattr(tc, 'minute', 0)
                        for col_idx, dt_val in date_header_cols:
                            if col_idx >= len(row) or row[col_idx] is None: continue
                            try: val = float(row[col_idx])
                            except (ValueError, TypeError): continue
                            d_val = dt_val.date() if isinstance(dt_val, datetime) else dt_val
                            yield (ct, comp_id, kpi_name, d_val, hr, mn, val, batch_id)

            sheet_inserted = bulk_insert_core_component_rows(db, _row_gen())
            total_inserted += sheet_inserted
            comp_names = [c[1] for c in comp_cols] if layout == "A" else [str(data_rows[0][comp_id_col])[:20] if data_rows else "?"]
            if sheet_inserted == 0:
                errors.append(f"Sheet '{kpi_name}': 0 rows inserted (layout={layout}, date_col={date_col}, comp_cols={len(comp_cols)}, date_headers={len(date_header_cols)})")
            kpi_summary.append({
                "name": kpi_name,
                "component_type": auto_comp_type,
                "layout": layout,
                "components": comp_names,
                "rows": sheet_inserted,
            })
            app.logger.info(f"Core component upload: sheet '{kpi_name}' ({auto_comp_type}) "
                            f"components={comp_names} — {sheet_inserted} rows")

    except Exception as e:
        app.logger.error(f"Core component upload error: {e}")
        return jsonify({"error": f"Upload failed: {e}"}), 500
    finally:
        wb.close()

    clear_analytics_cache()
    return jsonify({
        "inserted": total_inserted,
        "kpis_processed": len(kpi_summary),
        "kpi_summary": kpi_summary,
        "errors": errors,
        "batch_id": batch_id,
    })


@app.route("/api/admin/core-component-kpi-status", methods=["GET"])
@jwt_required()
def admin_core_component_kpi_status():
    """Return status of core component KPI data."""
    from models import CoreComponentKpi
    try:
        with db.engine.connect() as conn:
            r = conn.execute(text("""
                SELECT COUNT(*) AS total_rows,
                       COUNT(DISTINCT component_type) AS unique_types,
                       COUNT(DISTINCT component_id) AS unique_components,
                       COUNT(DISTINCT kpi_name) AS unique_kpis,
                       MIN(date) AS date_from,
                       MAX(date) AS date_to
                FROM core_component_kpi
            """)).mappings().first()
            comp_types = [row["component_type"] for row in conn.execute(
                text("SELECT DISTINCT component_type FROM core_component_kpi ORDER BY component_type")
            ).mappings()]
            kpi_names = [{"kpi_name": row["kpi_name"], "component_type": row["component_type"]}
                         for row in conn.execute(
                text("SELECT DISTINCT kpi_name, component_type FROM core_component_kpi ORDER BY component_type, kpi_name")
            ).mappings()]
        return jsonify({
            "total_rows": r["total_rows"] if r else 0,
            "unique_types": r["unique_types"] if r else 0,
            "unique_components": r["unique_components"] if r else 0,
            "unique_kpis": r["unique_kpis"] if r else 0,
            "date_from": str(r["date_from"]) if r and r["date_from"] else None,
            "date_to": str(r["date_to"]) if r and r["date_to"] else None,
            "component_types": comp_types,
            "kpi_names": kpi_names,
        })
    except Exception:
        return jsonify({"total_rows": 0, "unique_types": 0, "unique_components": 0,
                        "unique_kpis": 0, "component_types": [], "kpi_names": []})


@app.route("/api/admin/delete-core-component-kpi", methods=["DELETE"])
@jwt_required()
def admin_delete_core_component_kpi():
    """Delete core component KPI data. Optional ?component_type=MME to delete only one component."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role not in ("admin", "cto"):
        return jsonify({"error": "Unauthorized"}), 403
    comp_type = request.args.get("component_type", "").strip().upper()
    try:
        with db.engine.connect() as conn:
            if comp_type and comp_type in ("MME", "SGW", "PGW", "HSS", "PCRF"):
                r = conn.execute(text("DELETE FROM core_component_kpi WHERE component_type = :ct"), {"ct": comp_type})
            else:
                r = conn.execute(text("DELETE FROM core_component_kpi"))
            conn.commit()
            deleted = r.rowcount
    except Exception:
        deleted = 0
    clear_analytics_cache()
    return jsonify({"deleted": deleted, "component_type": comp_type or "ALL"})


@app.route("/api/admin/core-component-status-by-type", methods=["GET"])
@jwt_required()
def admin_core_component_status_by_type():
    """Return per-component upload status for the 5 core components."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role not in ("admin", "cto"):
        return jsonify({"error": "Unauthorized"}), 403
    result = {}
    for ct in ("MME", "SGW", "PGW", "HSS", "PCRF"):
        try:
            from sqlalchemy import text as _t
            r = db.session.execute(_t(
                "SELECT COUNT(*) AS rows, COUNT(DISTINCT kpi_name) AS kpis, "
                "COUNT(DISTINCT component_id) AS nodes, MIN(date) AS date_from, MAX(date) AS date_to "
                "FROM core_component_kpi WHERE component_type = :ct"
            ), {"ct": ct}).fetchone()
            result[ct] = {
                "rows": int(r.rows or 0), "kpis": int(r.kpis or 0),
                "nodes": int(r.nodes or 0),
                "date_from": str(r.date_from) if r.date_from else None,
                "date_to": str(r.date_to) if r.date_to else None,
            }
        except Exception:
            result[ct] = {"rows": 0, "kpis": 0, "nodes": 0, "date_from": None, "date_to": None}
    return jsonify(result)


@app.route("/api/admin/delete-sites", methods=["DELETE"])
@jwt_required()
def admin_delete_sites():
    """Delete all telecom site data."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    count = TelecomSite.query.count()
    TelecomSite.query.delete()
    db.session.commit()
    clear_analytics_cache()
    return jsonify({"deleted": count})


@app.route("/api/admin/delete-kpi-site-level", methods=["DELETE"])
@jwt_required()
def admin_delete_kpi_site_level():
    """Delete all site-level KPI data."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    count = KpiData.query.filter_by(data_level="site").count()
    KpiData.query.filter_by(data_level="site").delete()
    db.session.commit()
    clear_analytics_cache()
    return jsonify({"deleted": count})


@app.route("/api/admin/delete-kpi-cell-level", methods=["DELETE"])
@jwt_required()
def admin_delete_kpi_cell_level():
    """Delete all cell-level KPI data."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    count = KpiData.query.filter_by(data_level="cell").count()
    KpiData.query.filter_by(data_level="cell").delete()
    db.session.commit()
    clear_analytics_cache()
    return jsonify({"deleted": count})


@app.route("/api/admin/uploaded-kpis", methods=["GET"])
@jwt_required()
def admin_uploaded_kpis():
    """Return list of uploaded KPI names with row counts, split by data level."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    site_kpis = db.session.query(
        KpiData.kpi_name, db.func.count(KpiData.id)
    ).filter(
        KpiData.data_level == "site",
        ~KpiData.kpi_name.in_(SHARED_WORKBOOK_KPI_NAMES)
    ).group_by(KpiData.kpi_name).order_by(KpiData.kpi_name).all()

    cell_kpis = db.session.query(
        KpiData.kpi_name, db.func.count(KpiData.id)
    ).filter_by(data_level="cell").group_by(KpiData.kpi_name).order_by(KpiData.kpi_name).all()

    site_count = TelecomSite.query.count()
    return jsonify({
        "site_kpis": [{"name": r[0], "rows": r[1]} for r in site_kpis],
        "cell_kpis": [{"name": r[0], "rows": r[1]} for r in cell_kpis],
        "site_count": site_count,
    })


# ─── Flexible KPI Upload (Core / Revenue) ─────────────────────────────────────

# Known display labels for Core KPI columns (case-insensitive key → label)
_CORE_KPI_DISPLAY_LABELS = {
    "authentication success rate": "Auth Success Rate",
    "auth success rate":           "Auth Success Rate",
    "cpu utilization":             "CPU Usage",
    "cpu usage":                   "CPU Usage",
    "attach success rate":         "4G Attach Success",
    "4g attach success":           "4G Attach Success",
    "pdp bearer setup success rate": "4G Bearer Success",
    "pdp bearer sr":               "4G Bearer Success",
    "4g bearer success":           "4G Bearer Success",
}

def _flex_display_label(kpi_type, raw_name):
    """Return a human-friendly display label for a flexible KPI column."""
    if kpi_type == "core":
        return _CORE_KPI_DISPLAY_LABELS.get(raw_name.lower().strip(), raw_name)
    return raw_name


def _detect_col_type(values):
    """Given a list of raw cell values, return 'numeric', 'date', or 'text'.
    More permissive than before — if ANY sample parses as numeric and no
    samples are clearly text, treats the whole column as numeric. This fixes
    columns like 'Feb Total' / 'Mar Total' that were being silently
    classified as text when the first few rows happened to be None.
    """
    import numbers
    numeric_count = 0
    text_count = 0
    non_empty = 0
    for v in values:
        if v is None or v == "":
            continue
        non_empty += 1
        if isinstance(v, numbers.Number):
            numeric_count += 1
            continue
        if isinstance(v, str):
            s = v.replace(",", "").replace("%", "").strip()
            try:
                float(s)
                numeric_count += 1
                continue
            except ValueError:
                pass
        text_count += 1
    if non_empty == 0:
        # Don't guess — assume numeric so the column still gets stored
        # (revenue/business/transport uploads are overwhelmingly numeric KPIs).
        return "numeric"
    # Numeric wins unless we saw MORE text samples than numeric samples
    if numeric_count >= text_count and numeric_count > 0:
        return "numeric"
    return "text"


@app.route("/api/admin/upload-flexible-kpi", methods=["POST"])
@jwt_required()
def admin_upload_flexible_kpi():
    """Flexible KPI upload for Core or Revenue data.
    ?type=core | revenue
    Only Site_ID is mandatory; all other columns are auto-detected.
    Uploads APPEND to existing data (use Delete All to clear first).
    Uses chunked inserts for large files.
    """
    import uuid, io, datetime as _dt

    CHUNK_SIZE = 10_000  # flush every N records

    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    kpi_type = request.args.get("type", "").strip().lower()
    if kpi_type not in ("core", "revenue", "business", "transport"):
        return jsonify({"error": "Invalid type. Use ?type=core, ?type=revenue, ?type=business or ?type=transport"}), 400

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    fname = file.filename.lower()
    print(f"[FLEX UPLOAD] type={kpi_type}, file={fname}")

    def to_float(v):
        if v is None or v == "":
            return None
        if isinstance(v, (int, float)):
            return float(v)
        try:
            return float(str(v).replace(",", "").strip())
        except ValueError:
            return None

    def parse_date_header(h):
        if h is None:
            return None
        if isinstance(h, (_dt.datetime, _dt.date)):
            return h.date() if isinstance(h, _dt.datetime) else h
        s = str(h).strip()[:10]
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return _dt.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    def _flush_chunk(records):
        """Flush a chunk of records to DB."""
        if records:
            db.session.bulk_save_objects(records)
            db.session.flush()
        return []

    try:
        raw_bytes = file.read()

        if fname.endswith(".csv"):
            import csv as csv_mod
            text = raw_bytes.decode("utf-8-sig", errors="replace")
            reader = csv_mod.DictReader(io.StringIO(text))
            headers = reader.fieldnames or []
            rows = list(reader)

        elif fname.endswith((".xlsx", ".xls", ".xlsm")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
            sheet_names = wb.sheetnames

            if kpi_type == "core" and len(sheet_names) > 1:
                # Multi-sheet format: sheet name = KPI metric, col 0 = Site_ID, col 1+ = dates
                batch_id = str(uuid.uuid4())
                records = []
                total_inserted = 0
                skipped = 0
                unique_sites = set()
                all_kpi_cols = []

                for sheet_name in sheet_names:
                    ws = wb[sheet_name]
                    sheet_rows = list(ws.iter_rows(values_only=True))
                    if not sheet_rows:
                        continue
                    kpi_col_name = sheet_name.strip()
                    display_label = _flex_display_label(kpi_type, kpi_col_name)
                    all_kpi_cols.append(kpi_col_name)

                    header_row = sheet_rows[0]
                    date_headers = [parse_date_header(h) for h in header_row[1:]]

                    for row in sheet_rows[1:]:
                        if not row or row[0] is None:
                            skipped += 1
                            continue
                        site_id = str(row[0]).strip()
                        if not site_id:
                            skipped += 1
                            continue
                        unique_sites.add(site_id)
                        for i, date_val in enumerate(date_headers):
                            if date_val is None:
                                continue
                            raw_val = row[i + 1] if (i + 1) < len(row) else None
                            num = to_float(raw_val)
                            if num is None:
                                continue
                            records.append(FlexibleKpiUpload(
                                kpi_type=kpi_type,
                                upload_batch=batch_id,
                                site_id=site_id,
                                column_name=kpi_col_name,
                                column_type="numeric",
                                num_value=num,
                                str_value=display_label,
                                row_date=date_val,
                                kpi_name=kpi_col_name,
                            ))
                            if len(records) >= CHUNK_SIZE:
                                total_inserted += len(records)
                                records = _flush_chunk(records)

                wb.close()
                total_inserted += len(records)
                _flush_chunk(records)
                db.session.commit()
                _cache_clear("business_kpi", "technical_kpi")
                clear_analytics_cache()
                return jsonify({
                    "rows_in_file": total_inserted,
                    "records_inserted": total_inserted,
                    "unique_sites": len(unique_sites),
                    "columns_detected": all_kpi_cols,
                    "skipped_rows": skipped,
                })

            else:
                # Single-sheet fallback — auto-detect header row
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                wb.close()
                if not all_rows:
                    return jsonify({"error": "Empty file"}), 400

                # Find the header row: the row with the MOST header-like cells
                # (string values, at least one matching a site ID variant).
                SITE_ID_VARIANTS = {
                    "site_id", "site id", "siteid", "site name", "sitename", "site",
                    "site abs id", "abs site id", "abs id", "absid", "siteabsid",
                }
                best_idx = 0
                best_score = -1
                for idx, row in enumerate(all_rows[:15]):
                    if row is None:
                        continue
                    has_site_id = False
                    str_cells = 0
                    for cell in row:
                        if cell is None or cell == "":
                            continue
                        cs = str(cell).strip()
                        str_cells += 1
                        norm = cs.lower().replace("_", " ").replace("-", " ")
                        if norm in SITE_ID_VARIANTS:
                            has_site_id = True
                    score = (str_cells if has_site_id else -1)
                    if score > best_score:
                        best_score = score
                        best_idx = idx
                header_idx = best_idx

                raw_headers = all_rows[header_idx]
                headers = [str(h).strip() if h is not None else "" for h in raw_headers]
                print(f"[FLEX UPLOAD] Detected header row at index {header_idx} ({len([h for h in headers if h])} non-empty cols): {headers[:20]}")
                rows = []
                for r in all_rows[header_idx + 1:]:
                    rows.append({headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))})
        else:
            return jsonify({"error": "Unsupported file format. Use .xlsx, .xls or .csv"}), 400

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to read file: {e}"}), 400

    # --- single-sheet path ---
    # PREFER "SITE ABS ID" as the canonical site key — it is the ONLY column
    # that matches across site-data / revenue / business / transport uploads.
    # Falls back to SITE ID / Site Name only if ABS ID isn't present.
    def _norm_h(h):
        return h.strip().lower().replace("_", " ").replace("-", " ")
    abs_id_col = next((h for h in headers if _norm_h(h) in ("site abs id", "siteabsid", "abs site id", "abssiteid", "abs id", "absid")), None)
    site_col = abs_id_col or next(
        (h for h in headers if _norm_h(h) in (
            "site id", "siteid", "site name", "sitename", "site"
        )),
        None
    )
    if not site_col:
        print(f"[FLEX UPLOAD] Missing Site ID/ABS ID. Headers found: {headers[:15]}")
        return jsonify({"error": f"Missing required column: SITE ABS ID or SITE ID. Found columns: {headers[:10]}"}), 400
    print(f"[FLEX UPLOAD] Using '{site_col}' as linking key (ABS ID preferred)")

    kpi_cols = [h for h in headers if h and h != site_col]
    if not kpi_cols:
        return jsonify({"error": "No KPI columns found besides Site_ID"}), 400

    col_type_map = {}
    for col in kpi_cols:
        # Larger sample + scan all rows if file is small. Fixes the case where
        # "Feb Total" / "Mar Total" happen to be None for the first 50 rows
        # (formula-backed cells not cached in openpyxl).
        sample_size = min(len(rows), 500)
        sample = [r.get(col) for r in rows[:sample_size]]
        col_type_map[col] = _detect_col_type(sample)
    # Debug: log the detected columns so field issues are easier to trace.
    try:
        _num_cols = [c for c, t in col_type_map.items() if t == "numeric"]
        _txt_cols = [c for c, t in col_type_map.items() if t == "text"]
        print(f"[FLEX UPLOAD] type={kpi_type} site_col='{site_col}' numeric_cols={_num_cols} text_cols={_txt_cols}")
    except Exception:
        pass

    # Append mode — no deletion of previous data
    batch_id = str(uuid.uuid4())
    records = []
    total_inserted = 0
    skipped = 0

    # ── Currency detection (revenue uploads only) ─────────────────────────
    # Default = USD. If any header contains a non-USD currency symbol, all
    # numeric values for that column are converted to USD using a fixed
    # rate table. If a header explicitly says "USD" or has no symbol, no
    # conversion is performed.
    CURRENCY_TO_USD = {
        "$": 1.0, "usd": 1.0, "us$": 1.0,
        "€": 1.10, "eur": 1.10,
        "£": 1.27, "gbp": 1.27,
        "¥": 0.0067, "jpy": 0.0067,
        "₹": 0.012, "inr": 0.012, "rs": 0.012,
        "៛": 0.00024, "khr": 0.00024,  # Cambodian Riel
        "฿": 0.028, "thb": 0.028,
        "₫": 0.000040, "vnd": 0.000040,
    }
    def _detect_rate(col_name):
        if kpi_type != "revenue":
            return 1.0
        cl = (col_name or "").lower()
        for sym, rate in CURRENCY_TO_USD.items():
            if sym in cl:
                return rate
        return 1.0  # default USD
    col_rate = {c: _detect_rate(c) for c in kpi_cols}

    for row in rows:
        site_id = str(row.get(site_col, "") or "").strip()
        if not site_id:
            skipped += 1
            continue
        for col in kpi_cols:
            raw_val = row.get(col)
            ctype = col_type_map[col]
            col_norm = col.strip()
            label = _flex_display_label(kpi_type, col_norm)
            if ctype == "numeric":
                num = to_float(raw_val)
                if num is not None:
                    num *= col_rate.get(col, 1.0)
                records.append(FlexibleKpiUpload(
                    kpi_type=kpi_type,
                    upload_batch=batch_id,
                    site_id=site_id,
                    column_name=col_norm,
                    column_type="numeric",
                    num_value=num,
                    str_value=label,
                ))
            else:
                records.append(FlexibleKpiUpload(
                    kpi_type=kpi_type,
                    upload_batch=batch_id,
                    site_id=site_id,
                    column_name=col_norm,
                    column_type="text",
                    num_value=None,
                    str_value=str(raw_val) if raw_val is not None else None,
                ))
            if len(records) >= CHUNK_SIZE:
                total_inserted += len(records)
                records = _flush_chunk(records)

    total_inserted += len(records)
    _flush_chunk(records)
    db.session.commit()

    # Drop any cached business/CTO KPI response so the next dashboard load
    # picks up the freshly uploaded rows.
    _cache_clear("business_kpi", "technical_kpi")
    clear_analytics_cache()

    # Invalidate priority/urgency bracket caches so new uploads are reflected immediately
    try:
        from network_issues import _REVENUE_BRACKETS_CACHE, _USERS_BRACKETS_CACHE
        _REVENUE_BRACKETS_CACHE["data"] = None
        _REVENUE_BRACKETS_CACHE["date"] = None
        _USERS_BRACKETS_CACHE["data"] = None
        _USERS_BRACKETS_CACHE["date"] = None
    except Exception:
        pass
    try:
        from change_workflow import _AUTO_BRACKETS
        _AUTO_BRACKETS["date"] = None
        _AUTO_BRACKETS["rev"] = None
        _AUTO_BRACKETS["usr"] = None
    except Exception:
        pass

    return jsonify({
        "rows_in_file": len(rows),
        "records_inserted": total_inserted,
        "unique_sites": len({r.get(site_col, "") for r in rows if r.get(site_col)}),
        "columns_detected": kpi_cols,
        "skipped_rows": skipped,
    })


@app.route("/api/admin/flexible-kpi-status", methods=["GET"])
@jwt_required()
def admin_flexible_kpi_status():
    """Return record counts + column list for a flexible KPI type."""
    from sqlalchemy import func as sa_func

    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    kpi_type = request.args.get("type", "").strip().lower()
    if kpi_type not in ("core", "revenue", "business", "transport"):
        return jsonify({"error": "Invalid type"}), 400

    total = FlexibleKpiUpload.query.filter_by(kpi_type=kpi_type).count()
    if total == 0:
        return jsonify({"unique_sites": 0, "total_rows": 0, "unique_columns": 0, "columns": []})

    unique_sites = db.session.query(sa_func.count(sa_func.distinct(FlexibleKpiUpload.site_id))).filter_by(kpi_type=kpi_type).scalar()

    col_rows = db.session.query(
        FlexibleKpiUpload.column_name,
        FlexibleKpiUpload.column_type,
        FlexibleKpiUpload.str_value,
    ).filter_by(kpi_type=kpi_type).distinct(
        FlexibleKpiUpload.column_name
    ).order_by(FlexibleKpiUpload.column_name, FlexibleKpiUpload.id.desc()).all()

    seen = {}
    for col_name, col_type, str_val in col_rows:
        if col_name not in seen:
            label = _flex_display_label(kpi_type, col_name)
            if col_type == "numeric" and str_val and str_val != col_name:
                label = str_val  # stored display label
            seen[col_name] = {"column_name": col_name, "column_label": label, "column_type": col_type}

    date_range = {}
    if kpi_type == "core":
        min_d = db.session.query(sa_func.min(FlexibleKpiUpload.row_date)).filter_by(kpi_type=kpi_type).scalar()
        max_d = db.session.query(sa_func.max(FlexibleKpiUpload.row_date)).filter_by(kpi_type=kpi_type).scalar()
        date_range = {
            "from": min_d.isoformat() if min_d and hasattr(min_d, 'isoformat') else str(min_d) if min_d else None,
            "to": max_d.isoformat() if max_d and hasattr(max_d, 'isoformat') else str(max_d) if max_d else None,
        }

    return jsonify({
        "unique_sites": unique_sites,
        "total_rows": total,
        "unique_columns": len(seen),
        "date_range": date_range,
        "columns": list(seen.values()),
    })


@app.route("/api/admin/delete-flexible-kpi", methods=["DELETE"])
@jwt_required()
def admin_delete_flexible_kpi():
    """Delete all flexible KPI records for a given type."""
    user = User.query.get(int(get_jwt_identity()))
    if not user or user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    kpi_type = request.args.get("type", "").strip().lower()
    if kpi_type not in ("core", "revenue", "business", "transport"):
        return jsonify({"error": "Invalid type"}), 400

    deleted = FlexibleKpiUpload.query.filter_by(kpi_type=kpi_type).delete()
    db.session.commit()
    _cache_clear("business_kpi", "technical_kpi")
    clear_analytics_cache()
    return jsonify({"deleted": deleted, "kpi_type": kpi_type})


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTS & ANALYTICS ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

SLA_DEFAULTS = {
    "sla_critical": {"value": "2", "description": "SLA target hours for Critical priority"},
    "sla_high": {"value": "4", "description": "SLA target hours for High priority"},
    "sla_medium": {"value": "8", "description": "SLA target hours for Medium priority"},
    "sla_low": {"value": "16", "description": "SLA target hours for Low priority"},
}


def get_sla_targets():
    targets = {}
    for key in ["sla_critical", "sla_high", "sla_medium", "sla_low"]:
        setting = SystemSetting.query.filter_by(key=key).first()
        priority = key.replace("sla_", "")
        targets[priority] = float(setting.value) if setting else float(SLA_DEFAULTS[key]["value"])
    return targets


def get_date_range(range_param):
    now = datetime.now(timezone.utc)
    if range_param == "7d":
        return now - timedelta(days=7)
    elif range_param == "90d":
        return now - timedelta(days=90)
    elif range_param == "12m":
        return now - timedelta(days=365)
    else:
        return now - timedelta(days=30)


def get_previous_period(range_param):
    now = datetime.now(timezone.utc)
    if range_param == "7d":
        return now - timedelta(days=14), now - timedelta(days=7)
    elif range_param == "90d":
        return now - timedelta(days=180), now - timedelta(days=90)
    elif range_param == "12m":
        return now - timedelta(days=730), now - timedelta(days=365)
    else:
        return now - timedelta(days=60), now - timedelta(days=30)


def calc_trend(current, previous):
    if previous == 0:
        return 100.0 if current > 0 else 0.0
    return round(((current - previous) / previous) * 100, 1)


@app.route("/api/reports/overview", methods=["GET"])
@jwt_required()
def reports_overview():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    range_param = request.args.get("range", "30d")
    start_date = get_date_range(range_param)
    prev_start, prev_end = get_previous_period(range_param)

    # Current period tickets
    current_tickets = Ticket.query.filter(Ticket.created_at >= start_date)
    resolved_current = current_tickets.filter(Ticket.status == "resolved").count()

    # Previous period
    prev_tickets = Ticket.query.filter(Ticket.created_at >= prev_start, Ticket.created_at < prev_end)
    resolved_prev = prev_tickets.filter(Ticket.status == "resolved").count()

    # Avg resolution time (current period)
    resolved_with_time = Ticket.query.filter(
        Ticket.created_at >= start_date,
        Ticket.status == "resolved",
        Ticket.resolved_at.isnot(None)
    ).all()
    if resolved_with_time:
        total_hours = sum(
            (t.resolved_at - t.created_at).total_seconds() / 3600
            for t in resolved_with_time
        )
        avg_resolution = round(total_hours / len(resolved_with_time), 1)
    else:
        avg_resolution = 0

    # Previous avg resolution
    prev_resolved_with_time = Ticket.query.filter(
        Ticket.created_at >= prev_start, Ticket.created_at < prev_end,
        Ticket.status == "resolved", Ticket.resolved_at.isnot(None)
    ).all()
    if prev_resolved_with_time:
        prev_total_hours = sum(
            (t.resolved_at - t.created_at).total_seconds() / 3600
            for t in prev_resolved_with_time
        )
        prev_avg_resolution = round(prev_total_hours / len(prev_resolved_with_time), 1)
    else:
        prev_avg_resolution = 0

    # CSAT
    current_feedback = Feedback.query.filter(Feedback.created_at >= start_date)
    total_fb = current_feedback.count()
    satisfied = current_feedback.filter(Feedback.rating >= 4).count()
    csat = round((satisfied / max(total_fb, 1)) * 100, 1)

    prev_feedback = Feedback.query.filter(Feedback.created_at >= prev_start, Feedback.created_at < prev_end)
    prev_total_fb = prev_feedback.count()
    prev_satisfied = prev_feedback.filter(Feedback.rating >= 4).count()
    prev_csat = round((prev_satisfied / max(prev_total_fb, 1)) * 100, 1)

    # SLA compliance
    sla_targets = get_sla_targets()
    all_resolved = Ticket.query.filter(
        Ticket.created_at >= start_date,
        Ticket.status == "resolved",
        Ticket.resolved_at.isnot(None)
    ).all()
    within_sla = 0
    for t in all_resolved:
        hours = (t.resolved_at - t.created_at).total_seconds() / 3600
        target = sla_targets.get(t.priority, 48)
        if hours <= target:
            within_sla += 1
    sla_compliance = round((within_sla / max(len(all_resolved), 1)) * 100, 1)

    prev_all_resolved = Ticket.query.filter(
        Ticket.created_at >= prev_start, Ticket.created_at < prev_end,
        Ticket.status == "resolved", Ticket.resolved_at.isnot(None)
    ).all()
    prev_within_sla = 0
    for t in prev_all_resolved:
        hours = (t.resolved_at - t.created_at).total_seconds() / 3600
        target = sla_targets.get(t.priority, 48)
        if hours <= target:
            prev_within_sla += 1
    prev_sla = round((prev_within_sla / max(len(prev_all_resolved), 1)) * 100, 1)

    # Resolution trends (monthly)
    resolution_trends = db.session.query(
        db.func.date_trunc("month", Ticket.resolved_at).label("month"),
        db.func.avg(
            db.func.extract("epoch", Ticket.resolved_at - Ticket.created_at) / 3600
        ).label("avg_hours"),
        db.func.count(Ticket.id).label("volume")
    ).filter(
        Ticket.resolved_at.isnot(None),
        Ticket.created_at >= start_date
    ).group_by("month").order_by("month").all()

    # Weekly volume
    weekly_volume = db.session.query(
        db.func.extract("dow", Ticket.created_at).label("dow"),
        db.func.count(Ticket.id).label("opened")
    ).filter(Ticket.created_at >= start_date).group_by("dow").order_by("dow").all()

    weekly_resolved = db.session.query(
        db.func.extract("dow", Ticket.resolved_at).label("dow"),
        db.func.count(Ticket.id).label("resolved")
    ).filter(
        Ticket.resolved_at.isnot(None),
        Ticket.resolved_at >= start_date
    ).group_by("dow").order_by("dow").all()

    day_names = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    opened_map = {int(r[0]): r[1] for r in weekly_volume}
    resolved_map = {int(r[0]): r[1] for r in weekly_resolved}
    weekly_data = [
        {"day": day_names[i], "opened": opened_map.get(i, 0), "resolved": resolved_map.get(i, 0)}
        for i in range(7)
    ]

    # Category breakdown
    categories = db.session.query(
        Ticket.category, db.func.count(Ticket.id)
    ).filter(Ticket.created_at >= start_date).group_by(Ticket.category).all()

    # Priority distribution
    priorities = db.session.query(
        Ticket.priority, db.func.count(Ticket.id)
    ).filter(Ticket.created_at >= start_date).group_by(Ticket.priority).all()

    return jsonify({
        "total_resolved": resolved_current,
        "resolved_trend": calc_trend(resolved_current, resolved_prev),
        "avg_resolution_hours": avg_resolution,
        "resolution_trend": calc_trend(avg_resolution, prev_avg_resolution),
        "csat_score": csat,
        "csat_trend": calc_trend(csat, prev_csat),
        "sla_compliance": sla_compliance,
        "sla_trend": calc_trend(sla_compliance, prev_sla),
        "resolution_trends": [
            {
                "month": r[0].strftime("%b %Y") if r[0] else "",
                "avg_hours": round(float(r[1] or 0), 1),
                "volume": r[2]
            } for r in resolution_trends
        ],
        "weekly_volume": weekly_data,
        "category_breakdown": [{"name": c[0] or "Other", "count": c[1]} for c in categories],
        "priority_distribution": [{"priority": p[0], "count": p[1]} for p in priorities],
    })


@app.route("/api/reports/agents", methods=["GET"])
@jwt_required()
def reports_agents():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    range_param = request.args.get("range", "30d")
    start_date = get_date_range(range_param)

    managers = User.query.filter(User.role.in_(["manager"])).all()
    agents_data = []

    for mgr in managers:
        assigned = Ticket.query.filter(
            Ticket.assigned_to == mgr.id,
            Ticket.created_at >= start_date
        )
        resolved = assigned.filter(Ticket.status == "resolved").count()
        pending = assigned.filter(Ticket.status.in_(["pending", "in_progress"])).count()
        escalated = assigned.filter(Ticket.status == "escalated").count()

        resolved_tickets = Ticket.query.filter(
            Ticket.assigned_to == mgr.id,
            Ticket.status == "resolved",
            Ticket.resolved_at.isnot(None),
            Ticket.created_at >= start_date
        ).all()

        if resolved_tickets:
            avg_time = round(sum(
                max(0, (t.resolved_at - t.created_at).total_seconds() / 3600)
                for t in resolved_tickets
            ) / len(resolved_tickets), 1)
        else:
            avg_time = 0

        agent_feedback = db.session.query(db.func.avg(Feedback.rating)).join(
            Ticket, Feedback.chat_session_id == Ticket.chat_session_id
        ).filter(
            Ticket.assigned_to == mgr.id,
            Feedback.rating > 0,
            Feedback.created_at >= start_date
        ).scalar()

        agents_data.append({
            "id": mgr.id,
            "name": mgr.name,
            "resolved": resolved,
            "pending": pending,
            "escalated": escalated,
            "avg_resolution_hours": avg_time,
            "avg_rating": round(float(agent_feedback or 0), 1),
        })

    top_performer = max(agents_data, key=lambda x: x["resolved"], default=None)
    fastest = min(
        [a for a in agents_data if a["avg_resolution_hours"] > 0],
        key=lambda x: x["avg_resolution_hours"], default=None
    )
    highest_rated = max(
        [a for a in agents_data if a["avg_rating"] > 0],
        key=lambda x: x["avg_rating"], default=None
    )

    return jsonify({
        "agents": agents_data,
        "total_agents": len(managers),
        "top_performer": {"name": top_performer["name"], "resolved": top_performer["resolved"]} if top_performer else None,
        "fastest_agent": {"name": fastest["name"], "hours": fastest["avg_resolution_hours"]} if fastest else None,
        "highest_rated": {"name": highest_rated["name"], "rating": highest_rated["avg_rating"]} if highest_rated else None,
    })


@app.route("/api/reports/csat", methods=["GET"])
@jwt_required()
def reports_csat():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    range_param = request.args.get("range", "30d")
    start_date = get_date_range(range_param)
    prev_start, prev_end = get_previous_period(range_param)

    # Current CSAT
    current_fb = Feedback.query.filter(Feedback.created_at >= start_date)
    total_responses = current_fb.count()
    satisfied = current_fb.filter(Feedback.rating >= 4).count()
    csat = round((satisfied / max(total_responses, 1)) * 100, 1)

    prev_fb = Feedback.query.filter(Feedback.created_at >= prev_start, Feedback.created_at < prev_end)
    prev_total = prev_fb.count()
    prev_satisfied = prev_fb.filter(Feedback.rating >= 4).count()
    prev_csat = round((prev_satisfied / max(prev_total, 1)) * 100, 1)

    avg_rating = db.session.query(db.func.avg(Feedback.rating)).filter(
        Feedback.created_at >= start_date, Feedback.rating > 0
    ).scalar() or 0

    # Response rate
    resolved_tickets = Ticket.query.filter(
        Ticket.created_at >= start_date, Ticket.status == "resolved"
    ).count()
    response_rate = round((total_responses / max(resolved_tickets, 1)) * 100, 1)

    # Monthly CSAT trend
    monthly_csat = db.session.query(
        db.func.date_trunc("month", Feedback.created_at).label("month"),
        db.func.count(Feedback.id).label("total"),
        db.func.count(sql_case((Feedback.rating >= 4, 1))).label("satisfied")
    ).filter(Feedback.created_at >= start_date).group_by("month").order_by("month").all()

    # Feedback distribution (1-5 stars)
    distribution = db.session.query(
        Feedback.rating, db.func.count(Feedback.id)
    ).filter(
        Feedback.created_at >= start_date, Feedback.rating > 0
    ).group_by(Feedback.rating).order_by(Feedback.rating).all()

    dist_map = {r[0]: r[1] for r in distribution}
    feedback_dist = [{"stars": i, "count": dist_map.get(i, 0)} for i in range(1, 6)]

    # Response volume trend
    volume_trend = db.session.query(
        db.func.date_trunc("month", Feedback.created_at).label("month"),
        db.func.count(Feedback.id).label("count")
    ).filter(Feedback.created_at >= start_date).group_by("month").order_by("month").all()

    return jsonify({
        "current_csat": csat,
        "csat_trend": calc_trend(csat, prev_csat),
        "total_responses": total_responses,
        "responses_trend": calc_trend(total_responses, prev_total),
        "avg_rating": round(float(avg_rating), 1),
        "response_rate": min(response_rate, 100),
        "csat_monthly": [
            {
                "month": m[0].strftime("%b %Y") if m[0] else "",
                "csat": round((m[2] / max(m[1], 1)) * 100, 1)
            } for m in monthly_csat
        ],
        "feedback_distribution": feedback_dist,
        "response_volume": [
            {"month": v[0].strftime("%b %Y") if v[0] else "", "count": v[1]}
            for v in volume_trend
        ],
    })


@app.route("/api/reports/sla", methods=["GET"])
@jwt_required()
def reports_sla():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    range_param = request.args.get("range", "30d")
    start_date = get_date_range(range_param)
    prev_start, prev_end = get_previous_period(range_param)
    sla_targets = get_sla_targets()

    resolved = Ticket.query.filter(
        Ticket.created_at >= start_date,
        Ticket.status == "resolved",
        Ticket.resolved_at.isnot(None)
    ).all()

    within = 0
    near_breach = 0
    breached = 0
    first_response_times = []

    priority_stats = {}
    for p in ["critical", "high", "medium", "low"]:
        priority_stats[p] = {"target": sla_targets.get(p, 48), "times": [], "within": 0, "breached": 0}

    for t in resolved:
        hours = (t.resolved_at - t.created_at).total_seconds() / 3600
        target = sla_targets.get(t.priority, 48)
        first_response_times.append(hours)

        if t.priority in priority_stats:
            priority_stats[t.priority]["times"].append(hours)
            if hours <= target:
                priority_stats[t.priority]["within"] += 1
            else:
                priority_stats[t.priority]["breached"] += 1

        if hours <= target:
            within += 1
        elif hours <= target * 1.0 and hours > target * 0.8:
            near_breach += 1
        else:
            pct = hours / target if target > 0 else 999
            if pct > 1.0:
                breached += 1
            elif pct > 0.8:
                near_breach += 1
            else:
                within += 1

    total = max(len(resolved), 1)
    compliance_pct = round((within / total) * 100, 1)
    near_pct = round((near_breach / total) * 100, 1)
    breached_pct = round((breached / total) * 100, 1)
    avg_first_response = round(sum(first_response_times) / max(len(first_response_times), 1), 1)

    # Previous period compliance
    prev_resolved = Ticket.query.filter(
        Ticket.created_at >= prev_start, Ticket.created_at < prev_end,
        Ticket.status == "resolved", Ticket.resolved_at.isnot(None)
    ).all()
    prev_within = 0
    for t in prev_resolved:
        hours = (t.resolved_at - t.created_at).total_seconds() / 3600
        target = sla_targets.get(t.priority, 48)
        if hours <= target:
            prev_within += 1
    prev_compliance = round((prev_within / max(len(prev_resolved), 1)) * 100, 1)

    # SLA targets with actual averages
    sla_target_list = []
    for p in ["critical", "high", "medium", "low"]:
        ps = priority_stats[p]
        avg_actual = round(sum(ps["times"]) / max(len(ps["times"]), 1), 1) if ps["times"] else 0
        sla_target_list.append({
            "priority": p,
            "target_hours": ps["target"],
            "actual_hours": avg_actual,
            "status": "within" if avg_actual <= ps["target"] else "breached",
            "total": len(ps["times"]),
        })

    # Monthly breach trend
    monthly_trend = db.session.query(
        db.func.date_trunc("month", Ticket.resolved_at).label("month"),
        Ticket.priority,
        Ticket.resolved_at,
        Ticket.created_at
    ).filter(
        Ticket.resolved_at.isnot(None),
        Ticket.created_at >= start_date
    ).all()

    month_data = {}
    for t in monthly_trend:
        month_key = t[0].strftime("%b %Y") if t[0] else "Unknown"
        if month_key not in month_data:
            month_data[month_key] = {"compliant": 0, "near_breach": 0, "breached": 0, "total": 0}
        hours = (t[2] - t[3]).total_seconds() / 3600
        target = sla_targets.get(t[1], 48)
        month_data[month_key]["total"] += 1
        pct_of_target = hours / target if target > 0 else 999
        if pct_of_target <= 0.8:
            month_data[month_key]["compliant"] += 1
        elif pct_of_target <= 1.0:
            month_data[month_key]["near_breach"] += 1
        else:
            month_data[month_key]["breached"] += 1

    breach_trend = []
    for month_key, data in sorted(month_data.items()):
        t = max(data["total"], 1)
        breach_trend.append({
            "month": month_key,
            "compliant": round((data["compliant"] / t) * 100, 1),
            "near_breach": round((data["near_breach"] / t) * 100, 1),
            "breached": round((data["breached"] / t) * 100, 1),
        })

    return jsonify({
        "compliance_percentage": compliance_pct,
        "compliance_trend": calc_trend(compliance_pct, prev_compliance),
        "near_breach_percentage": near_pct,
        "breached_percentage": breached_pct,
        "avg_first_response": avg_first_response,
        "sla_targets": sla_target_list,
        "breach_trend": breach_trend,
        "within_count": within,
        "near_breach_count": near_breach,
        "breached_count": breached,
    })


@app.route("/api/reports/export", methods=["GET"])
@jwt_required()
def reports_export():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if user.role not in ("manager", "admin"):
        return jsonify({"error": "Unauthorized"}), 403

    fmt = request.args.get("format", "csv")
    section = request.args.get("section", "overview")
    range_param = request.args.get("range", "30d")
    start_date = get_date_range(range_param)

    if fmt == "csv":
        import io
        import csv
        from flask import Response

        output = io.StringIO()
        writer = csv.writer(output)

        if section == "overview":
            tickets = Ticket.query.filter(Ticket.created_at >= start_date).all()
            writer.writerow(["Reference", "Category", "Priority", "Status", "Created", "Resolved", "Resolution Hours"])
            for t in tickets:
                hours = ""
                if t.resolved_at and t.created_at:
                    hours = round((t.resolved_at - t.created_at).total_seconds() / 3600, 1)
                writer.writerow([t.reference_number, t.category, t.priority, t.status,
                                t.created_at.isoformat() if t.created_at else "",
                                t.resolved_at.isoformat() if t.resolved_at else "", hours])

        elif section == "agents":
            managers = User.query.filter(User.role.in_(["manager"])).all()
            writer.writerow(["Agent", "Resolved", "Pending", "Escalated", "Avg Hours", "Rating"])
            for mgr in managers:
                assigned = Ticket.query.filter(Ticket.assigned_to == mgr.id, Ticket.created_at >= start_date)
                resolved = assigned.filter(Ticket.status == "resolved").count()
                pending = assigned.filter(Ticket.status.in_(["pending", "in_progress"])).count()
                escalated = assigned.filter(Ticket.status == "escalated").count()
                writer.writerow([mgr.name, resolved, pending, escalated, 0, 0])

        elif section == "csat":
            feedbacks = Feedback.query.filter(Feedback.created_at >= start_date).all()
            writer.writerow(["User", "Rating", "Comment", "Date"])
            for f in feedbacks:
                writer.writerow([f.user.name if f.user else "", f.rating, f.comment,
                                f.created_at.isoformat() if f.created_at else ""])

        elif section == "sla":
            tickets = Ticket.query.filter(
                Ticket.created_at >= start_date, Ticket.status == "resolved",
                Ticket.resolved_at.isnot(None)
            ).all()
            sla_targets = get_sla_targets()
            writer.writerow(["Reference", "Priority", "Target Hours", "Actual Hours", "Status"])
            for t in tickets:
                hours = round((t.resolved_at - t.created_at).total_seconds() / 3600, 1)
                target = sla_targets.get(t.priority, 48)
                status = "Within SLA" if hours <= target else "Breached"
                writer.writerow([t.reference_number, t.priority, target, hours, status])

        response = Response(output.getvalue(), mimetype="text/csv")
        response.headers["Content-Disposition"] = f"attachment; filename=report_{section}_{range_param}.csv"
        return response

    return jsonify({"error": "PDF export is handled client-side"}), 400


# ═══════════════════════════════════════════════════════════════════════════════
# HUMAN AGENT ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/agent/status", methods=["PUT"])
@jwt_required()
def agent_toggle_status():
    """Toggle human agent online/offline status."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or user.role != "human_agent":
        return jsonify({"error": "Unauthorized"}), 403
    data = request.json or {}
    if "is_online" in data:
        user.is_online = bool(data["is_online"])
    else:
        user.is_online = not user.is_online
    db.session.commit()
    return jsonify({"is_online": user.is_online, "message": f"Status set to {'online' if user.is_online else 'offline'}"})


@app.route("/api/manager/status", methods=["PUT"])
@jwt_required()
def manager_toggle_status():
    """Toggle manager online/offline status."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or user.role != "manager":
        return jsonify({"error": "Unauthorized"}), 403
    data = request.json or {}
    if "is_online" in data:
        user.is_online = bool(data["is_online"])
    else:
        user.is_online = not user.is_online
    db.session.commit()
    return jsonify({"is_online": user.is_online, "message": f"Status set to {'online' if user.is_online else 'offline'}"})


@app.route("/api/agent/dashboard", methods=["GET"])
@jwt_required()
def agent_dashboard():
    """Return KPIs for the human agent."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or user.role != "human_agent":
        return jsonify({"error": "Unauthorized"}), 403

    # ── Mock dashboard for demo user ──
    from mock_dashboard import MOCK_EMAIL, get_mock_dashboard
    if user.email == MOCK_EMAIL:
        return jsonify(get_mock_dashboard(
            agent_name=user.name or "Agent",
            agent_location=getattr(user, "location", "Gurgaon"),
            agent_domain=getattr(user, "domain", "broadband"),
        ))

    now = datetime.now(timezone.utc)

    # Helper: make any datetime UTC-aware (DB columns are stored as naive UTC)
    def _utc(dt):
        if dt is None:
            return None
        return dt if dt.tzinfo is not None else dt.replace(tzinfo=timezone.utc)

    from sqlalchemy import or_
    my_tickets = Ticket.query.filter(
        or_(Ticket.assigned_to == user_id, Ticket.escalated_by == user_id)
    ).all()

    # Backfill first_response_at from first agent chat message (one-time)
    backfilled = False
    for t in my_tickets:
        if not t.first_response_at and t.chat_session_id:
            first_agent_msg = ChatMessage.query.filter(
                ChatMessage.session_id == t.chat_session_id,
                ChatMessage.sender == 'agent'
            ).order_by(ChatMessage.created_at.asc()).first()
            if first_agent_msg and first_agent_msg.created_at:
                t.first_response_at = first_agent_msg.created_at
                backfilled = True
    if backfilled:
        db.session.commit()

    resolved = [t for t in my_tickets if t.status == "resolved"]
    total = len(my_tickets)
    resolved_count = len(resolved)
    open_count = len([t for t in my_tickets if t.status not in ("resolved", "closed")])

    # MTTR – Mean Time To Resolve (hours)
    resolve_times = []
    for t in resolved:
        ra = _utc(t.resolved_at)
        ca = _utc(t.created_at)
        if ra and ca:
            resolve_times.append(max(0, (ra - ca).total_seconds() / 3600))
    mttr = round(sum(resolve_times) / len(resolve_times), 2) if resolve_times else 0

    # SLA Compliance Rate
    sla_ok = 0
    for t in resolved:
        dl = _utc(t.sla_deadline)
        ra = _utc(t.resolved_at)
        if dl and ra and ra <= dl:
            sla_ok += 1
    sla_compliance = round((sla_ok / max(resolved_count, 1)) * 100, 1)

    # First Contact Resolution (tickets resolved without reopening – simplified: resolved in 1st attempt)
    fcr = round((resolved_count / max(total, 1)) * 100, 1)

    # CSAT – average rating from feedbacks linked to agent's resolved sessions
    session_ids = [t.chat_session_id for t in my_tickets if t.chat_session_id]
    feedbacks = Feedback.query.filter(
        Feedback.chat_session_id.in_(session_ids),
        Feedback.rating > 0
    ).all() if session_ids else []
    csat = round(sum(f.rating for f in feedbacks) / max(len(feedbacks), 1), 2) if feedbacks else 0
    csat_pct = round((len([f for f in feedbacks if f.rating >= 4]) / max(len(feedbacks), 1)) * 100, 1)

    # ── Reopen Rate ────────────────────────────────────────────────────────
    # Check reopened_count field AND last_reopened_at as secondary indicator
    reopened_tickets = [t for t in resolved
                        if (getattr(t, 'reopened_count', 0) or 0) > 0
                        or getattr(t, 'last_reopened_at', None) is not None]
    reopen_rate = round((len(reopened_tickets) / max(resolved_count, 1)) * 100, 1)

    # ── H/S Incident Resolution Time ────────────────────────────────────
    # Priority order: critical/high → medium → all resolved
    hs_resolution_time = 0.0
    for prio_filter in [("critical", "high"), ("medium",), None]:
        hs_times = []
        pool = resolved if prio_filter is not None else resolved
        for t in pool:
            if prio_filter is not None and t.priority not in prio_filter:
                continue
            ra = _utc(t.resolved_at)
            ca = _utc(t.created_at)
            if ra and ca:
                hs_times.append(max(0, (ra - ca).total_seconds() / 3600))
        if hs_times:
            hs_resolution_time = round(sum(hs_times) / len(hs_times), 2)
            break
    # Absolute fallback: use MTTR which is known to be computed
    if hs_resolution_time == 0 and mttr > 0:
        hs_resolution_time = mttr

    # ── H/S Incident Response Time ──────────────────────────────────────
    # Try first_response_at, then estimate from first chat message, then from MTTR
    hs_response_time = 0.0
    # Method 1: Use first_response_at if populated
    resp_times = []
    for t in my_tickets:
        fra = _utc(getattr(t, 'first_response_at', None))
        ca = _utc(t.created_at)
        if fra and ca:
            resp_times.append(max(0, (fra - ca).total_seconds() / 3600))
    if resp_times:
        hs_response_time = round(sum(resp_times) / len(resp_times), 2)
    else:
        # Method 2: Estimate from first agent message in chat session
        est_resp_times = []
        for t in my_tickets:
            if t.chat_session_id and t.created_at:
                first_agent_msg = ChatMessage.query.filter(
                    ChatMessage.session_id == t.chat_session_id,
                    ChatMessage.sender == 'agent'
                ).order_by(ChatMessage.created_at.asc()).first()
                if first_agent_msg and first_agent_msg.created_at:
                    ca = _utc(t.created_at)
                    fra = _utc(first_agent_msg.created_at)
                    if ca and fra and fra >= ca:
                        est_resp_times.append(max(0, (fra - ca).total_seconds() / 3600))
        if est_resp_times:
            hs_response_time = round(sum(est_resp_times) / len(est_resp_times), 2)
        else:
            # Method 3: Estimate as 15% of resolution time
            hs_response_time = round(hs_resolution_time * 0.15, 2)

    # ── Complaint Resolution Time ────────────────────────────────────────
    complaint_resolution_time = mttr

    # ── RCA Timely Completion ────────────────────────────────────────────
    rca_completion = sla_compliance

    # ── Avg Open Ticket Age ──────────────────────────────────────────────
    # Count all non-resolved tickets including escalated, pending, in_progress
    aging_hours = []
    for t in my_tickets:
        if t.status not in ("resolved", "closed"):
            ca = _utc(t.created_at)
            if ca:
                elapsed = (now - ca).total_seconds() / 3600
                if elapsed >= 0:
                    aging_hours.append(elapsed)
    avg_aging = max(0.0, round(sum(aging_hours) / len(aging_hours), 2)) if aging_hours else 0

    # Monthly trend – tickets created vs resolved per month (last 6 months)
    monthly_created = {}
    monthly_resolved = {}
    for t in my_tickets:
        cr = _utc(t.created_at)
        if cr:
            key = cr.strftime("%b %Y")
            monthly_created[key] = monthly_created.get(key, 0) + 1
    for t in resolved:
        ra = _utc(t.resolved_at) or _utc(t.created_at)
        if ra:
            key = ra.strftime("%b %Y")
            monthly_resolved[key] = monthly_resolved.get(key, 0) + 1
    all_months = sorted(set(list(monthly_created.keys()) + list(monthly_resolved.keys())))[-6:]
    monthly_trend = [
        {"month": m, "created": monthly_created.get(m, 0), "resolved": monthly_resolved.get(m, 0)}
        for m in all_months
    ]

    # Priority distribution of my tickets
    priority_dist = {}
    for t in my_tickets:
        priority_dist[t.priority] = priority_dist.get(t.priority, 0) + 1
    priority_chart = [{"name": k, "value": v} for k, v in priority_dist.items()]

    # SLA compliance by priority
    sla_by_priority = {}
    for t in resolved:
        p = t.priority
        if p not in sla_by_priority:
            sla_by_priority[p] = {"total": 0, "ok": 0}
        sla_by_priority[p]["total"] += 1
        dl = _utc(t.sla_deadline)
        ra = _utc(t.resolved_at)
        if dl and ra and ra <= dl:
            sla_by_priority[p]["ok"] += 1
    sla_priority_chart = [
        {"priority": p, "compliance": round((v["ok"] / max(v["total"], 1)) * 100, 1)}
        for p, v in sla_by_priority.items()
    ]

    # ── Advanced Dashboard Data ──────────────────────────────────────────

    # Sentiment distribution from feedback ratings
    sentiment_dist = [
        {"name": "Excellent", "value": len([f for f in feedbacks if f.rating == 5])},
        {"name": "Good",      "value": len([f for f in feedbacks if f.rating == 4])},
        {"name": "Neutral",   "value": len([f for f in feedbacks if f.rating == 3])},
        {"name": "Poor",      "value": len([f for f in feedbacks if f.rating == 2])},
        {"name": "Bad",       "value": len([f for f in feedbacks if f.rating <= 1])},
    ]

    # Category-wise resolution rates
    cat_stats = {}
    for t in my_tickets:
        cat = t.category or "Uncategorized"
        if cat not in cat_stats:
            cat_stats[cat] = {"total": 0, "resolved": 0, "sla_ok": 0}
        cat_stats[cat]["total"] += 1
        if t.status == "resolved":
            cat_stats[cat]["resolved"] += 1
            dl2 = _utc(t.sla_deadline)
            ra2 = _utc(t.resolved_at)
            if dl2 and ra2 and ra2 <= dl2:
                cat_stats[cat]["sla_ok"] += 1
    category_resolution = sorted([
        {"category": k, "total": v["total"], "resolved": v["resolved"],
         "rate": round(v["resolved"] / max(v["total"], 1) * 100, 1),
         "sla_rate": round(v["sla_ok"] / max(v["total"], 1) * 100, 1)}
        for k, v in cat_stats.items()
    ], key=lambda x: x["total"], reverse=True)[:8]

    # ── Agent Efficiency Metrics (unique KPIs) ──────────────────────────
    # Messages-to-resolve ratio, avg touches per ticket, first response stats
    total_msgs_all = 0
    agent_msgs_all = 0
    tickets_with_msgs = 0
    for t in my_tickets:
        if t.chat_session_id:
            tm = ChatMessage.query.filter_by(session_id=t.chat_session_id).count()
            am = ChatMessage.query.filter_by(session_id=t.chat_session_id, sender='agent').count()
            total_msgs_all += tm
            agent_msgs_all += am
            if tm > 0:
                tickets_with_msgs += 1
    avg_msgs_per_ticket = round(total_msgs_all / max(tickets_with_msgs, 1), 1)
    agent_msg_ratio = round(agent_msgs_all / max(total_msgs_all, 1) * 100, 1)

    # First response stats
    fr_times = []
    for t in my_tickets:
        fra = _utc(getattr(t, 'first_response_at', None))
        ca = _utc(t.created_at)
        if fra and ca:
            fr_times.append(max(0, (fra - ca).total_seconds() / 3600))
    avg_first_resp = round(sum(fr_times) / max(len(fr_times), 1), 1) if fr_times else 0
    fastest_resp = round(min(fr_times), 1) if fr_times else 0

    efficiency_metrics = {
        "avg_msgs_per_ticket": avg_msgs_per_ticket,
        "agent_msg_pct": agent_msg_ratio,
        "ai_msg_pct": round(100 - agent_msg_ratio, 1),
        "avg_first_response_hrs": avg_first_resp,
        "fastest_response_hrs": fastest_resp,
        "tickets_with_response": len(fr_times),
        "total_conversations": total_msgs_all,
        "resolution_rate": round(resolved_count / max(total, 1) * 100, 1),
    }

    # Customer tier distribution
    tier_stats = {}
    for t in my_tickets:
        tier = (t.user.user_type or "bronze") if t.user else "bronze"
        tier = tier.capitalize()
        if tier not in tier_stats:
            tier_stats[tier] = {"total": 0, "resolved": 0, "avg_time": []}
        tier_stats[tier]["total"] += 1
        if t.status == "resolved":
            tier_stats[tier]["resolved"] += 1
            ra3 = _utc(t.resolved_at)
            ca3 = _utc(t.created_at)
            if ra3 and ca3:
                tier_stats[tier]["avg_time"].append(max(0, (ra3 - ca3).total_seconds() / 3600))
    customer_tiers = [
        {"tier": k, "total": v["total"], "resolved": v["resolved"],
         "rate": round(v["resolved"] / max(v["total"], 1) * 100, 1),
         "avg_hours": round(sum(v["avg_time"]) / max(len(v["avg_time"]), 1), 1)}
        for k, v in tier_stats.items()
    ]

    # ── AI Chatbot vs Agent Comparison ─────────────────────────────────────
    # Flow: Customer -> AI Chatbot -> (resolved) OR -> Escalate to Agent -> (resolved)
    # AI Resolved = sessions closed without creating a ticket
    # Escalated = sessions that created a ticket (chatbot couldn't resolve)
    # Agent Resolved = escalated tickets resolved by this agent

    # Total chat sessions system-wide
    total_sessions = ChatSession.query.count()
    # Sessions that led to a ticket (escalated to human)
    escalated_session_ids = set(t.chat_session_id for t in Ticket.query.filter(
        Ticket.chat_session_id.isnot(None)
    ).all())
    total_escalated = len(escalated_session_ids)
    # AI self-resolved = sessions that never created a ticket
    ai_self_resolved = max(0, total_sessions - total_escalated)

    # Agent resolution times (for this agent's resolved tickets)
    agent_resolve_times = []
    for t in resolved:
        ra4 = _utc(t.resolved_at); ca4 = _utc(t.created_at)
        if ra4 and ca4:
            agent_resolve_times.append(max(0, (ra4 - ca4).total_seconds() / 3600))
    agent_avg_resolution = round(sum(agent_resolve_times) / max(len(agent_resolve_times), 1), 1)

    # AI avg resolution time (from chat sessions that resolved without ticket)
    ai_resolved_sessions = ChatSession.query.filter(
        ChatSession.status == 'resolved',
        ~ChatSession.id.in_(escalated_session_ids) if escalated_session_ids else ChatSession.id > 0
    ).all()
    ai_resolve_times = []
    for s in ai_resolved_sessions:
        ca_s = _utc(s.created_at)
        ra_s = _utc(s.resolved_at)
        if ca_s and ra_s:
            ai_resolve_times.append(max(0, (ra_s - ca_s).total_seconds() / 3600))
    ai_avg_resolution = round(sum(ai_resolve_times) / max(len(ai_resolve_times), 1), 1)

    # Rates
    ai_resolution_rate = round(ai_self_resolved / max(total_sessions, 1) * 100, 1)
    escalation_rate = round(total_escalated / max(total_sessions, 1) * 100, 1)
    agent_resolution_rate = round(resolved_count / max(total, 1) * 100, 1)

    ai_vs_agent = {
        "total_conversations": total_sessions,
        "ai_resolved": ai_self_resolved,
        "ai_resolution_rate": ai_resolution_rate,
        "ai_avg_time": ai_avg_resolution,
        "escalated_to_agent": total_escalated,
        "escalation_rate": escalation_rate,
        "agent_resolved": resolved_count,
        "agent_resolution_rate": agent_resolution_rate,
        "agent_avg_time": agent_avg_resolution,
    }

    # Weekly activity heatmap (7 days x 24 hours) — convert to IST for display
    IST_OFFSET = timedelta(hours=5, minutes=30)
    heatmap = [[0] * 24 for _ in range(7)]
    for t in my_tickets:
        ca6 = _utc(t.created_at)
        if ca6:
            ist = ca6 + IST_OFFSET
            heatmap[ist.weekday()][ist.hour] += 1
    heatmap_resolved = [[0] * 24 for _ in range(7)]
    for t in resolved:
        ra6 = _utc(t.resolved_at)
        if ra6:
            ist = ra6 + IST_OFFSET
            heatmap_resolved[ist.weekday()][ist.hour] += 1

    # Agent performance score — computed AFTER perf_radar is built (below)
    # Placeholder — will be overwritten after radar calculation
    perf_score = 0

    # Agent badges
    badges = []
    if sla_compliance >= 95:
        badges.append({"tag": "SLA Champion", "icon": "shield"})
    if csat >= 4.0 and len(feedbacks) >= 3:
        badges.append({"tag": "Customer Expert", "icon": "star"})
    if mttr > 0 and mttr <= 12:
        badges.append({"tag": "Speed Resolver", "icon": "zap"})
    if reopen_rate == 0 and resolved_count > 5:
        badges.append({"tag": "Zero Reopen", "icon": "check"})
    if fcr >= 85:
        badges.append({"tag": "First Touch Pro", "icon": "target"})
    if resolved_count >= 30:
        badges.append({"tag": "Volume Leader", "icon": "trending"})
    if len(hs_times) > 0 and hs_resolution_time <= 8:
        badges.append({"tag": "Crisis Handler", "icon": "alert"})

    # ── Performance DNA Radar (6 dimensions, 0-100 each) ────────────────
    # Each dimension is a real metric scored out of 100

    # 1. Speed: based on avg resolution time vs weighted SLA target
    #    100 = resolved at 0% of SLA, 0 = resolved at 200%+ of SLA
    speed_ratios = []
    for t in resolved:
        ra, ca = _utc(t.resolved_at), _utc(t.created_at)
        sla_h = t.sla_hours or 8
        if ra and ca:
            actual = max((ra - ca).total_seconds() / 3600, 0)
            speed_ratios.append(min(actual / sla_h, 2.0))  # cap at 2x SLA
    avg_speed_ratio = sum(speed_ratios) / max(len(speed_ratios), 1) if speed_ratios else 1.0
    speed_score = round(max(0, min((1 - avg_speed_ratio / 2) * 100, 100)), 1)

    # 2. Quality: FCR — % of tickets resolved without reopening
    zero_reopen = sum(1 for t in resolved if (t.reopened_count or 0) == 0)
    quality_score = round((zero_reopen / max(resolved_count, 1)) * 100, 1)

    # 3. SLA: % of resolved tickets within SLA deadline
    sla_score = round(sla_compliance, 1)

    # 4. Satisfaction: % of feedbacks rated 4+
    satisfaction_score = round(csat_pct, 1) if feedbacks else 0

    # 5. Responsiveness: based on avg first response time vs SLA
    #    100 = responded instantly, 0 = responded after SLA deadline
    resp_ratios = []
    for t in my_tickets:
        fra = _utc(getattr(t, 'first_response_at', None))
        ca = _utc(t.created_at)
        sla_h = t.sla_hours or 8
        if fra and ca and sla_h > 0:
            resp_h = max((fra - ca).total_seconds() / 3600, 0)
            resp_ratios.append(min(resp_h / sla_h, 1.0))
    avg_resp_ratio = sum(resp_ratios) / max(len(resp_ratios), 1) if resp_ratios else 0.5
    responsiveness_score = round(max(0, (1 - avg_resp_ratio) * 100), 1)

    # 6. Workload: throughput — resolved per week (benchmarked: 10/week = 100)
    weeks_active = max((now - min((_utc(t.created_at) for t in my_tickets), default=now)).days / 7, 1)
    throughput_per_week = resolved_count / weeks_active
    workload_score = round(min(throughput_per_week / 10 * 100, 100), 1)

    perf_radar = [
        {"axis": "Speed",          "value": speed_score,          "detail": f"Avg {round(avg_speed_ratio*100)}% of SLA used"},
        {"axis": "Quality",        "value": quality_score,        "detail": f"{zero_reopen}/{resolved_count} zero-reopen"},
        {"axis": "SLA",            "value": sla_score,            "detail": f"{sla_ok}/{resolved_count} within SLA"},
        {"axis": "Satisfaction",   "value": satisfaction_score,    "detail": f"{csat}/5 avg rating"},
        {"axis": "Responsiveness", "value": responsiveness_score,  "detail": f"Avg {round(avg_resp_ratio*100)}% of SLA to respond"},
        {"axis": "Workload",       "value": workload_score,       "detail": f"{round(throughput_per_week,1)} resolved/week"},
    ]
    # Weighted composite performance score from radar dimensions
    _weights = {"Speed": 0.20, "Quality": 0.20, "SLA": 0.20, "Satisfaction": 0.15, "Responsiveness": 0.15, "Workload": 0.10}
    perf_score = round(sum(d["value"] * _weights.get(d["axis"], 0.15) for d in perf_radar), 1)

    # ── Issue Hotspot (subcategory breakdown with volume + status) ──────
    subcat_stats = {}
    for t in my_tickets:
        sc = t.subcategory or t.category or "Other"
        # Shorten long names
        if len(sc) > 30:
            sc = sc.split(" - ")[0] if " - " in sc else sc[:28] + ".."
        if sc not in subcat_stats:
            subcat_stats[sc] = {"total": 0, "resolved": 0, "open": 0}
        subcat_stats[sc]["total"] += 1
        if t.status == "resolved":
            subcat_stats[sc]["resolved"] += 1
        elif t.status not in ("resolved", "closed"):
            subcat_stats[sc]["open"] += 1
    issue_hotspots = sorted(
        [{"name": k, **v} for k, v in subcat_stats.items()],
        key=lambda x: x["total"], reverse=True
    )

    # ── Zone / Region distribution (3-tier fallback) ────────────────────
    #
    # Tier 1 (Dynamic):  state_province + country from ChatSession DB fields
    #                     Set by frontend when customer shares location.
    #                     Works for ANY country — no hardcoding.
    #
    # Tier 2 (Geocoding): If DB fields empty, use geopy Nominatim to resolve
    #                     city name or lat/lng → state/country dynamically.
    #                     Works for ANY country. Needs internet.
    #
    # Tier 3 (Hardcoded): If geocoding fails (network restricted), use a
    #                     static city → state lookup as last resort.

    # --- Tier 2 & 3: Geocoding fallbacks (cached) ---
    _geo_cache = getattr(app, '_geo_cache', {})
    app._geo_cache = _geo_cache

    def _try_geopy(location_input):
        """Tier 2: Free geocoding via geopy/Nominatim. Works for any country. Needs internet."""
        try:
            from geopy.geocoders import Nominatim
            geocoder = Nominatim(user_agent="telecom_cch", timeout=3)
            loc = geocoder.geocode(location_input, language="en", exactly_one=True, addressdetails=True)
            if loc and loc.raw and "address" in loc.raw:
                addr = loc.raw["address"]
                state = addr.get("state") or addr.get("province") or addr.get("region") or ""
                country = addr.get("country", "")
                if state:
                    return (state, country)
        except Exception:
            pass
        return (None, None)

    def _try_openai(location_input):
        """Tier 3: Azure OpenAI LLM geocoding. Works in restricted networks where OpenAI is whitelisted."""
        try:
            import json as _json
            resp = client.chat.completions.create(
                model=DEPLOYMENT_NAME,
                messages=[{
                    "role": "user",
                    "content": f'For the location "{location_input}", respond with ONLY a JSON object: {{"state": "<state/province name>", "country": "<country name>"}}. No other text.'
                }],
                temperature=0,
                max_tokens=60,
            )
            text = resp.choices[0].message.content.strip()
            if "{" in text:
                text = text[text.index("{"):text.rindex("}") + 1]
            data = _json.loads(text)
            state = data.get("state", "")
            country = data.get("country", "")
            if state:
                return (state, country)
        except Exception:
            pass
        return (None, None)

    def _try_geocode(city_name, lat=None, lng=None):
        """Resolve location → (state, country) using Tier 2 then Tier 3."""
        location_input = (city_name or "").strip()
        if not location_input and lat and lng:
            location_input = f"{lat}, {lng}"
        if not location_input:
            return (None, None)

        cache_key = f"geo:{location_input.lower()}"
        if cache_key in _geo_cache:
            return _geo_cache[cache_key]

        # Tier 2: Try geopy (free, works for any country)
        result = _try_geopy(location_input)

        # Tier 3: If geopy failed, try Azure OpenAI
        if not result[0]:
            result = _try_openai(location_input)

        _geo_cache[cache_key] = result
        return result

    # --- Tier 3: DB-driven fallback — build city→(state,country) from telecom_sites ---
    CITY_FALLBACK = {}
    try:
        from sqlalchemy import text as _sa_text
        _fb_rows = db.session.execute(_sa_text(
            "SELECT DISTINCT LOWER(city) AS city, state, province, commune, country "
            "FROM telecom_sites WHERE city IS NOT NULL AND city != '' LIMIT 500"
        )).fetchall()
        for _fbr in _fb_rows:
            _c = str(_fbr.city or "").strip().lower()
            _s = str(_fbr.province or _fbr.state or "").strip()
            _cn = str(_fbr.country or "").strip()
            if _c and _s:
                CITY_FALLBACK[_c] = (_s, _cn)
        # Also add commune → province mapping
        _com_rows = db.session.execute(_sa_text(
            "SELECT DISTINCT LOWER(commune) AS commune, province, country "
            "FROM telecom_sites WHERE commune IS NOT NULL AND commune != '' LIMIT 500"
        )).fetchall()
        for _cr in _com_rows:
            _cm = str(_cr.commune or "").strip().lower()
            _pv = str(_cr.province or "").strip()
            _cn = str(_cr.country or "").strip()
            if _cm and _pv:
                CITY_FALLBACK[_cm] = (_pv, _cn)
    except Exception:
        pass

    # Cache: state_val -> (centroid_lat, centroid_lng) computed from telecom_sites.
    # When chat-session lat/lng is wrong (e.g. customer in India but site is in
    # Cambodia), we fall back to the resolved state's actual centroid so the map
    # circle lands on the correct province.
    _state_centroid_cache = {}
    def _state_centroid(state_val):
        if not state_val: return (None, None)
        if state_val in _state_centroid_cache:
            return _state_centroid_cache[state_val]
        try:
            from sqlalchemy import text as _sa_text
            r = db.session.execute(_sa_text(
                "SELECT AVG(latitude) AS la, AVG(longitude) AS lo "
                "FROM telecom_sites WHERE state = :s OR province = :s OR city = :s"
            ), {"s": state_val}).fetchone()
            la = float(r.la) if r and r.la is not None else None
            lo = float(r.lo) if r and r.lo is not None else None
            if la == 0 and lo == 0: la = lo = None
        except Exception:
            la = lo = None
        _state_centroid_cache[state_val] = (la, lo)
        return (la, lo)

    # --- Resolve each ticket ---
    location_counts = {}
    country_set = set()
    for t in my_tickets:
        state_val, country_val = None, None
        lat, lng, city = None, None, ""
        site_lat = site_lng = None  # coordinates of the resolved telecom site

        if t.chat_session_id:
            cs = db.session.get(ChatSession, t.chat_session_id)
            if cs:
                lat = cs.latitude
                lng = cs.longitude
                city = (cs.location_description or "").strip()

                # Tier 1: DB fields (state_province, country)
                state_val = (getattr(cs, 'state_province', None) or "").strip() or None
                country_val = (getattr(cs, 'country', None) or "").strip() or None

        # Tier 1.5: For tickets with lat/lng, find nearest telecom site
        # and use its state/country from telecom_sites table
        if lat and lng:
            try:
                nearest = find_nearest_sites(lat, lng, n=1)
                if nearest and len(nearest) > 0:
                    site = nearest[0]
                    site_state = site.get('state', '') if isinstance(site, dict) else ''
                    site_city = site.get('city', '') if isinstance(site, dict) else ''
                    site_lat = site.get('latitude') if isinstance(site, dict) else None
                    site_lng = site.get('longitude') if isinstance(site, dict) else None
                    if not state_val:
                        if site_state:
                            state_val = site_state
                        elif site_city:
                            state_val = site_city
                    # Get country from telecom_sites table
                    if site_state and not country_val:
                        try:
                            from sqlalchemy import text as _sa_text
                            row = db.session.execute(_sa_text(
                                "SELECT country FROM telecom_sites WHERE state = :s LIMIT 1"
                            ), {"s": site_state}).fetchone()
                            if row and row.country:
                                country_val = row.country
                        except Exception:
                            pass
            except Exception:
                pass

        # Skip tickets with no location info at all
        if not state_val and not city and not (lat and lng):
            continue

        # Tier 2: Geocoding (if Tier 1 didn't resolve)
        if not state_val and (city or (lat and lng)):
            geo_state, geo_country = _try_geocode(city, lat, lng)
            if geo_state:
                state_val = geo_state
                country_val = country_val or geo_country

        # Tier 3: Hardcoded fallback (if Tier 2 also failed)
        if not state_val and city:
            city_lower = city.lower()
            fb = CITY_FALLBACK.get(city_lower)
            if not fb:
                for key, val in CITY_FALLBACK.items():
                    if key in city_lower or city_lower in key:
                        fb = val
                        break
            if fb:
                state_val, country_val = fb[0], country_val or fb[1]
            else:
                state_val = city  # last resort: use city name as zone

        if not state_val:
            continue

        if country_val:
            country_set.add(country_val)

        if state_val not in location_counts:
            location_counts[state_val] = {"total": 0, "resolved": 0, "lats": [], "lngs": []}
        location_counts[state_val]["total"] += 1
        if t.status == "resolved":
            location_counts[state_val]["resolved"] += 1

        # Accumulate coordinates for centroid. Priority order:
        #   1. State centroid computed from telecom_sites WHERE state=state_val
        #      (definitive — every site with this state averages to a point
        #      inside the correct country, so a Takeo ticket plots in Cambodia
        #      even if the customer's chat-session lat/lng was in India).
        #   2. Resolved-telecom-site lat/lng — only used if no state centroid is
        #      available (state has no rows in telecom_sites).
        #   3. Chat-session lat/lng — last resort.
        plot_lat = plot_lng = None
        cla, clo = _state_centroid(state_val)
        if cla is not None and clo is not None:
            plot_lat, plot_lng = cla, clo
        if (plot_lat is None or plot_lng is None) and site_lat is not None and site_lng is not None:
            plot_lat, plot_lng = site_lat, site_lng
        if (plot_lat is None or plot_lng is None) and lat and lng:
            try:
                _la, _lo = float(lat), float(lng)
                if _la != 0 and _lo != 0:
                    plot_lat, plot_lng = _la, _lo
            except (ValueError, TypeError):
                pass
        if plot_lat is not None and plot_lng is not None:
            try:
                location_counts[state_val]["lats"].append(float(plot_lat))
                location_counts[state_val]["lngs"].append(float(plot_lng))
            except (ValueError, TypeError):
                pass

    state_data = []
    for loc, v in sorted(location_counts.items(), key=lambda x: x[1]["total"], reverse=True):
        avg_lat = round(sum(v["lats"]) / len(v["lats"]), 6) if v["lats"] else None
        avg_lng = round(sum(v["lngs"]) / len(v["lngs"]), 6) if v["lngs"] else None
        state_data.append({
            "state": loc, "total": v["total"], "resolved": v["resolved"],
            "rate": round(v["resolved"] / max(v["total"], 1) * 100, 1),
            "lat": avg_lat, "lng": avg_lng,
        })
    zone_data = [{"zone": s["state"], **{k: s[k] for k in ("total", "resolved", "rate")}} for s in state_data]
    # Derive country from tickets or fall back to most common country in telecom_sites DB
    if len(country_set) == 1:
        detected_country = list(country_set)[0]
    elif not country_set:
        try:
            from sqlalchemy import text as _t
            _cr = db.session.execute(_t("SELECT country FROM telecom_sites WHERE country IS NOT NULL AND country != '' GROUP BY country ORDER BY COUNT(*) DESC LIMIT 1")).fetchone()
            detected_country = _cr[0] if _cr else ""
        except Exception:
            detected_country = ""
    else:
        detected_country = "Multiple"

    # ── SLA Risk Predictor (enhanced) ────────────────────────────────────
    sla_risk_items = []
    sla_by_priority = {}  # priority -> {total, breached, critical, warning, safe}
    total_open_sla = 0
    total_within_sla = 0
    for t in my_tickets:
        if t.status not in ("resolved", "closed"):
            dl7 = _utc(t.sla_deadline)
            if not dl7:
                # Backfill: assign default SLA if missing
                sla_targets = get_sla_targets()
                sla_h = sla_targets.get(t.priority or "medium", 8)
                cr = _utc(t.created_at) or now
                t.sla_hours = sla_h
                t.sla_deadline = cr + timedelta(hours=sla_h)
                db.session.add(t)
                dl7 = _utc(t.sla_deadline)
            if dl7:
                remaining = (dl7 - now).total_seconds() / 3600
                total_sla = t.sla_hours or 24
                pct_elapsed = round(min(max(((total_sla - remaining) / total_sla) * 100, 0), 100), 1)
                risk = "breached" if remaining <= 0 else "critical" if pct_elapsed >= 87.5 else "warning" if pct_elapsed >= 62.5 else "safe"
                total_open_sla += 1
                if risk in ("safe", "warning"):
                    total_within_sla += 1
                # Priority-wise breakdown
                pri = t.priority or "medium"
                if pri not in sla_by_priority:
                    sla_by_priority[pri] = {"total": 0, "breached": 0, "critical": 0, "warning": 0, "safe": 0}
                sla_by_priority[pri]["total"] += 1
                sla_by_priority[pri][risk] += 1
                sla_risk_items.append({
                    "ticket_id": t.id, "reference": t.reference_number,
                    "priority": t.priority, "pct_elapsed": pct_elapsed,
                    "remaining_hrs": round(remaining, 1),
                    "overdue_hrs": round(abs(remaining), 1) if remaining < 0 else 0,
                    "risk": risk,
                    "category": t.category or "",
                    "subcategory": t.subcategory or "",
                    "status": t.status,
                    "sla_hours": total_sla,
                    "sla_deadline": (dl7.replace(tzinfo=None).isoformat() + "Z") if dl7 and dl7.tzinfo else ((dl7.isoformat() + "Z") if dl7 else None),
                    "created_at": (t.created_at.replace(tzinfo=None).isoformat() + "Z") if t.created_at and t.created_at.tzinfo else ((t.created_at.isoformat() + "Z") if t.created_at else None),
                })
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
    sla_risk_items.sort(key=lambda x: x["remaining_hrs"])

    # SLA risk summary
    sla_risk_summary = {
        "safe": len([s for s in sla_risk_items if s["risk"] == "safe"]),
        "warning": len([s for s in sla_risk_items if s["risk"] == "warning"]),
        "critical": len([s for s in sla_risk_items if s["risk"] == "critical"]),
        "breached": len([s for s in sla_risk_items if s["risk"] == "breached"]),
    }
    # Overall SLA health percentage
    sla_health_pct = round((total_within_sla / max(total_open_sla, 1)) * 100, 1)
    # Priority distribution for chart
    sla_priority_dist = [
        {"priority": p, **v} for p, v in sla_by_priority.items()
    ]
    sla_priority_dist.sort(key=lambda x: {"critical": 0, "high": 1, "medium": 2, "low": 3}.get(x["priority"], 4))

    # ── Category Treemap (for visualization) ─────────────────────────────
    category_treemap = []
    for k, v in cat_stats.items():
        category_treemap.append({
            "name": k, "size": v["total"],
            "resolved": v["resolved"],
            "rate": round(v["resolved"] / max(v["total"], 1) * 100, 1),
        })
    category_treemap.sort(key=lambda x: x["size"], reverse=True)

    # ── Hourly ticket volume today (IST) ─────────────────────────────────
    now_ist = now + IST_OFFSET
    today_start_ist = now_ist.replace(hour=0, minute=0, second=0, microsecond=0)
    today_start_utc = today_start_ist - IST_OFFSET
    hourly_today = [0] * 24
    for t in my_tickets:
        ca7 = _utc(t.created_at)
        if ca7 and ca7 >= today_start_utc:
            ist_hour = (ca7 + IST_OFFSET).hour
            hourly_today[ist_hour] += 1
    hourly_data = [{"hour": f"{h:02d}:00", "tickets": hourly_today[h]} for h in range(24)]

    # ── AI Insights (auto-generated) ─────────────────────────────────────
    ai_insights = []
    # Most complained category
    if category_treemap:
        top_cat = category_treemap[0]
        ai_insights.append({"type": "info", "text": f"Most complaints: {top_cat['name']} ({top_cat['size']} tickets)"})
    # SLA risk
    if sla_risk_summary["critical"] > 0 or sla_risk_summary["breached"] > 0:
        ai_insights.append({"type": "warning", "text": f"{sla_risk_summary['critical'] + sla_risk_summary['breached']} tickets at SLA risk or breached"})
    # Best category
    best_cat = max(category_resolution, key=lambda x: x["rate"]) if category_resolution else None
    if best_cat and best_cat["rate"] > 0:
        ai_insights.append({"type": "success", "text": f"Best resolution: {best_cat['category']} at {best_cat['rate']}%"})
    # Aging insight
    if avg_aging > 24:
        ai_insights.append({"type": "warning", "text": f"Avg ticket age: {avg_aging:.0f}h - consider prioritizing older tickets"})
    # CSAT insight
    if csat >= 4.0:
        ai_insights.append({"type": "success", "text": f"Customer satisfaction at {csat}/5 - above target"})
    elif csat > 0:
        ai_insights.append({"type": "info", "text": f"CSAT at {csat}/5 - focus on customer experience"})

    # ── Predictive Workload Forecast (7 days) ──────────────────────────────
    from collections import defaultdict
    import calendar
    dow_ticket_counts = defaultdict(list)  # day_of_week -> list of daily counts
    # Group tickets by (date, day_of_week) to get daily volumes
    date_counts = defaultdict(int)
    for t in my_tickets:
        ca8 = _utc(t.created_at)
        if ca8:
            date_counts[ca8.date()] += 1
    for dt_date, cnt in date_counts.items():
        dow_ticket_counts[dt_date.weekday()].append(cnt)
    # Calculate average per day-of-week
    dow_avg = {}
    for dow in range(7):
        vals = dow_ticket_counts.get(dow, [])
        dow_avg[dow] = round(sum(vals) / max(len(vals), 1), 1) if vals else 0
    # Project next 7 days
    from datetime import timedelta as _td
    forecast = []
    for i in range(1, 8):
        future = (now + _td(days=i))
        dow = future.weekday()
        day_name = calendar.day_abbr[dow]
        forecast.append({
            "day": f"{day_name} {future.strftime('%d/%m')}",
            "predicted": dow_avg.get(dow, 0),
            "capacity": user.bandwidth_capacity or 10,
        })

    # ── Burndown: tickets to resolve this week to hit targets ────────────
    target_resolution_rate = 90.0
    tickets_needed = max(0, round(total * target_resolution_rate / 100) - resolved_count)
    burndown = {
        "target_rate": target_resolution_rate,
        "current_resolved": resolved_count,
        "needed": tickets_needed,
        "total": total,
        "current_rate": round(resolved_count / max(total, 1) * 100, 1),
    }

    return jsonify({
        "kpis": {
            "mttr": mttr,
            "sla_compliance_rate": sla_compliance,
            "first_contact_resolution": fcr,
            "csat": csat,
            "csat_pct": csat_pct,
            "reopen_rate": reopen_rate,
            "hs_incident_resolution_time": hs_resolution_time,
            "hs_incident_response_time": hs_response_time,
            "complaint_resolution_time": complaint_resolution_time,
            "rca_timely_completion": rca_completion,
            "avg_aging_hours": avg_aging,
        },
        "summary": {
            "total_tickets": total,
            "resolved": resolved_count,
            "open": open_count,
            "total_feedback": len(feedbacks),
        },
        "monthly_trend": monthly_trend,
        "priority_chart": priority_chart,
        "sla_priority_chart": sla_priority_chart,
        "sentiment": sentiment_dist,
        "category_resolution": category_resolution,
        "efficiency_metrics": efficiency_metrics,
        "customer_tiers": customer_tiers,
        "ai_vs_agent": ai_vs_agent,
        "heatmap": heatmap,
        "heatmap_resolved": heatmap_resolved,
        "performance_score": perf_score,
        "badges": badges,
        "perf_radar": perf_radar,
        "agent_name": user.name,
        "agent_location": user.location or "",
        "agent_domain": user.domain or "",
        "issue_hotspots": issue_hotspots,
        "zone_data": zone_data,
        "state_data": state_data,
        "detected_country": detected_country,
        "sla_risk": sla_risk_items[:15],
        "sla_risk_summary": sla_risk_summary,
        "sla_health_pct": sla_health_pct,
        "sla_priority_dist": sla_priority_dist,
        "sla_total_open": total_open_sla,
        "category_treemap": category_treemap,
        "hourly_today": hourly_data,
        "ai_insights": ai_insights,
        "forecast": forecast,
        "burndown": burndown,
        "recent_feedbacks": [
            {
                "rating": f.rating,
                "comment": f.comment,
                "customer": f.user.name if f.user else "",
                "session_id": f.chat_session_id,
                "created_at": f.created_at.isoformat() if f.created_at else None,
                "subprocess": f.chat_session.subprocess_name if f.chat_session else "",
            }
            for f in sorted(feedbacks, key=lambda x: x.created_at or datetime.min, reverse=True)[:10]
        ],
    })


@app.route("/api/agent/tickets", methods=["GET"])
@jwt_required()
def agent_tickets():
    """Return tickets assigned to the current human agent."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or user.role != "human_agent":
        return jsonify({"error": "Unauthorized"}), 403
    from sqlalchemy import or_
    tickets = (Ticket.query
               .filter(or_(Ticket.assigned_to == user_id, Ticket.escalated_by == user_id))
               .order_by(Ticket.created_at.desc())
               .all())
    return jsonify({"tickets": [t.to_dict() for t in tickets]})


@app.route("/api/agent/tickets/<int:ticket_id>/resolve", methods=["PUT"])
@jwt_required()
def agent_resolve_ticket(ticket_id):
    """Mark a ticket as resolved by the agent."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or user.role not in ("human_agent", "manager", "expert"):
        return jsonify({"error": "Unauthorized"}), 403
    ticket = Ticket.query.get(ticket_id)
    if not ticket:
        return jsonify({"error": "Ticket not found"}), 404
    if ticket.assigned_to != user_id:
        return jsonify({"error": "Ticket not assigned to you"}), 403

    data = request.json or {}
    ticket.status = "resolved"
    ticket.resolved_at = datetime.now(timezone.utc)
    resolution_notes = data.get("resolution_notes", "")
    if resolution_notes:
        ticket.resolution_notes = resolution_notes

    # Check SLA breach
    if ticket.sla_deadline:
        dl = ticket.sla_deadline if ticket.sla_deadline.tzinfo else ticket.sla_deadline.replace(tzinfo=timezone.utc)
        if ticket.resolved_at > dl:
            ticket.sla_breached = True

    # ── Add a bot message to the chat session so the customer sees it ──
    chat_session = None
    if ticket.chat_session_id:
        chat_session = ChatSession.query.get(ticket.chat_session_id)
        if chat_session:
            resolve_text = (
                f"Great news! Your support ticket ({ticket.reference_number}) has been resolved by "
                f"{user.name}."
            )
            if resolution_notes:
                resolve_text += f"\n\nResolution: {resolution_notes}"
            resolve_text += (
                "\n\nThank you for your patience. If you need further assistance, "
                "you can return to the Main Menu or exit the chat."
            )
            bot_msg = ChatMessage(
                session_id=ticket.chat_session_id,
                sender="bot",
                content=resolve_text,
            )
            db.session.add(bot_msg)
            chat_session.status = "resolved"

    db.session.commit()

    # Push session_updated via WebSocket so customer UI updates instantly
    if chat_session:
        socketio.emit("session_updated", {
            "session_id": ticket.chat_session_id,
            "status": "resolved",
        }, room=f"session_{ticket.chat_session_id}")

    # ── Notify customer via Email ──
    customer_user = User.query.get(ticket.user_id)
    if customer_user and customer_user.email:
        try:
            notes_row = f"<tr><td style='padding:8px 0;color:#64748b;width:140px;'>Resolution</td><td style='padding:8px 0;color:#1e293b;'>{resolution_notes}</td></tr>" if resolution_notes else ""
            html_body = f"""
            <div style="font-family:'Segoe UI',Arial,sans-serif;max-width:600px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,0.1);">
              <div style="background:#00338d;padding:24px 30px;text-align:center;">
                <h1 style="color:#fff;margin:0;font-size:20px;">Ticket Resolved</h1>
                <p style="color:rgba(255,255,255,0.8);margin:6px 0 0;font-size:13px;">Your support request has been successfully addressed</p>
              </div>
              <div style="padding:28px 30px;">
                <p style="margin:0 0 20px;font-size:15px;color:#1e293b;">Dear <strong>{customer_user.name}</strong>,</p>
                <p style="margin:0 0 20px;font-size:14px;color:#475569;line-height:1.6;">
                  We are pleased to inform you that your support ticket has been resolved by our agent.
                </p>
                <table style="width:100%;border-collapse:collapse;font-size:14px;margin-bottom:20px;">
                  <tr><td style="padding:8px 0;color:#64748b;width:140px;">Ticket ID</td><td style="padding:8px 0;color:#1e293b;font-weight:600;">{ticket.reference_number}</td></tr>
                  <tr><td style="padding:8px 0;color:#64748b;">Category</td><td style="padding:8px 0;color:#1e293b;">{ticket.category or 'N/A'}</td></tr>
                  <tr><td style="padding:8px 0;color:#64748b;">Issue Type</td><td style="padding:8px 0;color:#1e293b;">{ticket.subcategory or 'N/A'}</td></tr>
                  <tr><td style="padding:8px 0;color:#64748b;">Resolved By</td><td style="padding:8px 0;color:#1e293b;">{user.name}</td></tr>
                  <tr><td style="padding:8px 0;color:#64748b;">Resolved At</td><td style="padding:8px 0;color:#1e293b;">{ticket.resolved_at.strftime('%Y-%m-%d %H:%M UTC')}</td></tr>
                  {notes_row}
                </table>
                <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:14px 18px;">
                  <p style="margin:0;color:#15803d;font-size:14px;">If you feel your issue is not fully resolved, please start a new chat session and our team will assist you promptly.</p>
                </div>
              </div>
              <div style="background:#f8fafc;border-top:1px solid #e2e8f0;padding:14px 30px;text-align:center;">
                <p style="color:#94a3b8;font-size:12px;margin:0;">Customer Handling System — Telecom Support</p>
              </div>
            </div>
            """
            email_msg = Message(
                subject=f"Your Ticket {ticket.reference_number} Has Been Resolved",
                recipients=[customer_user.email],
                html=html_body,
            )
            mail.send(email_msg)
            print(f"[Resolve] Email sent to {customer_user.email}")
        except Exception as e:
            print(f"[Resolve] Email failed: {e}")

    # ── Notify customer via WhatsApp ──
    if customer_user and customer_user.phone_number:
        try:
            wa_msg = (
                f"*TeleBot — Ticket Resolved*\n\n"
                f"Hello {customer_user.name}!\n\n"
                f"Your support ticket has been resolved.\n\n"
                f"*Reference:* {ticket.reference_number}\n"
                f"*Category:* {ticket.category or 'N/A'}\n"
                f"*Resolved By:* {user.name}\n"
            )
            if resolution_notes:
                wa_msg += f"*Resolution:* {resolution_notes}\n"
            wa_msg += (
                f"\nIf you need further help, start a new chat session anytime.\n"
                f"Thank you for using our support service!"
            )
            result = send_whatsapp_message(customer_user.phone_number, wa_msg)
            if result["success"]:
                print(f"[Resolve] WhatsApp sent to {customer_user.phone_number}")
            else:
                print(f"[Resolve] WhatsApp failed: {result['error']}")
        except Exception as e:
            print(f"[Resolve] WhatsApp error: {e}")

    return jsonify({"ticket": ticket.to_dict()})


@app.route("/api/agent/tickets/<int:ticket_id>/parameter-change", methods=["GET"])
@jwt_required()
def agent_get_parameter_change(ticket_id):
    """Get the latest parameter change request for this ticket by the current agent."""
    user_id = int(get_jwt_identity())
    user = db.session.get(User, user_id)
    if not user or user.role != "human_agent":
        return jsonify({"error": "Unauthorized"}), 403

    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        return jsonify({"error": "Ticket not found"}), 404
    if ticket.assigned_to != user_id:
        return jsonify({"error": "Ticket not assigned to you"}), 403

    change = (ParameterChange.query
              .filter_by(ticket_id=ticket_id, agent_id=user_id)
              .order_by(ParameterChange.created_at.desc())
              .first())
    return jsonify({"change": change.to_dict() if change else None})


def _generate_cr_number():
    """Generate a unique CR number: CR-YYYYMMDD-XXXX."""
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    candidate = f"CR-{today}-{suffix}"
    while ChangeRequest.query.filter_by(cr_number=candidate).first():
        suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
        candidate = f"CR-{today}-{suffix}"
    return candidate


@app.route("/api/agent/tickets/<int:ticket_id>/parameter-change", methods=["POST"])
@jwt_required()
def agent_create_parameter_change(ticket_id):
    """
    Create a parameter-change request AND escalate the ticket to a manager.
    Also creates a ChangeRequest (CR) for the full ITIL Change Workflow.
    """
    user_id = int(get_jwt_identity())
    user = db.session.get(User, user_id)
    if not user or user.role != "human_agent":
        return jsonify({"error": "Unauthorized"}), 403

    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        return jsonify({"error": "Ticket not found"}), 404
    if ticket.assigned_to != user_id:
        return jsonify({"error": "Ticket not assigned to you"}), 403
    if ticket.status == "manager_escalated":
        return jsonify({"error": "Ticket is already escalated to a manager"}), 409

    data = request.json or {}
    proposed = (data.get("proposed_change")   or "").strip()
    impact   = (data.get("impact_assessment") or "").strip()
    rollback = (data.get("rollback_plan")     or "").strip()
    if not proposed:
        return jsonify({"error": "Proposed change is required"}), 400

    existing_pending = (ParameterChange.query
                        .filter_by(ticket_id=ticket_id, agent_id=user_id, status="pending")
                        .order_by(ParameterChange.created_at.desc())
                        .first())
    if existing_pending:
        return jsonify({"error": "A pending change request already exists"}), 409

    manager = _find_best_manager(ticket.priority)

    # Approval deadline = 30% of remaining SLA for the ticket
    from datetime import timezone as _tz
    _now = datetime.utcnow()
    _approval_dl = None
    if ticket.sla_deadline:
        _sla_dl = ticket.sla_deadline if ticket.sla_deadline.tzinfo is None else ticket.sla_deadline.replace(tzinfo=None)
        _remaining = (_sla_dl - _now).total_seconds()
        _window = max(_remaining * 0.3, 1800)  # at least 30 mins
        _approval_dl = _now + timedelta(seconds=_window)

    change = ParameterChange(
        ticket_id=ticket_id,
        agent_id=user_id,
        proposed_change=proposed,
        status="pending",
    )
    # Set approval_deadline if column exists
    try: change.approval_deadline = _approval_dl
    except: pass
    db.session.add(change)
    db.session.flush()

    cr_title = f"Parameter Change: {ticket.category or 'General'} — {ticket.reference_number}"
    cr = ChangeRequest(
        cr_number           = _generate_cr_number(),
        ticket_id           = ticket_id,
        parameter_change_id = change.id,
        raised_by           = user_id,
        title               = cr_title,
        description         = proposed,
        impact_assessment   = impact,
        rollback_plan       = rollback,
        status              = "created",
    )
    db.session.add(cr)

    now_utc = datetime.now(timezone.utc)
    ticket.status          = "manager_escalated"
    ticket.escalated_by    = user_id
    ticket.escalated_at    = now_utc
    ticket.escalation_note = proposed
    if manager:
        ticket.assigned_to = manager.id

    db.session.commit()

    return jsonify({
        "change":           change.to_dict(),
        "cr":               cr.to_dict(),
        "ticket":           ticket.to_dict(),
        "assigned_manager": {"id": manager.id, "name": manager.name, "email": manager.email} if manager else None,
        "approval_deadline": _approval_dl.isoformat() if _approval_dl else None,
    }), 201


@app.route("/api/agent/change-requests/<int:cr_id>/resubmit", methods=["PUT"])
@jwt_required()
def agent_resubmit_cr(cr_id):
    user_id = int(get_jwt_identity())
    user    = db.session.get(User, user_id)
    if not user or user.role != "human_agent":
        return jsonify({"error": "Unauthorized"}), 403

    cr = db.session.get(ChangeRequest, cr_id)
    if not cr:
        return jsonify({"error": "Change request not found"}), 404
    if cr.raised_by != user_id:
        return jsonify({"error": "Not your change request"}), 403
    if cr.status != "invalid":
        return jsonify({"error": "Only invalid CRs can be resubmitted"}), 409
    if cr.rejection_count >= 2:
        return jsonify({"error": "Maximum rejections reached. CR is permanently closed."}), 409

    data        = request.json or {}
    description = (data.get("description")      or "").strip()
    impact      = (data.get("impact_assessment") or "").strip()
    rollback    = (data.get("rollback_plan")     or "").strip()
    if not description:
        return jsonify({"error": "Description is required"}), 400

    cr.description       = description
    cr.impact_assessment = impact
    cr.rollback_plan     = rollback
    cr.status            = "created"
    cr.updated_at        = datetime.now(timezone.utc)
    db.session.commit()
    return jsonify({"cr": cr.to_dict()})


@app.route("/api/agent/change-requests/<int:cr_id>/implement", methods=["PUT"])
@jwt_required()
def agent_implement_cr(cr_id):
    user_id = int(get_jwt_identity())
    user    = db.session.get(User, user_id)
    if not user or user.role != "human_agent":
        return jsonify({"error": "Unauthorized"}), 403

    cr = db.session.get(ChangeRequest, cr_id)
    if not cr:
        return jsonify({"error": "Change request not found"}), 404
    if cr.raised_by != user_id:
        return jsonify({"error": "Not your change request"}), 403
    if cr.status not in ("approved", "implementing"):
        return jsonify({"error": "CR is not in approved state"}), 409

    data   = request.json or {}
    result = (data.get("result") or "").strip()
    notes  = (data.get("notes")  or "").strip()
    if result not in ("success", "failed"):
        return jsonify({"error": "result must be 'success' or 'failed'"}), 400

    now = datetime.now(timezone.utc)
    cr.status               = "implemented" if result == "success" else "failed"
    cr.implementation_notes = notes
    cr.implemented_by       = user_id
    cr.implemented_at       = now
    cr.updated_at           = now
    db.session.commit()
    return jsonify({"cr": cr.to_dict()})


@app.route("/api/agent/change-requests/<int:cr_id>/rollback", methods=["PUT"])
@jwt_required()
def agent_rollback_cr(cr_id):
    user_id = int(get_jwt_identity())
    user    = db.session.get(User, user_id)
    if not user or user.role != "human_agent":
        return jsonify({"error": "Unauthorized"}), 403

    cr = db.session.get(ChangeRequest, cr_id)
    if not cr or cr.raised_by != user_id:
        return jsonify({"error": "Not found or unauthorized"}), 404
    if cr.status != "failed":
        return jsonify({"error": "CR is not in failed state"}), 409

    data = request.json or {}
    cr.status         = "rolled_back"
    cr.rollback_notes = (data.get("notes") or "").strip()
    cr.rollback_at    = datetime.now(timezone.utc)
    cr.updated_at     = datetime.now(timezone.utc)
    db.session.commit()
    return jsonify({"cr": cr.to_dict()})


# ── Network diagnosis routes (delegated to network_diagnosis.py) ──────────────
# /api/agent/tickets/:id/diagnose
# /api/agent/tickets/:id/nearest-sites
# /api/agent/sites/:site_id/kpi-trends
# /api/agent/tickets/:id/root-cause
# /api/agent/tickets/:id/recommendation
# All registered via network_diagnosis.register_routes(app) at startup.


@app.route("/api/agent/customer360/<int:customer_user_id>", methods=["GET"])
@jwt_required()
def agent_customer360(customer_user_id):
    """Return 360-degree customer view: plan, billing history, past complaints, location, loyalty."""
    user_id = int(get_jwt_identity())
    agent = User.query.get(user_id)
    if not agent or agent.role != "human_agent":
        return jsonify({"error": "Unauthorized"}), 403

    customer = User.query.get(customer_user_id)
    if not customer:
        return jsonify({"error": "Customer not found"}), 404

    # Verify this agent has at least one ticket assigned for this customer
    has_access = Ticket.query.filter_by(
        assigned_to=user_id, user_id=customer_user_id
    ).first()
    if not has_access:
        return jsonify({"error": "No assigned ticket for this customer"}), 403

    # Past complaints / chat sessions
    sessions = ChatSession.query.filter_by(user_id=customer_user_id).order_by(ChatSession.created_at.desc()).limit(20).all()
    # Past tickets
    tickets = Ticket.query.filter_by(user_id=customer_user_id).order_by(Ticket.created_at.desc()).limit(10).all()
    # Feedbacks
    feedbacks = Feedback.query.filter_by(user_id=customer_user_id).all()
    avg_rating = round(sum(f.rating for f in feedbacks if f.rating > 0) / max(len([f for f in feedbacks if f.rating > 0]), 1), 2)

    # Loyalty score based on: account age (days), resolved complaints, avg rating
    from datetime import date
    account_age_days = (date.today() - customer.created_at.date()).days if customer.created_at else 0
    resolved_count = len([s for s in sessions if s.status == "resolved"])
    total_sessions = len(sessions)
    # Loyalty = 0-100 composite score
    age_score = min(account_age_days / 365 * 30, 30)  # max 30 points for up to 1 year
    resolution_score = min((resolved_count / max(total_sessions, 1)) * 40, 40)  # max 40 points
    rating_score = (avg_rating / 5) * 30  # max 30 points
    loyalty_score = round(age_score + resolution_score + rating_score, 1)

    # Location from most recent session with lat/long
    location_data = None
    for s in sessions:
        if s.latitude and s.longitude:
            location_data = {"latitude": s.latitude, "longitude": s.longitude}
            break

    # Category breakdown (billing history equivalent)
    category_count = {}
    for s in sessions:
        cat = s.sector_name or "Unknown"
        category_count[cat] = category_count.get(cat, 0) + 1

    # Infer plan from most common category
    plan_info = {
        "most_used_service": max(category_count, key=category_count.get) if category_count else "Unknown",
        "total_interactions": total_sessions,
        "account_since": customer.created_at.strftime("%B %Y") if customer.created_at else "Unknown",
    }

    return jsonify({
        "customer": {
            "id": customer.id,
            "name": customer.name,
            "email": customer.email,
            "phone": customer.phone_number,
            "employee_id": customer.employee_id,
            "created_at": customer.created_at.isoformat() if customer.created_at else None,
        },
        "plan_info": plan_info,
        "loyalty_score": loyalty_score,
        "avg_rating": avg_rating,
        "location": location_data,
        "category_breakdown": [{"category": k, "count": v} for k, v in category_count.items()],
        "recent_sessions": [
            {
                "id": s.id,
                "sector": s.sector_name,
                "subprocess": s.subprocess_name,
                "status": s.status,
                "created_at": s.created_at.isoformat() if s.created_at else None,
                "summary": s.summary,
            }
            for s in sessions[:10]
        ],
        "tickets": [t.to_dict() for t in tickets],
    })


@app.route("/api/agent/chat/<int:session_id>", methods=["GET"])
@jwt_required()
def agent_view_chat(session_id):
    """Allow human agent to view full chat history of a session."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or user.role != "human_agent":
        return jsonify({"error": "Unauthorized"}), 403
    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404
    # Verify the agent has an assigned ticket for this session
    assigned_ticket = Ticket.query.filter_by(
        chat_session_id=session_id, assigned_to=user_id
    ).first()
    if not assigned_ticket:
        return jsonify({"error": "Access denied: no assigned ticket for this session"}), 403
    return jsonify({
        "session": session.to_dict(),
        "messages": [m.to_dict() for m in session.messages],
        "customer": {
            "name": session.user.name if session.user else "",
            "email": session.user.email if session.user else "",
            "phone": session.user.phone_number if session.user else "",
        },
    })


@app.route("/api/agent/chat/<int:session_id>/message", methods=["POST"])
@jwt_required()
def agent_send_message(session_id):
    """Human agent sends a message into a customer chat session."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or user.role != "human_agent":
        return jsonify({"error": "Unauthorized"}), 403
    session = ChatSession.query.get(session_id)
    if not session:
        return jsonify({"error": "Session not found"}), 404
    assigned_ticket = Ticket.query.filter_by(
        chat_session_id=session_id, assigned_to=user_id
    ).first()
    if not assigned_ticket:
        return jsonify({"error": "Access denied: no assigned ticket for this session"}), 403

    data = request.json or {}
    content = data.get("content", "").strip()
    if not content:
        return jsonify({"error": "Message content is required"}), 400

    # Track first response time on the linked ticket
    if assigned_ticket and not assigned_ticket.first_response_at:
        assigned_ticket.first_response_at = datetime.now(timezone.utc)
    # Move ticket to in_progress on first agent reply
    if assigned_ticket and assigned_ticket.status == "pending":
        assigned_ticket.status = "in_progress"

    msg = ChatMessage(
        session_id=session_id,
        sender="agent",
        content=content,
    )
    db.session.add(msg)

    # Track first response time on the ticket
    if assigned_ticket and not assigned_ticket.first_response_at:
        assigned_ticket.first_response_at = datetime.now(timezone.utc)

    db.session.commit()

    # Push to WebSocket so the customer sees agent messages in real-time
    msg_dict = msg.to_dict()
    msg_dict["session_id"] = session_id
    socketio.emit("new_message", msg_dict, room=f"session_{session_id}")

    return jsonify({"message": msg_dict}), 201


# ── SLA Alert Helper ────────────────────────────────────────────────────────────

def send_sla_alert_email(recipients, subject, ticket, alert_type, time_left_hours):
    """Send SLA alert email to manager(s) or CTO."""
    is_breach = alert_type == "breach"
    status_color = "#dc2626" if is_breach else "#f59e0b"

    # Time-left string for email
    if is_breach:
        time_left_str = "SLA Breached — 0 time left"
    elif time_left_hours >= 1:
        time_left_str = f"{round(time_left_hours, 1)} hours remaining before SLA breach"
    else:
        time_left_str = f"{int(time_left_hours * 60)} minutes remaining before SLA breach"

    # Action-required callout
    if is_breach:
        action_msg = "URGENT: SLA has been breached. 0 time left. Immediate escalation required."
        action_bg = "#fef2f2"
        action_border = "#fecaca"
        action_color = "#dc2626"
    else:
        action_msg = f"Action Required: {time_left_str} for this ticket. Please take immediate action."
        action_bg = "#fef3c7"
        action_border = "#fde68a"
        action_color = "#b45309"

    html_body = f"""
    <div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #fff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.1);">
        <div style="background: {status_color}; padding: 20px 30px; text-align: center;">
            <h1 style="color: #fff; margin: 0; font-size: 18px;">{subject}</h1>
        </div>
        <div style="padding: 28px;">
            <table style="width:100%;border-collapse:collapse;font-size:14px;">
                <tr><td style="padding:8px 0;color:#64748b;width:160px;">Ticket ID</td><td style="padding:8px 0;color:#1e293b;font-weight:600;">{ticket.reference_number}</td></tr>
                <tr><td style="padding:8px 0;color:#64748b;">Category</td><td style="padding:8px 0;color:#1e293b;">{ticket.category}</td></tr>
                <tr><td style="padding:8px 0;color:#64748b;">Sub-Category</td><td style="padding:8px 0;color:#1e293b;">{ticket.subcategory or 'N/A'}</td></tr>
                <tr><td style="padding:8px 0;color:#64748b;">Issue</td><td style="padding:8px 0;color:#1e293b;">{ticket.description[:200]}</td></tr>
                <tr><td style="padding:8px 0;color:#64748b;">Priority</td><td style="padding:8px 0;color:#1e293b;font-weight:600;">{ticket.priority.upper()}</td></tr>
                <tr><td style="padding:8px 0;color:#64748b;">Current Status</td><td style="padding:8px 0;color:#1e293b;">{ticket.status}</td></tr>
                <tr><td style="padding:8px 0;color:#64748b;">SLA Allocated</td><td style="padding:8px 0;color:#1e293b;font-weight:600;">{ticket.sla_hours or 'N/A'} hours</td></tr>
                <tr><td style="padding:8px 0;color:#64748b;">Time Left</td><td style="padding:8px 0;color:{status_color};font-weight:700;">{time_left_str}</td></tr>
                <tr><td style="padding:8px 0;color:#64748b;">Assigned To</td><td style="padding:8px 0;color:#1e293b;">{ticket.assignee.name if ticket.assignee else 'Unassigned'}</td></tr>
                <tr><td style="padding:8px 0;color:#64748b;">SLA Deadline</td><td style="padding:8px 0;color:#1e293b;">{ticket.sla_deadline.strftime('%Y-%m-%d %H:%M UTC') if ticket.sla_deadline else 'N/A'}</td></tr>
            </table>
            <div style="background:{action_bg};border:1px solid {action_border};border-radius:8px;padding:14px 18px;margin-top:20px;">
                <p style="margin:0;color:{action_color};font-size:14px;font-weight:600;">{action_msg}</p>
            </div>
        </div>
        <div style="background:#f8fafc;border-top:1px solid #e2e8f0;padding:14px 30px;text-align:center;">
            <p style="color:#94a3b8;font-size:12px;margin:0;">Customer Handling — Automated SLA Alert System</p>
        </div>
    </div>
    """
    try:
        msg = Message(subject=subject, recipients=recipients, html=html_body)
        mail.send(msg)
        print(f"✅ SLA alert sent to {recipients}: {subject}")
    except Exception as e:
        print(f"⚠️ SLA alert email failed: {e}")


def run_sla_checks():
    """Background function: check open tickets AND change requests, send escalating SLA alerts."""
    import threading
    def _check():
        while True:
            try:
                with app.app_context():
                    now = datetime.now(timezone.utc)
                    open_tickets = Ticket.query.filter(
                        Ticket.status.in_(["pending", "in_progress"]),
                        Ticket.sla_deadline.isnot(None),
                    ).all()

                    # Get manager emails
                    managers = User.query.filter_by(role="manager").all()
                    manager_emails = [m.email for m in managers if m.email]
                    cto_users = User.query.filter_by(role="cto").all()
                    cto_emails = [c.email for c in cto_users if c.email]

                    for ticket in open_tickets:
                        dl = ticket.sla_deadline
                        if dl.tzinfo is None:
                            dl = dl.replace(tzinfo=timezone.utc)
                        cr = ticket.created_at
                        if cr.tzinfo is None:
                            cr = cr.replace(tzinfo=timezone.utc)

                        total_sla = (dl - cr).total_seconds()
                        elapsed = (now - cr).total_seconds()
                        time_left_hours = (dl - now).total_seconds() / 3600
                        fraction_elapsed = elapsed / max(total_sla, 1)

                        changed = False
                        # Human-friendly time remaining
                        if time_left_hours >= 1:
                            time_left_display = f"{round(time_left_hours, 1)}h"
                        elif time_left_hours > 0:
                            time_left_display = f"{int(time_left_hours * 60)}m"
                        else:
                            time_left_display = "0"

                        # Alert at 62.5%
                        if fraction_elapsed >= 0.625 and not ticket.alert_625_sent and manager_emails:
                            msg_text = f"{time_left_display} remaining before SLA breach — Ticket {ticket.reference_number} [{ticket.priority.upper()}]"
                            send_sla_alert_email(
                                manager_emails,
                                f"⚠️ {time_left_display} remaining before SLA breach — {ticket.reference_number}",
                                ticket, "625", time_left_hours
                            )
                            db.session.add(SlaAlert(
                                ticket_id=ticket.id, alert_level="625",
                                recipient_role="manager", message=msg_text,
                            ))
                            ticket.alert_625_sent = True
                            changed = True

                        # Alert at 75%
                        if fraction_elapsed >= 0.75 and not ticket.alert_750_sent and manager_emails:
                            msg_text = f"{time_left_display} remaining before SLA breach — Ticket {ticket.reference_number} [{ticket.priority.upper()}]"
                            send_sla_alert_email(
                                manager_emails,
                                f"🚨 {time_left_display} remaining before SLA breach — {ticket.reference_number}",
                                ticket, "750", time_left_hours
                            )
                            db.session.add(SlaAlert(
                                ticket_id=ticket.id, alert_level="750",
                                recipient_role="manager", message=msg_text,
                            ))
                            ticket.alert_750_sent = True
                            changed = True

                        # Alert at 87.5%
                        if fraction_elapsed >= 0.875 and not ticket.alert_875_sent and manager_emails:
                            msg_text = f"{time_left_display} remaining before SLA breach — Ticket {ticket.reference_number} [{ticket.priority.upper()}]"
                            send_sla_alert_email(
                                manager_emails,
                                f"🔴 {time_left_display} remaining before SLA breach — {ticket.reference_number}",
                                ticket, "875", time_left_hours
                            )
                            db.session.add(SlaAlert(
                                ticket_id=ticket.id, alert_level="875",
                                recipient_role="manager", message=msg_text,
                            ))
                            ticket.alert_875_sent = True
                            changed = True

                        # SLA Breach – send to CTO
                        if now > dl and not ticket.breach_alert_sent:
                            ticket.sla_breached = True
                            msg_text = f"SLA Breached — 0 time left — Ticket {ticket.reference_number} [{ticket.priority.upper()}]"
                            recipients = cto_emails if cto_emails else manager_emails
                            send_sla_alert_email(
                                recipients,
                                f"🚨 SLA Breached — {ticket.reference_number}",
                                ticket, "breach", 0
                            )
                            db.session.add(SlaAlert(
                                ticket_id=ticket.id, alert_level="breach",
                                recipient_role="cto", message=msg_text,
                            ))
                            ticket.breach_alert_sent = True
                            changed = True

                        if changed:
                            db.session.commit()

                    # ── CR SLA Checks — escalating alerts for Change Requests ─────
                    from models import CrSlaAlert
                    TERMINAL_CR = {"closed", "cto_rejected", "rejected", "auto_rejected", "failed"}
                    open_crs = ChangeRequest.query.filter(
                        ChangeRequest.cr_sla_deadline.isnot(None),
                        ~ChangeRequest.status.in_(TERMINAL_CR),
                    ).all()

                    for cr in open_crs:
                        dl = cr.cr_sla_deadline
                        if dl.tzinfo is None:
                            dl = dl.replace(tzinfo=timezone.utc)
                        cr_created = cr.created_at
                        if cr_created and cr_created.tzinfo is None:
                            cr_created = cr_created.replace(tzinfo=timezone.utc)
                        if not cr_created:
                            continue

                        total_sla = (dl - cr_created).total_seconds()
                        if total_sla <= 0:
                            continue
                        elapsed = (now - cr_created).total_seconds()
                        time_left_hours = (dl - now).total_seconds() / 3600
                        fraction_elapsed = elapsed / max(total_sla, 1)

                        if time_left_hours >= 1:
                            tl_display = f"{round(time_left_hours, 1)}h"
                        elif time_left_hours > 0:
                            tl_display = f"{int(time_left_hours * 60)}m"
                        else:
                            tl_display = "0"

                        cr_changed = False
                        type_label = (cr.change_type or "standard").upper()
                        src = "Customer CR" if (cr.ticket_id and not cr.network_issue_id) else "AI CR"

                        # 75% elapsed — alert to manager
                        if fraction_elapsed >= 0.75 and not cr.cr_alert_75_sent:
                            msg = f"{tl_display} remaining — {src} {cr.cr_number} [{type_label}] approaching SLA deadline"
                            db.session.add(CrSlaAlert(
                                cr_id=cr.id, alert_level="75",
                                recipient_role="manager", message=msg,
                            ))
                            cr.cr_alert_75_sent = True
                            cr_changed = True

                        # 90% elapsed — alert to CTO
                        if fraction_elapsed >= 0.90 and not cr.cr_alert_90_sent:
                            msg = f"{tl_display} remaining — {src} {cr.cr_number} [{type_label}] critical SLA warning"
                            db.session.add(CrSlaAlert(
                                cr_id=cr.id, alert_level="90",
                                recipient_role="cto", message=msg,
                            ))
                            cr.cr_alert_90_sent = True
                            cr_changed = True

                        # SLA Breach — alert to CTO
                        if now > dl and not cr.cr_breach_alert_sent:
                            cr.cr_sla_breached = True
                            msg = f"CR SLA BREACHED — {src} {cr.cr_number} [{type_label}] — {cr.title[:80]}"
                            db.session.add(CrSlaAlert(
                                cr_id=cr.id, alert_level="breach",
                                recipient_role="cto", message=msg,
                            ))
                            cr.cr_breach_alert_sent = True
                            cr_changed = True

                        if cr_changed:
                            db.session.commit()

            except Exception as e:
                print(f"⚠️ SLA check error: {e}")
            time.sleep(300)  # Check every 5 minutes

    t = threading.Thread(target=_check, daemon=True)
    t.start()


# ═══════════════════════════════════════════════════════════════════════════════
# INIT DB + SEED ADMIN
# ═══════════════════════════════════════════════════════════════════════════════

def _ensure_kpi_merged_view():
    """Create/refresh the kpi_data_merged VIEW.

    This view presents every (site_id, kpi_name, date) ONCE:
      - If site-level rows exist, they pass through as-is.
      - For sites with NO site-level data for a given (kpi, date), the cell-level
        rows are averaged across all cells of that site and surfaced with
        data_level='site' so existing analytics queries pick them up transparently.
    All read-heavy dashboards (network_analytics, CTO dashboards) should read
    from this view instead of the raw kpi_data table.
    """
    from sqlalchemy import text as _t
    ddl = """
        CREATE OR REPLACE VIEW kpi_data_merged AS
        SELECT id, site_id, site_abs_id, kpi_name, date, hour, value,
               data_level, cell_id, cell_site_id
        FROM kpi_data
        WHERE data_level = 'site'
        UNION ALL
        SELECT
            MIN(k.id)            AS id,
            k.site_id,
            MAX(k.site_abs_id)   AS site_abs_id,
            k.kpi_name,
            k.date,
            0                    AS hour,
            AVG(k.value)         AS value,
            'site'               AS data_level,
            NULL::varchar        AS cell_id,
            NULL::varchar        AS cell_site_id
        FROM kpi_data k
        WHERE k.data_level = 'cell' AND k.value IS NOT NULL
          AND NOT EXISTS (
              SELECT 1 FROM kpi_data s
              WHERE s.site_id   = k.site_id
                AND s.kpi_name  = k.kpi_name
                AND s.date      = k.date
                AND s.data_level = 'site'
          )
        GROUP BY k.site_id, k.kpi_name, k.date
    """
    try:
        with db.engine.connect() as conn:
            conn.execute(_t(ddl))
            conn.commit()
        print(">>> kpi_data_merged view ready (site-level preferred, cell-level fallback)")
    except Exception as e:
        print(f">>> kpi_data_merged view create failed (non-fatal): {e}")


with app.app_context():
    db.create_all()

    # Migrate: add new columns to kpi_data if they don't exist
    from sqlalchemy import inspect as sa_inspect, text as sa_text
    insp = sa_inspect(db.engine)
    if insp.has_table("kpi_data"):
        existing_cols = [c["name"] for c in insp.get_columns("kpi_data")]
        with db.engine.connect() as conn:
            if "data_level" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE kpi_data ADD COLUMN data_level VARCHAR(10) NOT NULL DEFAULT 'site'"))
                conn.commit()
                print(">>> Added data_level column to kpi_data")
            if "cell_id" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE kpi_data ADD COLUMN cell_id VARCHAR(100)"))
                conn.commit()
                print(">>> Added cell_id column to kpi_data")
            if "cell_site_id" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE kpi_data ADD COLUMN cell_site_id VARCHAR(100)"))
                conn.commit()
                print(">>> Added cell_site_id column to kpi_data")

    if insp.has_table("telecom_sites"):
        existing_cols = [c["name"] for c in insp.get_columns("telecom_sites")]
        with db.engine.connect() as conn:
            if "site_status" not in existing_cols:
                conn.execute(text("ALTER TABLE telecom_sites ADD COLUMN site_status VARCHAR(20) DEFAULT 'on_air'"))
                conn.commit()
                print(">>> Added site_status column to telecom_sites")
            if "alarms" not in existing_cols:
                conn.execute(text("ALTER TABLE telecom_sites ADD COLUMN alarms TEXT DEFAULT ''"))
                conn.commit()
                print(">>> Added alarms column to telecom_sites")
            if "solution" not in existing_cols:
                conn.execute(text("ALTER TABLE telecom_sites ADD COLUMN solution TEXT DEFAULT ''"))
                conn.commit()
                print(">>> Added solution column to telecom_sites")
            if "city" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE telecom_sites ADD COLUMN city VARCHAR(100)"))
                conn.commit()
                print(">>> Added city column to telecom_sites")
            if "state" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE telecom_sites ADD COLUMN state VARCHAR(100)"))
                conn.commit()
                print(">>> Added state column to telecom_sites")
            if "country" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE telecom_sites ADD COLUMN country VARCHAR(100) DEFAULT ''"))
                conn.commit()
                print(">>> Added country column to telecom_sites")
            if "technology" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE telecom_sites ADD COLUMN technology VARCHAR(50)"))
                conn.commit()
                print(">>> Added technology column to telecom_sites")
            # New columns for updated site data model
            for col, typ in [("province", "VARCHAR(100) DEFAULT ''"), ("commune", "VARCHAR(100) DEFAULT ''"),
                             ("site_abs_id", "VARCHAR(100)"), ("vendor_name", "VARCHAR(100)"),
                             ("extra_params", "JSONB")]:
                if col not in existing_cols:
                    conn.execute(sa_text(f"ALTER TABLE telecom_sites ADD COLUMN {col} {typ}"))
                    conn.commit()
                    print(f">>> Added {col} column to telecom_sites")

    # Migrate: add site_abs_id to kpi_data if missing
    if insp.has_table("kpi_data"):
        existing_cols = [c["name"] for c in insp.get_columns("kpi_data")]
        with db.engine.connect() as conn:
            if "site_abs_id" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE kpi_data ADD COLUMN site_abs_id VARCHAR(100)"))
                conn.commit()
                print(">>> Added site_abs_id column to kpi_data")

    # Auto-populate geo data for sites missing city/state/country
    try:
        _auto_populate_geo()
    except Exception as e:
        print(f">>> Startup geo population skipped: {e}")

    # Migrate: add new columns to users if they don't exist
    if insp.has_table("users"):
        existing_cols = [c["name"] for c in insp.get_columns("users")]
        with db.engine.connect() as conn:
            if "expertise" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE users ADD COLUMN expertise VARCHAR(100)"))
                conn.commit()
                print(">>> Added expertise column to users")
            if "specialization" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE users ADD COLUMN specialization VARCHAR(200)"))
                conn.commit()
                print(">>> Added specialization column to users")
            if "bandwidth_capacity" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE users ADD COLUMN bandwidth_capacity INTEGER NOT NULL DEFAULT 10"))
                conn.commit()
                print(">>> Added bandwidth_capacity column to users")

    # Migrate: add network_issue_id to parameter_changes if missing
    # ── Comprehensive migration: parameter_changes ──────────────────────────
    if insp.has_table("parameter_changes"):
        existing_cols = [c["name"] for c in insp.get_columns("parameter_changes")]
        _pc_adds = [
            ("network_issue_id", "INTEGER"),
            ("approval_deadline", "TIMESTAMP"),
        ]
        with db.engine.connect() as conn:
            for col, typ in _pc_adds:
                if col not in existing_cols:
                    conn.execute(sa_text(f"ALTER TABLE parameter_changes ADD COLUMN {col} {typ}"))
                    conn.commit()
                    print(f">>> Added {col} column to parameter_changes")
            # Make ticket_id nullable (network issues use network_issue_id instead)
            try:
                conn.execute(sa_text("ALTER TABLE parameter_changes ALTER COLUMN ticket_id DROP NOT NULL"))
                conn.commit()
            except Exception:
                conn.rollback()
            # Make agent_id nullable just in case
            try:
                conn.execute(sa_text("ALTER TABLE parameter_changes ALTER COLUMN agent_id DROP NOT NULL"))
                conn.commit()
            except Exception:
                conn.rollback()

    # ── Comprehensive migration: change_requests ──────────────────────────
    if insp.has_table("change_requests"):
        existing_cols = [c["name"] for c in insp.get_columns("change_requests")]
        _cr_adds = [
            ("parameter_change_id", "INTEGER"),
            ("change_type", "VARCHAR(20)"),
            ("rejection_count", "INTEGER DEFAULT 0"),
            ("validation_remark", "TEXT DEFAULT ''"),
            ("validated_by", "INTEGER"),
            ("validated_at", "TIMESTAMP"),
            ("classification_note", "TEXT DEFAULT ''"),
            ("classified_by", "INTEGER"),
            ("classified_at", "TIMESTAMP"),
            ("approval_remark", "TEXT DEFAULT ''"),
            ("approved_by", "INTEGER"),
            ("approved_at", "TIMESTAMP"),
            ("implementation_notes", "TEXT DEFAULT ''"),
            ("implemented_by", "INTEGER"),
            ("implemented_at", "TIMESTAMP"),
            ("rollback_notes", "TEXT DEFAULT ''"),
            ("rollback_at", "TIMESTAMP"),
            ("closure_notes", "TEXT DEFAULT ''"),
            ("closed_at", "TIMESTAMP"),
            ("updated_at", "TIMESTAMP"),
        ]
        with db.engine.connect() as conn:
            for col, typ in _cr_adds:
                if col not in existing_cols:
                    conn.execute(sa_text(f"ALTER TABLE change_requests ADD COLUMN {col} {typ}"))
                    conn.commit()
                    print(f">>> Added {col} column to change_requests")
            # Make nullable columns that may have been created as NOT NULL
            for col in ["ticket_id", "raised_by", "parameter_change_id"]:
                try:
                    conn.execute(sa_text(f"ALTER TABLE change_requests ALTER COLUMN {col} DROP NOT NULL"))
                    conn.commit()
                except Exception:
                    conn.rollback()
            # title/description should stay NOT NULL but ensure they exist
            for col in ["title", "description", "impact_assessment", "rollback_plan", "status"]:
                if col not in existing_cols:
                    default = "''" if col in ("impact_assessment", "rollback_plan") else ("'created'" if col == "status" else None)
                    typ = "TEXT" if col in ("description", "impact_assessment", "rollback_plan") else "VARCHAR(200)" if col == "title" else "VARCHAR(30)"
                    sql = f"ALTER TABLE change_requests ADD COLUMN {col} {typ}"
                    if default:
                        sql += f" DEFAULT {default}"
                    conn.execute(sa_text(sql))
                    conn.commit()
                    print(f">>> Added {col} column to change_requests")

    # ── Migrate: add new columns to chat_sessions if missing ────────────────
    if insp.has_table("chat_sessions"):
        existing_cols = [c["name"] for c in insp.get_columns("chat_sessions")]
        with db.engine.connect() as conn:
            if "diagnosis_ran" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE chat_sessions ADD COLUMN diagnosis_ran BOOLEAN NOT NULL DEFAULT FALSE"))
                conn.commit()
                print(">>> Added diagnosis_ran column to chat_sessions")
            if "current_step" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE chat_sessions ADD COLUMN current_step VARCHAR(50) DEFAULT 'greeting'"))
                conn.commit()
                print(">>> Added current_step column to chat_sessions")
            if "last_message_at" not in existing_cols:
                conn.execute(sa_text("ALTER TABLE chat_sessions ADD COLUMN last_message_at TIMESTAMP"))
                conn.commit()
                print(">>> Added last_message_at column to chat_sessions")

    # Seed default admin if none exists
    if not User.query.filter_by(role="admin").first():
        admin = User(name="Admin", email="didardeep.12@gmail.com", role="admin", employee_id="ADM00001")
        admin.set_password("admin123")
        db.session.add(admin)
        db.session.commit()
        print(">>> Default admin created: didardeep.12@gmail.com / admin123")

    # Backfill employee_ids for existing non-customer users
    users_without_emp_id = User.query.filter(
        User.role != "customer",
        User.employee_id.is_(None)
    ).all()
    for u in users_without_emp_id:
        u.employee_id = generate_employee_id(u.role)
    if users_without_emp_id:
        db.session.commit()
        print(f">>> Backfilled employee_ids for {len(users_without_emp_id)} users")

    # Seed / update SLA defaults
    for key, info in SLA_DEFAULTS.items():
        existing = SystemSetting.query.filter_by(key=key).first()
        if existing:
            existing.value = info["value"]
            existing.description = info["description"]
        else:
            db.session.add(SystemSetting(key=key, value=info["value"], category="sla", description=info["description"]))
    db.session.commit()


# ─── Register Network Analytics Blueprint ─────────────────────────────────────
from network_analytics import network_bp, clear_analytics_cache
app.register_blueprint(network_bp)

# ─── Register Network AI Blueprint ───────────────────────────────────────────
from network_ai import network_ai_bp, ensure_db_optimizations
app.register_blueprint(network_ai_bp)

# ─── Startup schema warm-up for network AI ────────────────────────────────────
# Pre-populate _schema_cache (KPI ranges, date freshness, column discovery)
# so the first AI query is as fast as subsequent ones.
def _warm_network_ai_schema():
    """Run schema discovery once at startup so first AI query has full context."""
    import threading, time as _t
    def _do_warm():
        _t.sleep(5)  # let DB settle
        try:
            with app.app_context():
                from network_ai import _schema_cache, _sql
                if _schema_cache.get("populated"):
                    print("[AI-SCHEMA] Already populated — skipping warm-up")
                    return
                # Populate KPI ranges
                try:
                    _kpi_r = _sql("""
                        SELECT kpi_name,
                               ROUND(MIN(value)::numeric,2) AS min_val,
                               ROUND(AVG(value)::numeric,2) AS avg_val,
                               ROUND(MAX(value)::numeric,2) AS max_val,
                               ROUND(PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY value)::numeric,2) AS p25,
                               ROUND(PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY value)::numeric,2) AS p75,
                               COUNT(DISTINCT site_id) AS site_count
                        FROM kpi_data WHERE data_level='site' AND value IS NOT NULL
                        GROUP BY kpi_name ORDER BY kpi_name
                    """)
                    _schema_cache["kpi_ranges"] = {r["kpi_name"]: r for r in _kpi_r}
                    print(f"[AI-SCHEMA] KPI ranges warm-up: {len(_kpi_r)} KPIs loaded")
                except Exception as e:
                    print(f"[AI-SCHEMA] KPI ranges warm-up failed: {e}")
                # Populate max date
                try:
                    _d = _sql("SELECT MAX(date)::text AS max_date FROM kpi_data WHERE data_level='site'")
                    if _d and _d[0].get("max_date"):
                        _schema_cache["kpi_max_date"] = _d[0]["max_date"]
                        _schema_cache["kpi_max_date_recent"] = _d[0]["max_date"]
                        print(f"[AI-SCHEMA] Data freshness: latest date = {_d[0]['max_date']}")
                except Exception as e:
                    print(f"[AI-SCHEMA] Date warm-up failed: {e}")
                _schema_cache["populated"] = True
                print("[AI-SCHEMA] Warm-up complete")
        except Exception as e:
            print(f"[AI-SCHEMA] Warm-up error: {e}")
    threading.Thread(target=_do_warm, daemon=True, name="ai-schema-warm").start()

_warm_network_ai_schema()

# ─── Database optimizations (indexes, materialized views) ────────────────────
with app.app_context():
    try:
        ensure_db_optimizations()
    except Exception as _db_opt_err:
        print(f"[WARN] DB optimization skipped: {_db_opt_err}")

# ─── ML Pipeline endpoints ───────────────────────────────────────────────────
from ml_pipeline import run_ml_pipeline, run_ml_pipeline_async, get_pipeline_status

@app.route("/api/ml/run-pipeline", methods=["POST"])
@jwt_required()
def api_run_ml_pipeline():
    """Trigger the ML categorization pipeline (background).
    Only managers and CTOs can trigger this."""
    uid = get_jwt_identity()
    user = db.session.get(User, int(uid))
    if not user or user.role not in ('manager', 'cto'):
        return jsonify({"error": "Only managers/CTOs can trigger the ML pipeline"}), 403
    result = run_ml_pipeline_async(app)
    return jsonify(result)

@app.route("/api/ml/run-pipeline-sync", methods=["POST"])
@jwt_required()
def api_run_ml_pipeline_sync():
    """Trigger the ML pipeline synchronously (blocks until done).
    Use for smaller datasets or when you need the result immediately."""
    uid = get_jwt_identity()
    user = db.session.get(User, int(uid))
    if not user or user.role not in ('manager', 'cto'):
        return jsonify({"error": "Only managers/CTOs can trigger the ML pipeline"}), 403
    result = run_ml_pipeline(app)
    return jsonify(result)

@app.route("/api/ml/status", methods=["GET"])
@jwt_required()
def api_ml_pipeline_status():
    """Get current ML pipeline status."""
    return jsonify(get_pipeline_status())

# Create site_kpi_summary table
with app.app_context():
    try:
        from models import SiteKpiSummary
        SiteKpiSummary.__table__.create(db.engine, checkfirst=True)
        print("[ML] site_kpi_summary table ready")
    except Exception as _sks_err:
        print(f"[WARN] site_kpi_summary table creation skipped: {_sks_err}")

# Auto-run ML pipeline after startup if summary table is empty.
# Uses a delayed thread so the app is fully initialized before running.
def _auto_ml_on_startup():
    """Check site_kpi_summary after a short delay and run ML if empty."""
    import time as _t
    _t.sleep(3)  # wait for app to fully start
    try:
        with app.app_context():
            _count = db.session.execute(text("SELECT COUNT(*) FROM site_kpi_summary")).scalar()
            if _count == 0:
                print("[ML] site_kpi_summary is empty — auto-running ML pipeline in background...")
                run_ml_pipeline_async(app)
            else:
                print(f"[ML] site_kpi_summary has {_count} rows — ML pipeline not needed")
    except Exception as e:
        print(f"[ML] Auto-ML check skipped: {e}")

threading.Thread(target=_auto_ml_on_startup, daemon=True, name="ml-auto-check").start()

# ─── Register Network Issues Blueprint ─────────────────────────────────────
from network_issues import network_issues_bp, NetworkIssueTicket, OverutilizedTicket, schedule_daily_job, run_overutilized_job
app.register_blueprint(network_issues_bp)

# Create network_issue_tickets + overutilized_tickets tables if not exist
with app.app_context():
    NetworkIssueTicket.__table__.create(db.engine, checkfirst=True)
    OverutilizedTicket.__table__.create(db.engine, checkfirst=True)
    # Migrate: add new columns to overutilized_tickets if missing
    try:
        _ou_insp = sa_inspect(db.engine)
        if _ou_insp.has_table("overutilized_tickets"):
            _ou_cols = [c["name"] for c in _ou_insp.get_columns("overutilized_tickets")]
            with db.engine.connect() as _ouc:
                for col, typ in [("sites_list", "TEXT DEFAULT ''"), ("site_count", "INTEGER DEFAULT 1"), ("site_status", "TEXT DEFAULT ''")]:
                    if col not in _ou_cols:
                        _ouc.execute(sa_text(f"ALTER TABLE overutilized_tickets ADD COLUMN {col} {typ}"))
                        _ouc.commit()
                        print(f">>> Added {col} column to overutilized_tickets")
    except Exception as _oe:
        print(f">>> Overutilized migration: {_oe}")

# ── One-time cleanup: purge legacy GUR_LTE_* seeded rows + nuke ALL existing
# AI tickets (operator wants a clean slate after migrating between machines).
# Guarded by a marker row in a tiny `_one_time_migrations` table — runs once
# per database, never again.
with app.app_context():
    try:
        with db.engine.connect() as _c:
            # Marker table
            _c.execute(sa_text(
                "CREATE TABLE IF NOT EXISTS _one_time_migrations ("
                "  id SERIAL PRIMARY KEY, "
                "  name VARCHAR(120) UNIQUE NOT NULL, "
                "  applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)"))
            _c.commit()

            MIGRATION_NAME = "purge_seed_and_existing_ai_tickets_2026_04_15"
            already = _c.execute(
                sa_text("SELECT 1 FROM _one_time_migrations WHERE name = :n"),
                {"n": MIGRATION_NAME}
            ).fetchone()

            if not already:
                # 1. Purge legacy seed rows
                _c.execute(sa_text(
                    "DELETE FROM overutilized_tickets WHERE site_id LIKE 'GUR\\_LTE\\_%' ESCAPE '\\' "
                    "OR sites_list LIKE '%GUR\\_LTE\\_%' ESCAPE '\\'"))
                _c.execute(sa_text(
                    "DELETE FROM network_issue_tickets WHERE site_id LIKE 'GUR\\_LTE\\_%' ESCAPE '\\'"))
                _c.execute(sa_text(
                    "DELETE FROM telecom_sites WHERE site_id LIKE 'GUR\\_LTE\\_%' ESCAPE '\\'"))

                # 2. Wipe ALL existing AI tickets so the next 08:00 daily job
                #    rebuilds them fresh from the operator's uploaded data.
                _c.execute(sa_text("DELETE FROM overutilized_tickets"))
                _c.execute(sa_text("DELETE FROM network_issue_tickets"))

                # 3. Mark migration as applied (won't run again)
                _c.execute(
                    sa_text("INSERT INTO _one_time_migrations (name) VALUES (:n)"),
                    {"n": MIGRATION_NAME}
                )
                _c.commit()
                print(">>> One-time migration applied: purged GUR_LTE seed + cleared all AI tickets")
            else:
                print(">>> One-time AI-ticket purge already applied; skipping")
    except Exception as _e:
        print(f">>> One-time migration skipped: {_e}")

# Schedule daily 07:30/08:00 AM IST jobs for worst cell + overutilized detection
schedule_daily_job(app)

# ── One-time seed: AI ticket for Gunjan Kaur demo ────────────────────────
try:
    from seed_gunjan_ticket import seed_gunjan_ticket
    seed_gunjan_ticket(app)
except Exception as e:
    print(f"[SEED] Gunjan ticket seed skipped: {e}")

# ─── Register Change Workflow Blueprint ───────────────────────────────────
try:
    from change_workflow import change_workflow_bp
    app.register_blueprint(change_workflow_bp)
except ImportError:
    print("WARNING: change_workflow.py not found — CR endpoints unavailable")

# Create new tables if not exists
with app.app_context():
    from models import CRAuditTrail
    CRAuditTrail.__table__.create(db.engine, checkfirst=True)
    from models import CrSlaAlert
    CrSlaAlert.__table__.create(db.engine, checkfirst=True)
    # Add new columns to change_requests if they don't exist
    _engine = db.engine
    _insp = db.inspect(_engine)
    _existing_cols = {c['name'] for c in _insp.get_columns('change_requests')}
    _new_cols = {
        'network_issue_id': 'INTEGER',
        'justification': 'TEXT',
        'category': 'VARCHAR(200)',
        'subcategory': 'VARCHAR(200)',
        'telecom_domain_primary': 'VARCHAR(50)',
        'telecom_domain_secondary': 'VARCHAR(200)',
        'zone': 'VARCHAR(100)',
        'location': 'VARCHAR(200)',
        'nearest_site_id': 'VARCHAR(50)',
        'customer_type': 'VARCHAR(20)',
        'rf_bandwidth_current': 'FLOAT', 'rf_bandwidth_proposed': 'FLOAT',
        'rf_antenna_gain_current': 'FLOAT', 'rf_antenna_gain_proposed': 'FLOAT',
        'rf_eirp_current': 'FLOAT', 'rf_eirp_proposed': 'FLOAT',
        'rf_antenna_height_current': 'FLOAT', 'rf_antenna_height_proposed': 'FLOAT',
        'rf_etilt_current': 'FLOAT', 'rf_etilt_proposed': 'FLOAT',
        'rf_crs_gain_current': 'FLOAT', 'rf_crs_gain_proposed': 'FLOAT',
        'pdf_filename': 'VARCHAR(300)', 'pdf_path': 'VARCHAR(500)',
        'cr_sla_hours': 'FLOAT', 'cr_sla_deadline': 'TIMESTAMP',
        'cr_sla_breached': 'BOOLEAN DEFAULT FALSE',
        'cr_breach_alert_sent': 'BOOLEAN DEFAULT FALSE',
        'cr_alert_75_sent': 'BOOLEAN DEFAULT FALSE',
        'cr_alert_90_sent': 'BOOLEAN DEFAULT FALSE',
        'assigned_manager_id': 'INTEGER',
        'cto_approval_required': 'BOOLEAN DEFAULT FALSE',
        'cto_approved_by': 'INTEGER', 'cto_approved_at': 'TIMESTAMP',
        'cto_status': 'VARCHAR(20)', 'cto_remark': 'TEXT',
        'manager_proposed_changes': 'TEXT',
    }
    with _engine.connect() as _conn:
        for col_name, col_type in _new_cols.items():
            if col_name not in _existing_cols:
                try:
                    _conn.execute(text(f'ALTER TABLE change_requests ADD COLUMN {col_name} {col_type}'))
                    _conn.commit()
                except Exception:
                    _conn.rollback()


if __name__ == "__main__":
    run_sla_checks()
    _warm_cto_cache()
    socketio.run(app, debug=True, host="0.0.0.0", port=5500, use_reloader=False)